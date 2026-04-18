"""
╔══════════════════════════════════════════════════════════════════╗
║  SAVDOAI v25.3.2 — REESTR (RO'YXAT) PARSER                      ║
║                                                                  ║
║  Kunlik yetkazib berish reestri:                                ║
║  №, Sana, Do'kon, Telefon, TP, Balans, Summa                   ║
║                                                                  ║
║  Nakladnoydan farqi:                                             ║
║  - Nakladnoy: har klient uchun TOVAR RO'YXATI                  ║
║  - Reestr: har klient uchun JAMI SUMMA (bitta qator)            ║
╚══════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io
import logging
from collections import defaultdict
log = logging.getLogger(__name__)


def reestr_ekanligini_tekshir(data: bytes) -> bool:
    """Reestr formatini aniqlash — 'Реестр' yoki sarlavhalar bo'yicha."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        for sn in wb.sheetnames:
            if "Реестр" in sn or "реестр" in sn or "reestr" in sn.lower():
                wb.close()
                return True
            ws = wb[sn]
            n = 0
            for row in ws.iter_rows(values_only=True):
                n += 1
                if n > 5: break
                row_text = " ".join(str(v) for v in row if v)
                if "Торгов" in row_text and "Баланс" in row_text:
                    wb.close()
                    return True
        wb.close()
    except Exception:
        pass
    return False


def reestr_tahlil(data: bytes) -> dict:
    """
    Reestr faylini tahlil qilish.
    
    Qaytaradi:
    {
        "klientlar": [{ism, telefon, tp, balans, summa, sana}, ...],
        "jami_summa": ...,
        "jami_soni": ...,
        "tp_lar": {tp: {soni, jami}},
        "qarzli": [{ism, balans}, ...],
        "sanalar": [...],
    }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"xato": "openpyxl kerak"}
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        return {"xato": f"Excel: {e}"}

    klientlar = []
    tp_map = defaultdict(lambda: {"soni": 0, "jami": 0.0})
    qarzli = []
    sanalar = set()
    total_rows = 0
    jami_summa = 0.0

    for sn in wb.sheetnames:
        ws = wb[sn]
        header_found = False

        for row in ws.iter_rows(values_only=True):
            total_rows += 1
            if total_rows > 500_000: break

            cells = list(row) + [None] * 10
            a, b, c, d, e, f, g, h = cells[:8]

            # Sarlavha qatorini o'tkazish
            if not header_found:
                row_text = " ".join(str(v) for v in cells[:9] if v)
                if "Торгов" in row_text or "Дата" in row_text:
                    header_found = True
                continue

            # Total qator
            if c and "Total" in str(c):
                continue

            # Data qator: A=№, B=sana, C=do'kon, D=adres, E=tel, F=tp, G=balans, H=summa
            if a and str(a).strip().isdigit() and c and h:
                try:
                    summa = float(h)
                    ism = str(c).strip()[:60]
                    sana_s = str(b).strip() if b else ""
                    tp = str(f).strip() if f else ""
                    telefon = str(e).strip() if e else ""
                    
                    # Balans
                    balans = 0
                    balans_str = ""
                    if g is not None:
                        try:
                            balans = float(str(g).replace(",", "").replace(" ", ""))
                            balans_str = str(g)
                        except (ValueError, TypeError):
                            balans_str = str(g)

                    klientlar.append({
                        "ism": ism,
                        "telefon": telefon,
                        "tp": tp,
                        "balans": balans,
                        "balans_str": balans_str,
                        "summa": summa,
                        "sana": sana_s,
                    })

                    jami_summa += summa
                    if sana_s: sanalar.add(sana_s)
                    if tp:
                        tp_map[tp]["soni"] += 1
                        tp_map[tp]["jami"] += summa
                    if balans < 0:
                        qarzli.append({"ism": ism[:30], "balans": balans_str, "summa": summa})

                except (ValueError, TypeError):
                    pass

    wb.close()

    return {
        "klientlar": klientlar,
        "jami_summa": jami_summa,
        "jami_soni": len(klientlar),
        "tp_lar": dict(tp_map),
        "qarzli": qarzli,
        "sanalar": sorted(sanalar),
        "qator_soni": total_rows,
    }


def reestr_xulosa_matn(data: dict, fayl_nomi: str = "") -> str:
    """Bot uchun reestr xulosa matni."""
    if "xato" in data: return f"❌ {data['xato']}"
    if data["jami_soni"] == 0: return "📋 Reestr ma'lumoti topilmadi."

    m = "📊 *REESTR TAHLILI*\n━━━━━━━━━━━━━━━━━━━━━\n"
    if fayl_nomi:
        m += f"📁 {fayl_nomi[:40]}\n"
    if data["sanalar"]:
        m += f"📅 {', '.join(data['sanalar'][:3])}\n"

    m += (
        f"\n👥 *{data['jami_soni']}* klient\n"
        f"💰 *JAMI: {data['jami_summa']:,.0f}* so'm\n"
        f"📊 O'rtacha: {data['jami_summa']/max(data['jami_soni'],1):,.0f} so'm/klient\n"
    )

    # TP bo'yicha
    if data["tp_lar"]:
        m += "\n👤 *TP bo'yicha:*\n"
        for tp, info in sorted(data["tp_lar"].items(), key=lambda x: -x[1]["jami"]):
            m += f"  {tp[:25]}: {info['soni']} klient — *{info['jami']:,.0f}*\n"

    # Top klientlar
    top_k = sorted(data["klientlar"], key=lambda x: -x["summa"])
    m += "\n👑 *TOP KLIENTLAR:*\n"
    for i, k in enumerate(top_k[:7], 1):
        m += f"  {i}. {k['ism'][:28]}: *{k['summa']:,.0f}*\n"
    if len(top_k) > 7:
        m += f"  _...va yana {len(top_k)-7} ta_\n"

    # Qarzli
    if data["qarzli"]:
        sum(abs(q.get("balans", 0)) if isinstance(q.get("balans"), (int, float)) else 0 for q in data["qarzli"])
        m += f"\n⚠️ *Qarzli: {len(data['qarzli'])} ta*\n"
        for q in sorted(data["qarzli"], key=lambda x: float(str(x.get("balans","0")).replace(",","").replace(" ","")) if str(x.get("balans","0")).replace(",","").replace(" ","").replace("-","").isdigit() else 0)[:5]:
            m += f"  📝 {q['ism']}: {q['balans']}\n"

    if len(m) > 4000:
        m = m[:3950] + "\n\n_...qisqartirildi_"
    return m
