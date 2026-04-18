"""
╔══════════════════════════════════════════════════════════════╗
║  SAVDOAI — EXCEL CHAT ANALYZER v2                            ║
║  Excel faylni o'qib, AI orqali savollarga javob berish       ║
║                                                              ║
║  v2: Xom data emas, tuzilmali xulosa yuboriladi             ║
║  AI auditorday javob beradi                                  ║
╚══════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io
import logging
import os

log = logging.getLogger("savdoai.excel_chat")


def excel_parse(file_bytes: bytes, filename: str = "file.xlsx") -> dict:
    """Excel faylni o'qib, tuzilmali dict qaytaradi."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl kutubxonasi o'rnatilmagan"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"error": f"Fayl o'qib bo'lmadi: {e}"}

    result = {
        "filename": filename,
        "sheets_count": len(wb.sheetnames),
        "sheets": []
    }
    total_cells = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        sheet_data = {"name": sn, "rows": ws.max_row or 0, "cols": ws.max_column or 0, "data": []}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            row_data = {}
            for cell in row:
                if cell.value is not None:
                    row_data[cell.coordinate] = cell.value if isinstance(cell.value, (int, float)) else str(cell.value)
                    total_cells += 1
            if row_data:
                sheet_data["data"].append(row_data)
        result["sheets"].append(sheet_data)
    result["total_cells"] = total_cells
    wb.close()
    return result


def _excel_to_smart_text(parsed: dict, max_chars: int = 90000) -> str:
    """
    Excel ni AI uchun ODAM O'QIYDIGAN formatga aylantirish.
    Xom katak emas, jadval ko'rinishida.
    """
    try:
        import openpyxl
    except ImportError:
        return _excel_to_text_fallback(parsed, max_chars)

    # Agar original bytes saqlangan bo'lsa
    # Fallback: parsed dict dan ishlash
    lines = []
    lines.append(f"FAYL: {parsed['filename']}")
    lines.append(f"SHEETLAR SONI: {parsed['sheets_count']}")
    lines.append(f"SHEET NOMLARI: {', '.join(s['name'] for s in parsed['sheets'])}")
    lines.append("")

    for sheet in parsed["sheets"]:
        sn = sheet["name"]
        lines.append(f"{'='*60}")
        lines.append(f"SHEET: '{sn}' ({sheet['rows']} qator, {sheet['cols']} ustun)")
        lines.append(f"{'='*60}")

        if not sheet["data"]:
            lines.append("  (bo'sh)")
            continue

        # Ma'lumotlarni qator raqami bo'yicha guruhlash
        rows_by_num = {}
        for row_data in sheet["data"]:
            for coord, val in row_data.items():
                import re
                m = re.match(r'([A-Z]+)(\d+)', coord)
                if m:
                    col_letter = m.group(1)
                    row_num = int(m.group(2))
                    if row_num not in rows_by_num:
                        rows_by_num[row_num] = {}
                    rows_by_num[row_num][col_letter] = val

        # Sarlavha qatorlarini aniqlash va ma'lumotlarni jadval formatda yozish
        header_cols = {}  # {col_letter: header_name}
        
        for rn in sorted(rows_by_num.keys()):
            row = rows_by_num[rn]
            
            # Sarlavha qatorini aniqlash (ko'p matnli ustunlar)
            text_count = sum(1 for v in row.values() if isinstance(v, str) and len(str(v)) > 3)
            num_count = sum(1 for v in row.values() if isinstance(v, (int, float)) and v != 0)
            
            if text_count >= 3 and num_count <= 1:
                # Bu sarlavha qatori
                header_cols = {}
                for col, val in row.items():
                    if isinstance(val, str) and len(val.strip()) > 1:
                        header_cols[col] = val.strip()
                if header_cols:
                    lines.append("")
                    lines.append(f"  --- Jadval (qator {rn}) ---")
                    hdr_text = " | ".join(f"{c}:{n}" for c, n in sorted(header_cols.items()))
                    lines.append(f"  Sarlavhalar: {hdr_text}")
                continue

            # Ma'lumot qatori
            parts = []
            for col in sorted(row.keys()):
                val = row[col]
                col_name = header_cols.get(col, col)
                if isinstance(val, (int, float)):
                    if val == 0:
                        continue
                    if val == int(val):
                        parts.append(f"{col_name}={int(val):,}")
                    else:
                        parts.append(f"{col_name}={val:,.2f}")
                elif isinstance(val, str) and val.strip():
                    v = val.strip()
                    if len(v) <= 50:
                        parts.append(f"{col_name}={v}")
            
            if parts:
                lines.append(f"  qator {rn}: {' | '.join(parts)}")

    text = "\n".join(lines)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n... (qisqartirildi — juda katta fayl)"
    return text


def _excel_to_text_fallback(parsed: dict, max_chars: int = 80000) -> str:
    """Oddiy fallback"""
    lines = [f"FAYL: {parsed['filename']}", f"SHEETLAR: {parsed['sheets_count']}", ""]
    for sheet in parsed["sheets"]:
        lines.append(f"SHEET '{sheet['name']}':")
        for row_data in sheet["data"]:
            parts = [f"{k}={v}" for k, v in row_data.items()]
            if parts:
                lines.append("  " + " | ".join(parts))
    text = "\n".join(lines)
    return text[:max_chars]


async def excel_savol(parsed: dict, savol: str,
                      tarix: list[dict] | None = None,
                      analysis: dict | None = None) -> str:
    """Excel haqida savolga AI orqali javob berish."""
    # Analyzer natijasi bor bo'lsa — uni ishlatish (ancha aniq)
    if analysis and "error" not in analysis:
        from shared.services.excel_analyzer import format_analysis_for_ai
        excel_text = format_analysis_for_ai(analysis)
    else:
        excel_text = _excel_to_smart_text(parsed)

    system_prompt = """Sen SavdoAI — professional moliya auditorisan. Senda Excel fayl tahlil natijalari bor.

MUHIM QOIDALAR:
1. Senga TAYYOR HISOBLANGAN natijalar berilgan — shu raqamlarni ishlat, o'zing qayta hisoblama
2. Javobni MARKDOWN JADVAL formatda ber — | ustun1 | ustun2 | formatida
3. Raqamlarni 1,000,000 ko'rinishda formatlash
4. Faqat so'ralgan narsaga javob ber — ortiqcha ma'lumot keramas
5. 0 qiymatli qatorlarni KO'RSATMA
6. JAMI qatorni **qalin** qil
7. O'zbek tilida javob ber (agar savol o'zbekcha bo'lsa), Rus tilida (agar ruscha)
8. Qisqa sarlavha qo'y (## bilan)
9. Agar savol umumiy bo'lsa — eng muhim ko'rsatkichlarni ko'rsat
10. Har xodim/kategoriya FAQAT 1 MARTA ko'rinsin (dublikat yozma!)
11. Bir xil xodimning turli yozuvlarini BIRLASHTIR (masalan: "Нодира ойлик" + "нодира ойлик" + "Нодира Ойлик" = bitta "Нодира")
12. Agent to'lovlar alohida: "тричап агентлари", "Ватика агентлар" — bu xodim oylik emas, agent to'lov

ATAMA LUG'ATI:
- "накд пул" / "nakd" = Фактическая оплата (I ustun) — agentlar olib kelgan cash
- "перечисление" / "perechisleniye" = bank o'tkazma
- "карта" / "karta" = bank karta orqali to'lov
- "клик" / "klik" = Click.uz to'lov
- "доллар" / "$" / "dollar" = valyuta operatsiya
- "ойлик" / "oylik" = maosh/salary
- "скидка" / "skidka" = klientga chegirma
- "газ" / "gaz" = yoqilg'i xarajati
- "обед" / "obed" = ovqatlanish
- "тачка" = aravakash/yuk tashish
- "расход" / "rasxod" = xarajat (umumiy)
- "итого" = jami (kunlik xulosa)

JAVOB NAMUNASI (shu formatda ber):

## Oylik to'lovlari

| Xodim | Summa |
|---|---|
| Искандар | 7,732,000 so'm |
| Зафар | 5,210,000 so'm |
| Фарход | 4,884,000 so'm |
| **JAMI** | **17,826,000 so'm** |"""

    messages = [
        {"role": "user", "content": f"Excel fayl tahlil natijalari:\n\n{excel_text}"},
        {"role": "assistant", "content": "Tahlil natijalarini qabul qildim. Barcha ko'rsatkichlar tayyor. Savol bering — aniq javob beraman."}
    ]

    if tarix:
        for msg in tarix[-10:]:
            messages.append(msg)

    messages.append({"role": "user", "content": savol})

    answer = await _ask_gemini(system_prompt, messages)
    if not answer:
        answer = await _ask_claude(system_prompt, messages)
    if not answer:
        answer = "❌ AI javob bera olmadi. Keyinroq urinib ko'ring."

    return answer


async def _ask_gemini(system: str, messages: list[dict]) -> str | None:
    api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    if not api_key:
        return None
    try:
        from google import genai
        client = genai.Client(api_key=api_key)
        contents = []
        for m in messages:
            role = "user" if m["role"] == "user" else "model"
            contents.append(genai.types.Content(role=role, parts=[genai.types.Part(text=m["content"])]))
        response = client.models.generate_content(
            model=os.getenv("GEMINI_MODEL", "gemini-2.5-flash"),
            contents=contents,
            config=genai.types.GenerateContentConfig(
                system_instruction=system, temperature=0.1, max_output_tokens=4096),
        )
        return response.text if response.text else None
    except Exception as e:
        log.warning("Gemini excel_chat xato: %s", e)
        return None


async def _ask_claude(system: str, messages: list[dict]) -> str | None:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return None
    try:
        import anthropic
        client = anthropic.AsyncAnthropic(api_key=api_key)
        response = await client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=4096,
            system=system, messages=messages)
        return response.content[0].text if response.content else None
    except Exception as e:
        log.warning("Claude excel_chat xato: %s", e)
        return None
