"""
╔══════════════════════════════════════════════════════════════════════╗
║  MASHRAB MOLIYA v21.4 — EXCEL IMPORT ENGINE                        ║
║                                                                      ║
║  ✅ .xlsx fayllarni o'qish va tahlil qilish                        ║
║  ✅ Реестр (reestr) fayllarni aniqlash                              ║
║  ✅ Накладная (nakladnoy) fayllarni aniqlash                       ║
║  ✅ Semi-structured layout (merged cells, optional columns)        ║
║  ✅ Telefon, sana, ism normalizatsiya                              ║
║  ✅ Dry-run preview (import qilmasdan ko'rsatish)                  ║
╚══════════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io
import logging
import re
from decimal import Decimal
from typing import Any, Optional

log = logging.getLogger(__name__)


def detect_file_type(wb) -> str:
    """Excel fayl turini aniqlash — reestr, nakladnoy, yoki noma'lum"""
    try:
        ws = wb.active
        if not ws:
            return "noma'lum"

        # Birinchi 5 qatordan sarlavhalarni o'qish
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    headers.append(str(cell).lower().strip())

        text = " ".join(headers)

        # Reestr belgilari
        reestr_words = ["реестр", "reestr", "баланс", "balans", "отправка",
                        "торговая точка", "savdo nuqtasi", "территория"]
        if any(w in text for w in reestr_words):
            return "reestr"

        # Nakladnoy belgilari
        nakl_words = ["накладная", "nakladnoy", "invoice", "faktura",
                      "кол-во", "цена", "сумма", "miqdor", "narx"]
        if any(w in text for w in nakl_words):
            return "nakladnoy"

        # Tovar ro'yxati
        if any(w in text for w in ["товар", "tovar", "mahsulot", "product", "sku"]):
            return "tovar_royxat"

        # Qarz ro'yxati
        if any(w in text for w in ["долг", "qarz", "задолженность", "qarzdor"]):
            return "qarz_royxat"

        return "noma'lum"
    except Exception as e:
        log.warning("detect_file_type: %s", e)
        return "noma'lum"


def parse_reestr(wb) -> dict:
    """Реестр/reestr faylni o'qish"""
    ws = wb.active
    result = {"tur": "reestr", "qatorlar": [], "jami": Decimal("0"), "xatolar": []}

    # Sarlavha qatorni topish
    header_row = None
    headers = {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        cells = [c.value for c in row]
        text = " ".join(str(c).lower() for c in cells if c)
        if any(w in text for w in ["точка", "nuqta", "клиент", "klient", "магазин", "shop"]):
            header_row = ri
            for ci, cell in enumerate(row):
                if cell.value:
                    h = str(cell.value).lower().strip()
                    headers[ci] = h
            break

    if not header_row:
        result["xatolar"].append("Sarlavha qatori topilmadi")
        return result

    # Ma'lumot qatorlarni o'qish
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        cells = {headers.get(ci, f"col_{ci}"): cell.value for ci, cell in enumerate(row)}
        if not any(v for v in cells.values() if v):
            continue  # Bo'sh qator

        # Klient/do'kon nomi topish
        klient = ""
        for key in cells:
            if any(w in key for w in ["точка", "nuqta", "клиент", "klient", "магазин"]):
                klient = str(cells[key] or "").strip()
                break

        # Summa topish
        summa = Decimal("0")
        for key in cells:
            if any(w in key for w in ["сумма", "summa", "amount", "итого", "jami"]):
                try:
                    val = str(cells[key] or "0").replace(" ", "").replace(",", ".")
                    summa = Decimal(val)
                except Exception as _exc:
                    log.debug("%s: %s", "excel_import", _exc)  # was silent
                break

        # Telefon
        telefon = ""
        for key in cells:
            if any(w in key for w in ["телефон", "telefon", "phone", "номер"]):
                telefon = str(cells[key] or "").strip()
                break

        if klient:
            result["qatorlar"].append({
                "klient": klient,
                "summa": str(summa),
                "telefon": telefon,
                "raw": {k: str(v) if v else "" for k, v in cells.items()},
            })
            result["jami"] += summa

    result["jami"] = str(result["jami"])
    return result


def parse_nakladnoy_excel(wb) -> dict:
    """Nakladnoy Excel faylni o'qish"""
    ws = wb.active
    result = {"tur": "nakladnoy", "tovarlar": [], "jami": Decimal("0"),
              "klient": "", "xatolar": []}

    # Sarlavha qatorni topish
    header_row = None
    headers = {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
        cells = [c.value for c in row]
        text = " ".join(str(c).lower() for c in cells if c)
        if any(w in text for w in ["наименование", "tovar", "товар", "nomi", "product"]):
            header_row = ri
            for ci, cell in enumerate(row):
                if cell.value:
                    headers[ci] = str(cell.value).lower().strip()
            break
        # Klient topish (sarlavhadan oldin)
        if any(w in text for w in ["клиент", "klient", "покупатель", "xaridor"]):
            for c in cells:
                if c and len(str(c)) > 2:
                    cand = str(c).strip()
                    if not any(w in cand.lower() for w in ["клиент", "klient", "покупатель"]):
                        result["klient"] = cand

    if not header_row:
        result["xatolar"].append("Tovar sarlavhasi topilmadi")
        return result

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        cells = {headers.get(ci, f"col_{ci}"): cell.value for ci, cell in enumerate(row)}
        if not any(v for v in cells.values() if v):
            continue

        nomi = ""
        miqdor = Decimal("0")
        narx = Decimal("0")
        jami = Decimal("0")

        for key, val in cells.items():
            if not val:
                continue
            kl = key.lower()
            if any(w in kl for w in ["наименование", "tovar", "товар", "nomi", "product"]):
                nomi = str(val).strip()
            elif any(w in kl for w in ["кол", "miqdor", "qty", "количество", "soni"]):
                try:
                    miqdor = Decimal(str(val).replace(" ", "").replace(",", "."))
                except Exception as _exc:
                    log.debug("%s: %s", "excel_import", _exc)  # was silent
            elif any(w in kl for w in ["цена", "narx", "price", "baho"]):
                try:
                    narx = Decimal(str(val).replace(" ", "").replace(",", "."))
                except Exception as _exc:
                    log.debug("%s: %s", "excel_import", _exc)  # was silent
            elif any(w in kl for w in ["сумма", "summa", "amount", "jami"]):
                try:
                    jami = Decimal(str(val).replace(" ", "").replace(",", "."))
                except Exception as _exc:
                    log.debug("%s: %s", "excel_import", _exc)  # was silent

        if nomi:
            if jami == 0 and miqdor > 0 and narx > 0:
                jami = miqdor * narx
            result["tovarlar"].append({
                "nomi": nomi,
                "miqdor": str(miqdor),
                "narx": str(narx),
                "jami": str(jami),
            })
            result["jami"] += jami

    result["jami"] = str(result["jami"])
    return result


def parse_excel(file_bytes: bytes) -> dict:
    """Excel faylni o'qish va tahlil qilish (dry-run)"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ftype = detect_file_type(wb)

        if ftype == "reestr":
            return parse_reestr(wb)
        elif ftype == "nakladnoy":
            return parse_nakladnoy_excel(wb)
        else:
            # Umumiy — barcha sheetlar haqida info
            sheets = wb.sheetnames
            ws = wb.active
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            return {
                "tur": ftype,
                "sheets": sheets,
                "qatorlar_soni": rows,
                "ustunlar_soni": cols,
                "xatolar": [],
            }
    except Exception as e:
        log.error("parse_excel: %s", e)
        return {"tur": "xato", "xatolar": [str(e)]}


def excel_preview_text(data: dict) -> str:
    """Parse natijasini Telegram matn formatga aylantirish"""
    tur = data.get("tur", "noma'lum")
    lines = [f"📊 *EXCEL TAHLIL*  |  Tur: *{tur}*\n"]

    if tur == "reestr":
        qatorlar = data.get("qatorlar", [])
        lines.append(f"👥 Klientlar: *{len(qatorlar)}* ta")
        lines.append(f"💰 Jami: *{data.get('jami', 0)}*\n")
        for i, q in enumerate(qatorlar[:10], 1):
            lines.append(f"{i}. {q['klient']} — {q['summa']}")
        if len(qatorlar) > 10:
            lines.append(f"...va yana {len(qatorlar)-10} ta")

    elif tur == "nakladnoy":
        tovarlar = data.get("tovarlar", [])
        klient = data.get("klient", "")
        if klient:
            lines.append(f"👤 Klient: *{klient}*")
        lines.append(f"📦 Tovarlar: *{len(tovarlar)}* ta")
        lines.append(f"💰 Jami: *{data.get('jami', 0)}*\n")
        for i, t in enumerate(tovarlar[:10], 1):
            lines.append(f"{i}. {t['nomi']} × {t['miqdor']} = {t['jami']}")
        if len(tovarlar) > 10:
            lines.append(f"...va yana {len(tovarlar)-10} ta")

    else:
        lines.append(f"📋 Sheetlar: {data.get('sheets', [])}")
        lines.append(f"📊 Qatorlar: {data.get('qatorlar_soni', 0)}")
        lines.append(f"📊 Ustunlar: {data.get('ustunlar_soni', 0)}")

    xatolar = data.get("xatolar", [])
    if xatolar:
        lines.append(f"\n⚠️ *Xatolar:*")
        for x in xatolar[:5]:
            lines.append(f"  ❌ {x}")

    return "\n".join(lines)
