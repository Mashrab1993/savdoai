"""
╔══════════════════════════════════════════════════════════════════════════╗
║  MASHRAB MOLIYA — EXCEL AUDITOR v2.0                                     ║
║                                                                          ║
║  200% ANIQLIK — har bir qator, har bir raqam, har bir so'm             ║
║  AI EMAS — o'zimiz hisoblaymiz, XATOSIZ                                ║
║  Auditor darajasidagi xulosa + PDF hisobot                             ║
╚══════════════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io, json, logging, re, datetime
from typing import Optional
from collections import defaultdict
log = logging.getLogger(__name__)


def excel_toliq_oqi(data: bytes) -> dict:
    """Excel ni AUDITOR darajasida o'qish — har bir hujayra."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"xato": "openpyxl kerak"}

    result = {
        "sheetlar_soni": 0, "sheetlar": [], "umumiy_matn": "",
        "xulosa": {}, "kunlik": [], "klientlar": {}, "rasxod_tafsilot": [],
    }

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
        result["sheetlar_soni"] = len(wb.sheetnames)

        umumiy_parts = []
        jami = {"tovar": 0, "dolg": 0, "oplata": 0, "vozvrat": 0, "sales": 0, "rasxod": 0, "fakt_oplata": 0}
        rasxod_kategoriya = defaultdict(float)
        klientlar = defaultdict(lambda: {"dolg": 0, "oplata": 0, "vozvrat": 0})
        kunlik = []
        oylik = {}

        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue

            # JAMI qatorini DINAMIK topish (birinchi 10 qatorda "#" izlash)
            jami_row = None
            sarlavha_row = None
            data_start_idx = 0
            for ri in range(min(10, len(rows))):
                row_vals = rows[ri]
                first_cell = str(row_vals[0]).strip() if row_vals[0] else ""
                if first_cell == "#":
                    jami_row = row_vals
                    data_start_idx = ri + 1
                    if ri > 0:
                        sarlavha_row = rows[ri - 1]
                    break

            # "#" TOPILMASA — UNIVERSAL READER
            if not jami_row:
                # Sarlavha topish — matnli qator
                for ri in range(min(5, len(rows))):
                    row_vals = rows[ri]
                    text_cells = [str(c).strip() for c in row_vals if c and not str(c).strip().replace('.','').replace(',','').replace('-','').replace(' ','').isdigit()]
                    if len(text_cells) >= 2:
                        sarlavha_row = row_vals
                        data_start_idx = ri + 1
                        break
                
                # Barcha raqamlarni yig'ish
                sheet_raqamlar = []
                sheet_matn = []
                for row in rows[data_start_idx:]:
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    if not any(cells): continue
                    sheet_matn.append(" | ".join(cells))
                    for cell in cells:
                        try:
                            v = float(str(cell).replace(',','.').replace(' ',''))
                            # Telefon raqamlarni filtrlash (9+ raqamli)
                            if v != 0 and abs(v) < 100_000_000_000 and len(str(cell).replace(' ','').replace(',','').replace('.','')) < 12:
                                sheet_raqamlar.append(v)
                        except: pass

                sheet_data = {
                    "nom": sn, "qator_soni": len(rows),
                    "sarlavha": [str(c).strip() if c else "" for c in (sarlavha_row or [])],
                    "kun_data": {"sheet": sn, "tovar": 0, "dolg": 0, "oplata": 0, "vozvrat": 0, "sales": 0, "rasxod": 0, "fakt": 0},
                    "universal": True,
                    "raqamlar_soni": len(sheet_raqamlar),
                    "jami_raqam": sum(r for r in sheet_raqamlar if r > 0),
                }
                if sheet_raqamlar:
                    sheet_data["eng_katta"] = max(sheet_raqamlar)
                result["sheetlar"].append(sheet_data)

                sheet_text = f"=== {sn} ({len(rows)} qator) ===\n"
                for line in sheet_matn[:80]:
                    sheet_text += line + "\n"
                umumiy_parts.append(sheet_text)
                continue

            sarlavha = [str(c).strip() if c else "" for c in (sarlavha_row or [])]

            # JAMI qatorini o'qish
            kun_data = {"sheet": sn, "tovar": 0, "dolg": 0, "oplata": 0, "vozvrat": 0, "sales": 0, "rasxod": 0, "fakt": 0}
            jami_row_idx = 0
            if jami_row:
                vals = []
                for c in jami_row:
                    try:
                        vals.append(float(c) if isinstance(c, (int, float)) and c else 0)
                    except:
                        vals.append(0)
                while len(vals) < 10:
                    vals.append(0)

                kun_data["tovar"] = vals[1]
                kun_data["dolg"] = vals[2]
                kun_data["oplata"] = vals[3]
                kun_data["vozvrat"] = vals[4]
                kun_data["sales"] = vals[5]
                kun_data["rasxod"] = vals[6]
                if len(vals) > 8:
                    kun_data["fakt"] = vals[8]

                for k in ["tovar", "dolg", "oplata", "vozvrat", "sales", "rasxod"]:
                    jami[k] += kun_data[k]
                jami["fakt_oplata"] += kun_data["fakt"]

            kunlik.append(kun_data)

            # Oy bo'yicha guruhlash
            sana_match = re.search(r'(\d{2})[\.\-,](\d{2})[\.\-,](\d{4})', sn)
            if sana_match and kun_data.get("tovar", 0) > 0:
                oy_num = int(sana_match.group(2))
                oy_nomi = {1:"YANVAR",2:"FEVRAL",3:"MART",4:"APREL",5:"MAY",6:"IYUN",
                           7:"IYUL",8:"AVGUST",9:"SENTYABR",10:"OKTYABR",11:"NOYABR",12:"DEKABR"}.get(oy_num, f"OY-{oy_num}")
                oylik.setdefault(oy_nomi, {"tovar":0,"dolg":0,"oplata":0,"vozvrat":0,"sales":0,"rasxod":0,"kunlar":set()})
                for k2 in ["tovar","dolg","oplata","vozvrat","sales","rasxod"]:
                    oylik[oy_nomi][k2] += kun_data.get(k2, 0)
                oylik[oy_nomi]["kunlar"].add(sana_match.group(1) + "." + sana_match.group(2))

            # Tafsilot qatorlarini o'qish (4-qatordan boshlab)
            for row in rows[3:]:
                cells = [str(c).strip() if c is not None else "" for c in row]
                if not any(cells):
                    continue

                # Klient ismi (3-ustun = Оплаты по долгам, 4-ustun = Сумма возврата)
                for col_idx in [3, 4]:
                    if col_idx < len(cells) and cells[col_idx]:
                        ism = cells[col_idx].strip()
                        if ism and len(ism) > 2 and not ism.replace(',','').replace('.','').replace(' ','').isdigit():
                            if not ism.startswith("VAZVRAT") and not ism.startswith("CLICK"):
                                pass  # Klient nomi

                # Dolg (2-ustun)
                if len(cells) > 2 and cells[2]:
                    try:
                        dolg_val = float(cells[2].replace(',','.').replace(' ',''))
                        if dolg_val > 0 and len(cells) > 3 and cells[3]:
                            klient_ism = cells[3].strip()
                            if klient_ism and len(klient_ism) > 2:
                                klientlar[klient_ism]["dolg"] += dolg_val
                    except:
                        pass

                # Oplata (3-ustun raqam bo'lsa)
                if len(cells) > 3 and cells[3]:
                    try:
                        opl_val = float(cells[3].replace(',','.').replace(' ',''))
                        if opl_val > 0 and len(cells) > 3:
                            # Klient ismi keyingi ustunda
                            if len(cells) > 4 and cells[4]:
                                klient_ism = cells[4].strip()
                                if klient_ism and len(klient_ism) > 2:
                                    klientlar[klient_ism]["oplata"] += opl_val
                    except:
                        pass

                # Rasxod tafsilot (6-ustun = summa, 7-ustun = izoh)
                if len(cells) > 6 and cells[6]:
                    try:
                        rsx_val = float(cells[6].replace(',','.').replace(' ',''))
                        if rsx_val > 0:
                            izoh = cells[7].strip() if len(cells) > 7 else ""
                            result["rasxod_tafsilot"].append({"sheet": sn, "summa": rsx_val, "izoh": izoh})
                            # Kategoriya aniqlash
                            izoh_lower = izoh.lower()
                            if "abet" in izoh_lower:
                                rasxod_kategoriya["ABET"] += rsx_val
                            elif "gaz" in izoh_lower or "benzin" in izoh_lower:
                                rasxod_kategoriya["GAZ/BENZIN"] += rsx_val
                            elif "oylik" in izoh_lower or "maosh" in izoh_lower:
                                rasxod_kategoriya["OYLIK"] += rsx_val
                            elif "skidka" in izoh_lower or "chegirma" in izoh_lower:
                                rasxod_kategoriya["SKIDKA"] += rsx_val
                            elif "click" in izoh_lower:
                                rasxod_kategoriya["CLICK"] += rsx_val
                            elif "moy" in izoh_lower or "moyka" in izoh_lower:
                                rasxod_kategoriya["MOY/MOYKA"] += rsx_val
                            elif "qoldiq" in izoh_lower:
                                rasxod_kategoriya["QOLDIQ"] += rsx_val
                            else:
                                rasxod_kategoriya["BOSHQA"] += rsx_val
                    except:
                        pass

            # Sheet info
            result["sheetlar"].append({
                "nom": sn, "qator_soni": len(rows),
                "sarlavha": sarlavha,
                "kun_data": kun_data,
            })

            # Umumiy matn
            sheet_text = f"=== {sn} ({len(rows)} qator) ===\n"
            for row in rows[:60]:
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    sheet_text += " | ".join(cells) + "\n"
            umumiy_parts.append(sheet_text)

        # Oylik kunlar set → count
        oylik_clean = {}
        for oy, data in oylik.items():
            oylik_clean[oy] = {k: v for k, v in data.items() if k != "kunlar"}
            oylik_clean[oy]["kunlar_soni"] = len(data.get("kunlar", set()))

        result["xulosa"] = {
            "jami_tovar": jami["tovar"],
            "jami_dolg": jami["dolg"],
            "jami_oplata": jami["oplata"],
            "jami_vozvrat": jami["vozvrat"],
            "jami_sales": jami["sales"],
            "jami_rasxod": jami["rasxod"],
            "jami_fakt_oplata": jami["fakt_oplata"],
            "sof_foyda": jami["sales"] - jami["rasxod"],
            "rasxod_kategoriya": dict(rasxod_kategoriya),
            "klientlar_soni": len(klientlar),
            "oylik": oylik_clean,
        }
        result["kunlik"] = kunlik
        result["klientlar"] = dict(klientlar)
        result["umumiy_matn"] = "\n".join(umumiy_parts)

        wb.close()
        log.info("Excel AUDITOR: %d sheet, jami_sales=%.0f, jami_rasxod=%.0f",
                 result["sheetlar_soni"], jami["sales"], jami["rasxod"])

    except Exception as e:
        log.error("Excel: %s", e)
        result["xato"] = str(e)

    return result


def _pul(v):
    try:
        return f"{float(v):,.0f}"
    except:
        return "0"


def excel_xulosa_matn(r: dict, fayl_nomi: str) -> str:
    """AUDITOR darajasidagi xulosa — Telegram uchun."""
    t = f"📊 *MASHRAB MOLIYA*\n"
    t += f"📂 {fayl_nomi}\n"
    t += "━━━━━━━━━━━━━━━━━━━━━━━━\n\n"

    if r.get("xato"):
        return t + f"❌ {r['xato']}"

    x = r.get("xulosa", {})
    kunlar_set = set()
    for k in r.get("kunlik", []):
        if k.get("tovar", 0) > 0:
            # Sheet nomidan sanani ajratish: "06.01.2025 CHELAK" → "06.01.2025"
            sana_match = re.search(r'(\d{2}[\.\-]\d{2}[\.\-]\d{4})', k.get("sheet", ""))
            if sana_match:
                kunlar_set.add(sana_match.group(1))
            else:
                kunlar_set.add(k["sheet"][:10])  # Birinchi 10 belgi
    kunlar = len(kunlar_set) if kunlar_set else len([k for k in r.get("kunlik", []) if k.get("tovar", 0) > 0])

    t += f"📅 Davr: *{kunlar} ish kuni*\n"
    t += f"📑 Sheetlar: *{r.get('sheetlar_soni', 0)}* ta\n\n"

    has_savdo = x.get("jami_sales", 0) > 0 or x.get("jami_tovar", 0) > 0
    
    if has_savdo:
        t += "💰 *ASOSIY KO'RSATKICHLAR:*\n"
        t += f"  📦 Olingan tovar: *{_pul(x.get('jami_tovar', 0))}*\n"
        t += f"  💳 Sotuv (Sales): *{_pul(x.get('jami_sales', 0))}*\n"
        t += f"  📝 Qarzga berilgan: *{_pul(x.get('jami_dolg', 0))}*\n"
        t += f"  ✅ Qarz to'langan: *{_pul(x.get('jami_oplata', 0))}*\n"
        t += f"  ↩️ Qaytarilgan: *{_pul(x.get('jami_vozvrat', 0))}*\n"
        t += f"  📤 Rasxodlar: *{_pul(x.get('jami_rasxod', 0))}*\n\n"

    # Rasxod tafsilot
    kat = x.get("rasxod_kategoriya", {})
    if kat:
        t += "📊 *RASXOD TAFSILOTI:*\n"
        for k, v in sorted(kat.items(), key=lambda x: -x[1]):
            t += f"  {k}: {_pul(v)}\n"
        t += "\n"

    sof = x.get("sof_foyda", 0)
    has_savdo = x.get("jami_sales", 0) > 0
    
    if has_savdo:
        if sof > 0:
            t += f"✅ *SOF NATIJA: +{_pul(sof)}*\n"
        else:
            t += f"❌ *SOF NATIJA: {_pul(sof)}*\n"
    
    # Universal sheetlar (# formatsiz)
    uni_sheets = [s for s in r.get("sheetlar", []) if s.get("universal")]
    if uni_sheets and not has_savdo:
        t += "📊 *JADVALLAR:*\n"
        for sh in uni_sheets[:10]:
            t += f"  📋 {sh['nom']}: {sh['qator_soni']} qator"
            if sh.get("jami_raqam"):
                t += f", jami={_pul(sh['jami_raqam'])}"
            t += "\n"

    # Oy bo'yicha
    oylik = x.get("oylik", {})
    if oylik and len(oylik) > 1:
        t += "\n📅 *OYLAR BO'YICHA:*\n"
        for oy, data in oylik.items():
            t += f"  {oy}: sales={_pul(data.get('sales',0))}, {data.get('kunlar_soni',0)} kun\n"

    t += "\n💡 Savolingizni bering — *Mashrab Moliya* javob beradi!"
    return t.rstrip()


def savol_javob(r: dict, savol: str) -> str:
    """Har qanday savolga ANIQ javob — AI siz, XATOSIZ."""
    s = savol.lower().strip().strip('"').strip("'").strip("?").strip()
    x = r.get("xulosa", {})
    kat = x.get("rasxod_kategoriya", {})

    # RASXOD kategoriyalar
    ABET_SOZLAR = ["abet", "abed", "abid", "abyet"]
    GAZ_SOZLAR = ["gaz", "benzin", "yoqilgi", "gaz benzin"]
    OYLIK_SOZLAR = ["oylik", "maosh", "ish haqi"]

    for soz in ABET_SOZLAR:
        if soz in s:
            v = kat.get("ABET", 0)
            return f"📊 *MASHRAB MOLIYA*\n\n💰 *ABET:* {_pul(v)} so'm\n\n📋 {len([r for r in (r if isinstance(r, list) else [])]) if False else ''}" if v else f"📊 *MASHRAB MOLIYA*\n\n💰 *ABET:* {_pul(v)} so'm"

    for soz in GAZ_SOZLAR:
        if soz in s:
            v = kat.get("GAZ/BENZIN", 0)
            return f"📊 *MASHRAB MOLIYA*\n\n⛽ *GAZ/BENZIN:* {_pul(v)} so'm"

    for soz in OYLIK_SOZLAR:
        if soz in s:
            v = kat.get("OYLIK", 0)
            return f"📊 *MASHRAB MOLIYA*\n\n👤 *OYLIK:* {_pul(v)} so'm"

    if "skidka" in s or "chegirma" in s:
        v = kat.get("SKIDKA", 0)
        return f"📊 *MASHRAB MOLIYA*\n\n🏷 *SKIDKA:* {_pul(v)} so'm"

    if "click" in s:
        v = kat.get("CLICK", 0)
        return f"📊 *MASHRAB MOLIYA*\n\n📱 *CLICK:* {_pul(v)} so'm"

    # UMUMIY SAVOLLAR
    if any(k in s for k in ["rasxod", "xarajat", "расход", "chiqim", "sarf"]):
        t = "📊 *MASHRAB MOLIYA — RASXODLAR*\n\n"
        jami = 0
        for k, v in sorted(kat.items(), key=lambda x: -x[1]):
            t += f"  {k}: *{_pul(v)}*\n"
            jami += v
        t += f"\n  📌 *JAMI RASXOD: {_pul(jami)}*"
        return t

    if any(k in s for k in ["sotuv", "savdo", "sales", "jami summa", "tushum"]):
        return (f"📊 *MASHRAB MOLIYA — SOTUV*\n\n"
                f"  💳 Sotuv (Sales): *{_pul(x.get('jami_sales', 0))}*\n"
                f"  📦 Olingan tovar: *{_pul(x.get('jami_tovar', 0))}*\n"
                f"  ✅ Sof natija: *{_pul(x.get('sof_foyda', 0))}*")

    if any(k in s for k in ["qarz", "dolg", "долг", "qarzdor"]):
        return (f"📊 *MASHRAB MOLIYA — QARZLAR*\n\n"
                f"  📝 Qarzga berilgan: *{_pul(x.get('jami_dolg', 0))}*\n"
                f"  ✅ To'langan: *{_pul(x.get('jami_oplata', 0))}*\n"
                f"  📌 Farq: *{_pul(x.get('jami_dolg', 0) - x.get('jami_oplata', 0))}*")

    if any(k in s for k in ["vozvrat", "qaytarish", "qaytarilgan", "возврат"]):
        return f"📊 *MASHRAB MOLIYA*\n\n↩️ *Qaytarilgan: {_pul(x.get('jami_vozvrat', 0))}*"

    if any(k in s for k in ["tovar", "olingan", "kelgan", "товар"]):
        return f"📊 *MASHRAB MOLIYA*\n\n📦 *Olingan tovar: {_pul(x.get('jami_tovar', 0))}*"

    if any(k in s for k in ["foyda", "natija", "sof", "daromad", "profit"]):
        sof = x.get("sof_foyda", 0)
        emoji = "✅" if sof > 0 else "❌"
        return f"📊 *MASHRAB MOLIYA*\n\n{emoji} *Sof natija: {_pul(sof)}*\n\n  Sales: {_pul(x.get('jami_sales',0))}\n  Rasxod: {_pul(x.get('jami_rasxod',0))}"

    if any(k in s for k in ["klient", "mijoz", "клиент"]):
        kl = r.get("klientlar", {})
        if kl:
            t = f"📊 *MASHRAB MOLIYA — KLIENTLAR ({len(kl)} ta)*\n\n"
            for ism, data in sorted(kl.items(), key=lambda x: -x[1].get("dolg",0))[:15]:
                d = data.get("dolg", 0)
                o = data.get("oplata", 0)
                if d > 0 or o > 0:
                    t += f"  👤 {ism}: dolg={_pul(d)}, to'lov={_pul(o)}\n"
            return t.rstrip()
        return "📊 *MASHRAB MOLIYA*\n\nKlient ma'lumoti topilmadi."

    if any(k in s for k in ["kunlik", "kun bo'yicha", "har kun"]):
        t = "📊 *MASHRAB MOLIYA — KUNLIK*\n\n"
        for k in r.get("kunlik", [])[:20]:
            if k.get("tovar", 0) > 0:
                t += f"  📅 {k['sheet'][:25]}: sales={_pul(k['sales'])}\n"
        return t.rstrip()

    # OY BO'YICHA
    oylik = x.get("oylik", {})
    if any(k in s for k in ["oy bo'yicha", "oylik hisobot", "har oy", "oy"]) and "oylik" not in s.split()[-1:]:
        if oylik:
            t = "📊 *MASHRAB MOLIYA — OYLAR BO'YICHA*\n\n"
            jami_sales = 0
            for oy, data in oylik.items():
                sal = data.get("sales", 0)
                rsx = data.get("rasxod", 0)
                t += f"  📅 *{oy}* ({data.get('kunlar_soni',0)} kun):\n"
                t += f"     Tovar: {_pul(data.get('tovar',0))} | Sales: {_pul(sal)}\n"
                t += f"     Qarz: {_pul(data.get('dolg',0))} | Rasxod: {_pul(rsx)}\n\n"
                jami_sales += sal
            t += f"  📌 *BARCHA OYLAR JAMI: {_pul(jami_sales)}*"
            return t
    
    # Aniq oy nomi
    OY_NOMLAR = {"yanvar":1,"fevral":2,"mart":3,"aprel":4,"may":5,"iyun":6,
                 "iyul":7,"avgust":8,"sentyabr":9,"oktyabr":10,"noyabr":11,"dekabr":12}
    for oy_nom, oy_num in OY_NOMLAR.items():
        if oy_nom in s:
            oy_key = {1:"YANVAR",2:"FEVRAL",3:"MART",4:"APREL",5:"MAY",6:"IYUN",
                      7:"IYUL",8:"AVGUST",9:"SENTYABR",10:"OKTYABR",11:"NOYABR",12:"DEKABR"}.get(oy_num)
            if oy_key and oy_key in oylik:
                data = oylik[oy_key]
                t = f"📊 *MASHRAB MOLIYA — {oy_key}*\n\n"
                t += f"  📅 Ish kunlari: *{data.get('kunlar_soni',0)}*\n"
                t += f"  📦 Tovar: *{_pul(data.get('tovar',0))}*\n"
                t += f"  💳 Sales: *{_pul(data.get('sales',0))}*\n"
                t += f"  📝 Qarz: *{_pul(data.get('dolg',0))}*\n"
                t += f"  ✅ To'langan: *{_pul(data.get('oplata',0))}*\n"
                t += f"  ↩️ Qaytarish: *{_pul(data.get('vozvrat',0))}*\n"
                t += f"  📤 Rasxod: *{_pul(data.get('rasxod',0))}*\n"
                sof = data.get("sales",0) - data.get("rasxod",0)
                emoji = "✅" if sof > 0 else "❌"
                t += f"\n  {emoji} *Sof: {_pul(sof)}*"
                return t

    return None  # AI ga yuborish


async def excel_ai_savol(r: dict, savol: str, gemini_key: str = "") -> str:
    """Avval o'zimiz javob, topilmasa AI."""
    import asyncio, os

    # 1. O'zimiz javob — ANIQ, XATOSIZ
    javob = savol_javob(r, savol)
    if javob:
        return javob

    # 2. AI — murakkab savollar uchun
    if not gemini_key:
        gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if not gemini_key:
        return f"📊 *MASHRAB MOLIYA*\n\n🔍 '{savol}' bo'yicha javob topilmadi."

    matn = r.get("umumiy_matn", "")[:25000]
    x = r.get("xulosa", {})

    prompt = f"""Sen MASHRAB MOLIYA auditor tizimisan. Excel ma'lumotlari berilgan.
ANIQ raqamlar bilan javob ber. "Mashrab Moliya" nomidan javob ber.
Raqamlarni 1,234,567 formatda yoz. 

UMUMIY KO'RSATKICHLAR (SEN BILASAN):
- Jami tovar: {x.get('jami_tovar',0):,.0f}
- Jami sotuv: {x.get('jami_sales',0):,.0f}
- Jami qarz: {x.get('jami_dolg',0):,.0f}
- Jami to'langan: {x.get('jami_oplata',0):,.0f}
- Jami qaytarish: {x.get('jami_vozvrat',0):,.0f}
- Jami rasxod: {x.get('jami_rasxod',0):,.0f}
- Rasxod tafsilot: {json.dumps(x.get('rasxod_kategoriya',{}), ensure_ascii=False)}

EXCEL MA'LUMOTLARI:
{matn}

SAVOL: {savol}

Javobni "📊 MASHRAB MOLIYA" bilan boshla. Aniq raqamlar ber."""

    try:
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=gemini_key)
        loop = asyncio.get_event_loop()
        response = await asyncio.wait_for(
            loop.run_in_executor(None, lambda: client.models.generate_content(
                model="gemini-2.5-pro", contents=prompt,
                config=types.GenerateContentConfig(temperature=0.1, max_output_tokens=3000))),
            timeout=30)
        javob = (response.text or "").strip()
        if javob:
            # "AI JAVOB" o'rniga "MASHRAB MOLIYA"
            javob = javob.replace("AI JAVOB", "MASHRAB MOLIYA")
            javob = javob.replace("🤖 AI", "📊 Mashrab Moliya")
            return javob
    except Exception as e:
        log.debug("Excel AI: %s", e)

    return f"📊 *MASHRAB MOLIYA*\n\n🔍 '{savol}' bo'yicha javob topilmadi. Boshqacha so'rab ko'ring."


def _oddiy_izlash(r: dict, savol: str) -> str:
    """Fallback."""
    javob = savol_javob(r, savol)
    if javob:
        return javob
    return f"📊 *MASHRAB MOLIYA*\n\n🔍 '{savol}' — boshqacha so'rab ko'ring."


# ═══════════════════════════════════════════════════════════════
#  PDF HISOBOT — AUDITOR DARAJASIDA
# ═══════════════════════════════════════════════════════════════

def excel_pdf_hisobot(r: dict, fayl_nomi: str) -> bytes:
    """Auditor darajasidagi PDF hisobot yaratish."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18,
        textColor=colors.HexColor('#1a237e'), spaceAfter=6))
    styles.add(ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=11,
        textColor=colors.HexColor('#424242'), alignment=TA_CENTER, spaceAfter=12))
    styles.add(ParagraphStyle('SectionHeader', parent=styles['Heading2'], fontSize=13,
        textColor=colors.HexColor('#0d47a1'), spaceBefore=14, spaceAfter=6,
        borderPadding=(0, 0, 3, 0)))
    styles.add(ParagraphStyle('ValueBig', parent=styles['Normal'], fontSize=14,
        textColor=colors.HexColor('#1b5e20'), alignment=TA_RIGHT))
    styles.add(ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8,
        textColor=colors.HexColor('#9e9e9e'), alignment=TA_CENTER))

    story = []
    x = r.get("xulosa", {})
    kat = x.get("rasxod_kategoriya", {})

    # Ish kunlarini to'g'ri hisoblash
    kunlar_set = set()
    for k in r.get("kunlik", []):
        if k.get("tovar", 0) > 0:
            sana_match = re.search(r'(\d{2}[\.\-]\d{2}[\.\-]\d{4})', k.get("sheet", ""))
            if sana_match:
                kunlar_set.add(sana_match.group(1))
            else:
                kunlar_set.add(k["sheet"][:10])
    kunlar = len(kunlar_set) if kunlar_set else len([k for k in r.get("kunlik", []) if k.get("tovar", 0) > 0])

    # ═══ SARLAVHA ═══
    story.append(Paragraph("MASHRAB MOLIYA", styles['CustomTitle']))
    story.append(Paragraph("MOLIYAVIY HISOBOT", styles['SubTitle']))
    story.append(Spacer(1, 3*mm))

    # Info jadvali
    sana = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    info = [
        ["Fayl:", fayl_nomi],
        ["Sana:", sana],
        ["Davr:", f"{kunlar} ish kuni"],
        ["Sheetlar:", f"{r.get('sheetlar_soni', 0)} ta"],
    ]
    info_table = Table(info, colWidths=[80, 350])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#616161')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 8*mm))

    # ═══ ASOSIY KO'RSATKICHLAR ═══
    story.append(Paragraph("ASOSIY KO'RSATKICHLAR", styles['SectionHeader']))
    
    korsatkichlar = [
        ["Ko'rsatkich", "Summa (so'm)"],
        ["Olingan tovar", f"{x.get('jami_tovar', 0):,.0f}"],
        ["Sotuv (Sales)", f"{x.get('jami_sales', 0):,.0f}"],
        ["Qarzga berilgan", f"{x.get('jami_dolg', 0):,.0f}"],
        ["Qarz to'langan", f"{x.get('jami_oplata', 0):,.0f}"],
        ["Qaytarilgan", f"{x.get('jami_vozvrat', 0):,.0f}"],
        ["Rasxodlar", f"{x.get('jami_rasxod', 0):,.0f}"],
    ]
    
    t = Table(korsatkichlar, colWidths=[250, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fafafa')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 8*mm))

    # ═══ SOF NATIJA ═══
    sof = x.get("sof_foyda", 0)
    natija_rang = '#1b5e20' if sof > 0 else '#c62828'
    natija_belgi = "+" if sof > 0 else ""
    story.append(Paragraph(
        f'<font color="{natija_rang}" size="16"><b>SOF NATIJA: {natija_belgi}{sof:,.0f} so\'m</b></font>',
        styles['Normal']))
    story.append(Spacer(1, 8*mm))

    # ═══ RASXOD TAFSILOTI ═══
    if kat:
        story.append(Paragraph("RASXOD TAFSILOTI", styles['SectionHeader']))
        rsx_data = [["Kategoriya", "Summa (so'm)", "%"]]
        jami_rsx = sum(kat.values())
        for k, v in sorted(kat.items(), key=lambda x: -x[1]):
            foiz = (v / jami_rsx * 100) if jami_rsx > 0 else 0
            rsx_data.append([k, f"{v:,.0f}", f"{foiz:.1f}%"])
        rsx_data.append(["JAMI", f"{jami_rsx:,.0f}", "100%"])
        
        t2 = Table(rsx_data, colWidths=[200, 160, 60])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e65100')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fff3e0')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#fff8f0')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t2)
        story.append(Spacer(1, 8*mm))

    # ═══ KUNLIK TAFSILOT ═══
    kunlik_data = [k for k in r.get("kunlik", []) if k.get("tovar", 0) > 0]

    # OY BO'YICHA (agar 2+ oy bo'lsa)
    oylik = x.get("oylik", {})
    if oylik and len(oylik) >= 1:
        story.append(Paragraph("OYLAR BO'YICHA", styles['SectionHeader']))
        oy_table = [["Oy", "Kunlar", "Tovar", "Sales", "Qarz", "Rasxod", "Sof"]]
        for oy, data in oylik.items():
            sof_oy = data.get("sales",0) - data.get("rasxod",0)
            oy_table.append([
                oy, str(data.get("kunlar_soni",0)),
                f"{data.get('tovar',0):,.0f}", f"{data.get('sales',0):,.0f}",
                f"{data.get('dolg',0):,.0f}", f"{data.get('rasxod',0):,.0f}",
                f"{sof_oy:,.0f}",
            ])
        t_oy = Table(oy_table, colWidths=[65, 35, 80, 80, 65, 65, 75])
        t_oy.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a148c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e0e0e0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3e5f5')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t_oy)
        story.append(Spacer(1, 6*mm))

    if kunlik_data:
        story.append(Paragraph("KUNLIK TAFSILOT", styles['SectionHeader']))
        kun_table = [["Sana / Marshrut", "Tovar", "Sales", "Qarz", "Rasxod"]]
        for k in kunlik_data:
            nom = k["sheet"][:30]
            kun_table.append([
                nom,
                f"{k.get('tovar', 0):,.0f}",
                f"{k.get('sales', 0):,.0f}",
                f"{k.get('dolg', 0):,.0f}",
                f"{k.get('rasxod', 0):,.0f}",
            ])

        t3 = Table(kun_table, colWidths=[130, 90, 90, 80, 80])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d47a1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e0e0e0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4ff')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t3)

    # ═══ FOOTER ═══
    story.append(Spacer(1, 15*mm))
    story.append(Paragraph(
        f"Mashrab Moliya Auditor Tizimi | {sana} | Avtomatik yaratilgan hisobot",
        styles['Footer']))

    doc.build(story)
    return buf.getvalue()
