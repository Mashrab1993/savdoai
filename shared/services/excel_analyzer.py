"""
╔══════════════════════════════════════════════════════════════╗
║  SAVDOAI — EXCEL ANALYZER ENGINE v3                          ║
║  Excel faylni PROGRAMMATIK tahlil qilish                     ║
║  AI ga xom data emas, TAYYOR NATIJALAR yuboriladi            ║
║                                                              ║
║  Hisoblaydi:                                                 ║
║  ✅ Nakd pul kirim                                           ║
║  ✅ Perechisleniya (karta, klik)                             ║
║  ✅ Dollar operatsiyalar                                      ║
║  ✅ Oyliklar (har xodim alohida)                             ║
║  ✅ Rasxodlar (kategoriya bo'yicha)                          ║
║  ✅ Kunlik savdo                                             ║
║  ✅ Agentlar bo'yicha                                        ║
╚══════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io, re, logging
from collections import defaultdict
from typing import Optional

log = logging.getLogger("savdoai.excel_analyzer")


def analyze_trade_excel(file_bytes: bytes, filename: str = "") -> dict:
    """
    Savdo Excel faylni to'liq tahlil qilish.
    Natija: barcha ko'rsatkichlar hisoblangan dict.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl yo'q"}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"error": str(e)}

    result = {
        "filename": filename,
        "sheets": len(wb.sheetnames),
        "kunlar": [],          # kunlik xulosa
        "oyliklar": [],        # xodim oylik to'lovlari
        "rasxodlar_turi": {},  # kategoriya: summa
        "rasxodlar_list": [],  # har bir rasxod
        "dollar_ops": [],      # dollar operatsiyalar
        "perechisleniya": [],  # pul o'tkazmalar
        "agentlar": {},        # agent: {tovar, nakd, ...}
        "jami": {
            "savdo": 0, "nakd": 0, "dollar": 0, "dollar_sum": 0,
            "rasxod": 0, "oylik": 0, "qarz_berilgan": 0,
            "qarz_yigilgan": 0, "qaytarish": 0,
            "karta": 0, "perechisleniye": 0, "klik": 0,
        }
    }

    for sn in wb.sheetnames:
        ws = wb[sn]
        if not ws.max_row or ws.max_row < 5:
            continue

        kun_data = {
            "kun": sn, "savdo": 0, "nakd": 0, "dollar": 0,
            "rasxod": 0, "bloklar": []
        }

        # ═══ BLOKLARNI TOPISH ═══
        # Har sheet da 4-6 ta blok bo'ladi
        # Blok boshi: sarlavha qatori (B=Полученный товар...)
        # Keyin jami qator, keyin tafsilotlar

        blok_starts = []
        for r in range(1, ws.max_row + 1):
            b = ws.cell(row=r, column=2).value
            if b and isinstance(b, str) and "олученн" in b:
                blok_starts.append(r)

        for bi, header_row in enumerate(blok_starts):
            summary_row = header_row + 1
            # Agent nomini aniqlash
            agent_name = _detect_agent(ws, summary_row + 1)
            
            # Jami qator qiymatlari
            vals = {}
            for col, key in [(2,"tovar"),(3,"qarz"),(4,"qarz_tolov"),
                             (5,"qaytarish"),(6,"sales"),(7,"rasxod"),(9,"nakd")]:
                v = ws.cell(row=summary_row, column=col).value
                vals[key] = float(v) if isinstance(v, (int, float)) else 0

            # Nakd pul (I ustuni, yaxlitlangan)
            nakd_round = ws.cell(row=summary_row + 1, column=9).value
            nakd = float(nakd_round) if isinstance(nakd_round, (int, float)) and nakd_round > 0 else vals["nakd"]
            if nakd < 0:
                nakd = 0

            blok_info = {
                "agent": agent_name,
                "tovar": vals["tovar"],
                "qarz": vals["qarz"],
                "qarz_tolov": vals["qarz_tolov"],
                "qaytarish": vals["qaytarish"],
                "sales": vals["sales"],
                "rasxod": vals["rasxod"],
                "nakd": nakd,
            }
            kun_data["bloklar"].append(blok_info)
            kun_data["savdo"] += vals["sales"]
            kun_data["nakd"] += nakd

            # Agent statistika
            if agent_name and agent_name != "?" and vals["tovar"] > 0:
                if agent_name not in result["agentlar"]:
                    result["agentlar"][agent_name] = {
                        "tovar": 0, "nakd": 0, "sales": 0, "kunlar": 0
                    }
                result["agentlar"][agent_name]["tovar"] += vals["tovar"]
                result["agentlar"][agent_name]["nakd"] += nakd
                result["agentlar"][agent_name]["sales"] += vals["sales"]
                result["agentlar"][agent_name]["kunlar"] += 1

            # ═══ RASXODLAR TAHLILI ═══
            next_header = blok_starts[bi+1] if bi+1 < len(blok_starts) else (ws.max_row + 1)
            for r in range(summary_row + 1, min(next_header, summary_row + 25)):
                g = ws.cell(row=r, column=7).value  # Rasxod
                h = ws.cell(row=r, column=8).value  # Izoh
                if not g or not isinstance(g, (int, float)) or g <= 0:
                    continue
                if not h or not isinstance(h, str):
                    continue

                g_val = float(g)
                h_low = h.lower().strip()

                # Kategoriyalash
                kat = _categorize_expense(h_low)

                result["rasxodlar_list"].append({
                    "kun": sn, "agent": agent_name,
                    "summa": g_val, "izoh": h.strip(), "kat": kat
                })

                if kat not in result["rasxodlar_turi"]:
                    result["rasxodlar_turi"][kat] = 0
                result["rasxodlar_turi"][kat] += g_val

                # Maxsus kategoriyalar
                if kat == "dollar":
                    result["jami"]["dollar_sum"] += g_val
                elif kat == "karta":
                    result["jami"]["karta"] += g_val
                elif kat == "perechisleniye":
                    result["jami"]["perechisleniye"] += g_val
                    result["perechisleniya"].append({
                        "kun": sn, "agent": agent_name, "summa": g_val, "izoh": h.strip()
                    })
                elif kat == "klik":
                    result["jami"]["klik"] += g_val
                elif kat == "oylik":
                    result["oyliklar"].append({
                        "kun": sn, "xodim": h.strip(), "summa": g_val, "agent": agent_name
                    })
                    result["jami"]["oylik"] += g_val

        # ═══ ИТОГО BO'LIMI ═══
        itogo_nakd = 0
        itogo_dollar = 0
        itogo_rasxod = 0
        for r in range(max(1, ws.max_row - 40), ws.max_row + 1):
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            e = ws.cell(row=r, column=5).value

            if b and isinstance(b, str) and "итого" in b.lower():
                itogo_nakd = float(c) if isinstance(c, (int, float)) else 0
                itogo_dollar = float(d) if isinstance(d, (int, float)) else 0
                itogo_rasxod = float(e) if isinstance(e, (int, float)) else 0
                break

            # Oylik/rasxod yozuvlari (итого bo'limida)
            if b and isinstance(b, str) and c is None and e and isinstance(e, (int, float)) and e > 0:
                h_text = b.strip()
                if "ойлик" in h_text.lower() or (isinstance(ws.cell(row=r, column=6).value, str) and "ойлик" in ws.cell(row=r, column=6).value.lower()):
                    result["oyliklar"].append({
                        "kun": sn, "xodim": h_text, "summa": float(e), "agent": "итого"
                    })
                    result["jami"]["oylik"] += float(e)

            # Dollar (итого da D ustuni)
            if b and isinstance(b, str) and d and isinstance(d, (int, float)) and d > 0:
                if "итого" not in b.lower():
                    result["dollar_ops"].append({
                        "kun": sn, "xodim": b.strip(), "dollar": float(d)
                    })

        if itogo_nakd > 0:
            kun_data["savdo"] = itogo_nakd
            kun_data["dollar"] = itogo_dollar
            kun_data["rasxod"] = itogo_rasxod
            kun_data["has_itogo"] = True
        else:
            kun_data["has_itogo"] = False

        result["kunlar"].append(kun_data)
        # Faqat итого bor sheetlarni jami savdoga qo'shish
        # (итого yo'q sheetlar — sub-sheet, allaqachon asosiy sheetda hisoblangan)
        if kun_data["has_itogo"]:
            result["jami"]["savdo"] += kun_data["savdo"]
            result["jami"]["dollar"] += kun_data["dollar"]
            result["jami"]["rasxod"] += itogo_rasxod

    # Nakd pul jami
    result["jami"]["nakd"] = result["jami"]["savdo"] - result["jami"]["rasxod"]

    wb.close()
    return result


def _detect_agent(ws, name_row: int) -> str:
    """Agent nomini aniqlash — harflardan"""
    letters = []
    for r in range(name_row, min(name_row + 8, ws.max_row + 1)):
        b = ws.cell(row=r, column=2).value
        if b and isinstance(b, str) and len(b.strip()) <= 2:
            letters.append(b.strip())
        else:
            break
    name = "".join(letters).strip()
    
    agent_map = {
        "ФАРХОД": "Фарход", "ФАРХО": "Фарход", "ФАРХ": "Фарход",
        "ЗАФАР": "Зафар", "ЗАФА": "Зафар",
        "МИРЖАФАР": "Миржафар", "МИРЖА": "Миржафар", "МИРЖ": "Миржафар",
        "ИСКАНДАР": "Искандар", "ИСКАНДА": "Искандар", "ИСКАН": "Искандар",
        "МАШРАБ": "Машраб", "МАШРА": "Машраб",
        "УМИД": "Умид",
    }
    return agent_map.get(name, name if name else "?")


def _categorize_expense(h: str) -> str:
    """Rasxod kategoriyasi"""
    if "ойлик" in h or "расчёт" in h or "расчет" in h:
        return "oylik"
    if "скидк" in h:
        return "skidka"
    if "газ" in h or "бензин" in h or "заправк" in h:
        return "gaz"
    if "обед" in h or "ужин" in h or "тушлик" in h:
        return "obed"
    if "карт" in h and "визит" not in h:
        return "karta"
    if "$" in h or "доллар" in h:
        return "dollar"
    if "перечисл" in h:
        return "perechisleniye"
    if "клик" in h or "click" in h:
        return "klik"
    if "тачк" in h or "арав" in h:
        return "tachka"
    if "стоянк" in h:
        return "stoyanka"
    if "колдик" in h or "колдиг" in h or "қолдиқ" in h:
        return "qoldiq"
    if "возврат" in h:
        return "vozvrat"
    if "мойк" in h or "ремонт" in h or "балон" in h:
        return "mashina"
    return "boshqa"


def format_analysis_for_ai(analysis: dict) -> str:
    """Tahlil natijalarini AI uchun matn formatga aylantirish"""
    j = analysis["jami"]
    lines = []
    
    lines.append("═══ EXCEL TAHLIL NATIJALARI ═══")
    lines.append(f"Fayl: {analysis['filename']}")
    lines.append(f"Sheetlar: {analysis['sheets']}")
    lines.append("")
    
    # 1. UMUMIY
    lines.append("1. UMUMIY KO'RSATKICHLAR:")
    lines.append(f"   Jami savdo (итого): {j['savdo']:,.0f} so'm")
    lines.append(f"   Jami dollar: {j['dollar']:,.0f} $")
    lines.append(f"   Jami rasxod: {j['rasxod']:,.0f} so'm")
    lines.append(f"   Sof nakd pul: {j['nakd']:,.0f} so'm")
    lines.append(f"   Bank karta: {j['karta']:,.0f} so'm")
    lines.append(f"   Click.uz: {j['klik']:,.0f} so'm")
    lines.append(f"   Perechisleniye: {j['perechisleniye']:,.0f} so'm")
    lines.append(f"   Jami oylik: {j['oylik']:,.0f} so'm")
    lines.append("")
    
    # 2. KUNLIK
    lines.append("2. KUNLIK SAVDO (itogo bo'limidan):")
    for k in analysis["kunlar"]:
        if k["savdo"] > 0:
            lines.append(f"   {k['kun']}: savdo={k['savdo']:,.0f} | dollar={k['dollar']:,.0f}$ | rasxod={k['rasxod']:,.0f}")
    lines.append("")
    
    # 3. RASXOD KATEGORIYALARI
    lines.append("3. RASXOD KATEGORIYALARI:")
    kat_names = {
        "skidka": "Skidka (chegirma)", "gaz": "Gaz/benzin", "obed": "Obed/ужин",
        "karta": "Bank karta to'lov", "dollar": "Dollar ($) operatsiya",
        "oylik": "Oylik (maosh)", "perechisleniye": "Perechisleniye (o'tkazma)",
        "klik": "Click.uz to'lov", "tachka": "Tachka/arava",
        "stoyanka": "Stoyanka", "qoldiq": "Qoldiq (farq)", "vozvrat": "Qaytarish",
        "mashina": "Mashina ta'mir/moyка", "boshqa": "Boshqa xarajatlar"
    }
    for kat, summa in sorted(analysis["rasxodlar_turi"].items(), key=lambda x: -x[1]):
        name = kat_names.get(kat, kat)
        lines.append(f"   {name}: {summa:,.0f} so'm")
    lines.append(f"   JAMI: {j['rasxod']:,.0f} so'm")
    lines.append("")
    
    # 4. OYLIKLAR (dublikatlarni birlashtirish)
    lines.append("4. OYLIK TO'LOVLARI (birma-bir):")
    oylik_merged = defaultdict(lambda: {"summa": 0, "soni": 0, "kunlar": []})
    for o in analysis["oyliklar"]:
        # Ismni normalizatsiya
        key = _normalize_name(o["xodim"])
        oylik_merged[key]["summa"] += o["summa"]
        oylik_merged[key]["soni"] += 1
        oylik_merged[key]["kunlar"].append(o["kun"])
    for xodim, data in sorted(oylik_merged.items(), key=lambda x: -x[1]["summa"]):
        kunlar_str = ", ".join(data["kunlar"][:5])
        lines.append(f"   {xodim}: {data['summa']:,.0f} so'm ({data['soni']} to'lov, kunlar: {kunlar_str})")
    lines.append(f"   JAMI OYLIK: {j['oylik']:,.0f} so'm")
    lines.append("")
    
    # 5. DOLLAR (faqat haqiqiy dollarlar - 10000 dan kichik)
    lines.append("5. DOLLAR OPERATSIYALAR:")
    real_dollars = [d for d in analysis["dollar_ops"] if d["dollar"] < 10000]
    dollar_by_person = defaultdict(float)
    for d in real_dollars:
        dollar_by_person[d["xodim"]] += d["dollar"]
    for xodim, dollar in sorted(dollar_by_person.items(), key=lambda x: -x[1]):
        if dollar > 0:
            lines.append(f"   {xodim}: {dollar:,.0f} $")
    lines.append(f"   JAMI DOLLAR: {j['dollar']:,.0f} $")
    lines.append("")
    
    # 6. PERECHISLENIYA
    lines.append("6. PERECHISLENIYA (pul o'tkazma):")
    for p in analysis["perechisleniya"]:
        lines.append(f"   {p['kun']}: {p['agent']} — {p['summa']:,.0f} so'm")
    lines.append(f"   JAMI: {j['perechisleniye']:,.0f} so'm")
    lines.append("")
    
    # 7. AGENTLAR
    lines.append("7. AGENTLAR BO'YICHA:")
    for agent, data in sorted(analysis["agentlar"].items(), key=lambda x: -x[1]["sales"]):
        lines.append(f"   {agent}: tovar={data['tovar']:,.0f} | nakd={data['nakd']:,.0f} | {data['kunlar']} kun ishlagan")
    lines.append("")
    
    # 8. GAZ XARAJATLARI (tafsilot)
    lines.append("8. GAZ XARAJATLARI (kunlik):")
    gaz_items = [r for r in analysis["rasxodlar_list"] if r["kat"] == "gaz"]
    gaz_total = 0
    for g in gaz_items:
        lines.append(f"   {g['kun']}: {g['agent']} — {g['summa']:,.0f} so'm")
        gaz_total += g["summa"]
    lines.append(f"   JAMI GAZ: {gaz_total:,.0f} so'm")
    lines.append("")
    
    # 9. OBED XARAJATLARI
    lines.append("9. OBED XARAJATLARI:")
    obed_items = [r for r in analysis["rasxodlar_list"] if r["kat"] == "obed"]
    obed_total = sum(o["summa"] for o in obed_items)
    lines.append(f"   JAMI OBED: {obed_total:,.0f} so'm ({len(obed_items)} marta)")
    
    return "\n".join(lines)


def _normalize_name(name: str) -> str:
    """Xodim ismini normalizatsiya — dublikatlarni birlashtirish"""
    n = name.lower().strip()
    # Ойлик so'zini olib tashlash
    n = n.replace("ойлик", "").replace("расчёт", "").replace("расчет", "").strip()
    # Ака/опа olib tashlash
    n = n.replace(" ака", "").replace(" опа", "").strip()
    
    # Tanilgan ismlar
    name_map = {
        "фарход": "Фарход",
        "зафар": "Зафар", 
        "искандар": "Искандар", "искаандар": "Искандар",
        "нодира": "Нодира",
        "миржафар": "Миржафар",
        "машраб": "Машраб",
        "дилдора": "Дилдора",
        "рахматилло": "Рахматилло",
        "умид": "Умид",
        "алимардон": "Алимардон",
        "насриддин": "Насриддин ака",
        "нариддин": "Насриддин ака",
        "шахноза": "Шахноза апа",
        "дилшод сладус агент": "Дилшод (агент)",
        "сурайё": "Сурайё",
        "давлат тричап": "Давлат (тричап)",
        "тричап агентлари": "Тричап агентлари",
        "ватика агентлар": "Ватика агентлари",
        "клик нодира": "Нодира (Click)",
    }
    
    for key, val in name_map.items():
        if key in n:
            return val
    
    # Birinchi harfni katta qilish
    return name.strip().capitalize() if name.strip() else name
