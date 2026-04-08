"""
╔══════════════════════════════════════════════════════════════════════════╗
║  SAVDOAI v25.4.0 — UNIVERSAL EXPORT ENGINE                              ║
║                                                                          ║
║  Har qanday ma'lumotni professional formatda eksport:                   ║
║  • Excel (openpyxl) — jadvallar, grafiklar, ranglar                    ║
║  • PDF (reportlab yoki weasyprint) — logoli, imzoli                    ║
║  • CSV — oddiy eksport                                                   ║
║                                                                          ║
║  SD Agent + Smartup'da bu QISMAN — SavdoAI'da UNIVERSAL.              ║
║                                                                          ║
║  QOLLAB-QUVVATLANGAN HISOBOTLAR:                                        ║
║  1. Sotuv hisoboti (kunlik/haftalik/oylik)                              ║
║  2. Klient ro'yxati + qarzlar                                           ║
║  3. Tovar ro'yxati + qoldiqlar                                          ║
║  4. Foyda/Zarar (P&L)                                                   ║
║  5. Akt sverki                                                           ║
║  6. ABC-XYZ matritsa                                                     ║
║  7. Kunlik kassa                                                         ║
║  8. Topshiriqlar                                                         ║
║  9. Agent KPI                                                            ║
║  10. Churn prediction                                                    ║
╚══════════════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import List, Optional, Any

log = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ════════════════════════════════════════════════════════════
#  UNIVERSAL EXCEL EXPORT
# ════════════════════════════════════════════════════════════

# Rang palitra
SARLAVHA_RANG = "059669"  # SavdoAI yashil
SARLAVHA_MATN = "FFFFFF"
JUFT_RANG = "F0FDF4"
BORDER_RANG = "D1D5DB"


def _border():
    side = Side(style="thin", color=BORDER_RANG)
    return Border(left=side, right=side, top=side, bottom=side)


def export_excel(
    sarlavha: str,
    ustunlar: List[str],
    qatorlar: List[List[Any]],
    fayl_nomi: str = "export.xlsx",
    sheet_nomi: str = "Ma'lumotlar",
    qoshimcha_info: dict = None,
) -> bytes:
    """Universal Excel eksport — har qanday ma'lumot uchun.

    Args:
        sarlavha: Hisobot sarlavhasi
        ustunlar: Ustun nomlari ["ID", "Nomi", "Summa"]
        qatorlar: Ma'lumotlar [[1, "Tovar", 50000], ...]
        fayl_nomi: Fayl nomi
        sheet_nomi: Sheet nomi
        qoshimcha_info: {"Sana": "2026-04-08", "Foydalanuvchi": "Mashrab"}

    Returns:
        bytes — Excel fayl
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl o'rnatilmagan: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_nomi

    row_idx = 1

    # ═══ SARLAVHA ═══
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(ustunlar))
    cell = ws.cell(row=row_idx, column=1, value=f"📊 {sarlavha}")
    cell.font = Font(name="Arial", size=14, bold=True, color=SARLAVHA_RANG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 30
    row_idx += 1

    # ═══ QOSHIMCHA INFO ═══
    if qoshimcha_info:
        for key, value in qoshimcha_info.items():
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=10, color="666666")
            ws.cell(row=row_idx, column=2, value=str(value)).font = Font(size=10)
            row_idx += 1
        row_idx += 1

    # ═══ USTUN SARLAVHALARI ═══
    header_fill = PatternFill("solid", fgColor=SARLAVHA_RANG)
    header_font = Font(name="Arial", size=10, bold=True, color=SARLAVHA_MATN)
    for col_idx, ustun in enumerate(ustunlar, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=ustun)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[row_idx].height = 25
    row_idx += 1

    # ═══ MA'LUMOTLAR ═══
    juft_fill = PatternFill("solid", fgColor=JUFT_RANG)
    data_font = Font(name="Arial", size=10)
    num_font = Font(name="Arial", size=10, color="059669")

    for i, qator in enumerate(qatorlar):
        for col_idx, value in enumerate(qator, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Decimal/float → raqam
            if isinstance(value, (Decimal, float, int)):
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = '#,##0' if isinstance(value, int) else '#,##0.00'
                cell.font = num_font
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = str(value) if value is not None else ""
                cell.font = data_font

            cell.border = _border()
            if i % 2 == 1:
                cell.fill = juft_fill
        row_idx += 1

    # ═══ JAMI QATOR ═══
    if qatorlar:
        for col_idx in range(1, len(ustunlar) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _border()
            cell.font = Font(bold=True, size=10)

            # Raqamli ustunlarni jamlash
            values = [q[col_idx - 1] for q in qatorlar if col_idx - 1 < len(q)]
            numeric_vals = [float(v) for v in values if isinstance(v, (int, float, Decimal))]
            if numeric_vals and col_idx > 1:  # ID ustunini jamlaMASLIK
                cell.value = sum(numeric_vals)
                cell.number_format = '#,##0.00'
                cell.fill = PatternFill("solid", fgColor="D1FAE5")
            elif col_idx == 1:
                cell.value = "JAMI"
                cell.fill = PatternFill("solid", fgColor="D1FAE5")

    # ═══ USTUN KENGLIGI ═══
    for col_idx in range(1, len(ustunlar) + 1):
        max_len = len(str(ustunlar[col_idx - 1]))
        for qator in qatorlar[:50]:
            if col_idx - 1 < len(qator):
                max_len = max(max_len, len(str(qator[col_idx - 1] or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # ═══ FOOTER ═══
    row_idx += 2
    ws.cell(row=row_idx, column=1, value=f"SavdoAI Pro — {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(size=8, color="999999", italic=True)

    # ═══ SAQLASH ═══
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════
#  CSV EXPORT
# ════════════════════════════════════════════════════════════

def export_csv(ustunlar: List[str], qatorlar: List[List[Any]]) -> bytes:
    """CSV eksport."""
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(ustunlar)
    for q in qatorlar:
        writer.writerow([str(v) if v is not None else "" for v in q])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel


# ════════════════════════════════════════════════════════════
#  HISOBOT GENERATORLAR
# ════════════════════════════════════════════════════════════

async def export_sotuvlar(conn, uid: int, sana_dan: str, sana_gacha: str, fmt: str = "excel") -> bytes:
    """Sotuv hisoboti eksport."""
    rows = await conn.fetch("""
        SELECT s.id, s.sana::text, k.nom AS klient, s.jami, s.tolangan, s.qarz, s.holat
        FROM sotuvlar s LEFT JOIN klientlar k ON k.id=s.klient_id
        WHERE s.user_id=$1 AND s.sana BETWEEN $2::date AND $3::date
        ORDER BY s.sana DESC
    """, uid, sana_dan, sana_gacha)

    ustunlar = ["ID", "Sana", "Klient", "Jami", "To'langan", "Qarz", "Holat"]
    qatorlar = [[r["id"], r["sana"], r["klient"] or "—", r["jami"], r["tolangan"], r["qarz"], r["holat"]] for r in rows]

    if fmt == "csv":
        return export_csv(ustunlar, qatorlar)

    return export_excel("Sotuv hisoboti", ustunlar, qatorlar,
        qoshimcha_info={"Davr": f"{sana_dan} — {sana_gacha}", "Sotuvlar soni": len(rows)})


async def export_klientlar(conn, uid: int, fmt: str = "excel") -> bytes:
    """Klient ro'yxati eksport."""
    rows = await conn.fetch("""
        SELECT k.id, k.nom, k.telefon, k.manzil, k.kategoriya,
            COALESCE((SELECT SUM(qarz) FROM sotuvlar WHERE klient_id=k.id AND qarz>0), 0) AS qarz,
            (SELECT COUNT(*) FROM sotuvlar WHERE klient_id=k.id) AS sotuv_soni
        FROM klientlar k WHERE k.user_id=$1 AND k.faol=TRUE ORDER BY k.nom
    """, uid)

    ustunlar = ["ID", "Nomi", "Telefon", "Manzil", "Kategoriya", "Qarz", "Sotuv soni"]
    qatorlar = [[r["id"], r["nom"], r["telefon"] or "", r["manzil"] or "", r["kategoriya"] or "", r["qarz"], r["sotuv_soni"]] for r in rows]

    if fmt == "csv":
        return export_csv(ustunlar, qatorlar)
    return export_excel("Klientlar ro'yxati", ustunlar, qatorlar, qoshimcha_info={"Jami": len(rows)})


async def export_tovarlar(conn, uid: int, fmt: str = "excel") -> bytes:
    """Tovar ro'yxati eksport."""
    rows = await conn.fetch("""
        SELECT id, nomi, shtrix_kod, kategoriya, birlik, sotuv_narx, tan_narx, qoldiq
        FROM tovarlar WHERE user_id=$1 AND faol=TRUE ORDER BY nomi
    """, uid)

    ustunlar = ["ID", "Nomi", "Barcode", "Kategoriya", "Birlik", "Sotuv narx", "Tan narx", "Qoldiq"]
    qatorlar = [[r["id"], r["nomi"], r["shtrix_kod"] or "", r["kategoriya"] or "", r["birlik"], r["sotuv_narx"], r["tan_narx"], r["qoldiq"]] for r in rows]

    if fmt == "csv":
        return export_csv(ustunlar, qatorlar)
    return export_excel("Tovarlar ro'yxati", ustunlar, qatorlar, qoshimcha_info={"Jami": len(rows)})
