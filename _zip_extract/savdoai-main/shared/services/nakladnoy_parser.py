"""
╔══════════════════════════════════════════════════════════════════╗
║  SAVDOAI v25.3.2 — NAKLADNOY PARSER (MEGA SCALE)                ║
║  10,000+ nakladnoy | 5,000+ tovar | 100,000+ qator             ║
║  read_only=True | O(n) single-pass | defaultdict O(1)          ║
╚══════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io, logging
from collections import defaultdict
log = logging.getLogger(__name__)


def nakladnoy_ekanligini_tekshir(data: bytes) -> bool:
    """Tezkor — faqat 20 qator o'qiydi."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            n = 0
            for row in ws.iter_rows(values_only=True):
                n += 1
                if n > 20: break
                if any(v and "Накладная" in str(v) for v in row):
                    wb.close(); return True
        wb.close()
    except Exception:
        pass
    return False


def nakladnoy_tahlil(data: bytes, max_rows: int = 500_000) -> dict:
    """
    MEGA SCALE parser — 10K nakladnoy, 5K tovar, 500K qator.
    Single-pass O(n), read_only stream, defaultdict O(1).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"xato": "openpyxl kerak"}
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        return {"xato": f"Excel: {e}"}

    invoices = []
    cur = None
    cur_items = 0
    cur_jami = 0.0
    tovar_map = defaultdict(lambda: [0.0, 0.0, 0])  # [miqdor, jami, count]
    hududlar = set()
    tp_set = set()
    firma = ""
    sana = ""
    total_rows = 0
    qarz_list = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            total_rows += 1
            if total_rows > max_rows: break

            cells = list(row) + [None] * 6
            a, b, c, d, e, f = cells[:6]

            # ── YANGI NAKLADNOY? ──
            nakl_found = False
            for v in (a, b, c, d, e, f):
                if v and "Накладная" in str(v):
                    if cur is not None and cur_items > 0:
                        cur["soni"] = cur_items
                        cur["jami"] = cur_jami
                        invoices.append(cur)
                    cur = {"klient":"","jami":0,"soni":0,"hudud":"","tp":"","tel":"","balans":""}
                    cur_items = 0; cur_jami = 0
                    s = str(v)
                    if "от " in s:
                        try: sana = s.split("от")[1].strip()[:10]
                        except Exception: pass
                    nakl_found = True; break
            if nakl_found or cur is None:
                continue

            # ── METADATA ──
            bs = str(b) if b else ""
            es = str(e) if e else ""

            if "Кому:" in bs:
                cur["klient"] = bs.split("Кому:")[1].strip()[:60]
            elif bs.startswith("Тел:") and "998" in bs:
                cur["tel"] = bs.replace("Тел:","").strip()
            elif "Баланс" in bs and ":" in bs:
                bal = bs.split(":")[1].strip()
                cur["balans"] = bal
                if "-" in bal and cur["klient"]:
                    qarz_list.append((cur["klient"][:30], bal))

            if "Территория:" in es:
                h = es.split(":")[1].strip()
                cur["hudud"] = h; hududlar.add(h)
            elif "ТП:" in es:
                t = es.split(":")[1].strip()
                cur["tp"] = t; tp_set.add(t)
            elif "Фирма:" in es and not firma:
                firma = es.split(":")[1].strip()

            # ── TOVAR QATORI ──
            if a is not None and b and c is not None:
                a_s = str(a).strip()
                if a_s.isdigit() and int(a_s) > 0:
                    try:
                        miqdor = float(c)
                        narx = float(e) if e else 0
                        jami_t = float(f) if f else miqdor * narx
                        nomi = str(b).strip()[:60]
                        tm = tovar_map[nomi]
                        tm[0] += miqdor; tm[1] += jami_t; tm[2] += 1
                        cur_items += 1; cur_jami += jami_t
                    except (ValueError, TypeError):
                        pass

            # ── ITOG ──
            if b and "Итог" in str(b) and f:
                try: cur_jami = float(f)
                except (ValueError, TypeError): pass

    if cur is not None and cur_items > 0:
        cur["soni"] = cur_items; cur["jami"] = cur_jami
        invoices.append(cur)
    wb.close()

    return {
        "nakladnoylar": invoices,
        "jami_summa": sum(inv["jami"] for inv in invoices),
        "jami_soni": len(invoices),
        "jami_pozitsiya": sum(inv["soni"] for inv in invoices),
        "tovar_xillari": len(tovar_map),
        "tovarlar": dict(tovar_map),
        "hududlar": sorted(hududlar),
        "tp": sorted(tp_set),
        "firma": firma,
        "sana": sana,
        "qator_soni": total_rows,
        "qarz_klientlar": qarz_list[:100],
    }


def nakladnoy_xulosa_matn(data: dict, fayl_nomi: str = "") -> str:
    """Bot xulosa — 4096 char limit, 10K+ nakladnoy uchun."""
    if "xato" in data: return f"❌ {data['xato']}"
    if data["jami_soni"] == 0: return "📋 Nakladnoy topilmadi."

    m = f"📋 *NAKLADNOY TAHLILI*\n━━━━━━━━━━━━━━━━━━━━━\n"
    if fayl_nomi: m += f"📁 {fayl_nomi[:40]}\n"
    if data["sana"]: m += f"📅 {data['sana']}\n"
    if data["firma"]: m += f"🏢 {data['firma'][:40]}\n"
    if data["hududlar"]: m += f"📍 {', '.join(data['hududlar'][:5])}\n"
    if data["tp"]: m += f"👤 TP: {', '.join(data['tp'][:5])}\n"

    m += (
        f"\n📊 *{data['jami_soni']:,}* nakladnoy"
        f" | *{data['tovar_xillari']:,}* xil tovar\n"
        f"💰 *JAMI: {data['jami_summa']:,.0f}* so'm\n"
        f"📄 {data.get('qator_soni',0):,} qator o'qildi\n"
    )

    # Top klientlar
    top_k = sorted(data["nakladnoylar"], key=lambda x: -x["jami"])
    m += "\n👑 *TOP KLIENTLAR:*\n"
    for i, inv in enumerate(top_k[:7], 1):
        k = inv["klient"][:28] if inv["klient"] else "—"
        m += f"  {i}. {k}: *{inv['jami']:,.0f}*\n"
    if len(top_k) > 7:
        m += f"  _...va yana {len(top_k)-7:,} ta_\n"

    # Top tovarlar
    top_t = sorted(data["tovarlar"].items(), key=lambda x: -x[1][1])
    m += "\n🏆 *TOP TOVARLAR:*\n"
    for i, (nomi, vals) in enumerate(top_t[:7], 1):
        m += f"  {i}. {nomi[:30]}: {vals[0]:,.0f} — *{vals[1]:,.0f}*\n"
    if len(top_t) > 7:
        m += f"  _...va yana {len(top_t)-7:,} xil_\n"

    # Qarzli
    qarz = data.get("qarz_klientlar", [])
    if qarz:
        m += f"\n⚠️ *Qarzli: {len(qarz)} ta*\n"
        for k, bal in qarz[:3]:
            m += f"  📝 {k}: {bal}\n"
        if len(qarz) > 3:
            m += f"  _...va yana {len(qarz)-3} ta_\n"

    if len(m) > 4000:
        m = m[:3950] + "\n\n_...qisqartirildi_"
    return m
