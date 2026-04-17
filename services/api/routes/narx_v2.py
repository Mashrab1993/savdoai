"""
╔══════════════════════════════════════════════════════════════════════╗
║  SAVDOAI v25.4.0 — NARX TURLARI v2 API                              ║
║                                                                      ║
║  SalesDoc /settings/priceType + /settings/prices funksiyalari:       ║
║  • Narx turi: kod, nomi, turi (prodaja/zakup/prayslist), tavsif     ║
║  • Narxlar: 3 tab (sotish, olish, prayslist)                        ║
║  • Bulk markup (naenka): foiz bilan hammasiga qo'llash              ║
║  • Excel shablon download + import                                  ║
╚══════════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from shared.database.pool import rls_conn
from services.api.deps import get_uid

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/narx", tags=["Narx v2"])

ALLOWED_TURI = {"prodaja", "zakup", "prayslist"}


class NarxTuriIn(BaseModel):
    nomi: str = Field(..., min_length=1, max_length=100)
    kod: Optional[str] = Field(None, max_length=50)
    turi: str = Field("prodaja")
    tavsif: Optional[str] = None
    tolov_usuli: Optional[str] = Field(None, max_length=50)
    foiz_chegirma: float = 0
    klient_turi_id: Optional[int] = None
    tartib: int = 0
    faol: bool = True


@router.get("/turlari")
async def list_narx_turlari(
    turi: Optional[str] = Query(None),
    faol: Optional[bool] = Query(None),
    uid: int = Depends(get_uid),
):
    """Barcha narx turlari (SalesDoc /settings/priceType).

    Qaytaradi: har bir turi alohida. turi bermasa — barchasi.
    """
    if turi and turi not in ALLOWED_TURI:
        raise HTTPException(400, f"turi '{turi}' noto'g'ri")

    where = ["user_id = $1"]
    params: list = [uid]
    if turi:
        params.append(turi)
        where.append(f"turi = ${len(params)}")
    if faol is not None:
        params.append(faol)
        where.append(f"faol = ${len(params)}")

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT nt.*,
                   kt.nomi AS klient_turi_nomi,
                   (SELECT COUNT(*) FROM tovar_narxlari tn WHERE tn.narx_turi_id = nt.id) AS tovar_soni
            FROM narx_turlari nt
            LEFT JOIN klient_turlari kt ON kt.id = nt.klient_turi_id
            WHERE {' AND '.join(where)}
            ORDER BY nt.tartib, nt.nomi
        """, *params)

    return {
        "items": [
            {
                "id": r["id"],
                "kod": r["kod"],
                "nomi": r["nomi"],
                "turi": r["turi"],
                "tavsif": r["tavsif"],
                "tolov_usuli": r["tolov_usuli"],
                "foiz_chegirma": float(r["foiz_chegirma"] or 0),
                "klient_turi_id": r["klient_turi_id"],
                "klient_turi_nomi": r["klient_turi_nomi"],
                "tartib": r["tartib"],
                "faol": r["faol"],
                "tovar_soni": int(r["tovar_soni"] or 0),
                "oxirgi_narx_sanasi": r["oxirgi_narx_sanasi"].isoformat() if r["oxirgi_narx_sanasi"] else None,
                "yaratilgan": r["yaratilgan"].isoformat() if r["yaratilgan"] else None,
            }
            for r in rows
        ],
        "jami": len(rows),
    }


@router.post("/turi")
async def create_narx_turi(body: NarxTuriIn, uid: int = Depends(get_uid)):
    if body.turi not in ALLOWED_TURI:
        raise HTTPException(400, f"turi '{body.turi}' noto'g'ri")
    async with rls_conn(uid) as c:
        dup = await c.fetchval("""
            SELECT id FROM narx_turlari
            WHERE user_id=$1 AND LOWER(nomi)=LOWER($2) AND turi=$3
        """, uid, body.nomi, body.turi)
        if dup:
            raise HTTPException(409, f"'{body.nomi}' ({body.turi}) allaqachon mavjud")
        row = await c.fetchrow("""
            INSERT INTO narx_turlari
                (user_id, nomi, kod, turi, tavsif, tolov_usuli,
                 foiz_chegirma, klient_turi_id, tartib, faol)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
            RETURNING *
        """, uid, body.nomi, body.kod, body.turi, body.tavsif, body.tolov_usuli,
             body.foiz_chegirma, body.klient_turi_id, body.tartib, body.faol)
    return dict(row) | {
        "yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None,
        "oxirgi_narx_sanasi": row["oxirgi_narx_sanasi"].isoformat() if row["oxirgi_narx_sanasi"] else None,
    }


@router.put("/turi/{nt_id}")
async def update_narx_turi(nt_id: int, body: NarxTuriIn, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        exists = await c.fetchval(
            "SELECT id FROM narx_turlari WHERE id=$1 AND user_id=$2", nt_id, uid,
        )
        if not exists:
            raise HTTPException(404, "Narx turi topilmadi")

        row = await c.fetchrow("""
            UPDATE narx_turlari
            SET nomi=$1, kod=$2, turi=$3, tavsif=$4, tolov_usuli=$5,
                foiz_chegirma=$6, klient_turi_id=$7, tartib=$8, faol=$9
            WHERE id=$10 AND user_id=$11
            RETURNING *
        """, body.nomi, body.kod, body.turi, body.tavsif, body.tolov_usuli,
             body.foiz_chegirma, body.klient_turi_id, body.tartib, body.faol,
             nt_id, uid)
    return dict(row) | {
        "yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None,
        "oxirgi_narx_sanasi": row["oxirgi_narx_sanasi"].isoformat() if row["oxirgi_narx_sanasi"] else None,
    }


@router.delete("/turi/{nt_id}")
async def delete_narx_turi(nt_id: int, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        row = await c.fetchrow(
            "DELETE FROM narx_turlari WHERE id=$1 AND user_id=$2 RETURNING id, nomi, turi",
            nt_id, uid,
        )
    if not row:
        raise HTTPException(404, "Narx turi topilmadi")
    return {"ok": True, "deleted": dict(row)}


# ═══════════════════════════════════════════════════════════════════
# BULK MARKUP — "Ustanovit naenku"
# ═══════════════════════════════════════════════════════════════════

class MarkupIn(BaseModel):
    narx_turi_id: int
    foiz: float = Field(..., description="Markup % (masalan 20 = +20%)")
    faqat_boshsiz: bool = Field(True, description="Faqat narxi yo'q tovarlarga qo'llash")


@router.post("/markup")
async def apply_markup(body: MarkupIn, uid: int = Depends(get_uid)):
    """Narx turi uchun bazaviy (olish_narxi)ga foiz qo'shib narx o'rnatish.

    SalesDoc "Установить наценку" funksiyasi.
    """
    async with rls_conn(uid) as c:
        nt = await c.fetchval(
            "SELECT id FROM narx_turlari WHERE id=$1 AND user_id=$2",
            body.narx_turi_id, uid,
        )
        if not nt:
            raise HTTPException(404, "Narx turi topilmadi")

        where = "ss.user_id=$1" if not body.faqat_boshsiz else """
            ss.user_id=$1 AND NOT EXISTS (
                SELECT 1 FROM tovar_narxlari tn
                WHERE tn.tovar_id=ss.id AND tn.narx_turi_id=$2
            )
        """
        if body.faqat_boshsiz:
            rows = await c.fetch(f"""
                SELECT ss.id, ss.nomi, COALESCE(ss.olish_narxi, 0) AS bazaviy
                FROM tovarlar ss
                WHERE {where}
            """, uid, body.narx_turi_id)
        else:
            rows = await c.fetch(f"""
                SELECT ss.id, ss.nomi, COALESCE(ss.olish_narxi, 0) AS bazaviy
                FROM tovarlar ss
                WHERE {where}
            """, uid)

        qoshildi = 0
        async with c.transaction():
            for r in rows:
                bazaviy = float(r["bazaviy"] or 0)
                if bazaviy <= 0:
                    continue
                yangi_narx = round(bazaviy * (1 + body.foiz / 100), 2)
                await c.execute("""
                    INSERT INTO tovar_narxlari (user_id, tovar_id, narx_turi_id, narx)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (tovar_id, narx_turi_id)
                    DO UPDATE SET narx = EXCLUDED.narx, yangilangan = NOW()
                """, uid, r["id"], body.narx_turi_id, yangi_narx)
                qoshildi += 1

            await c.execute(
                "UPDATE narx_turlari SET oxirgi_narx_sanasi = NOW() WHERE id=$1",
                body.narx_turi_id,
            )

    return {
        "ok": True,
        "narx_turi_id": body.narx_turi_id,
        "tovarlar_soni": qoshildi,
        "foiz": body.foiz,
    }


# ═══════════════════════════════════════════════════════════════════
# USTANOVIT SENY — tovarlar ro'yxati + bulk narx saqlash
# ═══════════════════════════════════════════════════════════════════

@router.get("/tovarlar/{narx_turi_id}")
async def list_tovarlar_narxlari(
    narx_turi_id: int,
    uid: int = Depends(get_uid),
):
    """Berilgan narx turi uchun barcha tovarlar + joriy narxlarini qaytaradi."""
    async with rls_conn(uid) as c:
        nt = await c.fetchrow(
            "SELECT id, nomi, turi FROM narx_turlari WHERE id=$1 AND user_id=$2",
            narx_turi_id, uid,
        )
        if not nt:
            raise HTTPException(404, "Narx turi topilmadi")

        rows = await c.fetch("""
            SELECT t.id, t.nomi, t.birlik,
                   COALESCE(t.olish_narxi, 0) AS olish_narxi,
                   COALESCE(t.sotish_narxi, 0) AS sotish_narxi,
                   COALESCE(tn.narx, 0) AS joriy_narx,
                   tn.id AS narx_yozuvi_id
            FROM tovarlar t
            LEFT JOIN tovar_narxlari tn
              ON tn.tovar_id = t.id AND tn.narx_turi_id = $1
            WHERE t.user_id = $2
            ORDER BY t.nomi
        """, narx_turi_id, uid)

    return {
        "narx_turi": dict(nt),
        "tovarlar": [
            {
                "id": r["id"],
                "nomi": r["nomi"],
                "birlik": r["birlik"],
                "olish_narxi": float(r["olish_narxi"] or 0),
                "sotish_narxi": float(r["sotish_narxi"] or 0),
                "joriy_narx": float(r["joriy_narx"] or 0),
                "bor_yoqligi": bool(r["narx_yozuvi_id"]),
            }
            for r in rows
        ],
    }


class BulkNarxIn(BaseModel):
    narx_turi_id: int
    narxlar: list[dict] = Field(..., description="[{tovar_id, narx}]")


@router.post("/bulk_set")
async def bulk_set_prices(body: BulkNarxIn, uid: int = Depends(get_uid)):
    """SalesDoc "Установить цены" — ko'p tovarga birdaniga narx o'rnatish.

    narxlar ichidagi {tovar_id, narx} lar upsert qilinadi.
    """
    async with rls_conn(uid) as c:
        nt = await c.fetchval(
            "SELECT id FROM narx_turlari WHERE id=$1 AND user_id=$2",
            body.narx_turi_id, uid,
        )
        if not nt:
            raise HTTPException(404, "Narx turi topilmadi")

        saqlangan = 0
        o_chirildi = 0
        async with c.transaction():
            for item in body.narxlar:
                try:
                    tid = int(item.get("tovar_id"))
                    narx_val = item.get("narx")
                except (TypeError, ValueError):
                    continue

                if narx_val is None or narx_val == "":
                    # Narx bo'sh => o'chirish (agar bor bo'lsa)
                    res = await c.execute("""
                        DELETE FROM tovar_narxlari
                        WHERE user_id=$1 AND tovar_id=$2 AND narx_turi_id=$3
                    """, uid, tid, body.narx_turi_id)
                    if "DELETE 1" in res:
                        o_chirildi += 1
                    continue

                try:
                    narx = float(narx_val)
                    if narx < 0:
                        continue
                except (TypeError, ValueError):
                    continue

                await c.execute("""
                    INSERT INTO tovar_narxlari (user_id, tovar_id, narx_turi_id, narx)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (tovar_id, narx_turi_id)
                    DO UPDATE SET narx = EXCLUDED.narx, yangilangan = NOW()
                """, uid, tid, body.narx_turi_id, narx)
                saqlangan += 1

            await c.execute(
                "UPDATE narx_turlari SET oxirgi_narx_sanasi = NOW() WHERE id=$1",
                body.narx_turi_id,
            )

    return {
        "ok": True,
        "narx_turi_id": body.narx_turi_id,
        "saqlandi": saqlangan,
        "o_chirildi": o_chirildi,
    }


# ═══════════════════════════════════════════════════════════════════
# EXCEL TEMPLATE DOWNLOAD + IMPORT
# ═══════════════════════════════════════════════════════════════════

@router.get("/template")
async def download_template(
    narx_turi_id: int = Query(..., description="Qaysi narx turi uchun shablon"),
    uid: int = Depends(get_uid),
):
    """SalesDoc "Скачать шаблон" — narx turi uchun Excel shablon yuklab olish."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl mavjud emas")

    async with rls_conn(uid) as c:
        nt = await c.fetchrow(
            "SELECT id, nomi, turi FROM narx_turlari WHERE id=$1 AND user_id=$2",
            narx_turi_id, uid,
        )
        if not nt:
            raise HTTPException(404, "Narx turi topilmadi")

        tovarlar = await c.fetch("""
            SELECT t.id, t.nomi, t.birlik, COALESCE(t.olish_narxi, 0) AS olish,
                   COALESCE(tn.narx, 0) AS joriy_narx
            FROM tovarlar t
            LEFT JOIN tovar_narxlari tn ON tn.tovar_id = t.id AND tn.narx_turi_id = $1
            WHERE t.user_id = $2
            ORDER BY t.nomi
        """, narx_turi_id, uid)

    wb = Workbook()
    ws = wb.active
    ws.title = (nt["nomi"] or "Narxlar")[:31]

    headers = ["ID", "Tovar nomi", "Birlik", "Olish narxi", f"{nt['nomi']} narx"]
    HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, t in enumerate(tovarlar, 2):
        ws.cell(row=i, column=1, value=t["id"])
        ws.cell(row=i, column=2, value=t["nomi"])
        ws.cell(row=i, column=3, value=t["birlik"] or "")
        ws.cell(row=i, column=4, value=float(t["olish"] or 0))
        ws.cell(row=i, column=5, value=float(t["joriy_narx"] or 0))

    for col in range(1, 6):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"narx_{nt['nomi']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/import")
async def import_prices(
    narx_turi_id: int = Query(...),
    file: UploadFile = File(...),
    uid: int = Depends(get_uid),
):
    """SalesDoc "Импорт цен" — Excel fayldan narxlarni import qilish.

    Format: A=Tovar ID, B=Nomi (ignore), C=Birlik (ignore),
            D=Olish narxi (ignore), E=Yangi narx.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl mavjud emas")

    async with rls_conn(uid) as c:
        nt = await c.fetchval(
            "SELECT id FROM narx_turlari WHERE id=$1 AND user_id=$2",
            narx_turi_id, uid,
        )
        if not nt:
            raise HTTPException(404, "Narx turi topilmadi")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(413, "Fayl juda katta (max 10MB)")

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel noto'g'ri: {e}")

    ws = wb.active
    yangi = 0
    xatolar: list[str] = []

    async with rls_conn(uid) as c:
        async with c.transaction():
            for row_idx in range(2, ws.max_row + 1):
                try:
                    tovar_id = ws.cell(row=row_idx, column=1).value
                    yangi_narx = ws.cell(row=row_idx, column=5).value
                    if not tovar_id or yangi_narx is None:
                        continue
                    try:
                        tid = int(tovar_id)
                        narx = float(yangi_narx)
                    except (ValueError, TypeError):
                        continue
                    if narx < 0:
                        continue

                    await c.execute("""
                        INSERT INTO tovar_narxlari (user_id, tovar_id, narx_turi_id, narx)
                        VALUES ($1, $2, $3, $4)
                        ON CONFLICT (tovar_id, narx_turi_id)
                        DO UPDATE SET narx = EXCLUDED.narx, yangilangan = NOW()
                    """, uid, tid, narx_turi_id, narx)
                    yangi += 1
                except Exception as e:
                    xatolar.append(f"Qator {row_idx}: {e}")
                    if len(xatolar) > 20:
                        break

            await c.execute(
                "UPDATE narx_turlari SET oxirgi_narx_sanasi = NOW() WHERE id=$1",
                narx_turi_id,
            )

    return {"ok": True, "import_qilindi": yangi, "xatolar": xatolar}
