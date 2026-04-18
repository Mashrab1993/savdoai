"""
╔══════════════════════════════════════════════════════════════╗
║  SAVDOAI — TOVARLAR CRUD ROUTELARI                          ║
║  main.py dan ajratildi — v25.3.2                            ║
╚══════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io
import base64
import logging

from fastapi import APIRouter, HTTPException, Depends, Request
from pydantic import BaseModel, Field

from shared.database.pool import rls_conn
from shared.utils import like_escape
from services.api.deps import get_uid, endpoint_rate_check

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1", tags=["Tovarlar"])


# ═══ PYDANTIC MODELS ═══

class TovarYaratSorov(BaseModel):
    nomi:             str   = Field(..., min_length=1, max_length=200)
    kategoriya:       str   = Field("Boshqa")
    birlik:           str   = Field("dona")
    olish_narxi:      float = Field(0, ge=0)
    sotish_narxi:     float = Field(0, ge=0)
    min_sotish_narxi: float = Field(0, ge=0)
    qoldiq:           float = Field(0, ge=0)
    min_qoldiq:       float = Field(0, ge=0)
    # SalesDoc-compatible maydonlar
    brend:            str | None   = None
    podkategoriya:    str | None   = None
    guruh:            str | None   = None
    ishlab_chiqaruvchi: str | None = None
    segment:          str | None   = None
    shtrix_kod:       str | None   = None
    artikul:          str | None   = None
    sap_kod:          str | None   = None
    kod:              str | None   = None
    ikpu_kod:         str | None   = None
    gtin:             str | None   = None
    hajm:             float | None = None
    ogirlik:          float | None = None
    blokda_soni:      int | None   = None
    korobkada_soni:   int | None   = None
    saralash:         int | None   = None
    yaroqlilik_muddati: int | None = None
    tavsif:           str | None   = None
    savdo_yonalishi:  str | None   = None


class TovarYangilaSorov(BaseModel):
    nomi:             str | None   = None
    kategoriya:       str | None   = None
    birlik:           str | None   = None
    olish_narxi:      float | None = None
    sotish_narxi:     float | None = None
    min_sotish_narxi: float | None = None
    qoldiq:           float | None = None
    min_qoldiq:       float | None = None
    # SalesDoc-compatible maydonlar
    brend:            str | None   = None
    podkategoriya:    str | None   = None
    guruh:            str | None   = None
    ishlab_chiqaruvchi: str | None = None
    segment:          str | None   = None
    shtrix_kod:       str | None   = None
    artikul:          str | None   = None
    sap_kod:          str | None   = None
    kod:              str | None   = None
    ikpu_kod:         str | None   = None
    gtin:             str | None   = None
    hajm:             float | None = None
    ogirlik:          float | None = None
    blokda_soni:      int | None   = None
    korobkada_soni:   int | None   = None
    saralash:         int | None   = None
    yaroqlilik_muddati: int | None = None
    tavsif:           str | None   = None
    savdo_yonalishi:  str | None   = None


class QoldiqYangilaSorov(BaseModel):
    qoldiq: float = Field(..., ge=0)


class TovarImportItem(BaseModel):
    nomi:         str
    kategoriya:   str   = "Boshqa"
    birlik:       str   = "dona"
    olish_narxi:  float = 0
    sotish_narxi: float = 0
    qoldiq:       float = 0


class TovarImportSorov(BaseModel):
    tovarlar: list[TovarImportItem]


# ═══ ENDPOINTS ═══

@router.get("/tovarlar")
async def tovarlar(
    limit: int = 20, offset: int = 0,
    kategoriya: str | None = None,
    brend: str | None = None,
    segment: str | None = None,
    ishlab_chiqaruvchi: str | None = None,
    savdo_yonalishi: str | None = None,
    qidiruv: str | None = None,
    kam_qoldiq: bool | None = None,
    sort: str = "kategoriya",
    uid: int = Depends(get_uid)
):
    """Tovarlar ro'yxati — SalesDoc-style filtrlash (brend, segment, qidiruv, kam qoldiq)"""
    limit = min(limit, 500)
    where = ["user_id = $1"]
    params: list = [uid]

    def add(col: str, val):
        params.append(val)
        where.append(f"{col} = ${len(params)}")

    if kategoriya:         add("kategoriya", kategoriya)
    if brend:              add("brend", brend)
    if segment:            add("segment", segment)
    if ishlab_chiqaruvchi: add("ishlab_chiqaruvchi", ishlab_chiqaruvchi)
    if savdo_yonalishi:    add("savdo_yonalishi", savdo_yonalishi)

    if qidiruv:
        q = f"%{like_escape(qidiruv)}%"
        params.append(q)
        where.append(
            f"(nomi ILIKE ${len(params)} OR shtrix_kod ILIKE ${len(params)} "
            f"OR artikul ILIKE ${len(params)} OR ikpu_kod ILIKE ${len(params)})"
        )
    if kam_qoldiq:
        where.append("min_qoldiq > 0 AND qoldiq <= min_qoldiq")

    sort_map = {
        "kategoriya": "kategoriya, nomi",
        "nomi":       "nomi",
        "narx":       "sotish_narxi DESC",
        "qoldiq":     "qoldiq DESC",
        "yangi":      "yaratilgan DESC",
    }
    order_by = sort_map.get(sort, "kategoriya, nomi")

    where_sql = " AND ".join(where)
    params.append(limit); params.append(offset)

    sql = f"""
        SELECT id, user_id, nomi, kategoriya, birlik, olish_narxi, sotish_narxi,
               min_sotish_narxi, qoldiq, min_qoldiq, yaratilgan,
               brend, podkategoriya, guruh, ishlab_chiqaruvchi, segment,
               shtrix_kod, artikul, sap_kod, kod, ikpu_kod, gtin,
               hajm, ogirlik, blokda_soni, korobkada_soni, saralash,
               yaroqlilik_muddati, tavsif, rasm_url, faol, savdo_yonalishi
        FROM tovarlar
        WHERE {where_sql}
        ORDER BY {order_by}
        LIMIT ${len(params)-1} OFFSET ${len(params)}
    """
    count_sql = f"SELECT COUNT(*) FROM tovarlar WHERE {where_sql}"

    async with rls_conn(uid) as c:
        rows = await c.fetch(sql, *params)
        total = await c.fetchval(count_sql, *params[:-2])
    return {"total": total, "items": [dict(r) for r in rows]}


@router.get("/tovarlar/facets")
async def tovarlar_facets(uid: int = Depends(get_uid)):
    """Filter dropdownlari uchun unikal qiymatlar — SalesDoc-style"""
    async with rls_conn(uid) as c:
        async def unique(col: str) -> list[str]:
            rows = await c.fetch(
                f"SELECT DISTINCT {col} AS v FROM tovarlar "
                f"WHERE user_id=$1 AND {col} IS NOT NULL AND {col} <> '' "
                f"ORDER BY {col}", uid
            )
            return [r["v"] for r in rows]
        total = await c.fetchval(
            "SELECT COUNT(*) FROM tovarlar WHERE user_id=$1", uid
        )
        kam_qoldiq = await c.fetchval(
            "SELECT COUNT(*) FROM tovarlar WHERE user_id=$1 "
            "AND min_qoldiq > 0 AND qoldiq <= min_qoldiq", uid
        )
        return {
            "jami":              int(total or 0),
            "kam_qoldiq":        int(kam_qoldiq or 0),
            "kategoriyalar":     await unique("kategoriya"),
            "brendlar":          await unique("brend"),
            "segmentlar":        await unique("segment"),
            "ishlab_chiqaruvchilar": await unique("ishlab_chiqaruvchi"),
            "savdo_yonalishlari":    await unique("savdo_yonalishi"),
        }


@router.get("/tovar/{tovar_id}")
async def tovar_bir(tovar_id: int, uid: int = Depends(get_uid)):
    """Bitta tovar to'liq ma'lumoti"""
    async with rls_conn(uid) as c:
        t = await c.fetchrow("""
            SELECT id, user_id, nomi, kategoriya, birlik, olish_narxi, sotish_narxi,
                   min_sotish_narxi, qoldiq, min_qoldiq, yaratilgan,
                   brend, podkategoriya, guruh, ishlab_chiqaruvchi, segment,
                   shtrix_kod, artikul, sap_kod, kod, ikpu_kod, ikpu_paket_kod,
                   ikpu_birlik_kod, gtin, hajm, ogirlik, blokda_soni, korobkada_soni,
                   saralash, yaroqlilik_muddati, tavsif, rasm_url, faol,
                   savdo_yonalishi, yangilangan
            FROM tovarlar WHERE id=$1
        """, tovar_id)
        if not t:
            raise HTTPException(404, "Tovar topilmadi")
        return dict(t)


@router.post("/tovar")
async def tovar_yarat(data: TovarYaratSorov, uid: int = Depends(get_uid)):
    """Yangi tovar yaratish — SalesDoc-compatible, barcha 40+ maydonlar saqlanadi"""
    from shared.cache.redis_cache import user_cache_tozala
    d = data.model_dump()
    d["nomi"] = d["nomi"].strip()
    async with rls_conn(uid) as c:
        tovar = await c.fetchrow("""
            INSERT INTO tovarlar (
                user_id, nomi, kategoriya, birlik,
                olish_narxi, sotish_narxi, min_sotish_narxi, qoldiq, min_qoldiq,
                brend, podkategoriya, guruh, ishlab_chiqaruvchi, segment,
                shtrix_kod, artikul, sap_kod, kod, ikpu_kod, gtin,
                hajm, ogirlik, blokda_soni, korobkada_soni,
                saralash, yaroqlilik_muddati, tavsif, savdo_yonalishi
            )
            VALUES (
                $1,$2,$3,$4,
                $5,$6,$7,$8,$9,
                COALESCE($10,''), COALESCE($11,''), COALESCE($12,''),
                COALESCE($13,''), COALESCE($14,''),
                COALESCE($15,''), COALESCE($16,''), COALESCE($17,''),
                COALESCE($18,''), COALESCE($19,''), COALESCE($20,''),
                COALESCE($21,1), COALESCE($22,1),
                COALESCE($23,1), COALESCE($24,1),
                COALESCE($25,500), COALESCE($26,0),
                COALESCE($27,''), COALESCE($28,'')
            )
            ON CONFLICT (user_id, lower(nomi)) DO UPDATE SET
                kategoriya       = EXCLUDED.kategoriya,
                birlik           = EXCLUDED.birlik,
                olish_narxi      = EXCLUDED.olish_narxi,
                sotish_narxi     = EXCLUDED.sotish_narxi,
                min_sotish_narxi = EXCLUDED.min_sotish_narxi,
                brend            = COALESCE(NULLIF(EXCLUDED.brend,''), tovarlar.brend),
                podkategoriya    = COALESCE(NULLIF(EXCLUDED.podkategoriya,''), tovarlar.podkategoriya),
                guruh            = COALESCE(NULLIF(EXCLUDED.guruh,''), tovarlar.guruh),
                ishlab_chiqaruvchi = COALESCE(NULLIF(EXCLUDED.ishlab_chiqaruvchi,''), tovarlar.ishlab_chiqaruvchi),
                segment          = COALESCE(NULLIF(EXCLUDED.segment,''), tovarlar.segment),
                shtrix_kod       = COALESCE(NULLIF(EXCLUDED.shtrix_kod,''), tovarlar.shtrix_kod),
                artikul          = COALESCE(NULLIF(EXCLUDED.artikul,''), tovarlar.artikul),
                sap_kod          = COALESCE(NULLIF(EXCLUDED.sap_kod,''), tovarlar.sap_kod),
                kod              = COALESCE(NULLIF(EXCLUDED.kod,''), tovarlar.kod),
                ikpu_kod         = COALESCE(NULLIF(EXCLUDED.ikpu_kod,''), tovarlar.ikpu_kod),
                gtin             = COALESCE(NULLIF(EXCLUDED.gtin,''), tovarlar.gtin),
                hajm             = EXCLUDED.hajm,
                ogirlik          = EXCLUDED.ogirlik,
                blokda_soni      = EXCLUDED.blokda_soni,
                korobkada_soni   = EXCLUDED.korobkada_soni,
                saralash         = EXCLUDED.saralash,
                yaroqlilik_muddati = EXCLUDED.yaroqlilik_muddati,
                tavsif           = COALESCE(NULLIF(EXCLUDED.tavsif,''), tovarlar.tavsif),
                savdo_yonalishi  = COALESCE(NULLIF(EXCLUDED.savdo_yonalishi,''), tovarlar.savdo_yonalishi),
                yangilangan      = NOW()
            RETURNING id, nomi
        """,
            uid, d["nomi"], d["kategoriya"], d["birlik"],
            d["olish_narxi"], d["sotish_narxi"], d["min_sotish_narxi"],
            d["qoldiq"], d["min_qoldiq"],
            d.get("brend"), d.get("podkategoriya"), d.get("guruh"),
            d.get("ishlab_chiqaruvchi"), d.get("segment"),
            d.get("shtrix_kod"), d.get("artikul"), d.get("sap_kod"),
            d.get("kod"), d.get("ikpu_kod"), d.get("gtin"),
            d.get("hajm"), d.get("ogirlik"),
            d.get("blokda_soni"), d.get("korobkada_soni"),
            d.get("saralash"), d.get("yaroqlilik_muddati"),
            d.get("tavsif"), d.get("savdo_yonalishi"),
        )
    await user_cache_tozala(uid)
    log.info("📦 Tovar yaratildi: %s (uid=%d)", data.nomi, uid)
    return {"id": tovar["id"], "nomi": tovar["nomi"], "status": "yaratildi"}


@router.put("/tovar/{tovar_id}")
async def tovar_yangilash(tovar_id: int, data: TovarYangilaSorov,
                          uid: int = Depends(get_uid)):
    """Tovar ma'lumotlarini yangilash"""
    from shared.cache.redis_cache import user_cache_tozala
    yangilar = {k: v for k, v in data.model_dump().items() if v is not None}
    if not yangilar:
        raise HTTPException(400, "Yangilash uchun kamida 1 ta maydon kerak")

    _RUXSAT = {
        "nomi", "kategoriya", "birlik", "olish_narxi", "sotish_narxi",
        "min_sotish_narxi", "qoldiq", "min_qoldiq",
        "brend", "podkategoriya", "guruh", "ishlab_chiqaruvchi", "segment",
        "shtrix_kod", "artikul", "sap_kod", "kod", "ikpu_kod", "gtin",
        "hajm", "ogirlik", "blokda_soni", "korobkada_soni",
        "saralash", "yaroqlilik_muddati", "tavsif", "savdo_yonalishi",
    }
    noma = set(yangilar.keys()) - _RUXSAT
    if noma:
        raise HTTPException(400, f"Ruxsat etilmagan maydon: {noma}")

    # yangilangan timestamp avtomatik yangilanishi uchun
    yangilar["yangilangan"] = "NOW()"  # SQL literal sentinel
    set_parts = []
    vals = []
    idx = 3
    for k, v in yangilar.items():
        if k == "yangilangan":
            set_parts.append("yangilangan = NOW()")
        else:
            set_parts.append(f"{k} = ${idx}")
            vals.append(v)
            idx += 1
    set_q = ", ".join(set_parts)

    async with rls_conn(uid) as c:
        result = await c.execute(
            f"UPDATE tovarlar SET {set_q} WHERE id=$1 AND user_id=$2",
            tovar_id, uid, *vals
        )
    if "UPDATE 0" in result:
        raise HTTPException(404, "Tovar topilmadi")
    await user_cache_tozala(uid)
    return {"id": tovar_id, "status": "yangilandi"}


@router.delete("/tovar/{tovar_id}")
async def tovar_ochirish(tovar_id: int, uid: int = Depends(get_uid)):
    """Tovarni o'chirish (agar sotuvda ishlatilmagan bo'lsa)"""
    from shared.cache.redis_cache import user_cache_tozala
    async with rls_conn(uid) as c:
        sotuv_bor = await c.fetchval(
            "SELECT EXISTS(SELECT 1 FROM chiqimlar WHERE tovar_id=$1)", tovar_id
        )
        if sotuv_bor:
            raise HTTPException(
                409, "Bu tovar sotuvlarda ishlatilgan — o'chirib bo'lmaydi. "
                     "Qoldiqni 0 ga o'zgartiring."
            )
        result = await c.execute(
            "DELETE FROM tovarlar WHERE id=$1 AND user_id=$2", tovar_id, uid
        )
    if "DELETE 0" in result:
        raise HTTPException(404, "Tovar topilmadi")
    await user_cache_tozala(uid)
    return {"id": tovar_id, "status": "ochirildi"}


@router.post("/tovar/{tovar_id}/qoldiq")
async def tovar_qoldiq_yangilash(tovar_id: int, data: QoldiqYangilaSorov,
                                  uid: int = Depends(get_uid)):
    """Inventarizatsiya — tovar qoldiqini yangilash"""
    from shared.cache.redis_cache import user_cache_tozala
    async with rls_conn(uid) as c:
        old = await c.fetchrow(
            "SELECT nomi, qoldiq FROM tovarlar WHERE id=$1 AND user_id=$2 FOR UPDATE",
            tovar_id, uid
        )
        if not old:
            raise HTTPException(404, "Tovar topilmadi")
        await c.execute(
            "UPDATE tovarlar SET qoldiq=$2 WHERE id=$1 AND user_id=$3",
            tovar_id, data.qoldiq, uid
        )
    await user_cache_tozala(uid)
    return {
        "id": tovar_id, "nomi": old["nomi"],
        "eski_qoldiq": float(old["qoldiq"]),
        "yangi_qoldiq": data.qoldiq, "status": "yangilandi",
    }


@router.get("/tovar/{tovar_id}/tarix")
async def tovar_tarix(tovar_id: int, limit: int = 20, uid: int = Depends(get_uid)):
    """Tovarning sotuv va kirim tarixi"""
    limit = min(limit, 100)
    async with rls_conn(uid) as c:
        tovar = await c.fetchrow("""
            SELECT nomi, kategoriya, birlik, olish_narxi, sotish_narxi, qoldiq
            FROM tovarlar WHERE id=$1 AND user_id=$2
        """, tovar_id, uid)
        if not tovar:
            raise HTTPException(404, "Tovar topilmadi")
        sotuvlar = await c.fetch("""
            SELECT ch.miqdor, ch.sotish_narxi, ch.jami, ch.sana, ss.klient_ismi
            FROM chiqimlar ch
            LEFT JOIN sotuv_sessiyalar ss ON ss.id = ch.sessiya_id
            WHERE ch.tovar_id=$1 AND ch.user_id=$2
            ORDER BY ch.sana DESC LIMIT $3
        """, tovar_id, uid, limit)
        kirimlar = await c.fetch("""
            SELECT miqdor, narx, jami, manba, sana
            FROM kirimlar WHERE tovar_id=$1 AND user_id=$2
            ORDER BY sana DESC LIMIT $3
        """, tovar_id, uid, limit)
        stats = await c.fetchrow("""
            SELECT COUNT(*) AS sotuv_soni,
                   COALESCE(SUM(miqdor), 0) AS jami_sotilgan,
                   COALESCE(SUM(jami), 0) AS jami_tushum
            FROM chiqimlar WHERE tovar_id=$1 AND user_id=$2
        """, tovar_id, uid)
    return {
        "tovar": dict(tovar),
        "sotuvlar": [dict(r) for r in sotuvlar],
        "kirimlar": [dict(r) for r in kirimlar],
        "statistika": {
            "sotuv_soni": int(stats["sotuv_soni"]),
            "jami_sotilgan": float(stats["jami_sotilgan"]),
            "jami_tushum": float(stats["jami_tushum"]),
        },
    }


@router.get("/tovar/shablon/excel")
async def tovar_shablon_excel(uid: int = Depends(get_uid)):
    """Tovar import qilish uchun shablon Excel fayli.

    28 ta ustun: SalesDoc-compatible hamma maydonlar.
    """
    import io as _io
    import base64 as _b64
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Tovarlar"

    headers = [
        # Asosiy
        ("nomi",             "Tovar nomi",        30,  True),
        ("kategoriya",       "Kategoriya",        18,  False),
        ("birlik",           "Birlik",            10,  False),
        ("olish_narxi",      "Olish narxi",       14,  False),
        ("sotish_narxi",     "Sotish narxi",      14,  False),
        ("min_sotish_narxi", "Min sotish",        14,  False),
        ("qoldiq",           "Qoldiq",            12,  False),
        ("min_qoldiq",       "Min qoldiq",        12,  False),
        # Identifikatsiya
        ("brend",            "Brend",             18,  False),
        ("ishlab_chiqaruvchi", "Ishlab chiqaruvchi", 22, False),
        ("podkategoriya",    "Podkategoriya",     18,  False),
        ("guruh",            "Guruh",             15,  False),
        ("segment",          "Segment",           14,  False),
        ("savdo_yonalishi",  "Savdo yo'nalishi",  16,  False),
        ("shtrix_kod",       "Shtrix kod (EAN)",  18,  False),
        ("gtin",             "GTIN",              15,  False),
        ("artikul",          "Artikul",           14,  False),
        ("sap_kod",          "SAP kod",           14,  False),
        ("kod",              "Ichki kod",         14,  False),
        ("ikpu_kod",         "IKPU kod",          16,  False),
        # O'lchamlar
        ("hajm",             "Hajm (l)",          10,  False),
        ("ogirlik",          "Og'irlik (kg)",     12,  False),
        ("blokda_soni",      "Blokda",            10,  False),
        ("korobkada_soni",   "Korobkada",         12,  False),
        ("saralash",         "Saralash",          10,  False),
        ("yaroqlilik_muddati", "Yaroq. muddat (kun)", 16, False),
        # Tavsif
        ("tavsif",           "Tavsif",            30,  False),
    ]

    header_fill     = PatternFill(start_color="0A819C", end_color="0A819C", fill_type="solid")
    Font(bold=True, color="FFFFFF", size=11)
    required_fill   = PatternFill(start_color="FFECB3", end_color="FFECB3", fill_type="solid")
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (key, label, width, required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=label + (" *" if required else ""))
        cell.fill = required_fill if required else header_fill
        cell.font = Font(bold=True, color="000000" if required else "FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[chr(64 + i) if i < 27 else "A" + chr(64 + i - 26)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Namunaviy qator (2-qator)
    example = [
        "Ariel Color 3kg", "Maishiy kimyo", "dona",
        32000, 42000, 40000, 50, 10,
        "Procter & Gamble", "P&G", "Kir yuvish", "Kukun", "Premium", "B2C",
        "4015600123456", "04015600123456", "PG-ARI-3", "1001234", "INT-001", "05031000001",
        3, 3, 6, 12, 500, 730,
        "Avtomatik kir mashinasi uchun konsentrat kukun",
    ]
    for col, v in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.border = border
        cell.font = Font(italic=True, color="666666", size=10)
        if col in (4, 5, 6, 7, 8, 21, 22, 23, 24, 25, 26):
            cell.alignment = Alignment(horizontal="right")

    # Izoh sahifasi
    info_ws = wb.create_sheet("Yo'riqnoma")
    info_lines = [
        ("SavdoAI Tovar Import — Yo'riqnoma", True),
        ("", False),
        ("1-qator: Ustun sarlavhalari — o'zgartirmang!", False),
        ("2-qator: Namunaviy qator — o'chirib yuboring", False),
        ("3-qatordan boshlab: o'z tovarlaringizni kiriting", False),
        ("", False),
        ("Majburiy ustunlar (*):", True),
        ("• nomi — tovar nomi", False),
        ("", False),
        ("Ixtiyoriy ustunlar: Qolganlarini istalgancha to'ldiring.", False),
        ("", False),
        ("Import qilish:", True),
        ("1. Bu Excel faylni to'ldiring", False),
        ("2. SavdoAI web panel → /products → Import tugmasini bosing", False),
        ("3. Faylni tanlang va yuklang", False),
        ("4. Tizim avtomatik tovarlarni qo'shadi va duplicate bo'lsa yangilaydi", False),
        ("", False),
        ("Shakllar:", True),
        ("• Narxlar — raqam (masalan: 42000)", False),
        ("• Qoldiq — son (masalan: 50 yoki 50.5)", False),
        ("• Shtrix kod — EAN-13 format (13 raqam)", False),
        ("• IKPU kod — 11 raqam (O'zbekiston soliq)", False),
        ("", False),
        ("Yordam: @savdoai_mashrab_bot /yordam", False),
    ]
    for idx, (text, bold) in enumerate(info_lines, 1):
        cell = info_ws.cell(row=idx, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12, color="0A819C")
    info_ws.column_dimensions["A"].width = 80

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    return {
        "filename": "SavdoAI_Tovar_Shablon.xlsx",
        "content_base64": _b64.b64encode(buf.getvalue()).decode(),
    }


@router.get("/tovar/export/excel")
async def tovar_excel_export(uid: int = Depends(get_uid)):
    """Tovarlar ro'yxatini Excel faylga export qilish"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT nomi, kategoriya, birlik,
                   olish_narxi, sotish_narxi, qoldiq, min_qoldiq
            FROM tovarlar WHERE user_id=$1 ORDER BY kategoriya, nomi
        """, uid)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tovarlar"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["Tovar nomi", "Kategoriya", "Birlik", "Olish narxi",
               "Sotish narxi", "Qoldiq", "Min qoldiq"]
    widths = [30, 18, 10, 15, 15, 12, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + i)].width = w

    num_fmt = '#,##0'
    for r_idx, row in enumerate(rows, 2):
        d = dict(row)
        ws.cell(row=r_idx, column=1, value=d["nomi"]).border = thin_border
        ws.cell(row=r_idx, column=2, value=d["kategoriya"]).border = thin_border
        ws.cell(row=r_idx, column=3, value=d["birlik"]).border = thin_border
        c4 = ws.cell(row=r_idx, column=4, value=float(d["olish_narxi"]))
        c4.number_format = num_fmt; c4.border = thin_border
        c5 = ws.cell(row=r_idx, column=5, value=float(d["sotish_narxi"]))
        c5.number_format = num_fmt; c5.border = thin_border
        c6 = ws.cell(row=r_idx, column=6, value=float(d["qoldiq"]))
        c6.number_format = '#,##0.###'; c6.border = thin_border
        c7 = ws.cell(row=r_idx, column=7, value=float(d["min_qoldiq"]))
        c7.number_format = '#,##0.###'; c7.border = thin_border

        if d["min_qoldiq"] > 0 and d["qoldiq"] <= d["min_qoldiq"]:
            red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0",
                                   fill_type="solid")
            for col in range(1, 8):
                ws.cell(row=r_idx, column=col).fill = red_fill

    ws.auto_filter.ref = f"A1:G{len(rows) + 1}"
    last = len(rows) + 2
    ws.cell(row=last, column=1, value="JAMI:").font = Font(bold=True)
    ws.cell(row=last, column=6, value=f"=SUM(F2:F{len(rows)+1})").font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return {
        "filename": "tovarlar.xlsx",
        "content_base64": base64.b64encode(buf.getvalue()).decode(),
        "tovar_soni": len(rows),
    }


@router.post("/tovar/import/excel")
async def tovar_import_excel(
    file_base64: str,
    uid: int = Depends(get_uid),
):
    """Excel faylidan tovarlarni import qilish.

    Shablondan (GET /tovar/shablon/excel) yuklab olingan formatdagi
    27 ustunli xlsx faylni qabul qiladi. Birinchi 1-qator sarlavha,
    2-qator namuna (o'chiriladi), 3-qatordan boshlab real tovarlar.
    """
    import io as _io
    import base64 as _b64
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl mavjud emas")

    try:
        content = _b64.b64decode(file_base64)
        wb = load_workbook(_io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Fayl o'qilmadi: {e}")

    ws = wb.active
    if not ws:
        raise HTTPException(400, "Bo'sh Excel fayli")

    # 2-qatordan boshlab (1-qator sarlavha)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(400, "Tovarlar topilmadi")

    # Shablon ustun tartibi (GET /tovar/shablon/excel bilan mos)
    fields = [
        "nomi", "kategoriya", "birlik",
        "olish_narxi", "sotish_narxi", "min_sotish_narxi",
        "qoldiq", "min_qoldiq",
        "brend", "ishlab_chiqaruvchi", "podkategoriya", "guruh",
        "segment", "savdo_yonalishi",
        "shtrix_kod", "gtin", "artikul", "sap_kod", "kod", "ikpu_kod",
        "hajm", "ogirlik", "blokda_soni", "korobkada_soni",
        "saralash", "yaroqlilik_muddati",
        "tavsif",
    ]

    yaratildi = yangilandi = 0
    xatolar: list[str] = []

    def _num(v, default=0):
        if v is None or v == "":
            return default
        try:
            return float(v)
        except Exception:
            return default

    def _str(v):
        if v is None:
            return ""
        return str(v).strip()

    async with rls_conn(uid) as c:
        for ridx, row in enumerate(rows, 2):
            # Namuna qatorini o'tkazib yuborish (italic, grey)
            if ridx == 2 and row and "Ariel" in str(row[0] or ""):
                continue

            d = dict(zip(fields, row))
            nomi = _str(d.get("nomi"))
            if not nomi:
                continue

            try:
                result = await c.fetchrow("""
                    INSERT INTO tovarlar (
                        user_id, nomi, kategoriya, birlik,
                        olish_narxi, sotish_narxi, min_sotish_narxi, qoldiq, min_qoldiq,
                        brend, ishlab_chiqaruvchi, podkategoriya, guruh,
                        segment, savdo_yonalishi,
                        shtrix_kod, gtin, artikul, sap_kod, kod, ikpu_kod,
                        hajm, ogirlik, blokda_soni, korobkada_soni,
                        saralash, yaroqlilik_muddati, tavsif
                    )
                    VALUES (
                        $1, $2, COALESCE($3, 'Boshqa'), COALESCE($4, 'dona'),
                        $5, $6, $7, $8, $9,
                        $10, $11, $12, $13, $14, $15,
                        $16, $17, $18, $19, $20, $21,
                        COALESCE($22, 1), COALESCE($23, 1),
                        COALESCE($24, 1), COALESCE($25, 1),
                        COALESCE($26, 500), COALESCE($27, 0),
                        $28
                    )
                    ON CONFLICT (user_id, lower(nomi)) DO UPDATE SET
                        kategoriya       = EXCLUDED.kategoriya,
                        birlik           = EXCLUDED.birlik,
                        olish_narxi      = CASE WHEN EXCLUDED.olish_narxi > 0
                                            THEN EXCLUDED.olish_narxi ELSE tovarlar.olish_narxi END,
                        sotish_narxi     = CASE WHEN EXCLUDED.sotish_narxi > 0
                                            THEN EXCLUDED.sotish_narxi ELSE tovarlar.sotish_narxi END,
                        min_sotish_narxi = COALESCE(NULLIF(EXCLUDED.min_sotish_narxi, 0), tovarlar.min_sotish_narxi),
                        min_qoldiq       = COALESCE(NULLIF(EXCLUDED.min_qoldiq, 0), tovarlar.min_qoldiq),
                        brend            = COALESCE(NULLIF(EXCLUDED.brend, ''), tovarlar.brend),
                        ishlab_chiqaruvchi = COALESCE(NULLIF(EXCLUDED.ishlab_chiqaruvchi, ''), tovarlar.ishlab_chiqaruvchi),
                        podkategoriya    = COALESCE(NULLIF(EXCLUDED.podkategoriya, ''), tovarlar.podkategoriya),
                        guruh            = COALESCE(NULLIF(EXCLUDED.guruh, ''), tovarlar.guruh),
                        segment          = COALESCE(NULLIF(EXCLUDED.segment, ''), tovarlar.segment),
                        savdo_yonalishi  = COALESCE(NULLIF(EXCLUDED.savdo_yonalishi, ''), tovarlar.savdo_yonalishi),
                        shtrix_kod       = COALESCE(NULLIF(EXCLUDED.shtrix_kod, ''), tovarlar.shtrix_kod),
                        gtin             = COALESCE(NULLIF(EXCLUDED.gtin, ''), tovarlar.gtin),
                        artikul          = COALESCE(NULLIF(EXCLUDED.artikul, ''), tovarlar.artikul),
                        sap_kod          = COALESCE(NULLIF(EXCLUDED.sap_kod, ''), tovarlar.sap_kod),
                        kod              = COALESCE(NULLIF(EXCLUDED.kod, ''), tovarlar.kod),
                        ikpu_kod         = COALESCE(NULLIF(EXCLUDED.ikpu_kod, ''), tovarlar.ikpu_kod),
                        hajm             = EXCLUDED.hajm,
                        ogirlik          = EXCLUDED.ogirlik,
                        blokda_soni      = EXCLUDED.blokda_soni,
                        korobkada_soni   = EXCLUDED.korobkada_soni,
                        saralash         = EXCLUDED.saralash,
                        yaroqlilik_muddati = EXCLUDED.yaroqlilik_muddati,
                        tavsif           = COALESCE(NULLIF(EXCLUDED.tavsif, ''), tovarlar.tavsif),
                        yangilangan      = NOW()
                    RETURNING (xmax = 0) AS yangi
                """,
                    uid, nomi,
                    _str(d.get("kategoriya")) or "Boshqa",
                    _str(d.get("birlik")) or "dona",
                    _num(d.get("olish_narxi")),
                    _num(d.get("sotish_narxi")),
                    _num(d.get("min_sotish_narxi")),
                    _num(d.get("qoldiq")),
                    _num(d.get("min_qoldiq")),
                    _str(d.get("brend")),
                    _str(d.get("ishlab_chiqaruvchi")),
                    _str(d.get("podkategoriya")),
                    _str(d.get("guruh")),
                    _str(d.get("segment")),
                    _str(d.get("savdo_yonalishi")),
                    _str(d.get("shtrix_kod")),
                    _str(d.get("gtin")),
                    _str(d.get("artikul")),
                    _str(d.get("sap_kod")),
                    _str(d.get("kod")),
                    _str(d.get("ikpu_kod")),
                    _num(d.get("hajm"), default=1),
                    _num(d.get("ogirlik"), default=1),
                    int(_num(d.get("blokda_soni"), default=1)),
                    int(_num(d.get("korobkada_soni"), default=1)),
                    int(_num(d.get("saralash"), default=500)),
                    int(_num(d.get("yaroqlilik_muddati"), default=0)),
                    _str(d.get("tavsif")),
                )
                if result and result["yangi"]:
                    yaratildi += 1
                else:
                    yangilandi += 1
            except Exception as e:
                xatolar.append(f"Qator #{ridx}: {nomi[:30]}: {str(e)[:80]}")
                if len(xatolar) >= 20:
                    break

    from shared.cache.redis_cache import user_cache_tozala
    await user_cache_tozala(uid)

    return {
        "jami":       yaratildi + yangilandi,
        "yaratildi":  yaratildi,
        "yangilandi": yangilandi,
        "xatolar":    xatolar[:20],
    }


@router.post("/tovar/import")
async def tovar_import(data: TovarImportSorov, request: Request,
                       uid: int = Depends(get_uid)):
    """Tovarlarni batch import qilish. Rate limit: 5/daqiqa."""
    await endpoint_rate_check(request, "import")
    from shared.cache.redis_cache import user_cache_tozala
    if not data.tovarlar:
        raise HTTPException(400, "Tovarlar ro'yxati bo'sh")
    if len(data.tovarlar) > 1000:
        raise HTTPException(400, "Maksimal 1000 ta tovar import qilish mumkin")

    yaratildi = yangilandi = 0
    xatolar = []

    async with rls_conn(uid) as c:
        for i, t in enumerate(data.tovarlar):
            nomi = t.nomi.strip()
            if not nomi:
                xatolar.append(f"#{i+1}: nom bo'sh")
                continue
            try:
                result = await c.fetchrow("""
                    INSERT INTO tovarlar
                        (user_id, nomi, kategoriya, birlik, olish_narxi, sotish_narxi, qoldiq)
                    VALUES ($1, $2, $3, $4, $5, $6, $7)
                    ON CONFLICT (user_id, lower(nomi)) DO UPDATE SET
                        kategoriya   = EXCLUDED.kategoriya,
                        birlik       = EXCLUDED.birlik,
                        olish_narxi  = CASE WHEN EXCLUDED.olish_narxi > 0
                                       THEN EXCLUDED.olish_narxi ELSE tovarlar.olish_narxi END,
                        sotish_narxi = CASE WHEN EXCLUDED.sotish_narxi > 0
                                       THEN EXCLUDED.sotish_narxi ELSE tovarlar.sotish_narxi END
                    RETURNING (xmax = 0) AS yangi
                """, uid, nomi, t.kategoriya, t.birlik,
                    t.olish_narxi, t.sotish_narxi, t.qoldiq)
                if result and result["yangi"]:
                    yaratildi += 1
                else:
                    yangilandi += 1
            except Exception as e:
                xatolar.append(f"#{i+1} {nomi}: {str(e)[:50]}")

    await user_cache_tozala(uid)
    return {
        "yaratildi": yaratildi, "yangilandi": yangilandi,
        "xatolar": xatolar[:20], "jami": yaratildi + yangilandi,
    }
