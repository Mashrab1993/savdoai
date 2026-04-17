"""
╔══════════════════════════════════════════════════════════════════════╗
║  SAVDOAI v25.4.0 — EKSPEDITOR + SKLAD + NAKLADNOY API               ║
║                                                                      ║
║  SalesDoc /settings/expeditors + /settings/warehouses + bulk         ║
║  nakladnoy generatsiya — Excel eksport bilan.                       ║
╚══════════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from shared.database.pool import rls_conn
from services.api.deps import get_uid

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1", tags=["Ekspeditor/Sklad"])


# ═══════════════════════════════════════════════════════════════════
# EKSPEDITORLAR
# ═══════════════════════════════════════════════════════════════════

class EkspeditorIn(BaseModel):
    ism: str = Field(..., min_length=1, max_length=200)
    telefon: Optional[str] = None
    mashina_nomi: Optional[str] = None
    mashina_raqami: Optional[str] = None
    faol: bool = True


@router.get("/ekspeditorlar")
async def list_ekspeditorlar(
    faol: Optional[bool] = None,
    uid: int = Depends(get_uid),
):
    where = ["user_id = $1"]
    params: list = [uid]
    if faol is not None:
        params.append(faol)
        where.append(f"faol = ${len(params)}")

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT e.*,
                   (SELECT COUNT(*) FROM sotuv_sessiyalar
                    WHERE user_id=$1 AND ekspeditor_id = e.id) AS sotuv_soni
            FROM ekspeditorlar e
            WHERE {' AND '.join(where)}
            ORDER BY e.faol DESC, e.ism
        """, *params)
    return {
        "items": [
            {
                "id": r["id"], "ism": r["ism"], "telefon": r["telefon"],
                "mashina_nomi": r["mashina_nomi"], "mashina_raqami": r["mashina_raqami"],
                "faol": r["faol"], "sotuv_soni": int(r["sotuv_soni"] or 0),
                "yaratilgan": r["yaratilgan"].isoformat() if r["yaratilgan"] else None,
            } for r in rows
        ],
        "jami": len(rows),
    }


@router.post("/ekspeditor")
async def create_ekspeditor(body: EkspeditorIn, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        row = await c.fetchrow("""
            INSERT INTO ekspeditorlar (user_id, ism, telefon, mashina_nomi, mashina_raqami, faol)
            VALUES ($1, $2, $3, $4, $5, $6) RETURNING *
        """, uid, body.ism, body.telefon, body.mashina_nomi, body.mashina_raqami, body.faol)
    return dict(row) | {"yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None}


@router.put("/ekspeditor/{eid}")
async def update_ekspeditor(eid: int, body: EkspeditorIn, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        row = await c.fetchrow("""
            UPDATE ekspeditorlar
            SET ism=$1, telefon=$2, mashina_nomi=$3, mashina_raqami=$4, faol=$5
            WHERE id=$6 AND user_id=$7 RETURNING *
        """, body.ism, body.telefon, body.mashina_nomi, body.mashina_raqami, body.faol, eid, uid)
    if not row:
        raise HTTPException(404, "Ekspeditor topilmadi")
    return dict(row) | {"yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None}


@router.delete("/ekspeditor/{eid}")
async def delete_ekspeditor(eid: int, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        r = await c.fetchrow(
            "DELETE FROM ekspeditorlar WHERE id=$1 AND user_id=$2 RETURNING id, ism",
            eid, uid,
        )
    if not r:
        raise HTTPException(404, "Topilmadi")
    return {"ok": True, "deleted": dict(r)}


# ═══════════════════════════════════════════════════════════════════
# SKLADLAR
# ═══════════════════════════════════════════════════════════════════

class SkladIn(BaseModel):
    nomi: str = Field(..., min_length=1, max_length=200)
    turi: Optional[str] = Field(None, description="asosiy, brak, aksiya...")
    kod: Optional[str] = None
    faol: bool = True


@router.get("/skladlar")
async def list_skladlar(faol: Optional[bool] = None, uid: int = Depends(get_uid)):
    where = ["user_id = $1"]
    params: list = [uid]
    if faol is not None:
        params.append(faol)
        where.append(f"faol = ${len(params)}")
    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT s.*,
                   (SELECT COUNT(*) FROM sotuv_sessiyalar
                    WHERE user_id=$1 AND sklad_id = s.id) AS sotuv_soni
            FROM skladlar s
            WHERE {' AND '.join(where)}
            ORDER BY s.faol DESC, s.nomi
        """, *params)
    return {
        "items": [
            {
                "id": r["id"], "nomi": r["nomi"], "turi": r["turi"], "kod": r["kod"],
                "faol": r["faol"], "sotuv_soni": int(r["sotuv_soni"] or 0),
                "yaratilgan": r["yaratilgan"].isoformat() if r["yaratilgan"] else None,
            } for r in rows
        ],
        "jami": len(rows),
    }


@router.post("/sklad")
async def create_sklad(body: SkladIn, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        row = await c.fetchrow("""
            INSERT INTO skladlar (user_id, nomi, turi, kod, faol)
            VALUES ($1, $2, $3, $4, $5) RETURNING *
        """, uid, body.nomi, body.turi, body.kod, body.faol)
    return dict(row) | {"yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None}


@router.put("/sklad/{sid}")
async def update_sklad(sid: int, body: SkladIn, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        row = await c.fetchrow("""
            UPDATE skladlar SET nomi=$1, turi=$2, kod=$3, faol=$4
            WHERE id=$5 AND user_id=$6 RETURNING *
        """, body.nomi, body.turi, body.kod, body.faol, sid, uid)
    if not row:
        raise HTTPException(404, "Sklad topilmadi")
    return dict(row) | {"yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None}


@router.delete("/sklad/{sid}")
async def delete_sklad(sid: int, uid: int = Depends(get_uid)):
    async with rls_conn(uid) as c:
        r = await c.fetchrow(
            "DELETE FROM skladlar WHERE id=$1 AND user_id=$2 RETURNING id, nomi",
            sid, uid,
        )
    if not r:
        raise HTTPException(404, "Topilmadi")
    return {"ok": True, "deleted": dict(r)}


# ═══════════════════════════════════════════════════════════════════
# NAKLADNOY REGISTR — bulk yaratish + Excel export
# ═══════════════════════════════════════════════════════════════════

class NakladnoyCreate(BaseModel):
    nomi: str = Field(..., min_length=1)
    sana: Optional[str] = None  # YYYY-MM-DD
    shogird_id: Optional[int] = None
    ekspeditor_id: Optional[int] = None
    sklad_id: Optional[int] = None
    sessiya_idlar: list[int] = Field(..., min_items=1)
    izoh: Optional[str] = None


@router.post("/nakladnoy_registr")
async def create_nakladnoy(body: NakladnoyCreate, uid: int = Depends(get_uid)):
    """Bulk tanlangan zayavkalar uchun nakladnoy registr yaratish.

    SalesDoc '[Nakladnaya]' bulk action — foydalanuvchi ☑️ bilan bir nechta
    zayavka tanlaydi va ularning hammasini bitta registrga yig'adi.
    """
    sana_obj = date.fromisoformat(body.sana) if body.sana else date.today()

    async with rls_conn(uid) as c:
        # Tanlangan sessiyalarning jamini hisoblash
        stats = await c.fetchrow("""
            SELECT COALESCE(SUM(jami), 0) AS jami_summa,
                   COALESCE(SUM(tolangan), 0) AS tolangan,
                   COUNT(*) AS soni
            FROM sotuv_sessiyalar
            WHERE user_id=$1 AND id = ANY($2::bigint[])
        """, uid, body.sessiya_idlar)

        if int(stats["soni"] or 0) == 0:
            raise HTTPException(400, "Tanlangan zayavka topilmadi")

        row = await c.fetchrow("""
            INSERT INTO nakladnoy_registrlari
                (user_id, nomi, sana, shogird_id, ekspeditor_id, sklad_id,
                 sessiya_idlar, jami_summa, tolangan, izoh)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
            RETURNING *
        """, uid, body.nomi, sana_obj, body.shogird_id, body.ekspeditor_id,
             body.sklad_id, body.sessiya_idlar, float(stats["jami_summa"] or 0),
             float(stats["tolangan"] or 0), body.izoh)

    return dict(row) | {
        "sessiya_soni": int(stats["soni"] or 0),
        "yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None,
        "sana": row["sana"].isoformat() if row["sana"] else None,
    }


@router.get("/nakladnoy_registrlari")
async def list_nakladnoy(
    sana_dan: Optional[str] = None,
    sana_gacha: Optional[str] = None,
    uid: int = Depends(get_uid),
):
    where = ["n.user_id = $1"]
    params: list = [uid]
    if sana_dan:
        params.append(date.fromisoformat(sana_dan))
        where.append(f"n.sana >= ${len(params)}")
    if sana_gacha:
        params.append(date.fromisoformat(sana_gacha))
        where.append(f"n.sana <= ${len(params)}")

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT n.*,
                   s.ism AS shogird_nomi,
                   e.ism AS ekspeditor_nomi,
                   sk.nomi AS sklad_nomi
            FROM nakladnoy_registrlari n
            LEFT JOIN shogirdlar s ON s.id = n.shogird_id
            LEFT JOIN ekspeditorlar e ON e.id = n.ekspeditor_id
            LEFT JOIN skladlar sk ON sk.id = n.sklad_id
            WHERE {' AND '.join(where)}
            ORDER BY n.sana DESC, n.id DESC
        """, *params)
    return {
        "items": [
            {
                "id": r["id"], "nomi": r["nomi"],
                "sana": r["sana"].isoformat() if r["sana"] else None,
                "shogird_id": r["shogird_id"], "shogird_nomi": r["shogird_nomi"],
                "ekspeditor_id": r["ekspeditor_id"], "ekspeditor_nomi": r["ekspeditor_nomi"],
                "sklad_id": r["sklad_id"], "sklad_nomi": r["sklad_nomi"],
                "sessiya_idlar": list(r["sessiya_idlar"] or []),
                "jami_summa": float(r["jami_summa"] or 0),
                "tolangan": float(r["tolangan"] or 0),
                "izoh": r["izoh"],
                "yaratilgan": r["yaratilgan"].isoformat() if r["yaratilgan"] else None,
            } for r in rows
        ],
    }


@router.get("/nakladnoy_registr/{rid}/excel")
async def export_nakladnoy_excel(
    rid: int,
    format: str = Query("reestr", description="reestr (3.0) | nakladnye (3.1) | sklad_zagruz (4.1)"),
    uid: int = Depends(get_uid),
):
    """Nakladnoy registrni SalesDoc formatlarida Excel yuklab olish.

    3 format:
    - reestr (3.0): Qisqa ro'yxat — bir mijoz bir qatorda
    - nakladnye (3.1): Har mijozga alohida nakladnoy + item list
    - sklad_zagruz (4.1): Ishlab chiqaruvchi bo'yicha guruhlangan sklad yuk
    """
    if format not in {"reestr", "nakladnye", "sklad_zagruz"}:
        raise HTTPException(400, "format kerak: reestr, nakladnye yoki sklad_zagruz")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl yo'q")

    async with rls_conn(uid) as c:
        reg = await c.fetchrow("""
            SELECT n.*, s.ism AS shogird_nomi, e.ism AS ekspeditor_nomi,
                   e.telefon AS ekspeditor_tel,
                   e.mashina_nomi, e.mashina_raqami,
                   sk.nomi AS sklad_nomi
            FROM nakladnoy_registrlari n
            LEFT JOIN shogirdlar s ON s.id = n.shogird_id
            LEFT JOIN ekspeditorlar e ON e.id = n.ekspeditor_id
            LEFT JOIN skladlar sk ON sk.id = n.sklad_id
            WHERE n.id=$1 AND n.user_id=$2
        """, rid, uid)
        if not reg:
            raise HTTPException(404, "Registr topilmadi")

        sessions = await c.fetch("""
            SELECT ss.id, ss.document_number, ss.tip_zayavki, ss.holat,
                   ss.sana, ss.jami, ss.tolangan,
                   COALESCE(ss.jami - ss.tolangan, 0) AS qarz,
                   k.ismi AS klient_ismi, k.telefon AS klient_tel,
                   k.manzil AS klient_manzil,
                   sh.ism AS shogird_nomi,
                   sh.telefon AS shogird_tel
            FROM sotuv_sessiyalar ss
            LEFT JOIN klientlar k ON k.id = ss.klient_id
            LEFT JOIN shogirdlar sh ON sh.id = ss.shogird_id
            WHERE ss.id = ANY($1::bigint[]) AND ss.user_id = $2
            ORDER BY sh.ism NULLS LAST, ss.sana
        """, list(reg["sessiya_idlar"] or []), uid)

        # Har sessiya uchun itemlar (nakladnye va sklad_zagruz formati uchun)
        items_by_session = {}
        if format in ("nakladnye", "sklad_zagruz"):
            items_rows = await c.fetch("""
                SELECT ch.sessiya_id, ch.tovar_nomi, ch.miqdor, ch.narx,
                       ch.jami AS summa, t.birlik, t.kategoriya
                FROM chiqimlar ch
                LEFT JOIN tovarlar t ON t.id = ch.tovar_id
                WHERE ch.sessiya_id = ANY($1::bigint[])
                ORDER BY ch.sessiya_id, ch.id
            """, list(reg["sessiya_idlar"] or []))
            for ir in items_rows:
                items_by_session.setdefault(ir["sessiya_id"], []).append(dict(ir))

    wb = Workbook()
    ws = wb.active
    HF = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    HFont = Font(bold=True, color="FFFFFF")
    BoldFont = Font(bold=True)

    if format == "reestr":
        # ═══ 3.0 REESTR — qisqa ro'yxat ═══
        ws.title = "Reestr 3.0"
        ws.cell(row=1, column=1, value=f"Reestr {reg['sana']} — {reg['nomi']}").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Agent: {reg['shogird_nomi'] or '—'}  |  Ekspeditor: {reg['ekspeditor_nomi'] or '—'}")

        headers = ["№", "Yetkazish sanasi", "Savdo nuqtasi", "Manzil", "Telefon",
                   "Agent (TP)", "Balans", "Summa", "Izoh"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = HFont; c.fill = HF; c.alignment = Alignment(horizontal="center")

        for i, s in enumerate(sessions, 5):
            ws.cell(row=i, column=1, value=i - 4)
            ws.cell(row=i, column=2, value=str(s["sana"])[:10] if s["sana"] else "")
            ws.cell(row=i, column=3, value=s["klient_ismi"] or "—")
            ws.cell(row=i, column=4, value=s["klient_manzil"] or "")
            ws.cell(row=i, column=5, value=s["klient_tel"] or "")
            ws.cell(row=i, column=6, value=s["shogird_nomi"] or "—")
            ws.cell(row=i, column=7, value=float(s["qarz"] or 0) or "")
            ws.cell(row=i, column=8, value=float(s["jami"] or 0))
            ws.cell(row=i, column=9, value="")

        # TOTAL
        tot_row = 5 + len(sessions)
        ws.cell(row=tot_row, column=1, value="JAMI:").font = BoldFont
        ws.cell(row=tot_row, column=8, value=float(reg["jami_summa"] or 0)).font = BoldFont
        ws.freeze_panes = "A5"

    elif format == "nakladnye":
        # ═══ 3.1 NAKLADNYE — har zayavkaga alohida invoice ═══
        ws.title = "Nakladnye 3.1"
        row = 1

        for s in sessions:
            # HEADER bloki
            ws.cell(row=row, column=3, value=f"Накладная №{s['document_number'] or s['id']}  от  {str(s['sana'])[:10]}").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=2, value=f"Kimga: {s['klient_ismi'] or '—'}")
            ws.cell(row=row, column=5, value=f"TP: {s['shogird_nomi'] or '—'}")
            row += 1
            ws.cell(row=row, column=2, value=f"Manzil: {s['klient_manzil'] or ''}")
            ws.cell(row=row, column=5, value=f"Tel(tp): {s['shogird_tel'] or ''}")
            row += 1
            ws.cell(row=row, column=2, value=f"Tel: {s['klient_tel'] or ''}")
            ws.cell(row=row, column=5, value=f"Ekspeditor: {reg['ekspeditor_nomi'] or '—'} ({reg['ekspeditor_tel'] or ''})")
            row += 1
            ws.cell(row=row, column=2, value=f"Balans klienta: {float(s['qarz'] or 0):,.0f}")
            row += 2

            # Items jadvali
            heads = ["№", "Nomi", "Miqdor", "Birlik", "Narx", "Summa"]
            for col, h in enumerate(heads, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = HFont; c.fill = HF
            row += 1

            items = items_by_session.get(s["id"], [])
            for i, item in enumerate(items, 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=item["tovar_nomi"] or "")
                ws.cell(row=row, column=3, value=float(item["miqdor"] or 0))
                ws.cell(row=row, column=4, value=item["birlik"] or "dona")
                ws.cell(row=row, column=5, value=float(item["narx"] or 0))
                ws.cell(row=row, column=6, value=float(item["summa"] or 0))
                row += 1

            # ITOGO
            ws.cell(row=row, column=1, value="Jami:").font = BoldFont
            ws.cell(row=row, column=6, value=float(s["jami"] or 0)).font = BoldFont
            row += 2  # bo'shliq blok orasida

        for col in range(1, 7):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = [5, 40, 10, 10, 12, 14][col - 1]

    else:  # sklad_zagruz
        # ═══ 4.1 SKLAD ZAGRUZ — ishlab chiqaruvchi bo'yicha guruh ═══
        ws.title = "Sklad Zagruz 4.1"
        ws.cell(row=1, column=1, value=f"Tovar-transport nakladnoy: {reg['nomi']}").font = Font(bold=True, size=14)
        ws.cell(row=1, column=7, value=str(reg["sana"])).font = BoldFont
        ws.cell(row=2, column=1, value=f"Avtomashina: {reg['mashina_nomi'] or '—'} {reg['mashina_raqami'] or ''}")
        ws.cell(row=3, column=1, value=f"Ekspeditor: {reg['ekspeditor_nomi'] or '—'}  |  Tel: {reg['ekspeditor_tel'] or ''}")

        heads = ["№", "Nomi", "Blok", "Miqdor", "Hajm", "Og'irlik", "Summa", "Qaytarish"]
        for col, h in enumerate(heads, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = HFont; c.fill = HF

        # Guruhlarga bo'lib chiqaramiz — item kategoriyasi bo'yicha
        groups: dict[str, list] = {}
        for s in sessions:
            for item in items_by_session.get(s["id"], []):
                k = item.get("kategoriya") or "Boshqa"
                groups.setdefault(k, []).append(item)

        row = 5
        for grp_name, items in groups.items():
            # Guruh nomi
            ws.cell(row=row, column=2, value=grp_name).font = Font(bold=True, color="4F46E5")
            row += 1

            tot_miqdor = 0.0
            tot_summa = 0.0
            for i, item in enumerate(items, 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=item["tovar_nomi"] or "")
                ws.cell(row=row, column=3, value=f"{float(item['miqdor'] or 0):.0f} bl")
                ws.cell(row=row, column=4, value=float(item["miqdor"] or 0))
                ws.cell(row=row, column=5, value="")
                ws.cell(row=row, column=6, value="")
                ws.cell(row=row, column=7, value=float(item["summa"] or 0))
                tot_miqdor += float(item["miqdor"] or 0)
                tot_summa += float(item["summa"] or 0)
                row += 1

            # Itogo guruh ichida
            ws.cell(row=row, column=2, value="Itogo").font = BoldFont
            ws.cell(row=row, column=3, value=f"{tot_miqdor:.0f} bl").font = BoldFont
            ws.cell(row=row, column=4, value=tot_miqdor).font = BoldFont
            ws.cell(row=row, column=7, value=tot_summa).font = BoldFont
            row += 2

        for col, w in enumerate([5, 40, 10, 10, 10, 10, 14, 14], 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # Auto-width (bir xil qilib)
    if format == "reestr":
        widths = {1: 5, 2: 15, 3: 35, 4: 35, 5: 18, 6: 20, 7: 12, 8: 14, 9: 20}
        for col, w in widths.items():
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname_prefix = {"reestr": "300_Reestr", "nakladnye": "310_Nakladnye", "sklad_zagruz": "410_SkladZagruz"}[format]
    fname = f"{fname_prefix}_{reg['nomi'].replace(' ', '_')}_{reg['sana']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
