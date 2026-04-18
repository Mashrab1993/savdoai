"""
╔══════════════════════════════════════════════════════════════╗
║  SAVDOAI — YANGI KUCHLI ROUTELAR v25.3.2                    ║
║  KPI, Loyalty, To'lov, Qarz Eslatma                        ║
╚══════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import logging

from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel, Field

from shared.database.pool import rls_conn
from services.api.deps import get_uid

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1", tags=["Yangi"])


# ═══ KPI ═══

@router.get("/kpi", tags=["KPI"])
async def kpi_olish(kunlar: int = 30, uid: int = Depends(get_uid)):
    """Agent KPI ko'rsatkichlari — reyting, badge, trend."""
    from shared.services.kpi_engine import agent_kpi
    async with rls_conn(uid) as c:
        return await agent_kpi(c, uid, kunlar=kunlar)


@router.get("/kpi/trend", tags=["KPI"])
async def kpi_trend(kunlar: int = 14, uid: int = Depends(get_uid)):
    """Kunlik sotuv trendi — grafik uchun."""
    from shared.services.kpi_engine import kunlik_trend
    async with rls_conn(uid) as c:
        return await kunlik_trend(c, uid, kunlar=kunlar)


@router.get("/kpi/leaderboard", tags=["KPI"])
async def kpi_leaderboard(kunlar: int = 30, uid: int = Depends(get_uid)):
    """Top sotuvchilar reytingi."""
    from shared.services.kpi_engine import leaderboard
    from shared.database.pool import get_pool
    async with get_pool().acquire() as c:
        return await leaderboard(c, kunlar=kunlar)


# ═══ QARZ ESLATMA ═══

@router.get("/qarz/eslatma", tags=["Qarz"])
async def qarz_eslatma_royxati(uid: int = Depends(get_uid)):
    """Eslatma kerak bo'lgan qarzlar ro'yxati."""
    from shared.services.qarz_eslatma import qarz_eslatma_royxati as _royxat
    async with rls_conn(uid) as c:
        return await _royxat(c, uid)


@router.get("/qarz/xulosa", tags=["Qarz"])
async def qarz_xulosa(uid: int = Depends(get_uid)):
    """Kunlik qarz xulosa — do'konchi uchun."""
    from shared.services.qarz_eslatma import kunlik_qarz_xulosa
    async with rls_conn(uid) as c:
        return await kunlik_qarz_xulosa(c, uid)


# ═══ LOYALTY ═══

@router.get("/loyalty/{klient_id}", tags=["Loyalty"])
async def loyalty_profil(klient_id: int, uid: int = Depends(get_uid)):
    """Klient loyalty profili — ball, daraja, keyingi daraja."""
    from shared.services.loyalty import klient_loyalty_profil
    async with rls_conn(uid) as c:
        return await klient_loyalty_profil(c, uid, klient_id)


class BallSarflashSorov(BaseModel):
    ball: int = Field(..., gt=0)
    izoh: str = Field("Chegirma")


@router.post("/loyalty/{klient_id}/sarflash", tags=["Loyalty"])
async def loyalty_sarflash(klient_id: int, data: BallSarflashSorov,
                            uid: int = Depends(get_uid)):
    """Klient ballini chegirmaga sarflash."""
    from shared.services.loyalty import klient_ball_sarflash
    async with rls_conn(uid) as c:
        natija = await klient_ball_sarflash(c, uid, klient_id, data.ball, data.izoh)
    if natija["status"] == "yetarli_emas":
        raise HTTPException(400,
            f"Ball yetarli emas. Mavjud: {natija['mavjud']}, so'ralgan: {natija['soralgan']}")
    return natija


# ═══ TO'LOV ═══

@router.get("/tolov/providerlar", tags=["To'lov"])
async def tolov_providerlar(uid: int = Depends(get_uid)):
    """Mavjud to'lov providerlar ro'yxati."""
    from shared.services.tolov_integratsiya import mavjud_providerlar
    return {"providerlar": mavjud_providerlar()}


class TolovLinkSorov(BaseModel):
    provider: str = Field("click", pattern="^(click|payme)$")
    summa: float = Field(..., gt=0)
    order_id: str = Field(..., min_length=1)
    tavsif: str = Field("")


@router.post("/tolov/link", tags=["To'lov"])
async def tolov_link_yaratish(data: TolovLinkSorov, uid: int = Depends(get_uid)):
    """To'lov linki yaratish (Click/Payme)."""
    from shared.services.tolov_integratsiya import get_provider, TolovSorov
    from decimal import Decimal
    provider = get_provider(data.provider)
    if not provider:
        raise HTTPException(400, f"'{data.provider}' sozlanmagan. CLICK_MERCHANT_ID yoki PAYME_MERCHANT_ID env kerak.")
    sorov = TolovSorov(
        order_id=data.order_id,
        summa=Decimal(str(data.summa)),
        tavsif=data.tavsif,
    )
    link = await provider.link_yaratish(sorov)
    return {"link": link, "provider": data.provider, "summa": data.summa}


# ═══ OMBOR PROGNOZ ═══

@router.get("/ombor/prognoz", tags=["Tovarlar"])
async def ombor_prognoz_v2(kunlar: int = 30, limit: int = 500,
                            kategoriya: str | None = None,
                            uid: int = Depends(get_uid)):
    """Tovar tugash prognozi — qolgan kunlar, buyurtma tavsiyasi."""
    from shared.services.ombor_prognoz import ombor_prognoz, kam_qoldiq_xulosa
    async with rls_conn(uid) as c:
        tovarlar = await ombor_prognoz(c, uid, kunlar=kunlar, limit=limit)
        xulosa = await kam_qoldiq_xulosa(c, uid)
    if kategoriya:
        tovarlar = [t for t in tovarlar if t.get("kategoriya") == kategoriya]
    return {"tovarlar": tovarlar, "xulosa": xulosa, "kunlar": kunlar}


@router.get("/ombor/prognoz/excel", tags=["Tovarlar"])
async def ombor_prognoz_excel(kunlar: int = 30, uid: int = Depends(get_uid)):
    """Ombor prognozini Excel faylga export qilish."""
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from shared.services.ombor_prognoz import ombor_prognoz

    async with rls_conn(uid) as c:
        tovarlar = await ombor_prognoz(c, uid, kunlar=kunlar, limit=5000)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Ombor prognoz {kunlar}k"

    headers = [
        "№", "Tovar", "Kategoriya", "Birlik",
        "Qoldiq", "Min qoldiq", "Kunlik sotuv", f"Sotilgan ({kunlar}k)",
        "Qolgan kun", "Holat", "Buyurtma", "Buyurtma narxi",
        "Ombor qiymati",
    ]
    widths = [5, 32, 18, 8, 10, 10, 12, 14, 12, 14, 12, 15, 15]

    header_fill = PatternFill(start_color="0A819C", end_color="0A819C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 30

    holat_color = {
        "tugagan":       "FFEBEE",
        "kam":           "FFF3E0",
        "xavfli":        "FFF9C4",
        "ogohlantirish": "FFFDE7",
        "diqqat":        "F1F8E9",
        "yaxshi":        "E8F5E9",
    }

    for idx, t in enumerate(tovarlar, 2):
        vals = [
            idx - 1,
            t.get("nomi") or "",
            t.get("kategoriya") or "",
            t.get("birlik") or "",
            t.get("qoldiq") or 0,
            t.get("min_qoldiq") or 0,
            t.get("kunlik_sotuv") or 0,
            t.get("sotilgan") or 0,
            t.get("qolgan_kun") if t.get("qolgan_kun") is not None else "∞",
            (t.get("holat") or "").upper(),
            t.get("buyurtma_miqdor") or 0,
            t.get("buyurtma_narx") or 0,
            t.get("ombor_qiymati") or 0,
        ]
        color = holat_color.get(t.get("holat") or "yaxshi", "FFFFFF")
        row_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=idx, column=col, value=v)
            cell.border = border
            cell.fill = row_fill
            if col in (5, 6, 7, 8, 11, 12, 13):
                cell.number_format = '#,##0.##'
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:M{len(tovarlar) + 1}"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return {
        "filename": f"Ombor_prognoz_{kunlar}kun.xlsx",
        "content_base64": base64.b64encode(buf.getvalue()).decode(),
        "soni": len(tovarlar),
    }


# ═══ AI BUSINESS ADVISOR ═══

@router.get("/advisor", tags=["AI"])
async def ai_advisor(uid: int = Depends(get_uid)):
    """AI biznes maslahat — aqlli insightlar."""
    from shared.services.ai_advisor import biznes_tahlil
    async with rls_conn(uid) as c:
        return await biznes_tahlil(c, uid)


# ═══ SUBSCRIPTION ═══

@router.get("/tarif", tags=["Subscription"])
async def tarif_info(uid: int = Depends(get_uid)):
    """Foydalanuvchi tarif ma'lumoti."""
    from shared.services.subscription import user_tarif_tekshir
    async with rls_conn(uid) as c:
        return await user_tarif_tekshir(c, uid)


# ═══ KLIENT SEGMENTATSIYA ═══

@router.get("/segment", tags=["Klientlar"])
async def klient_segmentatsiya_api(uid: int = Depends(get_uid)):
    """Klientlar RFM segmentatsiyasi — champion, loyal, at_risk va h.k."""
    from shared.services.klient_segment import klientlar_segmentatsiya
    async with rls_conn(uid) as c:
        return await klientlar_segmentatsiya(c, uid)


@router.get("/hisobot/oylik/batafsil", tags=["Hisobot"])
async def oylik_hisobot_api(yil: int = 2026, oy: int = 3,
                             uid: int = Depends(get_uid)):
    """Oylik batafsil hisobot — yil/oy bo'yicha top tovar, klient, trend."""
    from shared.services.oylik_hisobot import oylik_hisobot_data
    async with rls_conn(uid) as c:
        return await oylik_hisobot_data(c, uid, yil, oy)


# ═══ AI TALAB PROGNOZI (Starbucks Deep Brew ilhomi) ═══

@router.get("/forecast/demand", tags=["AI"])
async def demand_forecast_api(kunlar: int = 7, uid: int = Depends(get_uid)):
    """AI talab prognozi — tovar tugash bashorati va buyurtma tavsiyasi."""
    from shared.services.demand_forecast import talab_prognozi
    async with rls_conn(uid) as c:
        return await talab_prognozi(c, uid, kunlar)


# ═══ KLIENT CLV (Salesforce ilhomi) ═══

@router.get("/klient/clv", tags=["Klientlar"])
async def klient_clv_api(top: int = 20, uid: int = Depends(get_uid)):
    """Klient Lifetime Value — eng qimmat klientlar."""
    from shared.services.klient_clv import klient_clv
    async with rls_conn(uid) as c:
        return await klient_clv(c, uid, top)
