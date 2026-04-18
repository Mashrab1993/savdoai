"""
╔══════════════════════════════════════════════════════════════════════╗
║  SAVDOAI v25.4.0 — TOVAR KLASSIFIKATORLARI API                      ║
║                                                                      ║
║  7 turdagi klassifikator (kategoriya, subkat, gruppa, brend,         ║
║  ishlab_chiqaruvchi, segment, gruppa_kategoriya) — yagona endpoint.  ║
║                                                                      ║
║  SalesDoc /settings/view/productCategory funksiyalariga analog,     ║
║  lekin bitta API va bitta jadval — tezroq, toza va yengil.          ║
╚══════════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, validator
from shared.database.pool import rls_conn
from services.api.deps import get_uid

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/klassifikator", tags=["Klassifikator"])

ALLOWED_TURI = {
    "kategoriya", "subkategoriya", "gruppa", "brend",
    "ishlab_chiqaruvchi", "segment", "gruppa_kategoriya",
}


class KlfIn(BaseModel):
    turi: str = Field(..., description="Turi (kategoriya, brend, ...)")
    nomi: str = Field(..., min_length=1, max_length=200)
    kod: str | None = Field(None, max_length=50)
    davlat: str | None = Field(None, max_length=100)
    birlik_id: int | None = None
    parent_id: int | None = None
    tartib: int = 0
    faol: bool = True

    @validator("turi")
    def _check_turi(cls, v):
        if v not in ALLOWED_TURI:
            raise ValueError(f"turi '{v}' ruxsat etilmagan. Ruxsatetilgan: {sorted(ALLOWED_TURI)}")
        return v

    @validator("nomi")
    def _trim_nomi(cls, v):
        v = (v or "").strip()
        if not v:
            raise ValueError("Nomi bo'sh bo'lishi mumkin emas")
        return v


@router.get("")
async def list_klassifikatorlar(
    turi: str | None = Query(None, description="Filter: kategoriya/brend/..."),
    faol: bool | None = Query(None, description="Faqat faol/nofaol"),
    uid: int = Depends(get_uid),
):
    """Barcha klassifikatorlarni ro'yxatlaymiz. turi bermasa — barchasi."""
    if turi and turi not in ALLOWED_TURI:
        raise HTTPException(400, f"turi '{turi}' noto'g'ri")

    where = ["user_id = $1"]
    params: list = [uid]
    if turi:
        params.append(turi)
        where.append(f"turi = ${len(params)}")
    if faol is not None:
        params.append(faol)
        where.append(f"faol = ${len(params)}")

    sql = f"""
        SELECT k.*,
               p.nomi AS parent_nomi,
               b.nomi AS birlik_nomi,
               (SELECT COUNT(*) FROM tovarlar t
                WHERE t.user_id = k.user_id
                  AND CASE k.turi
                      WHEN 'kategoriya' THEN t.kategoriya_id = k.id
                      WHEN 'subkategoriya' THEN t.subkategoriya_id = k.id
                      WHEN 'gruppa' THEN t.gruppa_id = k.id
                      WHEN 'brend' THEN t.brend_id = k.id
                      WHEN 'ishlab_chiqaruvchi' THEN t.ishlab_chiqaruvchi_id = k.id
                      WHEN 'segment' THEN t.segment_id = k.id
                      WHEN 'gruppa_kategoriya' THEN t.gruppa_kategoriya_id = k.id
                      ELSE FALSE
                  END
               ) AS tovar_soni
        FROM tovar_klassifikatorlari k
        LEFT JOIN tovar_klassifikatorlari p ON p.id = k.parent_id
        LEFT JOIN birliklar b ON b.id = k.birlik_id
        WHERE {' AND '.join(where)}
        ORDER BY k.turi, k.tartib, k.nomi
    """

    async with rls_conn(uid) as c:
        rows = await c.fetch(sql, *params)

    result: dict[str, list] = {t: [] for t in ALLOWED_TURI}
    for r in rows:
        result[r["turi"]].append({
            "id": r["id"],
            "turi": r["turi"],
            "nomi": r["nomi"],
            "kod": r["kod"],
            "davlat": r["davlat"],
            "birlik_id": r["birlik_id"],
            "birlik_nomi": r["birlik_nomi"],
            "parent_id": r["parent_id"],
            "parent_nomi": r["parent_nomi"],
            "tartib": r["tartib"],
            "faol": r["faol"],
            "tovar_soni": int(r["tovar_soni"] or 0),
            "yaratilgan": r["yaratilgan"].isoformat() if r["yaratilgan"] else None,
        })

    totals = {t: len(v) for t, v in result.items()}
    return {"items": result, "totals": totals, "jami": sum(totals.values())}


@router.post("")
async def create_klassifikator(body: KlfIn, uid: int = Depends(get_uid)):
    """Yangi klassifikator qo'shamiz."""
    async with rls_conn(uid) as c:
        # Duplicate check
        dup = await c.fetchval("""
            SELECT id FROM tovar_klassifikatorlari
            WHERE user_id=$1 AND turi=$2 AND LOWER(nomi)=LOWER($3)
        """, uid, body.turi, body.nomi)
        if dup:
            raise HTTPException(409, f"'{body.nomi}' allaqachon mavjud ({body.turi})")

        row = await c.fetchrow("""
            INSERT INTO tovar_klassifikatorlari
                (user_id, turi, nomi, kod, davlat, birlik_id, parent_id, tartib, faol)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
            RETURNING *
        """, uid, body.turi, body.nomi, body.kod, body.davlat,
             body.birlik_id, body.parent_id, body.tartib, body.faol)

    return dict(row) | {
        "yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None,
        "yangilangan": row["yangilangan"].isoformat() if row["yangilangan"] else None,
    }


@router.put("/{klf_id}")
async def update_klassifikator(klf_id: int, body: KlfIn, uid: int = Depends(get_uid)):
    """Klassifikatorni yangilaymiz."""
    async with rls_conn(uid) as c:
        exists = await c.fetchval("""
            SELECT turi FROM tovar_klassifikatorlari WHERE id=$1 AND user_id=$2
        """, klf_id, uid)
        if not exists:
            raise HTTPException(404, "Klassifikator topilmadi")
        if exists != body.turi:
            raise HTTPException(400, "Turini o'zgartirib bo'lmaydi")

        # Duplicate check (boshqa yozuvda shu nom bormi)
        dup = await c.fetchval("""
            SELECT id FROM tovar_klassifikatorlari
            WHERE user_id=$1 AND turi=$2 AND LOWER(nomi)=LOWER($3) AND id != $4
        """, uid, body.turi, body.nomi, klf_id)
        if dup:
            raise HTTPException(409, f"'{body.nomi}' boshqa yozuvda band")

        row = await c.fetchrow("""
            UPDATE tovar_klassifikatorlari
            SET nomi=$1, kod=$2, davlat=$3, birlik_id=$4, parent_id=$5,
                tartib=$6, faol=$7, yangilangan=NOW()
            WHERE id=$8 AND user_id=$9
            RETURNING *
        """, body.nomi, body.kod, body.davlat, body.birlik_id,
             body.parent_id, body.tartib, body.faol, klf_id, uid)

    return dict(row) | {
        "yaratilgan": row["yaratilgan"].isoformat() if row["yaratilgan"] else None,
        "yangilangan": row["yangilangan"].isoformat() if row["yangilangan"] else None,
    }


@router.delete("/{klf_id}")
async def delete_klassifikator(klf_id: int, uid: int = Depends(get_uid)):
    """O'chiramiz (tovar *_id FK SET NULL bo'ladi)."""
    async with rls_conn(uid) as c:
        row = await c.fetchrow("""
            DELETE FROM tovar_klassifikatorlari
            WHERE id=$1 AND user_id=$2 RETURNING id, turi, nomi
        """, klf_id, uid)
    if not row:
        raise HTTPException(404, "Klassifikator topilmadi")
    return {"ok": True, "deleted": dict(row)}


@router.post("/reorder")
async def reorder_klassifikatorlar(
    items: list[dict] = Body(..., description="[{id, tartib}]"),
    uid: int = Depends(get_uid),
):
    """Tartiblashtirish: [{id, tartib}] ro'yxati."""
    if not items:
        return {"ok": True, "updated": 0}
    async with rls_conn(uid) as c:
        async with c.transaction():
            for it in items:
                await c.execute(
                    "UPDATE tovar_klassifikatorlari SET tartib=$1 WHERE id=$2 AND user_id=$3",
                    int(it.get("tartib", 0)), int(it["id"]), uid,
                )
    return {"ok": True, "updated": len(items)}


# ═══════════════════════════════════════════════════════════════════
# EXCEL EXPORT / IMPORT — SalesDoc format mos
# ═══════════════════════════════════════════════════════════════════

TURI_RU = {
    "kategoriya": "Категория товара",
    "subkategoriya": "Подкатегория",
    "gruppa": "Группа",
    "brend": "Бренд",
    "ishlab_chiqaruvchi": "Производитель",
    "segment": "Сегмент",
    "gruppa_kategoriya": "Группа категорий",
}


@router.get("/export")
async def export_xlsx(
    turi: str | None = Query(None),
    uid: int = Depends(get_uid),
):
    """SalesDoc formatida Excel eksport qiladi.

    Agar turi berilsa — faqat shu tur; yo'qligida — har tur alohida sheet.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl kutubxonasi mavjud emas")

    if turi and turi not in ALLOWED_TURI:
        raise HTTPException(400, f"turi '{turi}' noto'g'ri")

    turlari = [turi] if turi else list(ALLOWED_TURI)

    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    async with rls_conn(uid) as c:
        for t in turlari:
            rows = await c.fetch("""
                SELECT k.id, k.nomi, k.kod, k.davlat, k.tartib, k.faol,
                       b.nomi AS birlik_nomi,
                       p.nomi AS parent_nomi
                FROM tovar_klassifikatorlari k
                LEFT JOIN birliklar b ON b.id = k.birlik_id
                LEFT JOIN tovar_klassifikatorlari p ON p.id = k.parent_id
                WHERE k.user_id=$1 AND k.turi=$2
                ORDER BY k.tartib, k.nomi
            """, uid, t)

            ws = wb.create_sheet(title=TURI_RU.get(t, t)[:31])

            # Columns: bazaviylar barcha turlari uchun
            headers = ["ID", "Название", "Код", "Сортировка", "Активный"]
            if t == "kategoriya":
                headers.append("Ед. измерения")
            elif t == "subkategoriya":
                headers.append("Родительская категория")
            elif t == "ishlab_chiqaruvchi":
                headers.append("Страна")

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for i, r in enumerate(rows, 2):
                ws.cell(row=i, column=1, value=f"s{r['id']}")  # SavdoAI prefiksi
                ws.cell(row=i, column=2, value=r["nomi"])
                ws.cell(row=i, column=3, value=r["kod"] or "")
                ws.cell(row=i, column=4, value=int(r["tartib"] or 0))
                ws.cell(row=i, column=5, value="Ha" if r["faol"] else "Yo'q")
                if t == "kategoriya":
                    ws.cell(row=i, column=6, value=r["birlik_nomi"] or "")
                elif t == "subkategoriya":
                    ws.cell(row=i, column=6, value=r["parent_nomi"] or "")
                elif t == "ishlab_chiqaruvchi":
                    ws.cell(row=i, column=6, value=r["davlat"] or "")

            # Auto-width
            for col in range(1, len(headers) + 1):
                max_len = max(
                    (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
                    default=10,
                )
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 50)
            ws.freeze_panes = "A2"

    if not wb.sheetnames:
        raise HTTPException(404, "Ma'lumot yo'q")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"klassifikator_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    if turi:
        fname = f"{TURI_RU.get(turi, turi).replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/import")
async def import_xlsx(
    turi: str = Query(..., description="Import qilinayotgan tur"),
    file: UploadFile = File(...),
    uid: int = Depends(get_uid),
):
    """Excel fayldan import. Format: A=ID (ignored on new), B=Nomi, C=Kod, D=Sort, E=Aktiv, F=qo'shimcha.

    ID ustunidagi "s123" — SavdoAI yozuvi (update qilinadi); boshqalar yangi INSERT.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl mavjud emas")

    if turi not in ALLOWED_TURI:
        raise HTTPException(400, f"turi '{turi}' noto'g'ri")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(413, "Fayl juda katta (max 10MB)")

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel fayl noto'g'ri: {e}")

    ws = wb.active
    if ws.max_row < 2:
        raise HTTPException(400, "Excelda ma'lumot yo'q")

    # Birliklar lookup (agar kategoriya bo'lsa)
    birliklar_map: dict[str, int] = {}
    parent_map: dict[str, int] = {}

    yangi = 0
    yangilandi = 0
    xatolar: list[str] = []

    async with rls_conn(uid) as c:
        if turi == "kategoriya":
            for r in await c.fetch("SELECT id, nomi FROM birliklar WHERE user_id=$1", uid):
                birliklar_map[r["nomi"].lower().strip()] = r["id"]
        if turi == "subkategoriya":
            for r in await c.fetch(
                "SELECT id, nomi FROM tovar_klassifikatorlari WHERE user_id=$1 AND turi='kategoriya'",
                uid,
            ):
                parent_map[r["nomi"].lower().strip()] = r["id"]

        async with c.transaction():
            for row_idx in range(2, ws.max_row + 1):
                try:
                    id_val = ws.cell(row=row_idx, column=1).value
                    nomi = (ws.cell(row=row_idx, column=2).value or "")
                    kod = ws.cell(row=row_idx, column=3).value
                    tartib = ws.cell(row=row_idx, column=4).value or 0
                    aktiv_val = ws.cell(row=row_idx, column=5).value
                    qoshimcha = ws.cell(row=row_idx, column=6).value

                    if not isinstance(nomi, str):
                        nomi = str(nomi or "").strip()
                    else:
                        nomi = nomi.strip()
                    if not nomi:
                        continue

                    kod_str = None if kod is None or kod == "" else str(kod).strip()
                    try:
                        tartib_int = int(float(tartib or 0))
                    except (ValueError, TypeError):
                        tartib_int = 0

                    if isinstance(aktiv_val, bool):
                        faol = aktiv_val
                    elif isinstance(aktiv_val, str):
                        faol = aktiv_val.strip().lower() in ("ha", "yes", "1", "true", "да", "активный")
                    else:
                        faol = True

                    birlik_id = None
                    davlat = None
                    parent_id = None
                    if turi == "kategoriya" and qoshimcha:
                        birlik_id = birliklar_map.get(str(qoshimcha).lower().strip())
                    elif turi == "subkategoriya" and qoshimcha:
                        parent_id = parent_map.get(str(qoshimcha).lower().strip())
                    elif turi == "ishlab_chiqaruvchi" and qoshimcha:
                        davlat = str(qoshimcha).strip()

                    # ID orqali update
                    existing_id = None
                    if isinstance(id_val, str) and id_val.startswith("s"):
                        try:
                            cand_id = int(id_val[1:])
                            row = await c.fetchval(
                                "SELECT id FROM tovar_klassifikatorlari WHERE id=$1 AND user_id=$2 AND turi=$3",
                                cand_id, uid, turi,
                            )
                            existing_id = row
                        except ValueError:
                            pass

                    if not existing_id:
                        existing_id = await c.fetchval(
                            """SELECT id FROM tovar_klassifikatorlari
                               WHERE user_id=$1 AND turi=$2 AND LOWER(nomi)=LOWER($3)""",
                            uid, turi, nomi,
                        )

                    if existing_id:
                        await c.execute("""
                            UPDATE tovar_klassifikatorlari
                            SET nomi=$1, kod=$2, tartib=$3, faol=$4,
                                birlik_id=$5, davlat=$6, parent_id=$7, yangilangan=NOW()
                            WHERE id=$8 AND user_id=$9
                        """, nomi, kod_str, tartib_int, faol,
                             birlik_id, davlat, parent_id, existing_id, uid)
                        yangilandi += 1
                    else:
                        await c.execute("""
                            INSERT INTO tovar_klassifikatorlari
                                (user_id, turi, nomi, kod, tartib, faol,
                                 birlik_id, davlat, parent_id)
                            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                        """, uid, turi, nomi, kod_str, tartib_int, faol,
                             birlik_id, davlat, parent_id)
                        yangi += 1
                except Exception as e:
                    xatolar.append(f"Qator {row_idx}: {e}")
                    if len(xatolar) > 20:
                        xatolar.append("... (ko'p xato — to'xtatildi)")
                        break

    return {
        "ok": True,
        "turi": turi,
        "yangi_qoshildi": yangi,
        "yangilandi": yangilandi,
        "xatolar": xatolar,
    }
