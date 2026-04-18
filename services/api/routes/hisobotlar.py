"""
╔══════════════════════════════════════════════════════════════╗
║  SAVDOAI — HISOBOT ROUTELARI                                ║
║  Kunlik, haftalik, oylik, foyda, statistika                 ║
║  main.py dan ajratildi — v25.3.2                            ║
╚══════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import logging
from datetime import datetime, timezone
from decimal import Decimal

from fastapi import APIRouter, HTTPException, Depends

from shared.database.pool import rls_conn
from shared.cache.redis_cache import cache_ol, cache_yoz, TTL_HISOBOT
from shared.utils import like_escape
from services.api.deps import get_uid

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1", tags=["Hisobotlar"])


@router.get("/hisobot/haftalik")
async def hisobot_haftalik(uid: int = Depends(get_uid)):
    """7 kunlik hisobot"""
    cache_k = f"hisobot:haftalik:{uid}"
    cached = await cache_ol(cache_k)
    if cached:
        return cached
    async with rls_conn(uid) as c:
        ch = await c.fetchrow("""
            SELECT COUNT(*) n,
                   COALESCE(SUM(jami),0)     jami,
                   COALESCE(SUM(qarz),0)     qarz,
                   COALESCE(SUM(tolangan),0) tolangan
            FROM sotuv_sessiyalar
            WHERE sana >= NOW() - INTERVAL '7 days'
        """)
        kr = await c.fetchrow("""
            SELECT COUNT(*) n, COALESCE(SUM(jami),0) jami
            FROM kirimlar WHERE sana >= NOW() - INTERVAL '7 days'
        """)
        top_klientlar = [dict(r) for r in await c.fetch("""
            SELECT klient_ismi, SUM(jami) jami, COUNT(*) soni
            FROM sotuv_sessiyalar
            WHERE sana >= NOW() - INTERVAL '7 days' AND klient_ismi IS NOT NULL
            GROUP BY klient_ismi ORDER BY jami DESC LIMIT 5
        """)]
    result = {
        "davr": "7 kun",
        "sotuv": {"soni": int(ch["n"]), "jami": float(ch["jami"]),
                  "qarz": float(ch["qarz"])},
        "kirim": {"soni": int(kr["n"]), "jami": float(kr["jami"])},
        "top_klientlar": top_klientlar,
    }
    await cache_yoz(cache_k, result, TTL_HISOBOT * 6)
    return result


@router.get("/hisobot/oylik")
async def hisobot_oylik(uid: int = Depends(get_uid)):
    """30 kunlik hisobot"""
    cache_k = f"hisobot:oylik:{uid}"
    cached = await cache_ol(cache_k)
    if cached:
        return cached
    async with rls_conn(uid) as c:
        ch = await c.fetchrow("""
            SELECT COUNT(*) n,
                   COALESCE(SUM(jami),0) jami,
                   COALESCE(SUM(qarz),0) qarz,
                   COALESCE(SUM(tolangan),0) tolangan
            FROM sotuv_sessiyalar
            WHERE sana >= NOW() - INTERVAL '30 days'
        """)
        foyda = await c.fetchrow("""
            SELECT COALESCE(SUM(ch.jami - ch.miqdor*ch.olish_narxi),0) sof_foyda
            FROM chiqimlar ch WHERE sana >= NOW() - INTERVAL '30 days'
        """)
        top5_tovar = [dict(r) for r in await c.fetch("""
            SELECT tovar_nomi, SUM(miqdor) miqdor, SUM(jami) jami
            FROM chiqimlar WHERE sana >= NOW() - INTERVAL '30 days'
            GROUP BY tovar_nomi ORDER BY jami DESC LIMIT 5
        """)]
    result = {
        "davr": "30 kun",
        "sotuv": {"soni": int(ch["n"]), "jami": float(ch["jami"])},
        "sof_foyda": float(foyda["sof_foyda"] or 0),
        "top_tovarlar": top5_tovar,
    }
    await cache_yoz(cache_k, result, TTL_HISOBOT * 12)
    return result


@router.get("/hisobot/foyda")
async def hisobot_foyda(kunlar: int = 30, uid: int = Depends(get_uid)):
    """Foyda tahlili — sof foyda, xarajatlar, top foyda/zarar tovarlar."""
    async with rls_conn(uid) as c:
        foyda = await c.fetchrow("""
            SELECT
                COALESCE(SUM(ch.jami), 0) AS brutto,
                COALESCE(SUM(ch.miqdor * ch.olish_narxi), 0) AS tannarx,
                COALESCE(SUM(ch.jami - ch.miqdor * ch.olish_narxi), 0) AS sof_foyda
            FROM chiqimlar ch
            JOIN sotuv_sessiyalar ss ON ss.id = ch.sessiya_id
            WHERE ss.sana >= NOW() - make_interval(days => $1)
        """, kunlar)

        xarajat = await c.fetchval("""
            SELECT COALESCE(SUM(summa), 0)
            FROM xarajatlar
            WHERE user_id = $1
              AND NOT bekor_qilingan
              AND sana >= NOW() - make_interval(days => $2)
        """, uid, kunlar)

        top_foyda = await c.fetch("""
            SELECT ch.tovar_nomi,
                   SUM(ch.jami - ch.miqdor * ch.olish_narxi) AS foyda,
                   SUM(ch.miqdor) AS miqdor
            FROM chiqimlar ch
            JOIN sotuv_sessiyalar ss ON ss.id = ch.sessiya_id
            WHERE ss.sana >= NOW() - make_interval(days => $1) AND ch.olish_narxi > 0
            GROUP BY ch.tovar_nomi ORDER BY foyda DESC LIMIT 5
        """, kunlar)

        top_zarar = await c.fetch("""
            SELECT ch.tovar_nomi,
                   SUM(ch.jami - ch.miqdor * ch.olish_narxi) AS foyda,
                   SUM(ch.miqdor) AS miqdor
            FROM chiqimlar ch
            JOIN sotuv_sessiyalar ss ON ss.id = ch.sessiya_id
            WHERE ss.sana >= NOW() - make_interval(days => $1) AND ch.olish_narxi > 0
            GROUP BY ch.tovar_nomi
            HAVING SUM(ch.jami - ch.miqdor * ch.olish_narxi) < 0
            ORDER BY foyda ASC LIMIT 5
        """, kunlar)

    sof = float(foyda["sof_foyda"] or 0)
    xar = float(xarajat or 0)
    brutto = float(foyda["brutto"] or 0)
    return {
        "kunlar": kunlar,
        "brutto_sotuv": brutto,
        "tannarx": float(foyda["tannarx"] or 0),
        "sof_foyda": sof,
        "xarajatlar": xar,
        "toza_foyda": sof - xar,
        "margin_foiz": round(sof / brutto * 100, 1) if brutto > 0 else 0,
        "top_foyda": [{"nomi": r["tovar_nomi"], "foyda": float(r["foyda"]),
                       "miqdor": float(r["miqdor"])} for r in top_foyda],
        "top_zarar": [{"nomi": r["tovar_nomi"], "zarar": abs(float(r["foyda"])),
                       "miqdor": float(r["miqdor"])} for r in top_zarar],
    }


@router.get("/dashboard/summary", tags=["Dashboard"])
async def dashboard_summary(uid: int = Depends(get_uid)):
    """
    Dashboard uchun YAGONA aggregated endpoint.

    Avval dashboard 5-6 ta API chaqirar edi (statistika, dashboard,
    top, agentlar, heatmap). Endi hammasi BITTA endpoint'da — tezroq
    yuklanadi, kamroq server yuki.

    Response shape matches KpiGridPremium + AgentKpiBoard +
    SalesHeatmap + alert banners.
    """
    async with rls_conn(uid) as c:
        # 1. Bugungi sotuv
        bugun = await c.fetchrow("""
            SELECT COUNT(*) AS soni, COALESCE(SUM(jami), 0) AS jami
            FROM sotuv_sessiyalar
            WHERE user_id = $1
              AND (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
              AND COALESCE(holat, 'yangi') != 'bekor'
        """, uid)

        # 2. Haftalik
        hafta = await c.fetchrow("""
            SELECT COUNT(*) AS soni, COALESCE(SUM(jami), 0) AS jami
            FROM sotuv_sessiyalar
            WHERE user_id = $1 AND sana >= NOW() - interval '7 days'
        """, uid)

        # 3. Oylik
        oy = await c.fetchrow("""
            SELECT COUNT(*) AS soni, COALESCE(SUM(jami), 0) AS jami
            FROM sotuv_sessiyalar
            WHERE user_id = $1 AND sana >= NOW() - interval '30 days'
        """, uid)

        # 4. Qarzlar
        faol_qarz = await c.fetchval(
            "SELECT COALESCE(SUM(qolgan), 0) FROM qarzlar "
            "WHERE user_id=$1 AND yopildi=FALSE AND qolgan>0", uid) or 0

        muddat_otgan = await c.fetchval(
            "SELECT COUNT(*) FROM qarzlar "
            "WHERE user_id=$1 AND yopildi=FALSE AND qolgan>0 "
            "AND muddat IS NOT NULL AND muddat < NOW()", uid) or 0

        # 5. Faol klientlar (30 kun)
        faol_klient = await c.fetchval("""
            SELECT COUNT(DISTINCT klient_id)
            FROM sotuv_sessiyalar
            WHERE user_id = $1 AND sana >= NOW() - interval '30 days'
              AND klient_id IS NOT NULL
        """, uid) or 0

        # 6. Kam qoldiq
        kam_qoldiq_soni = await c.fetchval(
            "SELECT COUNT(*) FROM tovarlar "
            "WHERE user_id=$1 AND min_qoldiq > 0 AND qoldiq <= min_qoldiq", uid) or 0

        kam_qoldiq_tovarlar = await c.fetch("""
            SELECT id, nomi, qoldiq, min_qoldiq, birlik
            FROM tovarlar
            WHERE user_id=$1 AND min_qoldiq > 0 AND qoldiq <= min_qoldiq
            ORDER BY qoldiq ASC LIMIT 5
        """, uid)

        # 7. Bugun top 5
        top_bugun = await c.fetch("""
            SELECT c.tovar_nomi AS nomi,
                   SUM(c.miqdor) AS miqdor,
                   SUM(c.jami) AS jami
            FROM chiqimlar c
            JOIN sotuv_sessiyalar ss ON ss.id = c.sessiya_id
            WHERE ss.user_id = $1
              AND (ss.sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
            GROUP BY c.tovar_nomi
            ORDER BY SUM(c.jami) DESC LIMIT 5
        """, uid)

    def d(v):
        return float(v) if isinstance(v, Decimal) else v

    return {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "bugun":  {"soni": int(bugun["soni"]), "jami": d(bugun["jami"])},
        "hafta":  {"soni": int(hafta["soni"]), "jami": d(hafta["jami"])},
        "oy":     {"soni": int(oy["soni"]),    "jami": d(oy["jami"])},
        "faol_klientlar": int(faol_klient),
        "faol_qarz": d(faol_qarz),
        "muddat_otgan_qarz": int(muddat_otgan),
        "kam_qoldiq_soni": int(kam_qoldiq_soni),
        "kam_qoldiq_tovarlar": [dict(r) for r in kam_qoldiq_tovarlar],
        "top_bugun": [dict(r) for r in top_bugun],
    }


@router.get("/statistika", tags=["Dashboard"])
async def admin_statistika(uid: int = Depends(get_uid)):
    """Tizim statistikasi — admin uchun umumiy ko'rsatkichlar"""
    async with rls_conn(uid) as c:
        tovar_soni = await c.fetchval(
            "SELECT COUNT(*) FROM tovarlar WHERE user_id=$1", uid) or 0
        klient_soni = await c.fetchval(
            "SELECT COUNT(*) FROM klientlar WHERE user_id=$1", uid) or 0
        faol_qarz = await c.fetchval(
            "SELECT COALESCE(SUM(qolgan), 0) FROM qarzlar "
            "WHERE user_id=$1 AND yopildi=FALSE AND qolgan>0", uid) or 0
        bugun_sotuv = await c.fetchrow("""
            SELECT COUNT(*) AS soni, COALESCE(SUM(jami), 0) AS jami
            FROM sotuv_sessiyalar
            WHERE user_id=$1 AND (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
        """, uid)
        hafta_sotuv = await c.fetchrow("""
            SELECT COUNT(*) AS soni, COALESCE(SUM(jami), 0) AS jami
            FROM sotuv_sessiyalar
            WHERE user_id=$1 AND sana >= NOW() - interval '7 days'
        """, uid)
        oy_sotuv = await c.fetchrow("""
            SELECT COUNT(*) AS soni, COALESCE(SUM(jami), 0) AS jami
            FROM sotuv_sessiyalar
            WHERE user_id=$1 AND sana >= NOW() - interval '30 days'
        """, uid)
        kam_qoldiq = await c.fetchval(
            "SELECT COUNT(*) FROM tovarlar "
            "WHERE user_id=$1 AND min_qoldiq > 0 AND qoldiq <= min_qoldiq", uid) or 0
        muddat_otgan = await c.fetchval(
            "SELECT COUNT(*) FROM qarzlar "
            "WHERE user_id=$1 AND yopildi=FALSE AND qolgan>0 "
            "AND muddat IS NOT NULL AND muddat < NOW()", uid) or 0

        # Kam qoldiq tovarlar ro'yxati (top 10) — dashboard alert uchun
        kam_tovarlar = await c.fetch("""
            SELECT id, nomi, qoldiq, min_qoldiq, birlik, kategoriya
            FROM tovarlar
            WHERE user_id=$1 AND min_qoldiq > 0 AND qoldiq <= min_qoldiq
            ORDER BY qoldiq ASC
            LIMIT 10
        """, uid)

        # Eng ko'p sotilgan tovar bugun
        top_bugun = await c.fetch("""
            SELECT c.tovar_nomi AS nomi,
                   SUM(c.miqdor) AS miqdor,
                   SUM(c.jami) AS jami
            FROM chiqimlar c
            JOIN sotuv_sessiyalar ss ON ss.id = c.sessiya_id
            WHERE ss.user_id = $1
              AND (ss.sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
            GROUP BY c.tovar_nomi
            ORDER BY SUM(c.jami) DESC
            LIMIT 5
        """, uid)

    return {
        "tovar_soni": tovar_soni, "klient_soni": klient_soni,
        "faol_qarz": float(faol_qarz), "kam_qoldiq_soni": kam_qoldiq,
        "muddat_otgan_qarz": muddat_otgan,
        "bugun": {"soni": int(bugun_sotuv["soni"]), "jami": float(bugun_sotuv["jami"])},
        "hafta": {"soni": int(hafta_sotuv["soni"]), "jami": float(hafta_sotuv["jami"])},
        "oy":    {"soni": int(oy_sotuv["soni"]),    "jami": float(oy_sotuv["jami"])},
        "kam_qoldiq_tovarlar": [dict(r) for r in kam_tovarlar],
        "top_bugun": [dict(r) for r in top_bugun],
    }


# ════════════════════════════════════════════════════════════
#  REPORTS — SalesDoc-level reporting
# ════════════════════════════════════════════════════════════

@router.get("/hisobot/pnl")
async def hisobot_pnl(kunlar: int = 30, uid: int = Depends(get_uid)):
    """
    P&L (foyda-zarar) hisoboti — PnLReport komponentiga to'g'ridan
    to'g'ri mos format.

    Response:
      {
        "davr_nomi": "Oxirgi 30 kun",
        "tushum": 528640000,
        "tannarx": 356400000,
        "yalpi_foyda": 172240000,
        "operatsion_xarajatlar": 62180000,
        "sof_foyda": 110060000,
        "qaytarishlar": 2150000,
        "xarajat_kategoriyalar": [...],
        "prev": { "tushum": ..., "sof_foyda": ... }
      }
    """
    async with rls_conn(uid) as c:
        # Current period
        sotuv = await c.fetchrow("""
            SELECT
                COALESCE(SUM(ss.jami), 0)               AS tushum,
                COALESCE(SUM(
                    (SELECT SUM(ch.miqdor * ch.olish_narxi)
                     FROM chiqimlar ch WHERE ch.sessiya_id = ss.id)
                ), 0)                                    AS tannarx
            FROM sotuv_sessiyalar ss
            WHERE user_id = $1
              AND sana >= NOW() - make_interval(days => $2)
              AND COALESCE(holat, 'yangi') != 'bekor'
        """, uid, kunlar)

        xarajat = await c.fetchval("""
            SELECT COALESCE(SUM(summa), 0)
            FROM xarajatlar
            WHERE user_id = $1
              AND sana >= NOW() - make_interval(days => $2)
              AND COALESCE(bekor_qilingan, FALSE) = FALSE
        """, uid, kunlar) or 0

        try:
            qaytarish = await c.fetchval("""
                SELECT COALESCE(SUM(summa), 0)
                FROM qaytarishlar
                WHERE user_id = $1
                  AND sana >= NOW() - make_interval(days => $2)
            """, uid, kunlar) or 0
        except Exception:
            qaytarish = 0

        # Xarajat kategoriyalar breakdown
        try:
            xar_kat = await c.fetch("""
                SELECT
                    COALESCE(kategoriya_nomi, 'Boshqa') AS nomi,
                    SUM(summa)                          AS summa
                FROM xarajatlar
                WHERE user_id = $1
                  AND sana >= NOW() - make_interval(days => $2)
                  AND COALESCE(bekor_qilingan, FALSE) = FALSE
                GROUP BY kategoriya_nomi
                ORDER BY SUM(summa) DESC
            """, uid, kunlar)
        except Exception:
            xar_kat = []

        # Previous period (for delta)
        prev_sotuv = await c.fetchrow("""
            SELECT
                COALESCE(SUM(ss.jami), 0) AS tushum,
                COALESCE(SUM(
                    (SELECT SUM(ch.miqdor * ch.olish_narxi)
                     FROM chiqimlar ch WHERE ch.sessiya_id = ss.id)
                ), 0) AS tannarx
            FROM sotuv_sessiyalar ss
            WHERE user_id = $1
              AND sana >= NOW() - make_interval(days => $2)
              AND sana < NOW() - make_interval(days => $3)
              AND COALESCE(holat, 'yangi') != 'bekor'
        """, uid, kunlar * 2, kunlar)

        prev_xarajat = await c.fetchval("""
            SELECT COALESCE(SUM(summa), 0)
            FROM xarajatlar
            WHERE user_id = $1
              AND sana >= NOW() - make_interval(days => $2)
              AND sana < NOW() - make_interval(days => $3)
              AND COALESCE(bekor_qilingan, FALSE) = FALSE
        """, uid, kunlar * 2, kunlar) or 0

    tushum = float(sotuv["tushum"])
    tannarx = float(sotuv["tannarx"])
    yalpi = tushum - tannarx
    op_x = float(xarajat)
    sof = yalpi - op_x - float(qaytarish)

    prev_tushum = float(prev_sotuv["tushum"])
    prev_tannarx = float(prev_sotuv["tannarx"])
    prev_yalpi = prev_tushum - prev_tannarx
    prev_sof = prev_yalpi - float(prev_xarajat)

    return {
        "davr_nomi": f"Oxirgi {kunlar} kun",
        "tushum": tushum,
        "tannarx": tannarx,
        "yalpi_foyda": yalpi,
        "operatsion_xarajatlar": op_x,
        "sof_foyda": sof,
        "qaytarishlar": float(qaytarish),
        "xarajat_kategoriyalar": [
            {"nomi": str(r["nomi"]), "summa": float(r["summa"])}
            for r in xar_kat
        ],
        "prev": {
            "tushum": prev_tushum,
            "sof_foyda": prev_sof,
        },
    }


@router.get("/hisobot/heatmap")
async def hisobot_heatmap(kunlar: int = 30, uid: int = Depends(get_uid)):
    """
    Sotuv faolligi heatmap — 7 kun × 24 soat matritsa.

    Har hafta kuni va har soat uchun nechta sotuv bo'lganini hisoblab
    [7][24] matritsa qaytaradi. SalesHeatmap komponentiga to'g'ridan-
    to'g'ri ulash mumkin.

    Response:
      {
        "matrix": [[0,0,...,24 ta], ...7 ta],
        "metric": "soni",
        "jami": 1234,
        "kunlar": 30
      }
    """
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT
                EXTRACT(DOW FROM sana AT TIME ZONE 'Asia/Tashkent') AS dow,
                EXTRACT(HOUR FROM sana AT TIME ZONE 'Asia/Tashkent') AS soat,
                COUNT(*) AS soni
            FROM sotuv_sessiyalar
            WHERE user_id = $1
              AND sana >= NOW() - make_interval(days => $2)
              AND COALESCE(holat, 'yangi') != 'bekor'
            GROUP BY 1, 2
        """, uid, kunlar)

    # Build 7×24 matrix (DOW: 0=Sun in PG, we remap to 0=Mon)
    matrix = [[0] * 24 for _ in range(7)]
    jami = 0
    for r in rows:
        dow = int(r["dow"])  # PG: 0=Sun, 1=Mon, ..., 6=Sat
        # Remap to 0=Mon, 1=Tue, ..., 6=Sun
        mapped = (dow - 1) % 7
        soat = int(r["soat"])
        soni = int(r["soni"])
        matrix[mapped][soat] = soni
        jami += soni

    return {
        "matrix": matrix,
        "metric": "soni",
        "jami": jami,
        "kunlar": kunlar,
    }


@router.get("/hisobot/oylik-trend")
async def hisobot_oylik_trend(oylar: int = 6, uid: int = Depends(get_uid)):
    """Oxirgi N oylik sotuv trend — dashboard grafigi uchun."""
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT
                to_char(date_trunc('month', ss.sana), 'YYYY-MM') AS oy,
                to_char(date_trunc('month', ss.sana), 'Mon')     AS oy_nomi,
                COUNT(*)                                         AS soni,
                COALESCE(SUM(ss.jami), 0)                        AS sotuv,
                COALESCE(SUM(
                    (SELECT SUM(ch.jami - ch.miqdor * ch.olish_narxi)
                     FROM chiqimlar ch WHERE ch.sessiya_id = ss.id)
                ), 0)                                            AS foyda
            FROM sotuv_sessiyalar ss
            WHERE user_id = $1
              AND ss.sana >= NOW() - make_interval(months => $2)
            GROUP BY 1, 2
            ORDER BY 1
        """, uid, oylar)
    return [dict(r) for r in rows]


@router.get("/hisobot/top-klientlar")
async def hisobot_top_klientlar(kunlar: int = 30, limit: int = 20, uid: int = Depends(get_uid)):
    """
    Top N klientlar — SalesDoc /report/customer analogi.

    Oxirgi N kun ichida eng ko'p xarid qilgan klientlar.
    SalesPivotTable komponentiga to'g'ridan-to'g'ri mos keladi.
    """
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT
                k.ism AS key,
                COALESCE(SUM(ss.jami), 0)    AS jami,
                COUNT(ss.id)                  AS soni,
                COALESCE(SUM(
                    (SELECT SUM(ch.miqdor) FROM chiqimlar ch WHERE ch.sessiya_id = ss.id)
                ), 0)                         AS miqdor
            FROM sotuv_sessiyalar ss
            JOIN klientlar k ON k.id = ss.klient_id
            WHERE ss.user_id = $1
              AND ss.sana >= NOW() - make_interval(days => $2)
              AND COALESCE(ss.holat, 'yangi') != 'bekor'
              AND ss.klient_id IS NOT NULL
            GROUP BY k.ism
            ORDER BY SUM(ss.jami) DESC
            LIMIT $3
        """, uid, kunlar, limit)
    return [
        {
            "key":    r["key"],
            "jami":   float(r["jami"]),
            "soni":   int(r["soni"]),
            "miqdor": float(r["miqdor"]),
        }
        for r in rows
    ]


@router.get("/hisobot/top-tovarlar")
async def hisobot_top_tovarlar(kunlar: int = 30, limit: int = 20, uid: int = Depends(get_uid)):
    """
    Top N tovarlar — SalesDoc /report/volumeReport analogi.

    Oxirgi N kun ichida eng ko'p sotilgan tovarlar.
    SalesPivotTable komponentiga mos.
    """
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT
                c.tovar_nomi AS key,
                COALESCE(SUM(c.jami), 0)   AS jami,
                COUNT(DISTINCT c.sessiya_id) AS soni,
                COALESCE(SUM(c.miqdor), 0) AS miqdor
            FROM chiqimlar c
            JOIN sotuv_sessiyalar ss ON ss.id = c.sessiya_id
            WHERE ss.user_id = $1
              AND ss.sana >= NOW() - make_interval(days => $2)
              AND COALESCE(ss.holat, 'yangi') != 'bekor'
            GROUP BY c.tovar_nomi
            ORDER BY SUM(c.jami) DESC
            LIMIT $3
        """, uid, kunlar, limit)
    return [
        {
            "key":    r["key"],
            "jami":   float(r["jami"]),
            "soni":   int(r["soni"]),
            "miqdor": float(r["miqdor"]),
        }
        for r in rows
    ]


@router.get("/hisobot/ombor-holati")
async def hisobot_ombor_holati(uid: int = Depends(get_uid)):
    """
    Ombor holati summary — dashboard widget uchun.

    Response:
      {
        "jami_tovar": 245,
        "jami_qiymat": 125400000,
        "kam_qoldiq": 12,
        "tugagan": 5,
        "faol": 230,
        "kategoriya_soni": 8,
        "top_qimmat": [{nomi, qoldiq, qiymat}],
        "top_kam": [{nomi, qoldiq, min_qoldiq}]
      }
    """
    async with rls_conn(uid) as c:
        stats = await c.fetchrow("""
            SELECT
                COUNT(*) AS jami_tovar,
                COUNT(*) FILTER(WHERE faol = TRUE) AS faol,
                COUNT(*) FILTER(WHERE qoldiq <= 0) AS tugagan,
                COUNT(*) FILTER(WHERE min_qoldiq > 0 AND qoldiq <= min_qoldiq AND qoldiq > 0) AS kam_qoldiq,
                COUNT(DISTINCT kategoriya) AS kategoriya_soni,
                COALESCE(SUM(qoldiq * sotish_narxi), 0) AS jami_qiymat
            FROM tovarlar
            WHERE user_id = $1
        """, uid)

        top_qimmat = await c.fetch("""
            SELECT nomi, qoldiq, (qoldiq * sotish_narxi) AS qiymat
            FROM tovarlar
            WHERE user_id = $1 AND qoldiq > 0
            ORDER BY (qoldiq * sotish_narxi) DESC
            LIMIT 5
        """, uid)

        top_kam = await c.fetch("""
            SELECT nomi, qoldiq, min_qoldiq, birlik
            FROM tovarlar
            WHERE user_id = $1 AND min_qoldiq > 0 AND qoldiq <= min_qoldiq
            ORDER BY qoldiq ASC
            LIMIT 5
        """, uid)

    return {
        "jami_tovar":     int(stats["jami_tovar"]),
        "faol":           int(stats["faol"]),
        "tugagan":        int(stats["tugagan"]),
        "kam_qoldiq":     int(stats["kam_qoldiq"]),
        "kategoriya_soni": int(stats["kategoriya_soni"]),
        "jami_qiymat":    float(stats["jami_qiymat"]),
        "top_qimmat":     [dict(r) for r in top_qimmat],
        "top_kam":        [dict(r) for r in top_kam],
    }


@router.get("/hisobot/kunlik-trend")
async def hisobot_kunlik_trend(kunlar: int = 30, uid: int = Depends(get_uid)):
    """Oxirgi N kunlik sotuv — dashboard grafigi uchun (real-time)."""
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT
                to_char((ss.sana AT TIME ZONE 'Asia/Tashkent')::date, 'YYYY-MM-DD') AS sana,
                to_char((ss.sana AT TIME ZONE 'Asia/Tashkent')::date, 'DD')         AS kun,
                COUNT(*)                    AS soni,
                COALESCE(SUM(ss.jami), 0)   AS sotuv,
                COALESCE(SUM(ss.tolangan), 0) AS tolangan,
                COALESCE(SUM(ss.qarz), 0)   AS qarz
            FROM sotuv_sessiyalar ss
            WHERE user_id = $1
              AND ss.sana >= NOW() - make_interval(days => $2)
            GROUP BY 1, 2
            ORDER BY 1
        """, uid, kunlar)
    return [dict(r) for r in rows]


@router.get("/hisobot/agent")
async def hisobot_agent(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    uid: int = Depends(get_uid),
):
    """SalesDoc /report/agent analog — shogird bo'yicha sotuv hisoboti."""
    from datetime import date, timedelta
    sd = date.fromisoformat(sana_dan) if sana_dan else (date.today() - timedelta(days=30))
    sg = date.fromisoformat(sana_gacha) if sana_gacha else date.today()

    async with rls_conn(uid) as c:
        jami = await c.fetchrow("""
            SELECT
                COALESCE(SUM(jami), 0) AS obshchie_zakazy,
                COALESCE(SUM(jami) FILTER (WHERE holat IN ('otgruzka', 'yetkazildi')), 0) AS otgruzka,
                COALESCE(SUM(jami) FILTER (WHERE holat = 'yetkazildi'), 0) AS dostavka,
                COALESCE(SUM(tolangan), 0) AS oplaty,
                COALESCE(SUM(jami - tolangan) FILTER (WHERE jami > tolangan), 0) AS qarz,
                COUNT(DISTINCT klient_id) AS akb
            FROM sotuv_sessiyalar
            WHERE user_id = $1 AND sana::date BETWEEN $2 AND $3
              AND COALESCE(holat, 'yangi') != 'bekor'
        """, uid, sd, sg)

        agentlar: list[dict] = []
        try:
            rows = await c.fetch("""
                SELECT
                    s.id, s.ism,
                    COALESCE(SUM(ss.jami), 0) AS obshchie,
                    COALESCE(SUM(ss.jami) FILTER (WHERE ss.holat IN ('otgruzka','yetkazildi')), 0) AS otgruzka,
                    COALESCE(SUM(ss.jami) FILTER (WHERE ss.holat = 'yetkazildi'), 0) AS dostavka,
                    COALESCE(SUM(ss.tolangan), 0) AS oplaty,
                    COALESCE(SUM(ss.jami - ss.tolangan) FILTER (WHERE ss.jami > ss.tolangan), 0) AS qarz,
                    COUNT(DISTINCT ss.klient_id) AS akb
                FROM shogirdlar s
                LEFT JOIN sotuv_sessiyalar ss
                  ON ss.shogird_id = s.id
                  AND ss.sana::date BETWEEN $2 AND $3
                  AND COALESCE(ss.holat, 'yangi') != 'bekor'
                WHERE s.admin_uid = $1 AND s.faol = TRUE
                GROUP BY s.id, s.ism
                ORDER BY SUM(ss.jami) DESC NULLS LAST
            """, uid, sd, sg)
            agentlar = [
                {
                    "id": r["id"], "ism": r["ism"],
                    "obshchie": float(r["obshchie"] or 0),
                    "otgruzka": float(r["otgruzka"] or 0),
                    "dostavka": float(r["dostavka"] or 0),
                    "oplaty": float(r["oplaty"] or 0),
                    "qarz": float(r["qarz"] or 0),
                    "akb": int(r["akb"] or 0),
                } for r in rows
            ]
        except Exception:
            pass

    return {
        "davr": {"sana_dan": str(sd), "sana_gacha": str(sg)},
        "jami": {
            "obshchie_zakazy": float(jami["obshchie_zakazy"] or 0),
            "otgruzka": float(jami["otgruzka"] or 0),
            "dostavka": float(jami["dostavka"] or 0),
            "oplaty": float(jami["oplaty"] or 0),
            "qarz": float(jami["qarz"] or 0),
            "akb": int(jami["akb"] or 0),
        },
        "agentlar": agentlar,
    }


@router.get("/agentlar/bugungi-kpi")
async def agentlar_bugungi_kpi(uid: int = Depends(get_uid)):
    """
    Bugungi agent KPI hisoboti — SalesDoc supervisor dashboard shaklida.

    Agar shogirdlar jadvalida real agentlar bo'lsa, ularning bugungi
    ishlarini jamlab qaytaradi. Aks holda joriy foydalanuvchi o'zi
    "bosh agent" sifatida qaytariladi (bot-driven single-user flow).

    Response shape mos keladi services/web/components/dashboard/
    agent-kpi-board.tsx ga:
      [{ id, ism, reja, tashrif_soni, rejali_summa, rejali_soni,
         ofplan_summa, ofplan_soni, qaytarish }]
    """
    async with rls_conn(uid) as c:
        # 1. Shogirdlar bormi?
        try:
            shogirdlar = await c.fetch(
                "SELECT id, ism FROM shogirdlar WHERE admin_uid=$1 AND faol=TRUE ORDER BY id",
                uid,
            )
        except Exception:
            shogirdlar = []

        if shogirdlar:
            # Per-agent aggregate (sotuv_sessiyalar hali shogird_id'ga ega emas,
            # shuning uchun hozir placeholder — keyinroq join qilamiz)
            result = []
            for s in shogirdlar:
                result.append({
                    "id":           s["id"],
                    "ism":          s["ism"],
                    "reja":         0,
                    "tashrif_soni": 0,
                    "rejali_summa": 0,
                    "rejali_soni":  0,
                    "ofplan_summa": 0,
                    "ofplan_soni":  0,
                    "qaytarish":    0,
                })
            return result

        # 2. Bitta foydalanuvchili rejim — o'zini agent sifatida ko'rsatadi
        row = await c.fetchrow(
            """
            SELECT
                COUNT(*)                            AS soni,
                COALESCE(SUM(jami), 0)              AS summa
            FROM sotuv_sessiyalar
            WHERE user_id = $1
              AND (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
              AND COALESCE(holat, 'yangi') NOT IN ('bekor')
            """,
            uid,
        )
        klient_count = await c.fetchval(
            "SELECT COUNT(*) FROM klientlar WHERE user_id=$1", uid,
        ) or 0
        try:
            qaytarish = await c.fetchval(
                """
                SELECT COUNT(*) FROM qaytarishlar
                WHERE user_id=$1
                  AND (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
                """,
                uid,
            ) or 0
        except Exception:
            qaytarish = 0
        user_row = await c.fetchrow(
            """
            SELECT COALESCE(NULLIF(ism, ''), dokon_nomi, username, 'Agent') AS ism
            FROM users WHERE id=$1
            """,
            uid,
        )

    return [
        {
            "id":           uid,
            "ism":          (user_row["ism"] if user_row else "Agent"),
            "reja":         int(klient_count),
            "tashrif_soni": 0,          # visits not tracked in bot-only mode
            "rejali_summa": float(row["summa"] or 0),
            "rejali_soni":  int(row["soni"] or 0),
            "ofplan_summa": 0,
            "ofplan_soni":  0,
            "qaytarish":    int(qaytarish),
        }
    ]


@router.get("/audit-log")
async def audit_log_list(
    limit: int = 100,
    offset: int = 0,
    jadval: str | None = None,
    amal: str | None = None,
    uid: int = Depends(get_uid),
):
    """Audit log — foydalanuvchi amallarining tarixi.

    Har bir amal (bekor_qilish, yangilash va h.k.) log qilinadi.
    """
    limit = min(limit, 500)
    where = ["user_id = $1"]
    params: list = [uid]
    if jadval:
        params.append(jadval)
        where.append(f"jadval = ${len(params)}")
    if amal:
        params.append(amal)
        where.append(f"amal = ${len(params)}")
    where_sql = " AND ".join(where)

    params.append(limit); params.append(offset)

    async with rls_conn(uid) as c:
        try:
            rows = await c.fetch(f"""
                SELECT id, user_id, amal, jadval, yozuv_id,
                       eski, yangi, ip, manba, sana
                FROM audit_log
                WHERE {where_sql}
                ORDER BY sana DESC
                LIMIT ${len(params)-1} OFFSET ${len(params)}
            """, *params)
            stats = await c.fetchrow(f"""
                SELECT COUNT(*) AS jami,
                       COUNT(DISTINCT jadval) AS turli_jadval,
                       COUNT(DISTINCT amal) AS turli_amal
                FROM audit_log WHERE {where_sql}
            """, *params[:-2])
        except Exception as e:
            log.warning("audit_log: %s", e)
            return {"items": [], "stats": {}}

    return {
        "items": [dict(r) for r in rows],
        "stats": dict(stats) if stats else {},
    }


@router.get("/photo-reports")
async def photo_reports(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    qidiruv: str | None = None,
    agent_id: int | None = None,
    uid: int = Depends(get_uid),
):
    """Agentlar yuklagan rasmlar — checkin_out.foto_url orqali."""
    where = ["co.user_id = $1", "co.foto_url IS NOT NULL", "co.foto_url <> ''"]
    params: list = [uid]
    if sana_dan:
        params.append(sana_dan)
        where.append(f"co.vaqt >= ${len(params)}::timestamptz")
    if sana_gacha:
        params.append(sana_gacha)
        where.append(f"co.vaqt < ${len(params)}::timestamptz + interval '1 day'")
    if qidiruv:
        params.append(f"%{like_escape(qidiruv)}%")
        where.append(f"k.ism ILIKE ${len(params)}")
    if agent_id:
        params.append(agent_id)
        where.append(f"co.agent_id = ${len(params)}")
    where_sql = " AND ".join(where)

    async with rls_conn(uid) as c:
        try:
            rows = await c.fetch(f"""
                SELECT co.id, co.klient_id, co.turi, co.vaqt, co.foto_url,
                       co.izoh, co.latitude, co.longitude,
                       k.ism AS klient_nomi, k.manzil,
                       COALESCE(s.ism, 'Agent') AS agent_ismi,
                       co.agent_id
                FROM checkin_out co
                LEFT JOIN klientlar k ON k.id = co.klient_id
                LEFT JOIN shogirdlar s ON s.id = co.agent_id
                WHERE {where_sql}
                ORDER BY co.vaqt DESC
                LIMIT 500
            """, *params)

            stats = await c.fetchrow("""
                SELECT
                    COUNT(*)                                                      AS jami,
                    COUNT(*) FILTER (WHERE vaqt::date = CURRENT_DATE)             AS bugun,
                    COUNT(*) FILTER (WHERE vaqt >= NOW() - interval '7 days')     AS hafta,
                    COUNT(*) FILTER (WHERE vaqt >= NOW() - interval '30 days')    AS oy,
                    COUNT(DISTINCT agent_id)                                       AS agentlar_soni
                FROM checkin_out
                WHERE user_id = $1 AND foto_url IS NOT NULL AND foto_url <> ''
            """, uid)

            agents = await c.fetch("""
                SELECT DISTINCT s.id, s.ism
                FROM shogirdlar s
                JOIN checkin_out co ON co.agent_id = s.id AND co.user_id = $1
                    AND co.foto_url IS NOT NULL AND co.foto_url <> ''
                WHERE s.admin_uid = $1
                ORDER BY s.ism
            """, uid)
        except Exception as e:
            log.warning("photo-reports: %s", e)
            return {"items": [], "stats": {}, "agents": []}

    return {
        "items": [dict(r) for r in rows],
        "stats": dict(stats) if stats else {},
        "agents": [{"id": a["id"], "ism": a["ism"]} for a in agents],
    }


@router.get("/qaytarishlar")
async def qaytarishlar_list(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    qidiruv: str | None = None,
    limit: int = 200,
    uid: int = Depends(get_uid),
):
    """Qaytarishlar (vozvrat) ro'yxati + statistika."""
    where = ["user_id = $1"]
    params: list = [uid]
    if sana_dan:
        params.append(sana_dan)
        where.append(f"sana >= ${len(params)}::timestamptz")
    if sana_gacha:
        params.append(sana_gacha)
        where.append(f"sana < ${len(params)}::timestamptz + interval '1 day'")
    if qidiruv:
        params.append(f"%{like_escape(qidiruv)}%")
        where.append(
            f"(klient_ismi ILIKE ${len(params)} OR tovar_nomi ILIKE ${len(params)})"
        )
    params.append(limit)
    where_sql = " AND ".join(where)

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT id, chiqim_id, sessiya_id, klient_ismi, tovar_nomi,
                   miqdor, birlik, narx, jami, sabab, sana
            FROM qaytarishlar
            WHERE {where_sql}
            ORDER BY sana DESC
            LIMIT ${len(params)}
        """, *params)

        stats = await c.fetchrow(f"""
            SELECT COUNT(*)                  AS soni,
                   COALESCE(SUM(jami), 0)    AS jami_summa,
                   COALESCE(SUM(miqdor), 0)  AS jami_miqdor
            FROM qaytarishlar
            WHERE {where_sql}
        """, *params[:-1])

    return {
        "items": [dict(r) for r in rows],
        "stats": dict(stats) if stats else {},
    }


class QaytarishYarat(__import__("pydantic").BaseModel):
    chiqim_id: int | None = None
    klient_ismi: str | None = ""
    tovar_nomi: str
    miqdor: float
    birlik: str | None = "dona"
    narx: float
    sabab: str | None = ""


@router.post("/qaytarish")
async def qaytarish_yarat(data: QaytarishYarat, uid: int = Depends(get_uid)):
    """Yangi qaytarish (vozvrat) yaratish."""
    jami = data.miqdor * data.narx
    async with rls_conn(uid) as c:
        row = await c.fetchrow("""
            INSERT INTO qaytarishlar
                (user_id, chiqim_id, klient_ismi, tovar_nomi,
                 miqdor, birlik, narx, jami, sabab)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
            RETURNING id, sana
        """, uid, data.chiqim_id, data.klient_ismi or "",
            data.tovar_nomi, data.miqdor, data.birlik or "dona",
            data.narx, jami, data.sabab or "")

        # Tovar qoldig'ini qaytarish (agar chiqim_id orqali tovar_id aniqlanishi mumkin)
        if data.chiqim_id:
            await c.execute("""
                UPDATE tovarlar SET qoldiq = COALESCE(qoldiq, 0) + $1
                WHERE id = (SELECT tovar_id FROM chiqimlar WHERE id = $2)
                  AND user_id = $3
            """, data.miqdor, data.chiqim_id, uid)
    return {
        "id": row["id"],
        "jami": float(jami),
        "sana": row["sana"].isoformat(),
        "status": "yaratildi",
    }


@router.get("/reports/rfm")
async def report_rfm(uid: int = Depends(get_uid)):
    """Butun klient bazasi bo'yicha RFM segmentatsiya (web hisobot uchun).
    Qaytaradi: segmentlar bo'yicha klient soni + jami summa + har segmentdagi
    top 10 klient (drill-down uchun)."""
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            WITH rfm AS (
                SELECT
                    k.id, k.ism, k.telefon,
                    COUNT(ss.id)                                  AS frequency,
                    COALESCE(SUM(ss.jami), 0)                     AS monetary,
                    EXTRACT(EPOCH FROM (NOW() - MAX(ss.sana)))/86400 AS recency_days,
                    MAX(ss.sana)                                  AS oxirgi
                FROM klientlar k
                LEFT JOIN sotuv_sessiyalar ss ON ss.klient_id = k.id
                WHERE k.user_id = $1
                GROUP BY k.id
                HAVING COUNT(ss.id) > 0
            ),
            scored AS (
                SELECT
                    id, ism, telefon, frequency, monetary, recency_days, oxirgi,
                    NTILE(5) OVER (ORDER BY recency_days DESC) AS r,
                    NTILE(5) OVER (ORDER BY frequency)          AS f,
                    NTILE(5) OVER (ORDER BY monetary)           AS m
                FROM rfm
            )
            SELECT id, ism, telefon,
                   frequency, monetary, recency_days, oxirgi,
                   r, f, m,
                   CASE
                       WHEN (r + f + m) >= 13 THEN 'Champions'
                       WHEN (r + f + m) >= 10 THEN 'Loyal'
                       WHEN (r + f + m) >= 7  THEN 'Potential'
                       WHEN (r + f + m) >= 5  THEN 'At Risk'
                       ELSE 'Lost'
                   END AS segment
            FROM scored
            ORDER BY monetary DESC
        """, uid)

    seg_map: dict[str, dict] = {
        "Champions": {"soni": 0, "monetary": 0.0, "top": []},
        "Loyal":     {"soni": 0, "monetary": 0.0, "top": []},
        "Potential": {"soni": 0, "monetary": 0.0, "top": []},
        "At Risk":   {"soni": 0, "monetary": 0.0, "top": []},
        "Lost":      {"soni": 0, "monetary": 0.0, "top": []},
    }
    for r in rows:
        seg = r["segment"]
        d   = seg_map[seg]
        d["soni"] += 1
        d["monetary"] += float(r["monetary"] or 0)
        if len(d["top"]) < 10:
            d["top"].append({
                "id":           r["id"],
                "ism":          r["ism"],
                "telefon":      r["telefon"],
                "R":            int(r["r"]),
                "F":            int(r["f"]),
                "M":            int(r["m"]),
                "frequency":    int(r["frequency"]),
                "monetary":     float(r["monetary"] or 0),
                "recency_days": int(r["recency_days"] or 0),
                "oxirgi":       r["oxirgi"].isoformat() if r["oxirgi"] else None,
            })

    return {
        "jami_klient":  len(rows),
        "jami_summa":   sum(s["monetary"] for s in seg_map.values()),
        "segmentlar":   seg_map,
    }


@router.get("/reports/sales-detail")
async def report_sales_detail(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    kategoriya: str | None = None,
    klient: str | None = None,
    tovar: str | None = None,
    limit: int = 500,
    offset: int = 0,
    uid: int = Depends(get_uid),
):
    """Har bir sotuv qatori — tovar nomi, miqdor, narx, kategoriya, mijoz, sana.
    SalesDoc Sotuv detail hisoboti analog."""
    where = ["ch.user_id = $1"]
    params: list = [uid]

    def add(clause: str, val):
        params.append(val)
        where.append(clause)

    if sana_dan:
        add(f"ch.sana >= ${len(params)+1}::timestamptz", sana_dan)
    if sana_gacha:
        add(f"ch.sana < ${len(params)+1}::timestamptz + interval '1 day'", sana_gacha)
    if kategoriya:
        add(f"ch.kategoriya = ${len(params)+1}", kategoriya)
    if klient:
        add(f"lower(ch.klient_ismi) LIKE lower(${len(params)+1})", f"%{like_escape(klient)}%")
    if tovar:
        add(f"lower(ch.tovar_nomi) LIKE lower(${len(params)+1})", f"%{like_escape(tovar)}%")

    where_sql = " AND ".join(where)
    params.append(limit); params.append(offset)

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT
                ch.id, ch.sana, ch.sessiya_id,
                ch.tovar_nomi, ch.kategoriya, ch.birlik,
                ch.miqdor, ch.qaytarilgan,
                ch.olish_narxi, ch.sotish_narxi, ch.chegirma_foiz, ch.jami,
                ch.klient_ismi,
                (ch.sotish_narxi - ch.olish_narxi) * ch.miqdor AS foyda
            FROM chiqimlar ch
            WHERE {where_sql}
            ORDER BY ch.sana DESC, ch.id DESC
            LIMIT ${len(params)-1} OFFSET ${len(params)}
        """, *params)

        stats = await c.fetchrow(f"""
            SELECT
                COUNT(*)                                                 AS soni,
                COALESCE(SUM(ch.jami), 0)                                AS jami_summa,
                COALESCE(SUM((ch.sotish_narxi - ch.olish_narxi) * ch.miqdor), 0) AS jami_foyda,
                COALESCE(SUM(ch.miqdor), 0)                              AS jami_miqdor
            FROM chiqimlar ch
            WHERE {where_sql}
        """, *params[:-2])

    return {
        "items": [dict(r) for r in rows],
        "stats": dict(stats) if stats else {},
        "total": int(stats["soni"] or 0) if stats else 0,
    }


@router.get("/reports/sales-detail/excel")
async def report_sales_detail_excel(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    kategoriya: str | None = None,
    uid: int = Depends(get_uid),
):
    """Sotuv detail hisobotini Excel faylga export qilish."""
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    where = ["ch.user_id = $1"]
    params: list = [uid]
    if sana_dan:
        params.append(sana_dan)
        where.append(f"ch.sana >= ${len(params)}::timestamptz")
    if sana_gacha:
        params.append(sana_gacha)
        where.append(f"ch.sana < ${len(params)}::timestamptz + interval '1 day'")
    if kategoriya:
        params.append(kategoriya)
        where.append(f"ch.kategoriya = ${len(params)}")
    where_sql = " AND ".join(where)

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT ch.sana, ch.klient_ismi, ch.tovar_nomi, ch.kategoriya,
                   ch.miqdor, ch.birlik,
                   ch.olish_narxi, ch.sotish_narxi, ch.chegirma_foiz, ch.jami,
                   (ch.sotish_narxi - ch.olish_narxi) * ch.miqdor AS foyda
            FROM chiqimlar ch
            WHERE {where_sql}
            ORDER BY ch.sana DESC
            LIMIT 10000
        """, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sotuv detail"

    headers = ["Sana", "Mijoz", "Tovar", "Kategoriya", "Miqdor", "Birlik",
               "Olish", "Sotish", "Cheg %", "Jami", "Foyda"]
    widths  = [13, 25, 30, 15, 10, 8, 12, 12, 8, 14, 12]

    header_fill = PatternFill(start_color="0A819C", end_color="0A819C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[chr(64 + i)].width = w

    total_jami = 0.0
    total_foyda = 0.0
    for idx, r in enumerate(rows, 2):
        d = dict(r)
        sana_str = d["sana"].strftime("%d.%m.%Y") if d.get("sana") else ""
        vals = [
            sana_str, d["klient_ismi"] or "—", d["tovar_nomi"] or "",
            d["kategoriya"] or "", float(d["miqdor"] or 0), d["birlik"] or "",
            float(d["olish_narxi"] or 0), float(d["sotish_narxi"] or 0),
            float(d["chegirma_foiz"] or 0), float(d["jami"] or 0),
            float(d["foyda"] or 0),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=idx, column=col, value=v)
            cell.border = border
            if col in (5, 7, 8, 9, 10, 11):
                cell.number_format = '#,##0.##'
                cell.alignment = Alignment(horizontal="right")
        total_jami  += float(d["jami"] or 0)
        total_foyda += float(d["foyda"] or 0)

    total_row = len(rows) + 2
    for col in range(1, 12):
        ws.cell(row=total_row, column=col).fill = PatternFill(
            start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ws.cell(row=total_row, column=col).border = border
    ws.cell(row=total_row, column=1, value="JAMI").font = Font(bold=True)
    jc = ws.cell(row=total_row, column=10, value=total_jami)
    jc.font = Font(bold=True); jc.number_format = '#,##0'
    fc = ws.cell(row=total_row, column=11, value=total_foyda)
    fc.font = Font(bold=True, color="1B5E20"); fc.number_format = '#,##0'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return {
        "filename": f"Sotuv_detail_{sana_dan or 'barcha'}_{sana_gacha or ''}.xlsx",
        "content_base64": base64.b64encode(buf.getvalue()).decode(),
        "soni": len(rows),
        "jami_summa": total_jami,
        "jami_foyda": total_foyda,
    }


# ════════════════════════════════════════════════════════════
#  TAVSIYA QOLDIQ — SalesDoc stock/report uslubida
#  O'rtacha kunlik sotuv + necha kunga yetishi + tavsiya
# ════════════════════════════════════════════════════════════

@router.get("/hisobot/tavsiya-qoldiq")
async def tavsiya_qoldiq(
    kunlar: int = 30,
    uid: int = Depends(get_uid),
):
    """
    SalesDoc'dagi "Рекомендуемый запас" hisoboti.

    Har bir tovar uchun:
    - O'rtacha kunlik sotuv (oxirgi N kun)
    - Hozirgi qoldiq
    - Necha kunga yetishi
    - Tavsiya qoldiq (6, 10, 30 kunga)
    """
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            WITH sotuv_stats AS (
                SELECT
                    ch.tovar_id,
                    SUM(ch.miqdor) AS jami_sotilgan,
                    COUNT(DISTINCT (ch.sana AT TIME ZONE 'Asia/Tashkent')::date) AS sotuv_kunlari
                FROM chiqimlar ch
                WHERE ch.user_id = $1
                  AND ch.sana >= NOW() - make_interval(days => $2)
                GROUP BY ch.tovar_id
            )
            SELECT
                t.id,
                t.nomi,
                t.kategoriya,
                t.birlik,
                t.qoldiq,
                t.sotish_narxi,
                t.olish_narxi,
                COALESCE(ss.jami_sotilgan, 0) AS jami_sotilgan,
                COALESCE(ss.sotuv_kunlari, 0) AS sotuv_kunlari,
                CASE
                    WHEN COALESCE(ss.sotuv_kunlari, 0) > 0
                    THEN ROUND(ss.jami_sotilgan::numeric / ss.sotuv_kunlari, 1)
                    ELSE 0
                END AS kunlik_sotuv,
                CASE
                    WHEN COALESCE(ss.jami_sotilgan, 0) > 0 AND ss.sotuv_kunlari > 0
                    THEN ROUND(t.qoldiq::numeric / (ss.jami_sotilgan::numeric / ss.sotuv_kunlari), 1)
                    ELSE NULL
                END AS necha_kunga_yetadi
            FROM tovarlar t
            LEFT JOIN sotuv_stats ss ON ss.tovar_id = t.id
            WHERE t.user_id = $1 AND t.faol = TRUE
            ORDER BY
                CASE WHEN t.qoldiq <= 0 THEN 0
                     WHEN COALESCE(ss.jami_sotilgan, 0) > 0
                          AND ss.sotuv_kunlari > 0
                          AND t.qoldiq::numeric / (ss.jami_sotilgan::numeric / ss.sotuv_kunlari) < 7
                     THEN 1
                     ELSE 2
                END,
                t.nomi
        """, uid, kunlar)

    items = []
    for r in rows:
        kunlik = float(r["kunlik_sotuv"] or 0)
        qoldiq = float(r["qoldiq"] or 0)
        yetadi = float(r["necha_kunga_yetadi"]) if r["necha_kunga_yetadi"] is not None else None

        items.append({
            "id": r["id"],
            "nomi": r["nomi"],
            "kategoriya": r["kategoriya"],
            "birlik": r["birlik"],
            "qoldiq": qoldiq,
            "sotish_narxi": float(r["sotish_narxi"] or 0),
            "olish_narxi": float(r["olish_narxi"] or 0),
            "jami_sotilgan": float(r["jami_sotilgan"] or 0),
            "sotuv_kunlari": int(r["sotuv_kunlari"] or 0),
            "kunlik_sotuv": kunlik,
            "necha_kunga_yetadi": yetadi,
            "tavsiya_6_kun": max(0, round(kunlik * 6 - qoldiq, 1)),
            "tavsiya_10_kun": max(0, round(kunlik * 10 - qoldiq, 1)),
            "tavsiya_30_kun": max(0, round(kunlik * 30 - qoldiq, 1)),
            "holat": "tugagan" if qoldiq <= 0 else
                     "kritik" if yetadi is not None and yetadi < 3 else
                     "kam" if yetadi is not None and yetadi < 7 else
                     "yetarli",
        })

    return {
        "items": items,
        "jami_tovar": len(items),
        "tugagan": sum(1 for i in items if i["holat"] == "tugagan"),
        "kritik": sum(1 for i in items if i["holat"] == "kritik"),
        "kam": sum(1 for i in items if i["holat"] == "kam"),
        "yetarli": sum(1 for i in items if i["holat"] == "yetarli"),
        "kunlar": kunlar,
    }


# ════════════════════════════════════════════════════════════
#  VAN SELLING — Kunlik agent hisoboti
#  Har bir agentning bugungi sotuv faoliyati (SalesDoc uslubi)
# ════════════════════════════════════════════════════════════

@router.get("/hisobot/van-selling-kunlik")
async def van_selling_kunlik(
    sana: str | None = None,  # YYYY-MM-DD, default today
    uid: int = Depends(get_uid),
):
    """
    Van Selling kunlik hisobot — har bir agentning sotuv faoliyati.

    SalesDoc supervisor dashboard uchun: vizitlar, buyurtmalar,
    jami sotuv, foyda, naqd/qarz, qaytarishlar, yangi klientlar.

    Response:
      {
        "sana": "2026-04-13",
        "agentlar": [{ agent_id, agent_ismi, vizitlar_soni, ... }],
        "jami": { vizitlar, buyurtmalar, sotuv, foyda }
      }
    """
    if sana:
        try:
            datetime.strptime(sana, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="sana formati noto'g'ri, YYYY-MM-DD kerak")
        target_date = sana
    else:
        target_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    async with rls_conn(uid) as c:
        # 1. Agentlar bo'yicha sotuv sessiyalari
        agent_rows = await c.fetch("""
            SELECT
                ss.user_id                              AS agent_id,
                COALESCE(u.ism, u.dokon_nomi, u.username, 'Agent') AS agent_ismi,
                COUNT(DISTINCT ss.klient_id)            AS vizitlar_soni,
                COUNT(ss.id)                            AS buyurtmalar_soni,
                COALESCE(SUM(ss.jami), 0)               AS jami_sotuv,
                COALESCE(SUM(ss.tolangan), 0)           AS naqd_tulangan,
                COALESCE(SUM(ss.qarz), 0)               AS qarz_berilgan,
                MAX(ss.sana AT TIME ZONE 'Asia/Tashkent') AS oxirgi_vaqt
            FROM sotuv_sessiyalar ss
            LEFT JOIN users u ON u.id = ss.user_id
            WHERE (ss.sana AT TIME ZONE 'Asia/Tashkent')::date = $1::date
              AND COALESCE(ss.holat, 'yangi') != 'bekor'
            GROUP BY ss.user_id, u.ism, u.dokon_nomi, u.username
            ORDER BY SUM(ss.jami) DESC
        """, target_date)

        # 2. Foyda — chiqimlar orqali
        foyda_rows = await c.fetch("""
            SELECT
                ss.user_id AS agent_id,
                COALESCE(SUM((ch.sotish_narxi - ch.olish_narxi) * ch.miqdor), 0) AS foyda,
                COALESCE(SUM(ch.miqdor), 0) AS tovarlar_soni
            FROM chiqimlar ch
            JOIN sotuv_sessiyalar ss ON ss.id = ch.sessiya_id
            WHERE (ss.sana AT TIME ZONE 'Asia/Tashkent')::date = $1::date
              AND COALESCE(ss.holat, 'yangi') != 'bekor'
            GROUP BY ss.user_id
        """, target_date)
        foyda_map = {r["agent_id"]: r for r in foyda_rows}

        # 3. Qaytarishlar (jadval mavjud bo'lmasligi mumkin)
        try:
            qayt_rows = await c.fetch("""
                SELECT user_id AS agent_id,
                       COUNT(*) AS qaytarish_soni,
                       COALESCE(SUM(jami), 0) AS qaytarish_summa
                FROM qaytarishlar
                WHERE (sana AT TIME ZONE 'Asia/Tashkent')::date = $1::date
                GROUP BY user_id
            """, target_date)
            qayt_map = {r["agent_id"]: r for r in qayt_rows}
        except Exception:
            qayt_map = {}

        # 4. Yangi klientlar (bugun yaratilgan)
        try:
            yangi_rows = await c.fetch("""
                SELECT user_id AS agent_id, COUNT(*) AS yangi_klientlar
                FROM klientlar
                WHERE (yaratilgan AT TIME ZONE 'Asia/Tashkent')::date = $1::date
                GROUP BY user_id
            """, target_date)
            yangi_map = {r["agent_id"]: r for r in yangi_rows}
        except Exception:
            yangi_map = {}

    # Natijani yig'ish
    agentlar = []
    jami_vizitlar = 0
    jami_buyurtmalar = 0
    jami_sotuv = Decimal(0)
    jami_foyda = Decimal(0)

    for ar in agent_rows:
        aid = ar["agent_id"]
        fr = foyda_map.get(aid, {})
        qr = qayt_map.get(aid, {})
        yr = yangi_map.get(aid, {})

        vizitlar = int(ar["vizitlar_soni"] or 0)
        buyurtmalar = int(ar["buyurtmalar_soni"] or 0)
        sotuv = float(ar["jami_sotuv"] or 0)
        foyda = float(fr.get("foyda", 0) or 0)

        jami_vizitlar += vizitlar
        jami_buyurtmalar += buyurtmalar
        jami_sotuv += Decimal(str(sotuv))
        jami_foyda += Decimal(str(foyda))

        oxirgi = ar["oxirgi_vaqt"]
        oxirgi_str = oxirgi.strftime("%H:%M") if oxirgi else None

        agentlar.append({
            "agent_id":        aid,
            "agent_ismi":      ar["agent_ismi"],
            "vizitlar_soni":   vizitlar,
            "buyurtmalar_soni": buyurtmalar,
            "jami_sotuv":      sotuv,
            "jami_foyda":      foyda,
            "naqd_tulangan":   float(ar["naqd_tulangan"] or 0),
            "qarz_berilgan":   float(ar["qarz_berilgan"] or 0),
            "qaytarish_soni":  int(qr.get("qaytarish_soni", 0) or 0),
            "qaytarish_summa": float(qr.get("qaytarish_summa", 0) or 0),
            "tovarlar_soni":   int(fr.get("tovarlar_soni", 0) or 0),
            "yangi_klientlar": int(yr.get("yangi_klientlar", 0) or 0),
            "oxirgi_faollik":  oxirgi_str,
        })

    return {
        "sana": target_date,
        "agentlar": agentlar,
        "jami": {
            "vizitlar":    jami_vizitlar,
            "buyurtmalar": jami_buyurtmalar,
            "sotuv":       float(jami_sotuv),
            "foyda":       float(jami_foyda),
        },
    }


# ────────────────────────────────────────────────────────────────────
#  DEFEKT / QAYTARISH HISOBOTI
# ────────────────────────────────────────────────────────────────────
@router.get("/hisobot/defekt")
async def defekt_hisoboti(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    uid: int = Depends(get_uid),
):
    """Defekt/qaytarish hisoboti — qaytarilgan tovarlar, sabablar, statistika."""
    where = ["q.user_id = $1"]
    params: list = [uid]
    if sana_dan:
        params.append(sana_dan)
        where.append(f"q.sana >= ${len(params)}::timestamptz")
    if sana_gacha:
        params.append(sana_gacha)
        where.append(f"q.sana < ${len(params)}::timestamptz + interval '1 day'")
    where_sql = " AND ".join(where)

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT q.id, q.tovar_nomi, COALESCE(ch.kategoriya, 'Boshqa') AS kategoriya,
                   q.miqdor, q.narx, q.jami,
                   COALESCE(q.sabab, '') AS sabab,
                   COALESCE(q.klient_ismi, '') AS klient_ismi,
                   q.sana
            FROM qaytarishlar q
            LEFT JOIN chiqimlar ch ON ch.id = q.chiqim_id
            WHERE {where_sql}
            ORDER BY q.sana DESC
        """, *params)

        stats = await c.fetchrow(f"""
            SELECT COUNT(*)                  AS soni,
                   COALESCE(SUM(q.jami), 0)    AS jami_summa,
                   COALESCE(SUM(q.miqdor), 0)  AS jami_miqdor
            FROM qaytarishlar q
            WHERE {where_sql}
        """, *params)

    items = []
    for r in rows:
        items.append({
            "id":          r["id"],
            "tovar_nomi":  r["tovar_nomi"],
            "kategoriya":  r["kategoriya"],
            "miqdor":      float(Decimal(str(r["miqdor"]))) if r["miqdor"] else 0,
            "narx":        float(Decimal(str(r["narx"]))) if r["narx"] else 0,
            "jami":        float(Decimal(str(r["jami"]))) if r["jami"] else 0,
            "sabab":       r["sabab"],
            "klient_ismi": r["klient_ismi"],
            "sana":        r["sana"].strftime("%Y-%m-%d") if r["sana"] else None,
        })

    return {
        "items":      items,
        "jami_miqdor": float(Decimal(str(stats["jami_miqdor"]))) if stats else 0,
        "jami_summa":  float(Decimal(str(stats["jami_summa"]))) if stats else 0,
        "soni":        int(stats["soni"]) if stats else 0,
    }


# ════════════════════════════════════════════════════════════
#  EKSPEDITOR (DELIVERY AGENT) HISOBOTI
# ════════════════════════════════════════════════════════════

@router.get("/hisobot/ekspeditor")
async def ekspeditor_hisoboti(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    uid: int = Depends(get_uid),
):
    """
    Ekspeditor (delivery agent) hisoboti — yetkazib berish samaradorligi.

    sotuv_sessiyalar jadvalidagi holat = 'otgruzka' | 'yetkazildi'
    buyurtmalarni agent bo'yicha guruhlaydi.

    Parametrlar:
      sana_dan  — boshlanish sanasi (YYYY-MM-DD), default 30 kun oldin
      sana_gacha — tugash sanasi (YYYY-MM-DD), default bugun

    Response:
      {
        "items": [{ agent_id, agent_ismi, jami_buyurtma, yetkazilgan,
                     kutilmoqda, jami_summa, yetkazilgan_summa,
                     ortacha_vaqt_soat }],
        "jami": { buyurtma, yetkazilgan, kutilmoqda }
      }
    """
    from datetime import timedelta

    now = datetime.now(timezone.utc)

    if sana_gacha:
        try:
            datetime.strptime(sana_gacha, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="sana_gacha formati noto'g'ri, YYYY-MM-DD kerak")
        gacha = sana_gacha
    else:
        gacha = now.strftime("%Y-%m-%d")

    if sana_dan:
        try:
            datetime.strptime(sana_dan, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="sana_dan formati noto'g'ri, YYYY-MM-DD kerak")
        dan = sana_dan
    else:
        dan = (now - timedelta(days=30)).strftime("%Y-%m-%d")

    cache_k = f"hisobot:ekspeditor:{uid}:{dan}:{gacha}"
    cached = await cache_ol(cache_k)
    if cached:
        return cached

    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT
                ss.user_id                              AS agent_id,
                COALESCE(u.ism, u.dokon_nomi, u.username, 'Agent') AS agent_ismi,

                COUNT(ss.id)                            AS jami_buyurtma,

                COUNT(ss.id) FILTER (WHERE ss.holat = 'yetkazildi')
                                                        AS yetkazilgan,
                COUNT(ss.id) FILTER (WHERE ss.holat IN ('otgruzka','tasdiqlangan'))
                                                        AS kutilmoqda,

                COALESCE(SUM(ss.jami), 0)               AS jami_summa,
                COALESCE(SUM(ss.jami) FILTER (WHERE ss.holat = 'yetkazildi'), 0)
                                                        AS yetkazilgan_summa,

                -- O'rtacha yetkazish vaqti (otgruzka -> yetkazildi, soatlarda)
                COALESCE(
                    AVG(
                        EXTRACT(EPOCH FROM (ss.yetkazildi_vaqti - ss.otgruzka_vaqti)) / 3600.0
                    ) FILTER (
                        WHERE ss.yetkazildi_vaqti IS NOT NULL
                          AND ss.otgruzka_vaqti IS NOT NULL
                    ),
                    0
                )                                       AS ortacha_vaqt_soat

            FROM sotuv_sessiyalar ss
            LEFT JOIN users u ON u.id = ss.user_id
            WHERE ss.holat IN ('otgruzka','yetkazildi','tasdiqlangan')
              AND (ss.sana AT TIME ZONE 'Asia/Tashkent')::date
                    BETWEEN $1::date AND $2::date
            GROUP BY ss.user_id, u.ism, u.dokon_nomi, u.username
            ORDER BY COUNT(ss.id) DESC
        """, dan, gacha)

        jami_buyurtma = 0
        jami_yetkazilgan = 0
        jami_kutilmoqda = 0
        items = []

        for r in rows:
            jami_buyurtma += int(r["jami_buyurtma"])
            jami_yetkazilgan += int(r["yetkazilgan"])
            jami_kutilmoqda += int(r["kutilmoqda"])

            items.append({
                "agent_id":          int(r["agent_id"]),
                "agent_ismi":        r["agent_ismi"],
                "jami_buyurtma":     int(r["jami_buyurtma"]),
                "yetkazilgan":       int(r["yetkazilgan"]),
                "kutilmoqda":        int(r["kutilmoqda"]),
                "jami_summa":        float(r["jami_summa"]),
                "yetkazilgan_summa": float(r["yetkazilgan_summa"]),
                "ortacha_vaqt_soat": round(float(r["ortacha_vaqt_soat"]), 1),
            })

    result = {
        "sana_dan": dan,
        "sana_gacha": gacha,
        "items": items,
        "jami": {
            "buyurtma":    jami_buyurtma,
            "yetkazilgan": jami_yetkazilgan,
            "kutilmoqda":  jami_kutilmoqda,
        },
    }
    await cache_yoz(cache_k, result, TTL_HISOBOT)
    return result


# ═══════════════════════════════════════════════════════
#  VIZIT HISOBOTI
# ═══════════════════════════════════════════════════════

@router.get("/hisobot/vizitlar")
async def vizit_hisoboti(
    sana_dan: str | None = None,
    sana_gacha: str | None = None,
    uid: int = Depends(get_uid),
):
    """Agent vizitlari hisoboti — checkin_out yoki sotuv_sessiyalar asosida"""
    from datetime import timedelta

    now = datetime.now(timezone.utc)

    if sana_gacha:
        try:
            datetime.strptime(sana_gacha, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="sana_gacha formati noto'g'ri, YYYY-MM-DD kerak")
        gacha = sana_gacha
    else:
        gacha = now.strftime("%Y-%m-%d")

    if sana_dan:
        try:
            datetime.strptime(sana_dan, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="sana_dan formati noto'g'ri, YYYY-MM-DD kerak")
        dan = sana_dan
    else:
        dan = (now - timedelta(days=30)).strftime("%Y-%m-%d")

    cache_k = f"hisobot:vizitlar:{uid}:{dan}:{gacha}"
    cached = await cache_ol(cache_k)
    if cached:
        return cached

    async with rls_conn(uid) as c:
        # Jadval mavjudligini tekshirish
        has_checkin = await c.fetchval("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_name = 'checkin_out'
            )
        """)

        if has_checkin:
            # checkin_out jadvalidan vizit ma'lumotlari
            rows = await c.fetch("""
                WITH vizitlar AS (
                    SELECT
                        (ci.vaqt AT TIME ZONE 'Asia/Tashkent')::date AS sana,
                        COALESCE(u.ism, u.dokon_nomi, u.username, 'Agent') AS agent_ismi,
                        ci.klient_id,
                        ci.vaqt AT TIME ZONE 'Asia/Tashkent' AS mahalliy_vaqt
                    FROM checkin_out ci
                    LEFT JOIN users u ON u.id = ci.user_id
                    WHERE ci.turi = 'in'
                      AND (ci.vaqt AT TIME ZONE 'Asia/Tashkent')::date
                            BETWEEN $1::date AND $2::date
                ),
                sotuv_klientlar AS (
                    SELECT DISTINCT
                        (ss.sana AT TIME ZONE 'Asia/Tashkent')::date AS sana,
                        ss.klient_id
                    FROM sotuv_sessiyalar ss
                    WHERE (ss.sana AT TIME ZONE 'Asia/Tashkent')::date
                            BETWEEN $1::date AND $2::date
                      AND COALESCE(ss.holat, 'yangi') != 'bekor'
                )
                SELECT
                    v.sana,
                    v.agent_ismi,
                    COUNT(DISTINCT v.klient_id)      AS jami_vizit,
                    COUNT(DISTINCT v.klient_id)
                        FILTER (WHERE sk.klient_id IS NOT NULL)
                                                      AS sotuv_qilgan,
                    COUNT(DISTINCT v.klient_id)
                        FILTER (WHERE sk.klient_id IS NULL)
                                                      AS bosh_vizit,
                    MIN(v.mahalliy_vaqt)::time        AS birinchi_vizit,
                    MAX(v.mahalliy_vaqt)::time        AS oxirgi_vizit
                FROM vizitlar v
                LEFT JOIN sotuv_klientlar sk
                    ON sk.sana = v.sana AND sk.klient_id = v.klient_id
                GROUP BY v.sana, v.agent_ismi
                ORDER BY v.sana DESC
            """, dan, gacha)
        else:
            # Fallback: sotuv_sessiyalar dan distinct klient_id
            rows = await c.fetch("""
                SELECT
                    (ss.sana AT TIME ZONE 'Asia/Tashkent')::date AS sana,
                    COALESCE(u.ism, u.dokon_nomi, u.username, 'Agent') AS agent_ismi,
                    COUNT(DISTINCT ss.klient_id)     AS jami_vizit,
                    COUNT(DISTINCT ss.klient_id)     AS sotuv_qilgan,
                    0                                 AS bosh_vizit,
                    MIN(ss.sana AT TIME ZONE 'Asia/Tashkent')::time AS birinchi_vizit,
                    MAX(ss.sana AT TIME ZONE 'Asia/Tashkent')::time AS oxirgi_vizit
                FROM sotuv_sessiyalar ss
                LEFT JOIN users u ON u.id = ss.user_id
                WHERE (ss.sana AT TIME ZONE 'Asia/Tashkent')::date
                        BETWEEN $1::date AND $2::date
                  AND COALESCE(ss.holat, 'yangi') != 'bekor'
                GROUP BY (ss.sana AT TIME ZONE 'Asia/Tashkent')::date,
                         COALESCE(u.ism, u.dokon_nomi, u.username, 'Agent')
                ORDER BY sana DESC
            """, dan, gacha)

        items = []
        jami_vizitlar = 0
        jami_sotuv = 0

        for r in rows:
            birinchi = r["birinchi_vizit"]
            oxirgi = r["oxirgi_vizit"]

            # Ish vaqtini hisoblash (soatlarda)
            if birinchi and oxirgi:
                delta_seconds = (
                    oxirgi.hour * 3600 + oxirgi.minute * 60 + oxirgi.second
                ) - (
                    birinchi.hour * 3600 + birinchi.minute * 60 + birinchi.second
                )
                ish_vaqti = round(max(delta_seconds, 0) / 3600.0, 2)
            else:
                ish_vaqti = 0.0

            jv = int(r["jami_vizit"])
            sq = int(r["sotuv_qilgan"])
            jami_vizitlar += jv
            jami_sotuv += sq

            items.append({
                "sana":           str(r["sana"]),
                "agent_ismi":     r["agent_ismi"],
                "jami_vizit":     jv,
                "sotuv_qilgan":   sq,
                "bosh_vizit":     int(r["bosh_vizit"]),
                "birinchi_vizit": birinchi.strftime("%H:%M") if birinchi else None,
                "oxirgi_vizit":   oxirgi.strftime("%H:%M") if oxirgi else None,
                "ish_vaqti_soat": ish_vaqti,
            })

    result = {
        "sana_dan":  dan,
        "sana_gacha": gacha,
        "items": items,
        "jami": {
            "vizitlar":     jami_vizitlar,
            "sotuv_qilgan": jami_sotuv,
        },
    }
    await cache_yoz(cache_k, result, TTL_HISOBOT)
    return result
