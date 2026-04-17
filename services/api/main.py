"""
╔══════════════════════════════════════════════════════════════╗
║  SAVDOAI v25.3.2 — API SERVISI (FastAPI)                 ║
║  Bot va Web App uchun markaziy gateway                      ║
║                                                              ║
║  ✅ JWT autentifikatsiya                                     ║
║  ✅ Rate limiting (100 req/daqiqa)                          ║
║  ✅ Redis cache (5 daqiqa TTL)                              ║
║  ✅ PostgreSQL RLS (20,000 user izolyatsiyasi)              ║
║  ✅ CORS (Web uchun)                                        ║
║  ✅ /health Railway monitoring                              ║
╚══════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import os, sys, logging, time, hashlib, hmac, base64, json
from contextlib import asynccontextmanager
from typing import Optional, List

from fastapi import FastAPI, HTTPException, Depends, Request, status
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, validator
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from shared.database.pool import pool_init, pool_close, schema_init, rls_conn, get_pool
from shared.cache.redis_cache import redis_init, cache_ol, cache_yoz
from shared.cache.redis_cache import k_hisobot_kunlik, k_qarzlar, k_user, TTL_HISOBOT, TTL_USER
from shared.utils import like_escape
try:
    from shared.rag.vector_db import rag_init
except (ImportError, ModuleNotFoundError):
    rag_init = None

log = logging.getLogger(__name__)

# Sentry (ixtiyoriy — SENTRY_DSN bo'lsa uladi)
_SENTRY_DSN = os.getenv("SENTRY_DSN")
if _SENTRY_DSN:
    try:
        import sentry_sdk
        sentry_sdk.init(dsn=_SENTRY_DSN, traces_sample_rate=0.1)
        log.info("✅ Sentry ulandi")
    except ImportError:
        pass  # optional import

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-8s │ %(name)s │ %(message)s"
)

__version__ = "25.3.2"
_JWT_SECRET_RAW = os.getenv("JWT_SECRET", "")
_IS_PRODUCTION = bool(os.getenv("RAILWAY_ENVIRONMENT") or os.getenv("RAILWAY_SERVICE_ID"))
if not _JWT_SECRET_RAW:
    if _IS_PRODUCTION:
        log.critical("🚨 JWT_SECRET o'rnatilmagan! Production'da default secret ishlatilmaydi. Startup bekor.")
        sys.exit(1)
    else:
        _JWT_SECRET_RAW = "savdoai-dev-only-secret-NOT-FOR-PRODUCTION"
        log.warning("⚠️ JWT_SECRET o'rnatilmagan — faqat dev uchun default ishlatilmoqda.")
JWT_SECRET = _JWT_SECRET_RAW

# Process uptime — restart/crash ajratish (/live, /healthz, print_escpos trace)
from shared.observability.process_uptime import process_info  # noqa: E402

# ════════════════════════════════════════════════════════════
#  STARTUP / SHUTDOWN
# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
#  PYDANTIC MODELLAR (Request/Response validatsiya)
# ════════════════════════════════════════════════════════════

# ── Response modellar ──────────────────────────────────────

class HealthResponse(BaseModel):
    status: str = "ok"
    version: str
    service: str = "api"
    db_ping_ms: Optional[float] = None
    db_pool: Optional[str] = None
    latency_ms: Optional[float] = None


class DashboardResponse(BaseModel):
    bugun_sotuv_soni: int = 0
    bugun_sotuv_jami: float = 0
    bugun_yangi_qarz: float = 0
    jami_qarz: float = 0
    klient_soni: int = 0
    tovar_soni: int = 0
    kam_qoldiq_soni: int = 0


class TovarResponse(BaseModel):
    id: int
    nomi: str
    kategoriya: str = "Boshqa"
    birlik: str = "dona"
    olish_narxi: float = 0
    sotish_narxi: float = 0
    min_sotish_narxi: float = 0
    qoldiq: float = 0
    min_qoldiq: float = 0


class KlientResponse(BaseModel):
    id: int
    ism: str
    telefon: Optional[str] = None
    manzil: Optional[str] = None
    kredit_limit: float = 0
    jami_sotib: float = 0
    aktiv_qarz: float = 0


class SotuvResponse(BaseModel):
    sessiya_id: int
    status: str = "saqlandi"


class BildirishnomaTuri(BaseModel):
    tur: str
    darajasi: str
    matn: str
    klient: Optional[str] = None
    tovar: Optional[str] = None
    summa: Optional[float] = None
    qoldiq: Optional[float] = None
    soni: Optional[int] = None


class BildirishnomaResponse(BaseModel):
    items: List[BildirishnomaTuri] = []
    jami: int = 0


class FoydaResponse(BaseModel):
    kunlar: int
    brutto_sotuv: float
    tannarx: float
    sof_foyda: float
    xarajatlar: float
    toza_foyda: float
    margin_foiz: float
    top_foyda: List[dict] = []
    top_zarar: List[dict] = []


class StatistikaResponse(BaseModel):
    tovar_soni: int
    klient_soni: int
    faol_qarz: float
    kam_qoldiq_soni: int
    muddat_otgan_qarz: int
    bugun: dict
    hafta: dict
    oy: dict


# ── Request modellar ───────────────────────────────────────

class TovarModel(BaseModel):
    nomi:           str          = Field(..., min_length=1, max_length=200)
    miqdor:         float        = Field(..., gt=0)
    birlik:         str          = Field("dona")
    narx:           float        = Field(0, ge=0)
    jami:           float        = Field(0, ge=0)
    kategoriya:     str          = Field("Boshqa")
    chegirma_foiz:  float        = Field(0, ge=0, le=100)


class SotuvSo_rov(BaseModel):
    klient:         Optional[str] = None
    tovarlar:       List[TovarModel]
    jami_summa:     float         = Field(0, ge=0)
    tolangan:       float         = Field(0, ge=0)
    qarz:           float         = Field(0, ge=0)
    izoh:           Optional[str] = None


class KirimSo_rov(BaseModel):
    tovar_id:       Optional[int] = None
    tovar_nomi:     str   = Field(..., min_length=1)
    miqdor:         float = Field(..., gt=0)
    narx:           float = Field(0, ge=0)
    jami:           float = Field(0, ge=0)
    birlik:         str   = Field("dona")
    kategoriya:     str   = Field("Boshqa")
    manba:          Optional[str] = None
    izoh:           Optional[str] = None


class QarzTolashSo_rov(BaseModel):
    klient_ismi:    str   = Field(..., min_length=1)
    summa:          float = Field(..., gt=0)


class TelegramAuthSo_rov(BaseModel):
    user_id:        int
    ism:            Optional[str] = ""
    hash:           str


class PaginatsiyaResponse(BaseModel):
    total:          int
    page:           int
    limit:          int
    total_pages:    int
    has_next:       bool
    has_prev:       bool
    items:          list


@asynccontextmanager
async def lifespan(app: FastAPI):
    # ── Startup validation ──
    # ── Majburiy env'larni tekshirish (bo'lmasa ishlab chiqarish imkonsiz)
    missing = []
    for var in ["DATABASE_URL", "JWT_SECRET"]:
        if not os.getenv(var):
            missing.append(var)
    if missing:
        log.critical("❌ Kerakli env yo'q: %s", ", ".join(missing))
        sys.exit(1)

    dsn = os.environ.get("DATABASE_URL", "")

    # ── Format va xavfsizlik tekshiruvi (production'da o'zgartirish qiyin,
    # shuning uchun bootda aniqlaymiz, runtime'da emas)
    config_errors = []
    jwt_secret = os.getenv("JWT_SECRET", "")
    if len(jwt_secret) < 16:
        config_errors.append(
            f"JWT_SECRET juda qisqa ({len(jwt_secret)} belgi). Minimum 16 "
            "belgi kerak (HMAC xavfsizligi uchun). 32+ tavsiya."
        )
    if not (dsn.startswith("postgres://") or dsn.startswith("postgresql://")):
        config_errors.append(
            "DATABASE_URL format noto'g'ri — postgres:// yoki postgresql:// "
            "bilan boshlanishi kerak."
        )
    is_railway = bool(os.getenv("RAILWAY_ENVIRONMENT") or os.getenv("RAILWAY_SERVICE_ID"))
    if is_railway and not os.getenv("PRINT_SECRET", "").strip():
        config_errors.append(
            "PRINT_SECRET production'da majburiy (bot/API o'rtasida HMAC print "
            "token sync). Railway'da qo'shing."
        )
    if config_errors:
        log.critical("❌ Konfiguratsiya xato'lari:")
        for e in config_errors:
            log.critical("   • %s", e)
        sys.exit(1)

    # ── Ixtiyoriy sozlamalar haqida ogohlantirish
    optional_warns = []
    if not os.getenv("REDIS_URL"):
        optional_warns.append("REDIS_URL (cache o'chirilgan — tezlik ta'sir qiladi)")
    if not os.getenv("GOOGLE_API_KEY") and not os.getenv("GEMINI_API_KEY"):
        optional_warns.append("GEMINI_API_KEY (Ovoz STT, Rasm OCR ishlamaydi)")
    if not os.getenv("ANTHROPIC_API_KEY"):
        optional_warns.append("ANTHROPIC_API_KEY (Claude Brain-2 va Opus 4.7 audit o'chiq)")
    if not os.getenv("BOT_TOKEN"):
        optional_warns.append("BOT_TOKEN (Telegram bot o'chirilgan)")
    for w in optional_warns:
        log.warning("⚠️ %s", w)

    # Railway postgres:// → asyncpg postgresql://
    if dsn.startswith("postgres://"):
        dsn = dsn.replace("postgres://", "postgresql://", 1)
    r_url  = os.getenv("REDIS_URL", "")
    q_url  = os.getenv("QDRANT_URL")
    q_key  = os.getenv("QDRANT_API_KEY")

    await pool_init(dsn,
                    min_size=int(os.getenv("DB_MIN", "2")),
                    max_size=int(os.getenv("DB_MAX", "10")))

    try:
        await schema_init()
    except Exception as _schema_err:
        log.error("⚠️ Schema init xato (API davom etadi): %s", _schema_err)

    if r_url:
        try:
            await redis_init(r_url)
        except Exception as _r:
            log.warning("Redis ulana olmadi: %s", _r)

    if rag_init and q_url:
        try:
            rag_init(q_url, q_key)
        except Exception as _e:
            log.warning("Qdrant ulana olmadi (RAG o'chirildi): %s", _e)

    log.info("🚀 SavdoAI API v%s tayyor", __version__)
    log.info(
        "API boot ok: pid=%s port=%s env=%s print_secret_set=%s",
        os.getpid(),
        os.getenv("PORT", "8000"),
        os.getenv("RAILWAY_ENVIRONMENT") or "local",
        bool((os.getenv("PRINT_SECRET") or "").strip()),
    )
    yield
    await pool_close()
    log.info("API to'xtatildi")


app = FastAPI(
    title       = "SavdoAI Mashrab Moliya API",
    version     = __version__,
    description = "O'zbek bozori uchun AI-powered savdo boshqaruv tizimi REST API — 107 endpoint, CRUD, hisobotlar, foyda tahlili, real-time WebSocket",
    lifespan    = lifespan,
    docs_url    = "/docs",
    redoc_url   = "/redoc",
    openapi_tags = [
        {"name": "Auth",           "description": "Autentifikatsiya — login, token"},
        {"name": "Dashboard",      "description": "Bosh sahifa statistikasi"},
        {"name": "Tovarlar",       "description": "Tovarlar CRUD, import, export"},
        {"name": "Klientlar",      "description": "Klientlar CRUD, tarix"},
        {"name": "Sotuv",          "description": "Sotuv operatsiyalari"},
        {"name": "Qarz",           "description": "Qarzlar va to'lovlar"},
        {"name": "Hisobotlar",     "description": "Kunlik, haftalik, oylik, foyda"},
        {"name": "Xarajatlar",     "description": "Xarajatlar boshqaruvi"},
        {"name": "Kassa",          "description": "Kassa operatsiyalari"},
        {"name": "Narx",           "description": "Narx guruhlari va shaxsiy narxlar"},
        {"name": "Ledger",         "description": "SAP-grade buxgalteriya"},
        {"name": "Bildirishnoma",  "description": "Ogohlantirish va bildirishnomalar"},
        {"name": "Export",         "description": "Excel/PDF eksport"},
        {"name": "Monitoring",     "description": "Health check va metrikalar"},
    ],
)

# CORS — browser Origin must match exactly (no trailing slash). WEB_URL is the web app origin, not API.
def _cors_origin(url: str) -> str:
    u = (url or "").strip()
    return u.rstrip("/") if u else u


_web_cors_origins = [
    _cors_origin("https://savdoai-web-production.up.railway.app"),
    _cors_origin("http://localhost:3000"),
    _cors_origin("http://127.0.0.1:3000"),
    _cors_origin("https://mashrab-moliya.vercel.app"),
    _cors_origin("http://localhost:5173"),
    _cors_origin("http://localhost:8000"),
]
_web_url = _cors_origin(os.getenv("WEB_URL", "https://savdoai-web-production.up.railway.app"))
if _web_url and _web_url not in _web_cors_origins:
    _web_cors_origins.insert(0, _web_url)

# CORS only (no GZip here) — gzip has interfered with preflight/OPTIONS in some deployments.
# Explicit methods/headers avoid Starlette edge cases with "*" + credentials.
app.add_middleware(
    CORSMiddleware,
    allow_origins=_web_cors_origins,
    allow_origin_regex=r"https://savdoai[-\w]*\.up\.railway\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Har so'rovga X-Request-ID (502 da loglarni bog'lash)
from shared.middleware.request_id import RequestIDMiddleware  # noqa: E402

app.add_middleware(RequestIDMiddleware)

# ═══ v25.3.2 ROUTELAR ═══
try:
    from services.api.routes.kassa import router as kassa_router
    app.include_router(kassa_router, tags=["Kassa"])
    log.info("✅ Kassa moduli ulandi")
except Exception as e:
    log.warning("⚠️ Kassa moduli yuklanmadi: %s", e)

try:
    from services.api.routes.websocket import router as ws_router
    app.include_router(ws_router, tags=["Monitoring"])
    log.info("✅ WebSocket ulandi")
except Exception as e:
    log.warning("⚠️ WebSocket yuklanmadi: %s", e)

# Print HMAC: shared.services.print_session (lazy secret — import-time crash oldini olish)

try:
    from services.api.routes.printer import router as printer_router
    app.include_router(printer_router, tags=["Sotuv"])
    log.info("✅ Printer API ulandi")
except Exception as e:
    log.warning("⚠️ Printer API yuklanmadi: %s", e)

# ═══ v25.3.2 YANGI ROUTE MODULLARI ═══
try:
    from services.api.routes.tovarlar import router as tovarlar_router
    app.include_router(tovarlar_router, tags=["Tovarlar"])
    log.info("✅ Tovarlar moduli ulandi")
except Exception as e:
    log.warning("⚠️ Tovarlar moduli yuklanmadi: %s", e)

try:
    from services.api.routes.hisobotlar import router as hisobotlar_router
    app.include_router(hisobotlar_router, tags=["Hisobotlar"])
    log.info("✅ Hisobotlar moduli ulandi")
except Exception as e:
    log.warning("⚠️ Hisobotlar moduli yuklanmadi: %s", e)

try:
    from services.api.routes.klientlar import router as klientlar_router
    app.include_router(klientlar_router, tags=["Klientlar"])
    log.info("✅ Klientlar moduli ulandi")
except Exception as e:
    log.warning("⚠️ Klientlar moduli yuklanmadi: %s", e)

try:
    from services.api.routes.faktura import router as faktura_router
    app.include_router(faktura_router, tags=["Faktura"])
    log.info("✅ Faktura moduli ulandi")
except Exception as e:
    log.warning("⚠️ Faktura moduli yuklanmadi: %s", e)

try:
    from services.api.routes.monitoring import router as monitoring_router
    app.include_router(monitoring_router, tags=["Monitoring"])
    log.info("✅ Monitoring moduli ulandi")
except Exception as e:
    log.warning("⚠️ Monitoring moduli yuklanmadi: %s", e)

try:
    from services.api.routes.yangi import router as yangi_router
    app.include_router(yangi_router, tags=["Yangi"])
    log.info("✅ Yangi modullar ulandi (KPI, Loyalty, To'lov, Qarz eslatma)")
except Exception as e:
    log.warning("⚠️ Yangi modullar yuklanmadi: %s", e)

try:
    from services.api.routes.filial import router as filial_router
    app.include_router(filial_router, tags=["Filial"])
    log.info("✅ Multi-filial moduli ulandi")
except Exception as e:
    log.warning("⚠️ Multi-filial yuklanmadi: %s", e)

# ═══ YANGI MODULLAR (SalesDoc funksiyalari) ═══
try:
    from services.api.routes.aksiya import router as aksiya_router
    app.include_router(aksiya_router, tags=["Aksiya"])
    log.info("✅ Aksiya (chegirma/bonus) moduli ulandi")
except Exception as e:
    log.warning("⚠️ Aksiya moduli yuklanmadi: %s", e)

try:
    from services.api.routes.analitika import router as analitika_router
    app.include_router(analitika_router, tags=["Analitika"])
    log.info("✅ Analitika moduli ulandi")
except Exception as e:
    log.warning("⚠️ Analitika moduli yuklanmadi: %s", e)

try:
    from services.api.routes.config import router as config_router
    app.include_router(config_router, tags=["Config"])
    log.info("✅ Config moduli ulandi")
except Exception as e:
    log.warning("⚠️ Config moduli yuklanmadi: %s", e)

try:
    from services.api.routes.moliya_live import router as moliya_router
    app.include_router(moliya_router, tags=["Moliya"])
    log.info("✅ Moliya (P&L, balans) moduli ulandi")
except Exception as e:
    log.warning("⚠️ Moliya moduli yuklanmadi: %s", e)

try:
    from services.api.routes.tashrif import router as tashrif_router
    app.include_router(tashrif_router, tags=["Tashrif"])
    log.info("✅ Tashrif (check-in/out) moduli ulandi")
except Exception as e:
    log.warning("⚠️ Tashrif moduli yuklanmadi: %s", e)

try:
    from services.api.routes.tovarlar_v2 import router as tovarlar_v2_router
    app.include_router(tovarlar_v2_router, tags=["Tovarlar V2"])
    log.info("✅ Tovarlar V2 (kengaytirilgan) moduli ulandi")
except Exception as e:
    log.warning("⚠️ Tovarlar V2 yuklanmadi: %s", e)

try:
    from services.api.routes.export_calendar_notif import router as export_router
    app.include_router(export_router, tags=["Export"])
    log.info("✅ Export/Calendar/Notification moduli ulandi")
except Exception as e:
    log.warning("⚠️ Export moduli yuklanmadi: %s", e)

try:
    from services.api.routes.gps import router as gps_router
    app.include_router(gps_router, tags=["GPS"])
    log.info("✅ GPS moduli ulandi")
except Exception as e:
    log.warning("⚠️ GPS moduli yuklanmadi: %s", e)

try:
    from services.api.routes.enterprise import router as enterprise_router
    app.include_router(enterprise_router, tags=["Enterprise"])
    log.info("✅ Enterprise (vazifa, uskunalar) moduli ulandi")
except Exception as e:
    log.warning("⚠️ Enterprise moduli yuklanmadi: %s", e)

try:
    from services.api.routes.van_sverka import router as van_router
    app.include_router(van_router, tags=["Van Selling"])
    log.info("✅ Van Selling (marshrut yetkazish) moduli ulandi")
except Exception as e:
    log.warning("⚠️ Van Selling yuklanmadi: %s", e)

try:
    from services.api.routes.pro_features import router as pro_router
    app.include_router(pro_router, tags=["Pro"])
    log.info("✅ Pro Features (360, marshrut, leaderboard) ulandi")
except Exception as e:
    log.warning("⚠️ Pro Features yuklanmadi: %s", e)

try:
    from services.api.routes.ai_extras import router as ai_extras_router
    app.include_router(ai_extras_router)
    log.info("✅ AI Extras (Claude Opus 4.7, DeepSeek, Grok, v0) endpoint ulandi")
except Exception as e:
    log.warning("⚠️ AI Extras yuklanmadi: %s", e)

try:
    from services.api.routes.sync import router as sync_router
    app.include_router(sync_router)
    log.info("✅ Sync (delta) endpoint ulandi — offline-first")
except Exception as e:
    log.warning("⚠️ Sync yuklanmadi: %s", e)

try:
    from services.api.routes.sd_agent_gaps import router as sd_router
    app.include_router(sd_router, tags=["SD Agent"])
    log.info("✅ SD Agent (blok, tara, almashtirish) ulandi")
except Exception as e:
    log.warning("⚠️ SD Agent yuklanmadi: %s", e)

try:
    from services.api.routes.suppliers import router as suppliers_router
    app.include_router(suppliers_router, tags=["Suppliers"])
    log.info("✅ Suppliers/Purchase moduli ulandi")
except Exception as e:
    log.warning("⚠️ Suppliers yuklanmadi: %s", e)

try:
    from services.api.routes.sozlamalar import router as sozlamalar_router
    app.include_router(sozlamalar_router, tags=["Sozlamalar"])
    log.info("✅ Sozlamalar (Settings) moduli ulandi")
except Exception as e:
    log.warning("⚠️ Sozlamalar yuklanmadi: %s", e)

# ════════════════════════════════════════════════════════════
#  JWT + AUTH — deps.py dan import (shared bilan kassa/ws)
# ════════════════════════════════════════════════════════════

from services.api.deps import get_uid, jwt_tekshir


def jwt_yarat(user_id: int, ttl: int = 86400) -> str:
    """JWT token yaratish"""
    h64 = base64.urlsafe_b64encode(
        b'{"alg":"HS256","typ":"JWT"}'
    ).rstrip(b"=").decode()
    payload = json.dumps({"sub": str(user_id), "exp": int(time.time()) + ttl})
    p64 = base64.urlsafe_b64encode(payload.encode()).rstrip(b"=").decode()
    sig = base64.urlsafe_b64encode(
        hmac.new(JWT_SECRET.encode(), f"{h64}.{p64}".encode(), "sha256").digest()
    ).rstrip(b"=").decode()
    return f"{h64}.{p64}.{sig}"


# ════════════════════════════════════════════════════════════
#  ENDPOINTLAR
# ════════════════════════════════════════════════════════════

@app.get("/", include_in_schema=False)
async def root():
    """Landing page — API ishlayotganini ko'rsatadi"""
    from fastapi.responses import HTMLResponse
    html = f"""<!DOCTYPE html>
<html lang="uz">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SavdoAI Mashrab Moliya v{__version__}</title>
    <style>
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
               background:linear-gradient(135deg,#0f172a 0%,#1e293b 100%);
               color:#e2e8f0; min-height:100vh; display:flex; align-items:center; justify-content:center; }}
        .card {{ background:#1e293b; border:1px solid #334155; border-radius:16px;
                padding:48px; max-width:600px; width:90%; box-shadow:0 20px 60px rgba(0,0,0,0.4); }}
        h1 {{ font-size:28px; margin-bottom:8px; color:#38bdf8; }}
        .version {{ color:#94a3b8; font-size:14px; margin-bottom:24px; }}
        .status {{ display:flex; align-items:center; gap:8px; margin-bottom:24px;
                  padding:12px 16px; background:#0f172a; border-radius:8px; }}
        .dot {{ width:10px; height:10px; border-radius:50%; background:#22c55e; animation:pulse 2s infinite; }}
        @keyframes pulse {{ 0%,100%{{opacity:1}} 50%{{opacity:0.5}} }}
        .links {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; margin-top:24px; }}
        .links a {{ display:block; padding:12px 16px; background:#334155; border-radius:8px;
                   color:#38bdf8; text-decoration:none; text-align:center; font-size:14px;
                   transition:background 0.2s; }}
        .links a:hover {{ background:#475569; }}
        .stats {{ display:grid; grid-template-columns:repeat(3,1fr); gap:12px; margin-top:24px; }}
        .stat {{ text-align:center; padding:16px; background:#0f172a; border-radius:8px; }}
        .stat-num {{ font-size:24px; font-weight:700; color:#38bdf8; }}
        .stat-label {{ font-size:12px; color:#94a3b8; margin-top:4px; }}
    </style>
</head>
<body>
    <div class="card">
        <h1>SavdoAI Mashrab Moliya</h1>
        <div class="version">v{__version__} &mdash; AI-powered savdo boshqaruv tizimi</div>
        <div class="status">
            <div class="dot"></div>
            <span>API ishlayapti</span>
        </div>
        <div class="stats">
            <div class="stat"><div class="stat-num">107</div><div class="stat-label">API Endpoints</div></div>
            <div class="stat"><div class="stat-num">19</div><div class="stat-label">DB Jadvallar</div></div>
            <div class="stat"><div class="stat-num">2</div><div class="stat-label">AI Modellar</div></div>
        </div>
        <div class="links">
            <a href="/docs">API Docs</a>
            <a href="/health">Health Check</a>
            <a href="/redoc">ReDoc</a>
            <a href="https://t.me/savdoai_mashrab_bot">Telegram Bot</a>
        </div>
    </div>
</body>
</html>"""
    return HTMLResponse(content=html)
    """Redirect to Next.js Web Dashboard"""
    from fastapi.responses import RedirectResponse
    web_url = os.getenv("WEB_DASHBOARD_URL", "")
    if web_url:
        return RedirectResponse(url=web_url)
    return {"message": "Dashboard: WEB_DASHBOARD_URL sozlang"}


@app.post("/auth/telegram", tags=["Auth"])
async def auth_telegram(data: TelegramAuthSo_rov):
    """
    Telegram bot → API token olish.
    Bot har foydalanuvchi uchun bu endpoint orqali JWT oladi.
    """
    uid      = data.user_id
    bot_hash = data.hash
    expected = hmac.new(
        JWT_SECRET.encode(), str(uid).encode(), "sha256"
    ).hexdigest()[:32]

    if not hmac.compare_digest(bot_hash, expected):
        raise HTTPException(403, "Noto'g'ri hash")

    # User mavjudligini tekshirish/yaratish
    async with get_pool().acquire() as c:
        u = await c.fetchrow("SELECT id, ism, faol FROM users WHERE id=$1", uid)
        if not u:
            await c.execute(
                "INSERT INTO users(id,ism) VALUES($1,$2) ON CONFLICT DO NOTHING",
                uid, data.ism or ""
            )

    token = jwt_yarat(uid)
    return {"token": token, "user_id": uid}


# ════════════════════════════════════════════════════════════
#  TELEGRAM MINI APP (WebApp) AUTH
# ════════════════════════════════════════════════════════════


@app.post("/auth/webapp", tags=["Auth"])
async def auth_webapp(data: dict):
    """
    Telegram Mini App (WebApp) autentifikatsiya.
    Client Telegram.WebApp.initData yuboradi → server tekshiradi → JWT qaytaradi.

    Telegram WebApp initData validatsiya:
    1. initData parse → hash ajratish
    2. Qolgan fieldlarni alifbo tartibida saralash, \\n bilan birlashtirish
    3. secret_key = HMAC-SHA256("WebAppData", BOT_TOKEN)
    4. Tekshirish: HMAC-SHA256(secret_key, data_check_string) == hash
    """
    from urllib.parse import parse_qs

    init_data = data.get("initData", "")
    if not init_data:
        raise HTTPException(400, "initData bo'sh")

    bot_token = os.getenv("BOT_TOKEN", "")
    if not bot_token:
        raise HTTPException(503, "BOT_TOKEN o'rnatilmagan")

    # 1. Parse initData
    parsed = parse_qs(init_data, keep_blank_values=True)
    received_hash = parsed.pop("hash", [""])[0]
    if not received_hash:
        raise HTTPException(400, "hash topilmadi")

    # 2. data_check_string — alifbo tartibida saralangan key=value
    data_check_parts = []
    for key in sorted(parsed.keys()):
        val = parsed[key][0]
        data_check_parts.append(f"{key}={val}")
    data_check_string = "\n".join(data_check_parts)

    # 3. secret_key = HMAC-SHA256("WebAppData", BOT_TOKEN)
    secret_key = hmac.new(
        b"WebAppData", bot_token.encode(), "sha256"
    ).digest()

    # 4. Tekshirish
    computed_hash = hmac.new(
        secret_key, data_check_string.encode(), "sha256"
    ).hexdigest()

    if not hmac.compare_digest(computed_hash, received_hash):
        raise HTTPException(403, "initData tekshiruvi muvaffaqiyatsiz")

    # 5. auth_date tekshirish (24 soatdan eski bo'lmasligi kerak)
    auth_date = int(parsed.get("auth_date", ["0"])[0])
    if time.time() - auth_date > 86400:
        raise HTTPException(403, "initData muddati o'tgan")

    # 6. User ma'lumotlarini olish
    user_json = parsed.get("user", ["{}"])[0]
    try:
        tg_user = json.loads(user_json)
    except json.JSONDecodeError:
        raise HTTPException(400, "user field noto'g'ri")

    tg_id = tg_user.get("id")
    if not tg_id:
        raise HTTPException(400, "user.id topilmadi")

    # 7. User yaratish/olish
    ism = tg_user.get("first_name", "")
    username = tg_user.get("username", "")

    async with get_pool().acquire() as c:
        u = await c.fetchrow(
            "SELECT id, ism FROM users WHERE id=$1", int(tg_id)
        )
        if not u:
            await c.execute(
                "INSERT INTO users(id, ism, username) VALUES($1, $2, $3) ON CONFLICT DO NOTHING",
                int(tg_id), ism, username
            )

    token = jwt_yarat(int(tg_id))
    log.info("📱 Mini App auth: uid=%d ism=%s", tg_id, ism)
    return {"token": token, "user_id": tg_id, "ism": ism}


# ════════════════════════════════════════════════════════════
#  LOGIN/PAROL — Web panel uchun
# ════════════════════════════════════════════════════════════

def _parol_hash(parol: str) -> str:
    """Parolni xavfsiz hash qilish (PBKDF2-SHA256, salt bilan)"""
    import os as _os2
    salt = _os2.urandom(16).hex()
    h = hashlib.pbkdf2_hmac("sha256", parol.encode(), salt.encode(), 100_000).hex()
    return f"{salt}:{h}"


def _parol_tekshir(parol: str, stored: str) -> bool:
    """Saqlangan hash bilan parolni solishtirish"""
    if not stored or ":" not in stored:
        return False
    salt, h = stored.split(":", 1)
    return hashlib.pbkdf2_hmac("sha256", parol.encode(), salt.encode(), 100_000).hex() == h


def _telefon_tozala(tel: str) -> str:
    """Telefon raqamni normallashtirish: +998901234567 → 998901234567"""
    t = tel.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if t.startswith("+"): t = t[1:]
    if len(t) == 9 and t[0] in "3456789":
        t = "998" + t
    return t


class LoginSorov(BaseModel):
    login: Optional[str] = None
    telefon: Optional[str] = None
    parol: str = Field(..., min_length=1)


@app.post("/auth/login", tags=["Auth"])
async def auth_login(data: LoginSorov, request: Request):
    """
    Web panel kirish — login+parol YOKI telefon+parol.
    Admin do'konchilariga login/parol beradi.
    Rate limit: 5 urinish/daqiqa (brute-force himoya).
    """
    from services.api.deps import login_rate_check
    await login_rate_check(request)
    login = (data.login or "").strip().lower()
    telefon = (data.telefon or "").strip()
    parol = data.parol.strip()

    if not login and not telefon:
        raise HTTPException(400, "Login yoki telefon kiriting")

    async with get_pool().acquire() as c:
        if login:
            user = await c.fetchrow(
                "SELECT id, ism, to_liq_ism, username, telefon, dokon_nomi, segment, faol, login, parol_hash FROM users WHERE lower(login)=$1 AND faol=TRUE",
                login,
            )
        else:
            tel = _telefon_tozala(telefon)
            user = await c.fetchrow(
                "SELECT id, ism, to_liq_ism, username, telefon, dokon_nomi, segment, faol, login, parol_hash FROM users WHERE replace(replace(telefon,' ',''),'-','') LIKE '%' || $1 || '%' AND faol=TRUE LIMIT 1",
                tel[-9:],
            )

    if not user:
        raise HTTPException(401, "Login yoki parol noto'g'ri")

    if not user.get("parol_hash"):
        raise HTTPException(401, "Parol o'rnatilmagan. Admin bilan bog'laning.")

    if not _parol_tekshir(parol, user["parol_hash"]):
        raise HTTPException(401, "Login yoki parol noto'g'ri")

    token = jwt_yarat(user["id"])
    log.info("🔐 Web login: uid=%d login=%s", user["id"], login or telefon)
    return {"token": token, "user_id": user["id"]}


@app.post("/api/v1/admin/parol", tags=["Auth"])
async def admin_parol_qoy(data: dict, uid: int = Depends(get_uid)):
    """Admin: do'konchiga login/parol berish"""
    # Admin tekshiruvi
    admin_ids = os.getenv("ADMIN_IDS", "").split(",")
    if str(uid) not in [a.strip() for a in admin_ids]:
        raise HTTPException(403, "Faqat admin uchun")

    target_id = data.get("user_id")
    login = (data.get("login") or "").strip()
    parol = (data.get("parol") or "").strip()

    if not target_id or not parol:
        raise HTTPException(400, "user_id va parol kerak")
    if len(parol) < 4:
        raise HTTPException(400, "Parol kamida 4 belgi bo'lishi kerak")

    hashed = _parol_hash(parol)

    async with get_pool().acquire() as c:
        user = await c.fetchrow("SELECT id, ism FROM users WHERE id=$1", int(target_id))
        if not user:
            raise HTTPException(404, "Foydalanuvchi topilmadi")

        if login:
            # Login band emasligini tekshirish
            existing = await c.fetchrow(
                "SELECT id FROM users WHERE lower(login)=$1 AND id!=$2",
                login.lower(), int(target_id),
            )
            if existing:
                raise HTTPException(409, f"'{login}' login allaqachon band")
            await c.execute(
                "UPDATE users SET login=$1, parol_hash=$2, yangilangan=NOW() WHERE id=$3",
                login, hashed, int(target_id),
            )
        else:
            await c.execute(
                "UPDATE users SET parol_hash=$1, yangilangan=NOW() WHERE id=$2",
                hashed, int(target_id),
            )

    from shared.cache.redis_cache import user_cache_tozala
    await user_cache_tozala(int(target_id))
    log.info("🔐 Admin parol qo'ydi: target=%d login=%s", int(target_id), login)
    return {"ok": True, "user_id": int(target_id), "login": login or None}


@app.get("/api/v1/me", tags=["Auth"])
async def me(uid: int = Depends(get_uid)):
    """Joriy foydalanuvchi"""
    cached = await cache_ol(k_user(uid))
    if cached:
        safe = {k: v for k, v in cached.items() if k != "parol_hash"}
        return safe
    async with rls_conn(uid) as c:
        u = await c.fetchrow("SELECT id, ism, to_liq_ism, username, telefon, inn, manzil, dokon_nomi, segment, faol, obuna_tugash, til, plan, login, yaratilgan FROM users WHERE id=$1", uid)
        if not u:
            raise HTTPException(404, "Topilmadi")
        result = dict(u)
        await cache_yoz(k_user(uid), result, TTL_USER)
        return result


@app.get("/api/v1/dashboard", tags=["Dashboard"])
async def dashboard(uid: int = Depends(get_uid)):
    """Dashboard statistikasi (5 daqiqa cache)"""
    cache_k = f"dashboard:{uid}"
    cached  = await cache_ol(cache_k)
    if cached:
        return cached

    async with rls_conn(uid) as c:
        bugun = await c.fetchrow("""
            SELECT
                COUNT(*)              AS sotuv_soni,
                COALESCE(SUM(jami),0) AS sotuv_jami,
                COALESCE(SUM(qarz),0) AS yangi_qarz
            FROM sotuv_sessiyalar
            WHERE (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
        """)
        jami_qarz = await c.fetchval("""
            SELECT COALESCE(SUM(qolgan),0)
            FROM qarzlar WHERE yopildi=FALSE
        """)
        klient_soni = await c.fetchval("SELECT COUNT(*) FROM klientlar WHERE user_id=$1", uid)
        tovar_soni  = await c.fetchval("SELECT COUNT(*) FROM tovarlar WHERE user_id=$1", uid)
        kam_qoldiq  = await c.fetchval("""
            SELECT COUNT(*) FROM tovarlar
            WHERE user_id=$1 AND min_qoldiq>0 AND qoldiq<=min_qoldiq
        """, uid)
        # Kumulyativ (butun vaqt) sotuv — dashboardda total revenue uchun
        total_revenue = await c.fetchval("""
            SELECT COALESCE(SUM(jami), 0) FROM sotuv_sessiyalar WHERE user_id=$1
        """, uid) or 0
        # 30-kunlik faol klientlar soni
        active_clients = await c.fetchval("""
            SELECT COUNT(DISTINCT klient_id) FROM sotuv_sessiyalar
            WHERE user_id=$1 AND sana >= NOW() - interval '30 days'
              AND klient_id IS NOT NULL
        """, uid) or 0

        # Muddati o'tgan qarzlar
        overdue = await c.fetchrow("""
            SELECT COUNT(*) AS soni, COALESCE(SUM(qolgan), 0) AS jami
            FROM qarzlar
            WHERE yopildi=FALSE AND qolgan>0 AND muddat < CURRENT_DATE
        """)

        # Kutilayotgan xarajatlar (jadval yo'q bo'lsa 0 qaytariladi — sokin fallback)
        pending_exp = 0
        try:
            pending_exp = await c.fetchval("""
                SELECT COUNT(*) FROM xarajatlar
                WHERE admin_uid=$1 AND NOT tasdiqlangan AND NOT bekor_qilingan
            """, uid) or 0
        except Exception as _e:
            log.debug("dashboard: pending_exp olishda xato (fallback 0): %s", _e)

        # Faol shogirdlar (jadval yo'q bo'lsa 0 qaytariladi)
        active_app = 0
        try:
            active_app = await c.fetchval("""
                SELECT COUNT(*) FROM shogirdlar
                WHERE admin_uid=$1 AND faol=TRUE
            """, uid) or 0
        except Exception as _e:
            log.debug("dashboard: active_app olishda xato (fallback 0): %s", _e)

    result = {
        "bugun_sotuv_soni":  int(bugun["sotuv_soni"]),
        "bugun_sotuv_jami":  float(bugun["sotuv_jami"]),
        "bugun_yangi_qarz":  float(bugun["yangi_qarz"]),
        "jami_qarz":         float(jami_qarz or 0),
        "total_revenue":     float(total_revenue),
        "active_clients":    int(active_clients),
        "klient_soni":       klient_soni,
        "tovar_soni":        tovar_soni,
        "kam_qoldiq_soni":   kam_qoldiq,
        "overdue_count":     int(overdue["soni"]) if overdue else 0,
        "overdue_amount":    float(overdue["jami"]) if overdue else 0,
        "pending_expenses":  int(pending_exp),
        "active_apprentices": int(active_app),
    }
    await cache_yoz(cache_k, result, TTL_HISOBOT)
    return result


# ════════════════════════════════════════════════════════════
#  DASHBOARD V2 — KPI + Ombor + Qarz birga
# ════════════════════════════════════════════════════════════

@app.get("/api/v1/dashboard/v2", tags=["Dashboard"])
async def dashboard_v2(uid: int = Depends(get_uid)):
    """Yaxshilangan dashboard — barcha ko'rsatkichlar bitta so'rovda."""
    cache_k = f"dashboard_v2:{uid}"
    cached = await cache_ol(cache_k)
    if cached:
        return cached

    async with rls_conn(uid) as c:
        bugun = await c.fetchrow("""
            SELECT COUNT(*) soni, COALESCE(SUM(jami),0) jami,
                   COALESCE(SUM(qarz),0) qarz
            FROM sotuv_sessiyalar
            WHERE (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
        """)
        oy = await c.fetchrow("""
            SELECT COUNT(*) soni, COALESCE(SUM(jami),0) jami
            FROM sotuv_sessiyalar WHERE sana >= NOW() - interval '30 days'
        """)
        qarz_stat = await c.fetchrow("""
            SELECT COUNT(DISTINCT klient_ismi) klient_soni,
                   COALESCE(SUM(qolgan),0) jami,
                   COUNT(*) FILTER(WHERE muddat IS NOT NULL AND muddat < CURRENT_DATE) muddati_otgan
            FROM qarzlar WHERE yopildi=FALSE AND qolgan>0
        """)
        ombor = await c.fetchrow("""
            SELECT COUNT(*) jami, COUNT(*) FILTER(WHERE min_qoldiq>0 AND qoldiq<=min_qoldiq) kam,
                   COUNT(*) FILTER(WHERE qoldiq<=0) tugagan,
                   COALESCE(SUM(qoldiq*olish_narxi),0) qiymat
            FROM tovarlar WHERE user_id=$1
        """, uid)
        trend = [{"kun": str(r["kun"]), "soni": int(r["soni"]), "jami": float(r["jami"])}
                 for r in await c.fetch("""
            SELECT (sana AT TIME ZONE 'Asia/Tashkent')::date AS kun,
                   COUNT(*) soni, COALESCE(SUM(jami),0) jami
            FROM sotuv_sessiyalar WHERE sana >= NOW()-interval '7 days'
            GROUP BY kun ORDER BY kun
        """)]

    soni_30 = int(oy["soni"])
    result = {
        "bugun": {"soni": int(bugun["soni"]), "jami": float(bugun["jami"]), "qarz": float(bugun["qarz"])},
        "oy": {"soni": soni_30, "jami": float(oy["jami"])},
        "qarz": {"klient_soni": int(qarz_stat["klient_soni"]), "jami": float(qarz_stat["jami"]),
                  "muddati_otgan": int(qarz_stat["muddati_otgan"])},
        "ombor": {"jami": int(ombor["jami"]), "kam": int(ombor["kam"]),
                   "tugagan": int(ombor["tugagan"]), "qiymat": float(ombor["qiymat"])},
        "reyting": "A" if soni_30>=200 else ("B" if soni_30>=100 else ("C" if soni_30>=30 else "D")),
        "trend": trend,
    }
    await cache_yoz(cache_k, result, TTL_HISOBOT)
    return result
    """Faol qarzlar (10 daqiqa cache)"""
    cached = await cache_ol(k_qarzlar(uid))
    if cached:
        return cached
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT klient_ismi,
                   SUM(qolgan)  AS qolgan,
                   COUNT(*)     AS soni,
                   MIN(muddat)  AS muddat
            FROM qarzlar
            WHERE yopildi=FALSE AND qolgan>0
            GROUP BY klient_ismi
            ORDER BY qolgan DESC
        """)
    result = [dict(r) for r in rows]
    await cache_yoz(k_qarzlar(uid), result, TTL_HISOBOT * 2)
    return result


@app.get("/api/v1/qarzlar", tags=["Qarz"])
async def qarzlar(uid: int = Depends(get_uid)):
    """Faol qarzlar (10 daqiqa cache)"""
    cached = await cache_ol(k_qarzlar(uid))
    if cached:
        return cached
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT klient_ismi,
                   SUM(qolgan)  AS qolgan,
                   COUNT(*)     AS soni,
                   MIN(muddat)  AS muddat
            FROM qarzlar
            WHERE yopildi=FALSE AND qolgan>0
            GROUP BY klient_ismi
            ORDER BY qolgan DESC
        """)
    result = [dict(r) for r in rows]
    await cache_yoz(k_qarzlar(uid), result, TTL_HISOBOT * 2)
    return result


@app.get("/api/v1/hisobot/kunlik", tags=["Hisobotlar"])
async def hisobot_kunlik(uid: int = Depends(get_uid)):
    """Bugungi hisobot (5 daqiqa cache)"""
    cached = await cache_ol(k_hisobot_kunlik(uid))
    if cached:
        return cached
    async with rls_conn(uid) as c:
        kr = await c.fetchrow("""
            SELECT COUNT(*) n, COALESCE(SUM(jami),0) jami
            FROM kirimlar
            WHERE (sana AT TIME ZONE 'Asia/Tashkent')::date=CURRENT_DATE
        """)
        ch = await c.fetchrow("""
            SELECT COUNT(*) n,
                   COALESCE(SUM(jami),0)     jami,
                   COALESCE(SUM(qarz),0)     qarz,
                   COALESCE(SUM(tolangan),0) tolangan
            FROM sotuv_sessiyalar
            WHERE (sana AT TIME ZONE 'Asia/Tashkent')::date=CURRENT_DATE
        """)
        jami_qarz = await c.fetchval(
            "SELECT COALESCE(SUM(qolgan),0) FROM qarzlar WHERE yopildi=FALSE"
        )

    result = {
        "kirim":    {"soni": int(kr["n"]),  "jami": float(kr["jami"])},
        "sotuv":    {"soni": int(ch["n"]),  "jami": float(ch["jami"]),
                     "qarz": float(ch["qarz"])},
        "jami_qarz": float(jami_qarz or 0),
    }
    await cache_yoz(k_hisobot_kunlik(uid), result, TTL_HISOBOT)
    return result


# ════════════════════════════════════════════════════════════
#  KOGNITIV ENGINE PROXY
# ════════════════════════════════════════════════════════════

@app.post("/api/v1/tahlil", tags=["Sotuv"])
async def tahlil(data: dict, uid: int = Depends(get_uid)):
    """
    AI tahlil — Cognitive Engine ga proksi.
    RAG + Tool Calling + temperature=0.0
    """
    import httpx
    matn = data.get("matn", "").strip()
    if not matn:
        raise HTTPException(400, "Matn bo'sh")

    cog_url = os.getenv("COGNITIVE_URL", "")
    if not cog_url:
        raise HTTPException(503, "Kognitiv dvigatel ulanmagan")

    async with httpx.AsyncClient(timeout=35.0) as http:
        try:
            r = await http.post(
                f"{cog_url}/tahlil",
                json={"matn": matn, "uid": uid},
            )
            r.raise_for_status()
            return r.json()
        except httpx.TimeoutException:
            raise HTTPException(504, "AI timeout")
        except Exception as e:
            log.error("Kognitiv xato: %s", e)
            raise HTTPException(502, "AI xato")


# ════════════════════════════════════════════════════════════
#  SOTUV VA KIRIM ENDPOINTLARI
# ════════════════════════════════════════════════════════════

@app.post("/api/v1/sotuv", tags=["Sotuv"])
async def sotuv_saqlash(data: SotuvSo_rov, request: Request, uid: int = Depends(get_uid)):
    """Sotuv operatsiyasini saqlash. Rate limit: 30/daqiqa.

    1. Sessiya yaratish
    2. Har bir tovar uchun chiqim yozuvi + qoldiq kamaytirish
    3. Klient topish/yaratish
    4. Qarz saqlash (agar bor bo'lsa)
    """
    from services.api.deps import endpoint_rate_check
    await endpoint_rate_check(request, "sotuv")
    from shared.utils.hisob import sotuv_validatsiya, ai_hisob_tekshir
    from shared.cache.redis_cache import user_cache_tozala

    ok, xato = sotuv_validatsiya(data.model_dump())
    if not ok:
        raise HTTPException(400, "So'rov ma'lumotlari noto'g'ri")

    data_d = ai_hisob_tekshir(data.model_dump())

    async with rls_conn(uid) as c:
        async with c.transaction():
            # 1. Klient topish/yaratish
            klient_id = None
            klient_ismi = (data_d.get("klient") or "").strip()
            if klient_ismi:
                kl = await c.fetchrow("""
                    INSERT INTO klientlar (user_id, ism) VALUES ($1, $2)
                    ON CONFLICT (user_id, lower(ism))
                    DO UPDATE SET ism = klientlar.ism RETURNING id
                """, uid, klient_ismi)
                klient_id = kl["id"]

            jami = float(data_d.get("jami_summa", 0))
            tolangan = float(data_d.get("tolangan", 0))
            qarz_summa = float(data_d.get("qarz", 0))

            # 2. Sessiya yaratish
            sess_id = await c.fetchval("""
                INSERT INTO sotuv_sessiyalar
                    (user_id, klient_id, klient_ismi, jami, tolangan, qarz, izoh)
                VALUES ($1,$2,$3,$4,$5,$6,$7) RETURNING id
            """, uid, klient_id, klient_ismi or None,
                jami, tolangan, qarz_summa,
                data_d.get("izoh"),
            )

            # 3. Har bir tovar — chiqim + qoldiq
            tovarlar = data_d.get("tovarlar", [])
            for t in tovarlar:
                nomi = t.get("nomi", "").strip()
                miqdor = float(t.get("miqdor", 0))
                narx = float(t.get("narx", 0))
                t_jami = float(t.get("jami", 0)) or (miqdor * narx)
                birlik = t.get("birlik", "dona")

                # Tovar topish — avval exact match, keyin LIKE
                tovar = await c.fetchrow("""
                    SELECT id, nomi, olish_narxi, sotish_narxi FROM tovarlar
                    WHERE user_id=$1 AND lower(nomi) = lower($2)
                """, uid, nomi.strip())
                if not tovar:
                    tovar = await c.fetchrow("""
                        SELECT id, nomi, olish_narxi, sotish_narxi FROM tovarlar
                        WHERE user_id=$1 AND lower(nomi) LIKE lower($2)
                        ORDER BY length(nomi) ASC LIMIT 1
                    """, uid, f"%{like_escape(nomi)}%")

                tovar_id = tovar["id"] if tovar else None
                olish = float(tovar["olish_narxi"]) if tovar else 0

                # Chiqim yozuvi — foyda virtual hisoblanadi (schema'da bu ustun yo'q)
                await c.execute("""
                    INSERT INTO chiqimlar
                        (user_id, sessiya_id, klient_id, tovar_id, tovar_nomi, klient_ismi,
                         miqdor, birlik, sotish_narxi, jami, olish_narxi)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)
                """, uid, sess_id, klient_id, tovar_id, nomi, klient_ismi or None,
                    miqdor, birlik, narx, t_jami, olish,
                )

                # Qoldiq kamaytirish
                if tovar_id:
                    await c.execute("""
                        UPDATE tovarlar SET qoldiq = GREATEST(qoldiq - $2, 0)
                        WHERE id = $1 AND user_id = $3
                    """, tovar_id, miqdor, uid)

            # 4. Qarz saqlash
            if qarz_summa > 0 and klient_id:
                await c.execute("""
                    INSERT INTO qarzlar
                        (user_id, klient_id, klient_ismi, sessiya_id,
                         dastlabki_summa, qolgan, tolangan)
                    VALUES ($1,$2,$3,$4,$5,$5,0)
                """, uid, klient_id, klient_ismi, sess_id, qarz_summa)

            # 5. Klient jami_sotib + CRM statistika yangilash
            if klient_id and jami > 0:
                await c.execute("""
                    UPDATE klientlar
                    SET jami_sotib     = jami_sotib + $2,
                        jami_xaridlar  = COALESCE(jami_xaridlar, 0) + $2,
                        xarid_soni     = COALESCE(xarid_soni, 0) + 1,
                        oxirgi_sotuv   = NOW()
                    WHERE id = $1
                """, klient_id, jami)

    await user_cache_tozala(uid)

    # ── Ogohlantirish: zarar sotuv va kam qoldiq ──
    ogohlar = []
    async with rls_conn(uid) as c:
        # Zarar sotuv tekshiruvi
        for t in tovarlar:
            nomi = t.get("nomi", "").strip()
            narx_s = float(t.get("narx", 0))
            if not nomi or narx_s <= 0:
                continue
            tv = await c.fetchrow(
                "SELECT olish_narxi FROM tovarlar WHERE user_id=$1 AND lower(nomi)=lower($2)",
                uid, nomi
            )
            if tv and float(tv["olish_narxi"]) > 0 and narx_s < float(tv["olish_narxi"]):
                ogohlar.append(f"⚠️ {nomi}: sotish {narx_s:,.0f} < olish {float(tv['olish_narxi']):,.0f}")
        # Kam qoldiq tekshiruvi
        kam = await c.fetch("""
            SELECT nomi, qoldiq, min_qoldiq FROM tovarlar
            WHERE user_id=$1 AND min_qoldiq > 0 AND qoldiq <= min_qoldiq
            AND qoldiq > 0
            LIMIT 5
        """, uid)
        for r in kam:
            ogohlar.append(f"📦 {r['nomi']}: qoldiq {float(r['qoldiq'])}, min {float(r['min_qoldiq'])}")

    log.info("📤 Web sotuv: sessiya=%d tovarlar=%d jami=%.0f uid=%d",
             sess_id, len(tovarlar), jami, uid)
    result = {"sessiya_id": sess_id, "status": "saqlandi"}

    # ── Loyalty ball qo'shish ──
    if klient_id and jami > 0:
        try:
            from shared.services.loyalty import klient_ball_qoshish
            async with rls_conn(uid) as c_loyalty:
                loyalty = await klient_ball_qoshish(c_loyalty, uid, klient_id, jami, sess_id)
            if loyalty.get("ball", 0) > 0:
                ogohlar.append(
                    f"⭐ +{loyalty['ball']} bonus ball → "
                    f"jami {loyalty.get('jami_ball', 0)} "
                    f"({loyalty.get('daraja', {}).get('nomi', '')})"
                )
        except Exception:
            pass

    if ogohlar:
        result["ogohlar"] = ogohlar
    return result


@app.post("/api/v1/kirim", tags=["Sotuv"])
async def kirim_saqlash(data: KirimSo_rov, uid: int = Depends(get_uid)):
    """Tovar kirimini saqlash va avtomatik ombor qoldig'ini oshirish.

    1. kirimlar jadvaliga yozuv qo'shiladi (tarix uchun)
    2. tovarlar jadvalida qoldiq avtomatik oshiriladi
    3. Agar tovar mavjud bo'lmasa, yangi yaratiladi
    4. Olish narxi yangilanadi (moving average emas — oxirgi narx)
    """
    from shared.utils.hisob import kirim_validatsiya
    from shared.cache.redis_cache import user_cache_tozala

    ok, xato = kirim_validatsiya({
        "tovar_nomi": data.tovar_nomi,
        "miqdor":     data.miqdor,
        "narx":       data.narx,
    })
    if not ok:
        raise HTTPException(400, "So'rov ma'lumotlari noto'g'ri")

    async with rls_conn(uid) as c:
        async with c.transaction():
            kirim_id = await c.fetchval("""
                INSERT INTO kirimlar
                    (user_id, tovar_id, tovar_nomi, kategoriya, miqdor,
                     birlik, narx, jami, manba, izoh)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10) RETURNING id
            """, uid,
                getattr(data, "tovar_id", None),
                data.tovar_nomi,
                data.kategoriya,
                data.miqdor,
                data.birlik,
                data.narx,
                data.jami,
                data.manba,
                getattr(data, "izoh", None),
            )

            # Tovar qoldig'ini avtomatik oshirish
            tovar_id = getattr(data, "tovar_id", None)
            if tovar_id:
                await c.execute("""
                    UPDATE tovarlar
                    SET qoldiq = COALESCE(qoldiq, 0) + $1,
                        olish_narxi = CASE WHEN $2 > 0 THEN $2 ELSE olish_narxi END,
                        yangilangan = NOW()
                    WHERE id = $3 AND user_id = $4
                """, data.miqdor, data.narx, tovar_id, uid)
            else:
                # tovar_id berilmagan bo'lsa, nom bo'yicha yangilash (upsert)
                await c.execute("""
                    INSERT INTO tovarlar
                        (user_id, nomi, kategoriya, birlik, olish_narxi, qoldiq)
                    VALUES ($1, $2, $3, $4, $5, $6)
                    ON CONFLICT (user_id, lower(nomi)) DO UPDATE SET
                        qoldiq      = COALESCE(tovarlar.qoldiq, 0) + EXCLUDED.qoldiq,
                        olish_narxi = CASE WHEN EXCLUDED.olish_narxi > 0
                                           THEN EXCLUDED.olish_narxi
                                           ELSE tovarlar.olish_narxi END,
                        yangilangan = NOW()
                """, uid, data.tovar_nomi, data.kategoriya,
                    data.birlik, data.narx, data.miqdor)

    await user_cache_tozala(uid)
    return {"kirim_id": kirim_id, "status": "saqlandi"}


@app.post("/api/v1/kirim/import/excel", tags=["Sotuv"])
async def kirim_import_excel(
    file_base64: str,
    uid: int = Depends(get_uid),
):
    """Omborga dastlabki qoldiqlarni Excel fayldan bulk import qilish.

    Ustunlar (1-qator sarlavha):
    - nomi (majburiy)
    - kategoriya, birlik
    - miqdor (majburiy)
    - narx
    - manba, izoh
    """
    import io as _io, base64 as _b64
    from openpyxl import load_workbook
    try:
        content = _b64.b64decode(file_base64)
        wb = load_workbook(_io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Fayl o'qilmadi: {e}")

    ws = wb.active
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    norm = [(str(h or "").strip().lower()) for h in headers_row]

    def find_col(names: list[str]) -> int:
        for nm in names:
            for i, h in enumerate(norm):
                if nm in h:
                    return i
        return -1

    idx_nomi       = find_col(["nom", "tovar", "name"])
    idx_kateg      = find_col(["kateg", "category"])
    idx_birlik     = find_col(["birlik", "unit"])
    idx_miqdor     = find_col(["miqdor", "qty", "count", "kol"])
    idx_narx       = find_col(["narx", "price"])
    idx_manba      = find_col(["manba", "supplier", "post"])
    idx_izoh       = find_col(["izoh", "comment", "note"])

    if idx_nomi < 0 or idx_miqdor < 0:
        raise HTTPException(400,
            "Majburiy ustunlar topilmadi: 'nomi' va 'miqdor'")

    saved = 0
    errors: list[str] = []

    async with rls_conn(uid) as c:
        for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            def cell(i): return row[i] if 0 <= i < len(row) else None
            nomi = str(cell(idx_nomi) or "").strip()
            if not nomi:
                continue
            try:
                miqdor = float(cell(idx_miqdor) or 0)
                if miqdor <= 0:
                    continue
                kateg  = str(cell(idx_kateg)  or "Boshqa")
                birlik = str(cell(idx_birlik) or "dona")
                narx   = float(cell(idx_narx) or 0)
                jami   = miqdor * narx
                manba  = str(cell(idx_manba) or "") if idx_manba >= 0 else ""
                izoh   = str(cell(idx_izoh)  or "") if idx_izoh  >= 0 else ""

                async with c.transaction():
                    # Upsert tovar + oshirish
                    tv = await c.fetchrow("""
                        INSERT INTO tovarlar
                            (user_id, nomi, kategoriya, birlik, olish_narxi, qoldiq)
                        VALUES ($1, $2, $3, $4, $5, $6)
                        ON CONFLICT (user_id, lower(nomi)) DO UPDATE SET
                            qoldiq      = COALESCE(tovarlar.qoldiq, 0) + EXCLUDED.qoldiq,
                            olish_narxi = CASE WHEN EXCLUDED.olish_narxi > 0
                                               THEN EXCLUDED.olish_narxi
                                               ELSE tovarlar.olish_narxi END,
                            yangilangan = NOW()
                        RETURNING id
                    """, uid, nomi, kateg, birlik, narx, miqdor)

                    await c.execute("""
                        INSERT INTO kirimlar
                            (user_id, tovar_id, tovar_nomi, kategoriya,
                             miqdor, birlik, narx, jami, manba, izoh)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                    """, uid, tv["id"], nomi, kateg, miqdor, birlik,
                        narx, jami, manba or None, izoh or None)

                saved += 1
            except Exception as e:
                errors.append(f"Qator #{ridx}: {nomi[:30]}: {str(e)[:80]}")
                if len(errors) >= 20:
                    break

    from shared.cache.redis_cache import user_cache_tozala
    await user_cache_tozala(uid)

    return {
        "saved": saved,
        "errors": errors[:20],
    }


@app.get("/api/v1/kirimlar", tags=["Sotuv"])
async def kirimlar_royxati(
    limit: int = 50, offset: int = 0,
    sana_dan: Optional[str] = None,
    sana_gacha: Optional[str] = None,
    qidiruv: Optional[str] = None,
    kategoriya: Optional[str] = None,
    uid: int = Depends(get_uid),
):
    """Kirimlar ro'yxati — sana/kategoriya/qidiruv filtrlari bilan."""
    limit = min(limit, 500)
    where = ["user_id = $1"]
    params: list = [uid]
    if sana_dan:
        params.append(sana_dan)
        where.append(f"sana >= ${len(params)}::timestamptz")
    if sana_gacha:
        params.append(sana_gacha)
        where.append(f"sana < ${len(params)}::timestamptz + interval '1 day'")
    if kategoriya:
        params.append(kategoriya)
        where.append(f"kategoriya = ${len(params)}")
    if qidiruv:
        params.append(f"%{qidiruv}%")
        where.append(
            f"(tovar_nomi ILIKE ${len(params)} OR manba ILIKE ${len(params)})"
        )
    params.append(limit); params.append(offset)
    where_sql = " AND ".join(where)

    async with rls_conn(uid) as c:
        rows = await c.fetch(f"""
            SELECT id, tovar_id, tovar_nomi, kategoriya,
                   miqdor, birlik, narx, jami, manba, izoh, sana
            FROM kirimlar
            WHERE {where_sql}
            ORDER BY sana DESC
            LIMIT ${len(params)-1} OFFSET ${len(params)}
        """, *params)

        stats = await c.fetchrow(f"""
            SELECT COUNT(*)                  AS soni,
                   COALESCE(SUM(jami), 0)    AS jami_summa,
                   COALESCE(SUM(miqdor), 0)  AS jami_miqdor,
                   COUNT(DISTINCT tovar_id) FILTER (WHERE tovar_id IS NOT NULL)
                       AS turli_tovar
            FROM kirimlar
            WHERE {where_sql}
        """, *params[:-2])

    return {
        "items": [dict(r) for r in rows],
        "stats": dict(stats) if stats else {},
    }


@app.delete("/api/v1/kirim/{kirim_id}", tags=["Sotuv"])
async def kirim_ochir(kirim_id: int, uid: int = Depends(get_uid)):
    """Kirimni o'chirish va avtomatik qoldiqni kamaytirish."""
    async with rls_conn(uid) as c:
        async with c.transaction():
            kirim = await c.fetchrow(
                "SELECT tovar_id, miqdor FROM kirimlar WHERE id=$1 AND user_id=$2",
                kirim_id, uid)
            if not kirim:
                raise HTTPException(404, "Kirim topilmadi")
            if kirim["tovar_id"]:
                await c.execute(
                    "UPDATE tovarlar SET qoldiq = GREATEST(0, qoldiq - $1) "
                    "WHERE id = $2 AND user_id = $3",
                    kirim["miqdor"], kirim["tovar_id"], uid)
            await c.execute(
                "DELETE FROM kirimlar WHERE id=$1 AND user_id=$2",
                kirim_id, uid)
    return {"id": kirim_id, "status": "ochirildi"}


# ════════════════════════════════════════════════════════════
#  QARZ TO'LASH
# ════════════════════════════════════════════════════════════

@app.post("/api/v1/qarz/tolash", tags=["Qarz"])
async def qarz_tolash_endpoint(
    data: QarzTolashSo_rov,
    uid: int = Depends(get_uid)
):
    """Klientning qarzini to'lash (FIFO tartibida)"""
    from shared.utils.hisob import qarz_to_lash_hisob, D
    from shared.cache.redis_cache import user_cache_tozala
    from decimal import Decimal

    async with rls_conn(uid) as c:
        qarzlar = await c.fetch("""
            SELECT id, qolgan FROM qarzlar
            WHERE lower(klient_ismi) LIKE lower($1)
              AND yopildi=FALSE AND qolgan>0
            ORDER BY yaratilgan ASC
            FOR UPDATE
        """, f"%{like_escape(data.klient_ismi.strip())}%")

        if not qarzlar:
            raise HTTPException(404, f"'{data.klient_ismi}' uchun qarz topilmadi")

        summa     = Decimal(str(data.summa))
        qoldi     = summa
        tolandi_j = Decimal("0")

        for q in qarzlar:
            if qoldi <= 0: break
            r = qarz_to_lash_hisob(q["qolgan"], str(qoldi))
            tl = D(r["tolandi"]); yq = D(r["qolgan"])

            await c.execute("""
                UPDATE qarzlar
                SET tolangan=tolangan+$1, qolgan=$2,
                    yopildi=$3, yangilangan=NOW()
                WHERE id=$4
            """, tl, yq, yq == 0, q["id"])

            tolandi_j += tl
            qoldi     -= tl

        qolgan_jami = D(await c.fetchval("""
            SELECT COALESCE(SUM(qolgan),0) FROM qarzlar
            WHERE lower(klient_ismi) LIKE lower($1) AND yopildi=FALSE
        """, f"%{like_escape(data.klient_ismi.strip())}%") or 0)

    await user_cache_tozala(uid)
    return {
        "klient":      data.klient_ismi,
        "tolandi":     str(tolandi_j),
        "qolgan_qarz": str(qolgan_jami),
        "status":      "tolandi" if qolgan_jami == 0 else "qisman_tolandi",
    }


# ════════════════════════════════════════════════════════════
#  QIDIRUV
# ════════════════════════════════════════════════════════════

@app.get("/api/v1/search", tags=["Sotuv"])
async def search(
    q: str,
    tur: str = "barchasi",   # tovar | klient | barchasi
    limit: int = 10,
    uid: int = Depends(get_uid)
):
    """Global qidiruv — tovar va klientlar (relevance ranked)"""
    if len(q.strip()) < 2:
        raise HTTPException(400, "Kamida 2 belgi kiriting")
    limit = min(limit, 50)
    async with rls_conn(uid) as c:
        tovarlar_r, klientlar_r = [], []
        if tur in ("tovar","barchasi"):
            # Relevance ranked: exact → starts with → contains
            tovarlar_r = [dict(r) for r in await c.fetch("""
                SELECT id, nomi, kategoriya, qoldiq, sotish_narxi,
                    CASE
                        WHEN lower(nomi) = lower($1) THEN 3
                        WHEN lower(nomi) LIKE lower($3) THEN 2
                        ELSE 1
                    END AS relevance
                FROM tovarlar
                WHERE lower(nomi) LIKE lower($2)
                   OR lower(kategoriya) LIKE lower($2)
                ORDER BY relevance DESC, nomi LIMIT $4
            """, q.strip(), f"%{like_escape(q)}%",
                f"{like_escape(q)}%", limit)]
            # relevance ni olib tashlaymiz
            for t in tovarlar_r:
                t.pop("relevance", None)
        if tur in ("klient","barchasi"):
            klientlar_r = [dict(r) for r in await c.fetch("""
                SELECT id, ism, telefon, jami_sotib,
                    CASE
                        WHEN lower(ism) = lower($1) THEN 3
                        WHEN lower(ism) LIKE lower($3) THEN 2
                        ELSE 1
                    END AS relevance
                FROM klientlar
                WHERE lower(ism) LIKE lower($2)
                   OR (telefon IS NOT NULL AND telefon LIKE $2)
                ORDER BY relevance DESC, jami_sotib DESC LIMIT $4
            """, q.strip(), f"%{like_escape(q)}%",
                f"{like_escape(q)}%", limit)]
            for k in klientlar_r:
                k.pop("relevance", None)
    return {"tovarlar": tovarlar_r, "klientlar": klientlar_r,
            "jami": len(tovarlar_r) + len(klientlar_r)}
# ════════════════════════════════════════════════════════════
#  HISOBOTLAR (haftalik / oylik)
# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════

@app.post("/api/v1/export", tags=["Export"])
async def export_trigger(data: dict, request: Request, uid: int = Depends(get_uid)):
    """
    Excel/PDF eksportni Worker ga yuboradi.
    Katta hisobotlar background da tayyorlanadi.
    Rate limit: 3 so'rov/daqiqa.
    """
    from services.api.deps import endpoint_rate_check
    await endpoint_rate_check(request, "export")
    tur      = data.get("tur", "kunlik")       # kunlik | haftalik | oylik
    format_  = data.get("format", "excel")     # excel | pdf
    sana_dan   = data.get("sana_dan", "")
    sana_gacha = data.get("sana_gacha", "")

    if tur not in ("kunlik", "haftalik", "oylik"):
        raise HTTPException(400, "tur: kunlik | haftalik | oylik bo'lishi kerak")
    if format_ not in ("excel", "pdf"):
        raise HTTPException(400, "format: excel | pdf bo'lishi kerak")

    redis_url = os.getenv("REDIS_URL","")
    if not redis_url:
        raise HTTPException(503, "Worker ulanganda emas")

    try:
        from celery import Celery as _Celery
        _app = _Celery(broker=redis_url)
        task = _app.send_task(
            "tasks.katta_export",
            args=[uid, tur, sana_dan, sana_gacha, format_],
        )
        return {
            "task_id": task.id,
            "status":  "navbatda",
            "format":  format_,
            "tur":     tur,
        }
    except Exception as e:
        log.error("Export celery xato: %s", e)
        raise HTTPException(503, "Export navbati vaqtincha mavjud emas")

# ════════════════════════════════════════════════════════════
#  EXPORT NATIJA OLISH
# ════════════════════════════════════════════════════════════

@app.get("/api/v1/export/{task_id}", tags=["Export"])
async def export_natija(task_id: str, uid: int = Depends(get_uid)):
    """
    Export task natijasini tekshirish.
    task_id: /api/v1/export POST dan qaytgan task_id (UUID format)

    Holat: pending | processing | ready | error
    ready bo'lsa: file_url yoki file_bytes qaytaradi (bot uchun: file_url=None, download is via /export/file/{task_id})
    """
    # task_id UUID formatini tekshirish (injection himoyasi)
    import uuid as _uuid_mod
    try:
        _uuid_mod.UUID(task_id)
    except ValueError:
        raise HTTPException(400, "task_id noto'g'ri format")
    redis_url = os.getenv("REDIS_URL", "")
    if not redis_url:
        raise HTTPException(503, "Worker ulanganda emas")

    try:
        from celery import Celery as _Celery
        from celery.result import AsyncResult
        _app = _Celery(broker=redis_url, backend=redis_url)
        result = AsyncResult(task_id, app=_app)

        if result.state == "PENDING":
            return {"task_id": task_id, "holat": "kutilmoqda"}
        elif result.state == "STARTED" or result.state == "RETRY":
            return {"task_id": task_id, "holat": "bajarilmoqda"}
        elif result.state == "SUCCESS":
            res = result.result or {}
            status = res.get("status", "")
            if status == "tayyor" and res.get("content_b64"):
                return {
                    "task_id":  task_id,
                    "holat":    "tayyor",
                    "format":   res.get("format", "excel"),
                    "hajm_kb":  res.get("hajm_kb", 0),
                    "download": f"/api/v1/export/file/{task_id}",
                }
            elif status == "katta_fayl":
                return {
                    "task_id": task_id,
                    "holat":   "katta_fayl",
                    "xato":    res.get("xato", "Fayl hajmi katta"),
                }
            else:
                return {
                    "task_id": task_id,
                    "holat":   "xato",
                    "xato":    res.get("xato", "Export bajarilmadi"),
                }
        elif result.state == "FAILURE":
            # Celery task xato bilan tugadi
            err_info = str(result.info) if result.info else "Export bajarilmadi"
            return {
                "task_id": task_id,
                "holat":   "xato",
                "xato":    "Export bajarilmadi. Keyinroq urinib ko'ring.",
            }
        else:
            # REVOKED yoki noma'lum holat
            return {
                "task_id": task_id,
                "holat":   "xato",
                "xato":    f"Task holati: {result.state}",
            }
    except Exception as e:
        log.error("Export natija xato: %s", e)
        raise HTTPException(503, "Export holati tekshirib bo'lmadi")


@app.get("/api/v1/export/file/{task_id}", tags=["Export"])
async def export_file_yuklab(task_id: str, uid: int = Depends(get_uid)):
    """
    Tayyor export faylni yuklash.
    Auth: Bearer header YOKI ?token= query param (browser download uchun).
    Faqat faylni yaratgan foydalanuvchi uchun (uid tekshiruvi fayl nomida).
    """
    # task_id UUID formatini tekshirish (injection himoyasi)
    import uuid as _uuid_mod
    try:
        _uuid_mod.UUID(task_id)
    except ValueError:
        raise HTTPException(400, "task_id noto'g'ri format")
    redis_url = os.getenv("REDIS_URL", "")
    if not redis_url:
        raise HTTPException(503, "Worker ulanganda emas")

    try:
        from celery import Celery as _Celery
        from celery.result import AsyncResult
        import base64
        from fastapi.responses import Response
        _app = _Celery(broker=redis_url, backend=redis_url)
        result = AsyncResult(task_id, app=_app)
        if result.state != "SUCCESS":
            raise HTTPException(404, "Fayl tayyor emas yoki hali bajarilmoqda")
        res = result.result or {}
        if res.get("status") != "tayyor":
            raise HTTPException(404, "Export muvaffaqiyatsiz tugadi")

        content_b64 = res.get("content_b64", "")
        if not content_b64:
            raise HTTPException(404, "Fayl mazmuni topilmadi (muddati o'tgan bo'lishi mumkin)")

        # Security: task result user_id tekshirish — boshqa user faylini yuklab olishni oldini olish
        task_uid = res.get("user_id")
        if task_uid is not None and int(task_uid) != uid:
            log.warning("EXPORT SECURITY: uid=%d tried to download task for uid=%s", uid, task_uid)
            raise HTTPException(403, "Bu fayl sizga tegishli emas")

        format_ = res.get("format", "excel")
        ext     = "xlsx" if format_ == "excel" else "pdf"
        media   = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if ext == "xlsx" else "application/pdf"
        )
        filename = f"export_{uid}.{ext}"
        content  = base64.b64decode(content_b64)
        return Response(
            content=content,
            media_type=media,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        log.error("Export file yuklab xato: %s", e)
        raise HTTPException(500, "Fayl yuklanmadi")


# ════════════════════════════════════════════════════════════
#  GLOBAL XATO HANDLER
# ════════════════════════════════════════════════════════════

@app.exception_handler(Exception)
async def xato_handler(request: Request, exc: Exception):
    log.error("API xato: %s %s — %s", request.method, request.url.path, exc)
    return JSONResponse(status_code=500, content={"detail": "Ichki server xatosi"})


# ════════════════════════════════════════════════════════════
#  v25.3.2 — REQUEST TIMING MIDDLEWARE
# ════════════════════════════════════════════════════════════

@app.middleware("http")
async def timing_middleware(request: Request, call_next):
    """Har so'rovda X-Response-Time va X-Version header qo'shadi"""
    import time as _t
    start = _t.monotonic()
    response = await call_next(request)
    ms = round((_t.monotonic() - start) * 1000, 1)
    response.headers["X-Response-Time"] = f"{ms}ms"
    response.headers["X-Version"] = __version__
    response.headers["X-Powered-By"] = "SavdoAI"
    if ms > 1000:
        log.warning("SLOW: %s %s — %sms", request.method, request.url.path, ms)
    return response


# ═══ RATE LIMITER — IN-PROCESS (--workers 1 uchun) ═══
# TODO: Multi-worker da Redis-based rate limiting ga o'tish kerak.
_rate_buckets: dict = {}  # {ip: [timestamps]}
_rate_last_gc: float = 0.0  # oxirgi tozalash vaqti
RATE_LIMIT = int(os.environ.get("API_RATE_LIMIT", "60"))  # per minute
RATE_WINDOW = 60  # seconds
_RATE_GC_INTERVAL = 300  # har 5 daqiqada eski IPlarni tozalash
_RATE_MAX_IPS = 10_000  # max IP soni (DDoS himoya)

@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    """IP-based rate limiting — SAP-GRADE API himoya"""
    import time as _t
    global _rate_last_gc
    # CORS preflight must never be rate-limited (browser needs 200 + ACAO on OPTIONS).
    if request.method == "OPTIONS":
        return await call_next(request)
    # Health/readyz endpoints skip
    if request.url.path in ("/health", "/healthz", "/readyz", "/live"):
        return await call_next(request)
    # Android printer helper — bir nechta tezkor so'rov (yuklash + ack)
    if request.url.path.startswith("/api/print"):
        return await call_next(request)

    ip = request.client.host if request.client else "unknown"
    now = _t.time()

    # Periodic GC — eski IPlarni tozalash (memory leak oldini oladi)
    if now - _rate_last_gc > _RATE_GC_INTERVAL:
        dead_ips = [k for k, v in _rate_buckets.items()
                    if not v or (now - max(v)) > RATE_WINDOW * 2]
        for k in dead_ips:
            del _rate_buckets[k]
        _rate_last_gc = now

    # Max IP cap — DDoS himoya
    if len(_rate_buckets) > _RATE_MAX_IPS and ip not in _rate_buckets:
        return await call_next(request)  # yangi IP qo'shmaymiz

    # Clean old entries
    if ip in _rate_buckets:
        _rate_buckets[ip] = [t for t in _rate_buckets[ip] if now - t < RATE_WINDOW]
    else:
        _rate_buckets[ip] = []

    if len(_rate_buckets[ip]) >= RATE_LIMIT:
        log.warning("RATE LIMIT: %s — %d req/min", ip, len(_rate_buckets[ip]))
        from starlette.responses import JSONResponse
        return JSONResponse(
            {"error": "Rate limit exceeded", "limit": RATE_LIMIT, "window": "60s"},
            status_code=429,
            headers={"Retry-After": "60", "X-RateLimit-Limit": str(RATE_LIMIT)}
        )

    _rate_buckets[ip].append(now)
    response = await call_next(request)
    response.headers["X-RateLimit-Limit"] = str(RATE_LIMIT)
    response.headers["X-RateLimit-Remaining"] = str(RATE_LIMIT - len(_rate_buckets[ip]))
    return response


# ════════════════════════════════════════════════════════════
#  v25.3.2 — KASSA BOT PROXY (bot → API kassa)
# ════════════════════════════════════════════════════════════



# ════════════════════════════════════════════════════════════════
#  § LEDGER — SAP-GRADE Buxgalteriya
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/ledger/balans", tags=["Ledger"])
async def ledger_balans(uid: int = Depends(get_uid)):
    """Balans tekshiruvi — debit = credit"""
    from shared.services.ledger import balans_tekshir, hisob_balans, HisobTuri
    async with rls_conn(uid) as c:
        b = await balans_tekshir(c, uid)
        hisoblar = {}
        for h in HisobTuri:
            val = await hisob_balans(c, uid, h)
            hisoblar[h.value] = float(val)
    return {**b, "hisoblar": hisoblar}


@app.get("/api/v1/ledger/jurnal", tags=["Ledger"])
async def ledger_jurnal(uid: int = Depends(get_uid), limit: int = 50):
    """Jurnal yozuvlari"""
    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT jurnal_id, tur, sana, tavsif, jami_debit, jami_credit
            FROM jurnal_yozuvlar
            WHERE user_id = $1
            ORDER BY sana DESC LIMIT $2
        """, uid, min(limit, 200))
    return [dict(r) for r in rows]


@app.get("/api/v1/ledger/jurnal/{jurnal_id}", tags=["Ledger"])
async def ledger_jurnal_detail(jurnal_id: str, uid: int = Depends(get_uid)):
    """Bitta jurnal yozuvi batafsil (qatorlar bilan)"""
    async with rls_conn(uid) as c:
        header = await c.fetchrow("""
            SELECT id, jurnal_id, user_id, tur, sana, tavsif, jami_debit, jami_credit, manba_id, manba_jadval, idempotency_key, yaratilgan FROM jurnal_yozuvlar
            WHERE user_id = $1 AND jurnal_id = $2
        """, uid, jurnal_id)
        if not header:
            raise HTTPException(404, "Jurnal topilmadi")
        qatorlar = await c.fetch("""
            SELECT hisob, debit, credit, tavsif
            FROM jurnal_qatorlar WHERE jurnal_id = $1
        """, header["id"])
    return {"header": dict(header), "qatorlar": [dict(q) for q in qatorlar]}


@app.get("/api/v1/ledger/hisob/{hisob}", tags=["Ledger"])
async def ledger_hisob(hisob: str, uid: int = Depends(get_uid)):
    """Bitta hisob balansi va tarixi"""
    from shared.services.ledger import hisob_balans, HisobTuri
    try:
        h = HisobTuri(hisob)
    except ValueError:
        raise HTTPException(400, f"Noma'lum hisob: {hisob}")
    async with rls_conn(uid) as c:
        balans = await hisob_balans(c, uid, h)
        tarix = await c.fetch("""
            SELECT jy.jurnal_id, jy.sana, jy.tavsif, jq.debit, jq.credit
            FROM jurnal_qatorlar jq
            JOIN jurnal_yozuvlar jy ON jq.jurnal_id = jy.id
            WHERE jy.user_id = $1 AND jq.hisob = $2
            ORDER BY jy.sana DESC LIMIT 50
        """, uid, hisob)
    return {"hisob": hisob, "balans": float(balans), "tarix": [dict(r) for r in tarix]}


# ═══════════════════════════════════════════════════════════
#  SHOGIRD XARAJAT API (Web Dashboard uchun)
# ═══════════════════════════════════════════════════════════

@app.get("/api/v1/shogirdlar", tags=["Xarajatlar"])
async def api_shogirdlar(uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import shogirdlar_royxati
    async with rls_conn(uid) as c:
        return [dict(s) for s in await shogirdlar_royxati(c, uid)]

@app.get("/api/v1/shogird/dashboard", tags=["Xarajatlar"])
async def api_shogird_dashboard(uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import dashboard_data
    async with rls_conn(uid) as c:
        return await dashboard_data(c, uid)

@app.get("/api/v1/xarajatlar/bugungi", tags=["Xarajatlar"])
async def api_xarajatlar_bugungi(uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import kunlik_hisobot
    async with rls_conn(uid) as c:
        return await kunlik_hisobot(c, uid)

@app.get("/api/v1/xarajatlar/oylik", tags=["Xarajatlar"])
async def api_xarajatlar_oylik(uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import oylik_hisobot
    async with rls_conn(uid) as c:
        return await oylik_hisobot(c, uid)

@app.get("/api/v1/xarajatlar/kutilmoqda", tags=["Xarajatlar"])
async def api_kutilmoqda(uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import kutilmoqda_royxati
    async with rls_conn(uid) as c:
        return [dict(k) for k in await kutilmoqda_royxati(c, uid)]

@app.get("/api/v1/shogird/{shogird_id}/hisobot", tags=["Xarajatlar"])
async def api_shogird_hisobot(shogird_id: int, kunlar: int = 7, uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import shogird_hisobot
    async with rls_conn(uid) as c:
        return await shogird_hisobot(c, uid, shogird_id, kunlar)

@app.post("/api/v1/xarajat/{xarajat_id}/tasdiqlash", tags=["Xarajatlar"])
async def api_xarajat_tasdiq(xarajat_id: int, uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import xarajat_tasdiqlash
    async with rls_conn(uid) as c:
        ok = await xarajat_tasdiqlash(c, xarajat_id, uid)
    return {"ok": ok}

@app.post("/api/v1/xarajat/{xarajat_id}/bekor", tags=["Xarajatlar"])
async def api_xarajat_bekor(xarajat_id: int, uid: int = Depends(get_uid)):
    from shared.services.shogird_xarajat import xarajat_bekor
    async with rls_conn(uid) as c:
        ok = await xarajat_bekor(c, xarajat_id, uid)
    return {"ok": ok}


# ═══════════════════════════════════════════════════════════
#  NARX GURUH API (Web Dashboard uchun)
# ═══════════════════════════════════════════════════════════

@app.get("/api/v1/narx/guruhlar", tags=["Narx"])
async def api_narx_guruhlar(uid: int = Depends(get_uid)):
    from shared.services.smart_narx import guruhlar_royxati
    async with rls_conn(uid) as c:
        return [dict(g) for g in await guruhlar_royxati(c, uid)]

@app.post("/api/v1/narx/guruh", tags=["Narx"])
async def api_narx_guruh_yarat(data: dict, uid: int = Depends(get_uid)):
    from shared.services.smart_narx import guruh_yaratish
    nomi = data.get("nomi", "")
    if not nomi: raise HTTPException(400, "Guruh nomi kerak")
    async with rls_conn(uid) as c:
        gid = await guruh_yaratish(c, uid, nomi, data.get("izoh", ""))
    return {"id": gid, "nomi": nomi}

@app.post("/api/v1/narx/qoyish", tags=["Narx"])
async def api_narx_qoy(data: dict, uid: int = Depends(get_uid)):
    from shared.services.smart_narx import guruh_narx_qoyish, shaxsiy_narx_qoyish
    from decimal import Decimal
    narx = Decimal(str(data.get("narx", 0)))
    if narx <= 0: raise HTTPException(400, "Narx musbat bo'lishi kerak")
    async with rls_conn(uid) as c:
        if data.get("guruh_id"):
            await guruh_narx_qoyish(c, uid, int(data["guruh_id"]), int(data["tovar_id"]), narx)
        elif data.get("klient_id"):
            await shaxsiy_narx_qoyish(c, uid, int(data["klient_id"]), int(data["tovar_id"]), narx)
    return {"ok": True}

@app.post("/api/v1/narx/klient_guruh", tags=["Narx"])
async def api_klient_guruhga(data: dict, uid: int = Depends(get_uid)):
    from shared.services.smart_narx import klient_guruhga_qoyish
    async with rls_conn(uid) as c:
        await klient_guruhga_qoyish(c, uid, int(data["klient_id"]), int(data["guruh_id"]))
    return {"ok": True}


# ════════════════════════════════════════════════════════════════
#  TOVAR CRUD — Web panel uchun to'liq CRUD
# ════════════════════════════════════════════════════════════════


class TovarYaratSorov(BaseModel):
    nomi:             str   = Field(..., min_length=1, max_length=200)
    kategoriya:       str   = Field("Boshqa")
    birlik:           str   = Field("dona")
    olish_narxi:      float = Field(0, ge=0)
    sotish_narxi:     float = Field(0, ge=0)
    min_sotish_narxi: float = Field(0, ge=0)
    qoldiq:           float = Field(0, ge=0)
    min_qoldiq:       float = Field(0, ge=0)


class TovarYangilaSorov(BaseModel):
    nomi:             Optional[str]   = None
    kategoriya:       Optional[str]   = None
    birlik:           Optional[str]   = None
    olish_narxi:      Optional[float] = None
    sotish_narxi:     Optional[float] = None
    min_sotish_narxi: Optional[float] = None
    qoldiq:           Optional[float] = None
    min_qoldiq:       Optional[float] = None


class QoldiqYangilaSorov(BaseModel):
    qoldiq: float = Field(..., ge=0)
    ism:          Optional[str]   = None
    telefon:      Optional[str]   = None
    manzil:       Optional[str]   = None
    kredit_limit: Optional[float] = None
    eslatma:      Optional[str]   = None
# ════════════════════════════════════════════════════════════════


class XarajatSorov(BaseModel):
    kategoriya_nomi: str   = Field(..., min_length=1)
    summa:           float = Field(..., gt=0)
    izoh:            str   = Field("")
    shogird_id:      Optional[int] = None


@app.post("/api/v1/xarajat", tags=["Xarajatlar"])
async def api_xarajat_qoshish(data: XarajatSorov, uid: int = Depends(get_uid)):
    """Admin web paneldan xarajat qo'shish"""
    async with rls_conn(uid) as c:
        if data.shogird_id:
            # Shogird xarajati
            from shared.services.shogird_xarajat import xarajat_saqlash
            natija = await xarajat_saqlash(
                c, uid, data.shogird_id,
                data.kategoriya_nomi, data.summa, data.izoh
            )
        else:
            # Admin o'zi xarajat — shogirdsiz
            from shared.services.shogird_xarajat import _default_kategoriyalar
            await _default_kategoriyalar(c, uid)

            # Adminning o'zi uchun shogird yaratish (agar yo'q bo'lsa)
            admin_shogird = await c.fetchrow(
                "SELECT id FROM shogirdlar WHERE admin_uid=$1 AND telegram_uid=$1",
                uid
            )
            if not admin_shogird:
                admin_shogird = await c.fetchrow("""
                    INSERT INTO shogirdlar (admin_uid, telegram_uid, ism, lavozim)
                    VALUES ($1, $1, 'Admin', 'admin')
                    ON CONFLICT (admin_uid, telegram_uid) DO UPDATE SET faol=TRUE
                    RETURNING id
                """, uid)

            from shared.services.shogird_xarajat import xarajat_saqlash
            natija = await xarajat_saqlash(
                c, uid, admin_shogird["id"],
                data.kategoriya_nomi, data.summa, data.izoh
            )
            # Admin xarajatini avtomatik tasdiqlash
            if natija.get("id"):
                await c.execute(
                    "UPDATE xarajatlar SET tasdiqlangan=TRUE, tasdiq_vaqti=NOW() WHERE id=$1",
                    natija["id"]
                )
    log.info("💸 Xarajat qo'shildi: %s %s (uid=%d)", data.kategoriya_nomi, data.summa, uid)
    return {"status": "saqlandi", **natija}


# ════════════════════════════════════════════════════════════════
#  BILDIRISHNOMALAR — Web panel notification tizimi
# ════════════════════════════════════════════════════════════════


@app.get("/api/v1/bildirishnomalar", tags=["Bildirishnoma"])
async def bildirishnomalar(uid: int = Depends(get_uid)):
    """
    Web panel uchun bildirishnomalar:
    - Muddati o'tgan qarzlar
    - Kam qoldiqli tovarlar
    - Tasdiq kutayotgan xarajatlar
    """
    natija = {"items": [], "jami": 0}

    async with rls_conn(uid) as c:
        # 1. Muddati o'tgan qarzlar
        muddat_otgan = await c.fetch("""
            SELECT klient_ismi, COUNT(*) as soni,
                   SUM(qolgan) as jami_qarz
            FROM qarzlar
            WHERE yopildi=FALSE AND qolgan>0
              AND muddat IS NOT NULL AND muddat < NOW()
            GROUP BY klient_ismi
            ORDER BY jami_qarz DESC LIMIT 10
        """)
        for r in muddat_otgan:
            natija["items"].append({
                "tur": "qarz_muddati",
                "darajasi": "xavfli",
                "matn": f"{r['klient_ismi']}: {r['soni']} ta qarz muddati o'tgan ({float(r['jami_qarz']):,.0f} so'm)",
                "klient": r["klient_ismi"],
                "summa": float(r["jami_qarz"]),
            })

        # 2. Kam qoldiqli tovarlar
        kam_qoldiq = await c.fetch("""
            SELECT nomi, qoldiq, min_qoldiq FROM tovarlar
            WHERE min_qoldiq > 0 AND qoldiq <= min_qoldiq
            ORDER BY (qoldiq / NULLIF(min_qoldiq, 0)) ASC
            LIMIT 10
        """)
        for r in kam_qoldiq:
            natija["items"].append({
                "tur": "kam_qoldiq",
                "darajasi": "ogohlantirish",
                "matn": f"{r['nomi']}: qoldiq {float(r['qoldiq'])}, minimum {float(r['min_qoldiq'])}",
                "tovar": r["nomi"],
                "qoldiq": float(r["qoldiq"]),
            })

        # 3. Tasdiq kutayotgan xarajatlar
        kutilmoqda = await c.fetchval("""
            SELECT COUNT(*) FROM xarajatlar
            WHERE admin_uid=$1 AND NOT tasdiqlangan AND NOT bekor_qilingan
        """, uid)
        if kutilmoqda and kutilmoqda > 0:
            natija["items"].append({
                "tur": "xarajat_tasdiq",
                "darajasi": "info",
                "matn": f"{kutilmoqda} ta xarajat tasdiqlashni kutmoqda",
                "soni": kutilmoqda,
            })

    natija["jami"] = len(natija["items"])
    return natija


# ════════════════════════════════════════════════════════════════
#  TOVAR EXCEL EXPORT — Web panel uchun
# ════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════
#  SAVDOLAR — Sotuv sessiyalari ro'yxati (Web panel uchun)
# ════════════════════════════════════════════════════════════════


@app.get("/api/v1/savdolar", tags=["Sotuv"])
async def savdolar_royxati(
    limit: int = 20, offset: int = 0,
    klient: Optional[str] = None,
    sana_dan: Optional[str] = None,
    sana_gacha: Optional[str] = None,
    qarzdor: Optional[bool] = None,
    min_summa: Optional[float] = None,
    holat: Optional[str] = None,  # yangi/tasdiqlangan/otgruzka/yetkazildi/bekor
    sort: str = "sana",
    uid: int = Depends(get_uid),
):
    """
    Sotuv sessiyalari ro'yxati — SalesDoc-style filtrlash.
    Klient telefon va manzili klientlar jadvalidan olinadi.
    """
    limit = min(limit, 200)
    async with rls_conn(uid) as c:
        where_parts = []
        params: list = [limit, offset]
        idx = 3

        if klient:
            where_parts.append(
                f"(lower(ss.klient_ismi) LIKE lower(${idx}) "
                f"OR lower(COALESCE(k.ism,'')) LIKE lower(${idx}))"
            )
            params.append(f"%{like_escape(klient)}%")
            idx += 1

        if sana_dan:
            where_parts.append(f"ss.sana >= ${idx}::timestamptz")
            params.append(sana_dan)
            idx += 1

        if sana_gacha:
            where_parts.append(f"ss.sana < ${idx}::timestamptz + interval '1 day'")
            params.append(sana_gacha)
            idx += 1

        if qarzdor:
            where_parts.append("ss.qarz > 0")

        if min_summa is not None:
            where_parts.append(f"ss.jami >= ${idx}")
            params.append(min_summa)
            idx += 1

        if holat:
            where_parts.append(f"ss.holat = ${idx}")
            params.append(holat)
            idx += 1

        where_sql = (" AND " + " AND ".join(where_parts)) if where_parts else ""

        sort_map = {
            "sana":   "ss.sana DESC",
            "jami":   "ss.jami DESC",
            "qarz":   "ss.qarz DESC",
            "klient": "ss.klient_ismi ASC",
        }
        order_by = sort_map.get(sort, "ss.sana DESC")

        rows = await c.fetch(f"""
            SELECT ss.id, ss.klient_ismi, ss.jami, ss.tolangan, ss.qarz,
                   ss.izoh, ss.sana, ss.holat, ss.holat_yangilangan,
                   COALESCE(k.telefon, '')  AS telefon,
                   COALESCE(k.manzil, '')   AS manzil,
                   COALESCE(k.ism, '')      AS klient_nomi,
                   COUNT(ch.id)             AS tovar_soni
            FROM sotuv_sessiyalar ss
            LEFT JOIN chiqimlar ch ON ch.sessiya_id = ss.id
            LEFT JOIN klientlar  k ON k.id = ss.klient_id
            WHERE 1=1 {where_sql}
            GROUP BY ss.id, k.telefon, k.manzil, k.ism
            ORDER BY {order_by}
            LIMIT $1 OFFSET $2
        """, *params)

        total = await c.fetchval(f"""
            SELECT COUNT(*) FROM sotuv_sessiyalar ss
            LEFT JOIN klientlar k ON k.id = ss.klient_id
            WHERE 1=1 {where_sql}
        """, *params[2:])

        # Umumiy statistika (bugungi) + holat bo'yicha ajratish
        stats = await c.fetchrow("""
            SELECT
                COALESCE(SUM(jami), 0)     AS jami_tushum,
                COALESCE(SUM(tolangan), 0) AS tolangan,
                COALESCE(SUM(qarz), 0)     AS qarz,
                COUNT(*)                   AS soni,
                COUNT(*) FILTER (WHERE holat = 'yangi')         AS yangi,
                COUNT(*) FILTER (WHERE holat = 'tasdiqlangan')  AS tasdiqlangan,
                COUNT(*) FILTER (WHERE holat = 'otgruzka')      AS otgruzka,
                COUNT(*) FILTER (WHERE holat = 'yetkazildi')    AS yetkazildi,
                COUNT(*) FILTER (WHERE holat = 'bekor')         AS bekor
            FROM sotuv_sessiyalar
            WHERE (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
        """)

    return {
        "total": total or 0,
        "items": [dict(r) for r in rows],
        "stats": {
            "bugun_tushum": float(stats["jami_tushum"]),
            "bugun_tolangan": float(stats["tolangan"]),
            "bugun_qarz": float(stats["qarz"]),
            "bugun_soni": int(stats["soni"]),
            "holatlar": {
                "yangi":        int(stats["yangi"] or 0),
                "tasdiqlangan": int(stats["tasdiqlangan"] or 0),
                "otgruzka":     int(stats["otgruzka"] or 0),
                "yetkazildi":   int(stats["yetkazildi"] or 0),
                "bekor":        int(stats["bekor"] or 0),
            },
        },
    }


@app.get("/api/v1/sklad-qogozi/excel", tags=["Ombor"])
async def sklad_qogozi_excel(uid: int = Depends(get_uid)):
    """Sklad qog'ozi — ombor inventarizatsiyasi (barcha tovarlar + qoldiq + qiymat)."""
    import io, base64
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    async with rls_conn(uid) as c:
        rows = await c.fetch("""
            SELECT t.nomi, t.kategoriya, t.brend, t.shtrix_kod,
                   t.birlik, t.qoldiq, t.min_qoldiq,
                   t.olish_narxi, t.sotish_narxi,
                   (t.qoldiq * t.olish_narxi)   AS ombor_qiymati,
                   (t.qoldiq * t.sotish_narxi)  AS bozor_qiymati
            FROM tovarlar t
            WHERE t.user_id = $1
            ORDER BY t.kategoriya, t.nomi
        """, uid)
        user = await c.fetchrow(
            "SELECT ism, dokon_nomi, manzil FROM users WHERE id=$1", uid)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sklad qog'ozi"

    # Sarlavha qator
    ws.merge_cells("A1:K1")
    t = ws.cell(row=1, column=1, value="SKLAD QOG'OZI — OMBOR INVENTARIZATSIYASI")
    t.font = Font(bold=True, size=16, color="1B5E20")
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    dokon = (user and user["dokon_nomi"]) or "SavdoAI"
    sub = ws.cell(row=2, column=1,
                  value=f"{dokon} · {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    sub.font = Font(size=11, italic=True, color="666666")
    sub.alignment = Alignment(horizontal="center")

    # Jadval sarlavhasi (4-qator)
    headers = [
        "№", "Tovar", "Kategoriya", "Brend", "Shtrix kod",
        "Qoldiq", "Birlik", "Min q.",
        "Olish narxi", "Ombor qiymati", "Bozor qiymati",
    ]
    widths  = [5, 32, 18, 16, 16, 10, 8, 8, 14, 16, 16]

    header_fill = PatternFill(start_color="0A819C", end_color="0A819C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    start = 4
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=start, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[start].height = 28

    total_ombor = 0.0
    total_bozor = 0.0
    total_qoldiq = 0.0
    for idx, r in enumerate(rows, start + 1):
        d = dict(r)
        qoldiq = float(d["qoldiq"] or 0)
        min_q  = float(d["min_qoldiq"] or 0)
        ombor  = float(d["ombor_qiymati"] or 0)
        bozor  = float(d["bozor_qiymati"] or 0)
        vals = [
            idx - start,
            d["nomi"], d["kategoriya"] or "", d["brend"] or "",
            d["shtrix_kod"] or "",
            qoldiq, d["birlik"] or "", min_q,
            float(d["olish_narxi"] or 0),
            ombor, bozor,
        ]
        # Rang: tugagan=qizil, kam=sariq, normal=oq
        if qoldiq <= 0:
            fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        elif min_q > 0 and qoldiq <= min_q:
            fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        else:
            fill = None

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=idx, column=col, value=v)
            cell.border = border
            if fill:
                cell.fill = fill
            if col in (6, 8, 9, 10, 11):
                cell.number_format = '#,##0.##'
                cell.alignment = Alignment(horizontal="right")

        total_qoldiq += qoldiq
        total_ombor  += ombor
        total_bozor  += bozor

    # Total
    total_row = len(rows) + start + 1
    for col in range(1, 12):
        ws.cell(row=total_row, column=col).fill = PatternFill(
            start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ws.cell(row=total_row, column=col).border = border
    ws.cell(row=total_row, column=1, value="JAMI").font = Font(bold=True)
    q_cell = ws.cell(row=total_row, column=6, value=total_qoldiq)
    q_cell.font = Font(bold=True); q_cell.number_format = '#,##0.##'
    o_cell = ws.cell(row=total_row, column=10, value=total_ombor)
    o_cell.font = Font(bold=True, color="1B5E20"); o_cell.number_format = '#,##0'
    b_cell = ws.cell(row=total_row, column=11, value=total_bozor)
    b_cell.font = Font(bold=True, color="1B5E20"); b_cell.number_format = '#,##0'

    # Imzo joyi
    sig_row = total_row + 3
    ws.cell(row=sig_row, column=2, value="Ombor mudiri: ____________________").font = Font(size=10)
    ws.cell(row=sig_row, column=7, value="Buxgalter: ____________________").font = Font(size=10)
    ws.cell(row=sig_row + 1, column=2, value="Sana: ____________________").font = Font(size=10)

    ws.freeze_panes = f"A{start + 1}"
    ws.auto_filter.ref = f"A{start}:K{total_row}"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return {
        "filename": f"Sklad_qogozi_{datetime.now().strftime('%Y%m%d')}.xlsx",
        "content_base64": base64.b64encode(buf.getvalue()).decode(),
        "soni": len(rows),
        "jami_ombor_qiymati": total_ombor,
        "jami_bozor_qiymati": total_bozor,
    }


@app.post("/api/v1/nakladnoy/excel", tags=["Sotuv"])
async def nakladnoy_excel_batch(
    payload: dict,
    uid: int = Depends(get_uid),
):
    """Tanlangan buyurtmalar uchun multi-sheet nakladnoy Excel (SalesDoc 3.1 format)."""
    import io, base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sessiya_ids = payload.get("sessiya_ids") or []
    if not sessiya_ids:
        raise HTTPException(400, "sessiya_ids kerak")

    async with rls_conn(uid) as c:
        sessiyalar = await c.fetch("""
            SELECT ss.id, ss.klient_ismi, ss.jami, ss.tolangan, ss.qarz,
                   ss.izoh, ss.sana,
                   COALESCE(k.telefon, '') AS telefon,
                   COALESCE(k.manzil, '')  AS manzil,
                   COALESCE(k.ism, ss.klient_ismi, 'Mijoz') AS klient_nomi
            FROM sotuv_sessiyalar ss
            LEFT JOIN klientlar k ON k.id = ss.klient_id
            WHERE ss.user_id = $1 AND ss.id = ANY($2::bigint[])
            ORDER BY ss.id
        """, uid, sessiya_ids)

        if not sessiyalar:
            raise HTTPException(404, "Buyurtmalar topilmadi")

        # Har bir sessiya uchun tovarlarni birdaniga olish
        chiqimlar_map: dict = {}
        for ss_id in [s["id"] for s in sessiyalar]:
            rows = await c.fetch("""
                SELECT tovar_nomi, kategoriya, miqdor, birlik,
                       sotish_narxi, chegirma_foiz, jami
                FROM chiqimlar
                WHERE sessiya_id = $1 AND user_id = $2
                ORDER BY id
            """, ss_id, uid)
            chiqimlar_map[ss_id] = [dict(r) for r in rows]

    wb = Workbook()
    # Default sheet ni olib tashlash (birinchi nakladnoy bilan o'rnini oladi)
    default_sheet = wb.active
    first = True

    header_fill = PatternFill(start_color="0A819C", end_color="0A819C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, size=14, color="1B5E20")
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for s in sessiyalar:
        name = f"Nak#{s['id']}"[:28]
        if first:
            ws = default_sheet
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)

        # Title row
        ws.merge_cells("A1:F1")
        tcell = ws.cell(row=1, column=1, value=f"НАКЛАДНАЯ № {s['id']}")
        tcell.font = title_font
        tcell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 24

        sana_str = s["sana"].strftime("%d.%m.%Y %H:%M") if s.get("sana") else ""

        # Header info (2-5 qatorlar)
        info = [
            ("Sana:",    sana_str),
            ("Mijoz:",   s["klient_nomi"]),
            ("Telefon:", s["telefon"]),
            ("Manzil:",  s["manzil"]),
            ("Izoh:",    s["izoh"] or ""),
        ]
        for i, (k, v) in enumerate(info, 2):
            c_k = ws.cell(row=i, column=1, value=k)
            c_k.font = Font(bold=True)
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
            ws.cell(row=i, column=2, value=v)

        # Tovar jadvali (7-qatordan boshlab)
        headers = ["№", "Tovar", "Kategoriya", "Miqdor", "Narx", "Jami"]
        widths  = [5, 35, 18, 10, 14, 16]
        start_row = 8
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            ws.column_dimensions[chr(64 + col)].width = w

        chs = chiqimlar_map.get(s["id"], [])
        total = 0.0
        for idx, ch in enumerate(chs, start_row + 1):
            vals = [
                idx - start_row,
                ch["tovar_nomi"],
                ch["kategoriya"],
                f"{float(ch['miqdor']):.0f} {ch['birlik'] or ''}",
                float(ch["sotish_narxi"] or 0),
                float(ch["jami"] or 0),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=idx, column=col, value=v)
                cell.border = border
                if col in (5, 6):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
            total += float(ch["jami"] or 0)

        # Total row
        total_row = start_row + len(chs) + 1
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        c_tot = ws.cell(row=total_row, column=1, value="JAMI:")
        c_tot.font = Font(bold=True, size=12)
        c_tot.alignment = Alignment(horizontal="right")
        c_sum = ws.cell(row=total_row, column=6, value=total)
        c_sum.font = Font(bold=True, size=12, color="1B5E20")
        c_sum.number_format = '#,##0'
        c_sum.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

        # Imzo joyi
        sig_row = total_row + 3
        ws.cell(row=sig_row, column=1, value="Отпустил: ____________________").font = Font(size=10)
        ws.cell(row=sig_row, column=4, value="Принял: ____________________").font = Font(size=10)

        # Footer
        ws.cell(row=sig_row + 2, column=1,
                value=f"Сгенерировано: SavdoAI | {sana_str}").font = Font(size=8, italic=True, color="999999")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return {
        "filename": f"Nakladnoy_3.1_x{len(sessiyalar)}.xlsx",
        "content_base64": base64.b64encode(buf.getvalue()).decode(),
        "soni": len(sessiyalar),
    }


@app.get("/api/v1/savdolar/excel", tags=["Sotuv"])
async def savdolar_excel(
    sana_dan: Optional[str] = None,
    sana_gacha: Optional[str] = None,
    uid: int = Depends(get_uid),
):
    """SalesDoc Реестр 3.0 formatida Excel — sanalar bo'yicha buyurtmalar reestri."""
    import io, base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    async with rls_conn(uid) as c:
        where_parts = []
        params: list = []
        idx = 1
        if sana_dan:
            where_parts.append(f"ss.sana >= ${idx}::timestamptz")
            params.append(sana_dan); idx += 1
        if sana_gacha:
            where_parts.append(f"ss.sana < ${idx}::timestamptz + interval '1 day'")
            params.append(sana_gacha); idx += 1
        where_sql = (" AND " + " AND ".join(where_parts)) if where_parts else ""

        rows = await c.fetch(f"""
            SELECT ss.id, ss.klient_ismi, ss.jami, ss.tolangan, ss.qarz, ss.sana,
                   COALESCE(k.telefon, '') AS telefon,
                   COALESCE(k.manzil, '')  AS manzil,
                   COALESCE(k.jami_sotib, 0) AS balans
            FROM sotuv_sessiyalar ss
            LEFT JOIN klientlar k ON k.id = ss.klient_id
            WHERE 1=1 {where_sql}
            ORDER BY ss.sana DESC
        """, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр 3.0"

    headers = ["№", "Дата отгрузки", "Торгов. Точка", "Адрес", "Номер клиента",
               "Торгов. Пред.", "Баланс клиента", "Сумма", "Отметка"]
    widths  = [5, 13, 30, 35, 16, 14, 16, 16, 10]

    header_fill = PatternFill(start_color="0A819C", end_color="0A819C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 28

    total_sum = 0.0
    for idx2, r in enumerate(rows, 2):
        d = dict(r)
        sana_str = d["sana"].strftime("%d.%m.%Y") if d.get("sana") else ""
        vals = [
            idx2 - 1, sana_str, d["klient_ismi"] or "Mijoz",
            d["manzil"] or "", d["telefon"] or "", "SavdoAI Bot",
            float(d["balans"]), float(d["jami"]), ""
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=idx2, column=col, value=v)
            cell.border = border
            if col in (7, 8):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
        total_sum += float(d["jami"])

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=3, value="ИТОГО").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=8, value=total_sum)
    total_cell.font = Font(bold=True, color="1B5E20")
    total_cell.number_format = '#,##0'
    for col in range(1, 10):
        ws.cell(row=total_row, column=col).fill = PatternFill(
            start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
        )
        ws.cell(row=total_row, column=col).border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return {
        "filename": f"Реестр_3.0_{sana_dan or 'barcha'}.xlsx",
        "content_base64": base64.b64encode(buf.getvalue()).decode(),
        "soni": len(rows),
        "jami_summa": total_sum,
    }


@app.get("/api/v1/savdo/{sessiya_id}", tags=["Sotuv"])
async def savdo_tafsilot(sessiya_id: int, uid: int = Depends(get_uid)):
    """Bitta sotuv sessiyasi tafsiloti — tovarlar + klient + status bilan"""
    async with rls_conn(uid) as c:
        sess = await c.fetchrow("""
            SELECT ss.id, ss.klient_id, ss.klient_ismi, ss.jami, ss.tolangan,
                   ss.qarz, ss.izoh, ss.sana, ss.holat, ss.holat_yangilangan,
                   ss.otgruzka_vaqti, ss.yetkazildi_vaqti, ss.bekor_vaqti, ss.bekor_sabab,
                   COALESCE(k.telefon, '') AS klient_telefon,
                   COALESCE(k.manzil, '')  AS klient_manzil
            FROM sotuv_sessiyalar ss
            LEFT JOIN klientlar k ON k.id = ss.klient_id
            WHERE ss.id = $1 AND ss.user_id = $2
        """, sessiya_id, uid)
        if not sess:
            raise HTTPException(404, "Sotuv topilmadi")
        tovarlar = await c.fetch("""
            SELECT tovar_nomi, kategoriya, miqdor, birlik,
                   sotish_narxi, olish_narxi, chegirma_foiz, jami
            FROM chiqimlar WHERE sessiya_id=$1 AND user_id=$2 ORDER BY id
        """, sessiya_id, uid)
    return {**dict(sess), "tovarlar": [dict(r) for r in tovarlar]}


class SavdoHolatSorov(BaseModel):
    holat: str  # yangi / tasdiqlangan / otgruzka / yetkazildi / bekor
    sabab: Optional[str] = None


@app.put("/api/v1/savdo/{sessiya_id}/holat", tags=["Sotuv"])
async def savdo_holat_ozgartir(
    sessiya_id: int, data: SavdoHolatSorov, uid: int = Depends(get_uid)
):
    """Sotuv holatini o'zgartirish workflow: yangi → tasdiqlangan → otgruzka → yetkazildi.

    Yoki istalgan vaqtda bekor qilish mumkin.
    """
    ALLOWED = {"yangi", "tasdiqlangan", "otgruzka", "yetkazildi", "bekor"}
    if data.holat not in ALLOWED:
        raise HTTPException(400, f"Noto'g'ri holat. Ruxsat: {ALLOWED}")

    async with rls_conn(uid) as c:
        sess = await c.fetchrow(
            "SELECT id, holat FROM sotuv_sessiyalar WHERE id=$1 AND user_id=$2",
            sessiya_id, uid
        )
        if not sess:
            raise HTTPException(404, "Sotuv topilmadi")

        # Timestamp'ni holatga qarab belgilash
        extra_updates = []
        params: list = [data.holat, sessiya_id, uid]
        idx = 4
        if data.holat == "otgruzka":
            extra_updates.append("otgruzka_vaqti = NOW()")
        elif data.holat == "yetkazildi":
            extra_updates.append("yetkazildi_vaqti = NOW()")
        elif data.holat == "bekor":
            extra_updates.append("bekor_vaqti = NOW()")
            if data.sabab:
                extra_updates.append(f"bekor_sabab = ${idx}")
                params.insert(-2, data.sabab)
                idx += 1

        extras = (", " + ", ".join(extra_updates)) if extra_updates else ""
        await c.execute(f"""
            UPDATE sotuv_sessiyalar
            SET holat = $1, holat_yangilangan = NOW(){extras}
            WHERE id = $2 AND user_id = $3
        """, *params)

        # Bekor qilinganda — tovar qoldig'ini qaytarish
        if data.holat == "bekor" and sess["holat"] != "bekor":
            chiqimlar = await c.fetch(
                "SELECT tovar_id, miqdor FROM chiqimlar WHERE sessiya_id=$1", sessiya_id
            )
            for ch in chiqimlar:
                if ch["tovar_id"]:
                    await c.execute(
                        "UPDATE tovarlar SET qoldiq = qoldiq + $1 "
                        "WHERE id = $2 AND user_id = $3",
                        ch["miqdor"], ch["tovar_id"], uid
                    )

    log.info("📋 Sotuv #%d holat: %s → %s (uid=%d)",
             sessiya_id, sess["holat"], data.holat, uid)
    return {"id": sessiya_id, "holat": data.holat, "status": "yangilandi"}


# ════════════════════════════════════════════════════════════════
#  DASHBOARD TOP — Grafiklar uchun top tovar va top klient
# ════════════════════════════════════════════════════════════════


@app.get("/api/v1/dashboard/top", tags=["Dashboard"])
async def dashboard_top(kunlar: int = 30, uid: int = Depends(get_uid)):
    """Dashboard uchun top 5 tovar va top 5 klient (oxirgi N kun)"""
    from shared.cache.redis_cache import cache_ol, cache_yoz, TTL_HISOBOT
    cache_k = f"dashboard_top:{uid}:{kunlar}"
    cached = await cache_ol(cache_k)
    if cached:
        return cached

    async with rls_conn(uid) as c:
        top_tovar = await c.fetch("""
            SELECT ch.tovar_nomi AS nomi, SUM(ch.jami) AS jami, SUM(ch.miqdor) AS miqdor,
                   SUM((ch.sotish_narxi - ch.olish_narxi) * ch.miqdor) AS foyda
            FROM chiqimlar ch
            JOIN sotuv_sessiyalar ss ON ss.id = ch.sessiya_id
            WHERE ss.sana >= NOW() - make_interval(days => $1)
            GROUP BY ch.tovar_nomi
            ORDER BY jami DESC LIMIT 5
        """, kunlar)

        top_klient = await c.fetch("""
            SELECT klient_ismi AS ism, SUM(jami) AS jami, COUNT(*) AS soni,
                   SUM(qarz) AS qarz
            FROM sotuv_sessiyalar
            WHERE sana >= NOW() - make_interval(days => $1) AND klient_ismi IS NOT NULL
            GROUP BY klient_ismi
            ORDER BY jami DESC LIMIT 5
        """, kunlar)

        # Kunlik trend (oxirgi 7 kun)
        kunlik = await c.fetch("""
            SELECT (sana AT TIME ZONE 'Asia/Tashkent')::date AS kun,
                   COALESCE(SUM(jami), 0) AS sotuv,
                   COALESCE(SUM(qarz), 0) AS qarz
            FROM sotuv_sessiyalar
            WHERE sana >= NOW() - interval '7 days'
            GROUP BY kun ORDER BY kun
        """)

    result = {
        "top_tovar": [
            {"nomi": r["nomi"], "jami": float(r["jami"]), "miqdor": float(r["miqdor"]),
             "foyda": float(r["foyda"] or 0)}
            for r in top_tovar
        ],
        "top_klient": [
            {"ism": r["ism"], "jami": float(r["jami"]), "soni": int(r["soni"]),
             "qarz": float(r["qarz"] or 0)}
            for r in top_klient
        ],
        "kunlik_trend": [
            {"kun": str(r["kun"]), "sotuv": float(r["sotuv"]), "qarz": float(r["qarz"])}
            for r in kunlik
        ],
    }
    await cache_yoz(cache_k, result, TTL_HISOBOT)
    return result


# ════════════════════════════════════════════════════════════════
#  TOVAR EXCEL IMPORT — Web dan tovar yuklash
# ════════════════════════════════════════════════════════════════


class TovarImportItem(BaseModel):
    nomi:         str
    kategoriya:   str   = "Boshqa"
    birlik:       str   = "dona"
    olish_narxi:  float = 0
    sotish_narxi: float = 0
    qoldiq:       float = 0


class TovarImportSorov(BaseModel):
    tovarlar: List[TovarImportItem]
# ════════════════════════════════════════════════════════════════
#  STATISTIKA — Admin panel uchun tizim statistikasi
# ════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════
#  FOYDA TAHLILI — maxsus endpoint
# ════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════
#  QR-KOD — chek uchun QR kod generatsiya
# ════════════════════════════════════════════════════════════════


@app.get("/api/v1/qr/{sessiya_id}", tags=["Sotuv"])
async def qr_kod_generatsiya(sessiya_id: int, uid: int = Depends(get_uid)):
    """
    Sotuv sessiyasi uchun QR-kod SVG generatsiya.
    QR ichida chek URL bo'ladi — klient telefonida skanerlasa chek ko'rinadi.
    """
    import hashlib, io

    async with rls_conn(uid) as c:
        sess = await c.fetchrow(
            "SELECT id, klient_ismi, jami, sana FROM sotuv_sessiyalar WHERE id=$1 AND user_id=$2",
            sessiya_id, uid
        )
    if not sess:
        raise HTTPException(404, "Sotuv topilmadi")

    # QR mazmuni — chek URL yoki sotuv ma'lumoti
    base_url = os.getenv("PRINT_LANDING_BASE_URL", "")
    if base_url:
        qr_content = f"{base_url}/p/{sessiya_id}"
    else:
        qr_content = (
            f"SAVDOAI CHEK #{sessiya_id}\n"
            f"Klient: {sess['klient_ismi'] or '-'}\n"
            f"Jami: {float(sess['jami']):,.0f} so'm\n"
            f"Sana: {sess['sana']}"
        )

    # QR-kod SVG generatsiya
    qr_svg = None
    try:
        import qrcode
        import qrcode.image.svg
        qr = qrcode.QRCode(version=1, box_size=10, border=2,
                           error_correction=qrcode.constants.ERROR_CORRECT_M)
        qr.add_data(qr_content)
        qr.make(fit=True)
        factory = qrcode.image.svg.SvgPathImage
        img = qr.make_image(image_factory=factory)
        buf = io.BytesIO()
        img.save(buf)
        qr_svg = buf.getvalue().decode("utf-8")
    except ImportError:
        log.warning("qrcode kutubxonasi o'rnatilmagan — pip install qrcode[pil]")
    except Exception as e:
        log.warning("QR generatsiya xato: %s", e)

    qr_hash = hashlib.sha256(qr_content.encode()).hexdigest()[:8]

    result = {
        "sessiya_id": sessiya_id,
        "klient": sess["klient_ismi"],
        "jami": float(sess["jami"]),
        "qr_content": qr_content,
        "qr_hash": qr_hash,
    }
    if qr_svg:
        result["qr_svg"] = qr_svg
    return result


# ════════════════════════════════════════════════════════════════
#  KLIENT TARIXI — sotuv va qarz tarixi
# ════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════
#  PROFIL YANGILASH — Settings sahifasi uchun
# ════════════════════════════════════════════════════════════════


class ProfilYangilaSorov(BaseModel):
    ism:        Optional[str] = None
    dokon_nomi: Optional[str] = None
    telefon:    Optional[str] = None
    manzil:     Optional[str] = None
    inn:        Optional[str] = None
    til:        Optional[str] = None


@app.put("/api/v1/me", tags=["Auth"])
async def profil_yangilash(data: ProfilYangilaSorov, uid: int = Depends(get_uid)):
    """Foydalanuvchi profilini yangilash"""
    yangilar = {k: v for k, v in data.model_dump().items() if v is not None}
    if not yangilar:
        raise HTTPException(400, "Yangilash uchun kamida 1 ta maydon kerak")

    _RUXSAT = {"ism", "dokon_nomi", "telefon", "manzil", "inn", "til"}
    noma = set(yangilar.keys()) - _RUXSAT
    if noma:
        raise HTTPException(400, f"Ruxsat etilmagan maydon: {noma}")

    set_q = ", ".join(f"{k} = ${i+2}" for i, k in enumerate(yangilar.keys()))
    vals = list(yangilar.values())

    async with rls_conn(uid) as c:
        await c.execute(
            f"UPDATE users SET {set_q} WHERE id=$1",
            uid, *vals
        )
    from shared.cache.redis_cache import user_cache_tozala
    await user_cache_tozala(uid)
    log.info("👤 Profil yangilandi: uid=%d, maydonlar=%s", uid, list(yangilar.keys()))
    return {"status": "yangilandi", "maydonlar": list(yangilar.keys())}


@app.put("/api/v1/me/parol", tags=["Auth"])
async def parol_yangilash(data: dict, uid: int = Depends(get_uid)):
    """Foydalanuvchi parolini o'zgartirish"""
    eski = (data.get("eski_parol") or "").strip()
    yangi = (data.get("yangi_parol") or "").strip()
    if not yangi or len(yangi) < 4:
        raise HTTPException(400, "Yangi parol kamida 4 belgi bo'lishi kerak")

    async with rls_conn(uid) as c:
        user = await c.fetchrow(
            "SELECT parol_hash FROM users WHERE id=$1", uid
        )
        if not user:
            raise HTTPException(404, "Foydalanuvchi topilmadi")

        # Agar parol mavjud — eski parolni tekshirish
        if user.get("parol_hash"):
            if not _parol_tekshir(eski, user["parol_hash"]):
                raise HTTPException(401, "Eski parol noto'g'ri")

        new_hash = _parol_hash(yangi)
        await c.execute("UPDATE users SET parol_hash=$2 WHERE id=$1", uid, new_hash)

    log.info("🔐 Parol yangilandi: uid=%d", uid)
    return {"status": "yangilandi"}


# ════════════════════════════════════════════════════════════════
#  TOVAR TARIXI — sotuv, kirim, narx o'zgarish tarixi
# ════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════
#  FAKTURA (Hisob-faktura) CRUD
# ════════════════════════════════════════════════════════════════


class FakturaYaratSorov(BaseModel):
    klient_ismi: str = Field(..., min_length=1, max_length=200)
    tovarlar: list = Field(default_factory=list)
    jami_summa: float = Field(0, ge=0)
    bank_rekvizit: Optional[dict] = None
    izoh: Optional[str] = None
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/dokon/{dokon_id}/tovarlar", tags=["Mini-Do'kon"])
async def dokon_tovarlar(dokon_id: int, q: str = "", request: Request = None):
    """Ommaviy tovar katalog — auth kerak emas. Rate limit: 30 req/min.
    Faqat faol do'konchilarning tovarlarini ko'rsatadi.
    """
    if request:
        ip = request.client.host if request.client else "unknown"
        from shared.cache.redis_cache import rate_limit_tekshir
        if not await rate_limit_tekshir(f"dokon:{ip}", max_req=30, window_s=60):
            raise HTTPException(429, "Juda ko'p so'rov. 1 daqiqa kuting.")
    from shared.utils import like_escape
    async with get_pool().acquire() as c:
        dokon = await c.fetchval(
            "SELECT id FROM users WHERE id=$1 AND faol=TRUE", dokon_id
        )
        if not dokon:
            raise HTTPException(404, "Do'kon topilmadi")
        if q:
            rows = await c.fetch(
                "SELECT id, nomi, kategoriya, sotish_narxi, birlik, qoldiq "
                "FROM tovarlar WHERE user_id=$1 AND qoldiq > 0 "
                "AND LOWER(nomi) LIKE LOWER($2) ORDER BY nomi LIMIT 100",
                dokon_id, f"%{like_escape(q)}%",
            )
        else:
            rows = await c.fetch(
                "SELECT id, nomi, kategoriya, sotish_narxi, birlik, qoldiq "
                "FROM tovarlar WHERE user_id=$1 AND qoldiq > 0 "
                "ORDER BY nomi LIMIT 100",
                dokon_id,
            )
    return [dict(r) for r in rows]


@app.post("/api/v1/dokon/{dokon_id}/buyurtma", tags=["Mini-Do'kon"])
async def dokon_buyurtma(dokon_id: int, data: dict, request: Request = None):
    """Klient buyurtma yaratish — auth kerak emas. Rate limit: 5 req/min."""
    # Rate limit — spam himoyasi
    if request:
        ip = request.client.host if request.client else "unknown"
        from shared.cache.redis_cache import rate_limit_tekshir
        if not await rate_limit_tekshir(f"buyurtma:{ip}", max_req=5, window_s=60):
            raise HTTPException(429, "Juda ko'p so'rov. 1 daqiqa kuting.")
    tovarlar = data.get("tovarlar", [])
    if not tovarlar:
        raise HTTPException(400, "Tovarlar ro'yxati bo'sh")

    async with get_pool().acquire() as c:
        async with c.transaction():
            row = await c.fetchrow(
                "INSERT INTO buyurtmalar (user_id, klient_ismi, telefon, izoh) "
                "VALUES ($1, $2, $3, $4) RETURNING id",
                dokon_id,
                data.get("klient_ismi", ""),
                data.get("telefon", ""),
                data.get("izoh", ""),
            )
            buyurtma_id = row["id"]

            for t in tovarlar:
                tovar = await c.fetchrow(
                    "SELECT id, nomi, sotish_narxi FROM tovarlar "
                    "WHERE id=$1 AND user_id=$2",
                    t.get("id"), dokon_id,
                )
                if tovar:
                    await c.execute(
                        "INSERT INTO buyurtma_tovarlar "
                        "(buyurtma_id, tovar_id, nomi, miqdor, narx) "
                        "VALUES ($1, $2, $3, $4, $5)",
                        buyurtma_id,
                        tovar["id"],
                        tovar["nomi"],
                        t.get("miqdor", 1),
                        tovar["sotish_narxi"],
                    )

    # Do'konchiga Telegram xabar
    try:
        import httpx
        bot_token = os.getenv("BOT_TOKEN", "")
        if bot_token:
            matn = (
                f"🛒 *Yangi buyurtma!*\n\n"
                f"👤 {data.get('klient_ismi', 'Noma')}\n"
                f"📞 {data.get('telefon', '-')}\n"
                f"📦 {len(tovarlar)} ta tovar\n"
                f"📝 {data.get('izoh', '-')}"
            )
            async with httpx.AsyncClient() as client:
                await client.post(
                    f"https://api.telegram.org/bot{bot_token}/sendMessage",
                    json={"chat_id": dokon_id, "text": matn, "parse_mode": "Markdown"},
                    timeout=10,
                )
    except Exception as e:
        log.debug("Buyurtma telegram: %s", e)

    return {"buyurtma_id": buyurtma_id, "status": "yangi"}


# ════════════════════════════════════════════════════════════════
#  AI NARX TAVSIYA
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/narx/tavsiya", tags=["Narxlar"])
async def narx_tavsiya(uid: int = Depends(get_uid)):
    """AI narx tavsiyasi — foyda optimizatsiya."""
    from shared.services.ai_narx_tavsiya import narx_tavsiyalar
    async with rls_conn(uid) as c:
        return await narx_tavsiyalar(c, uid, limit=20)


# ════════════════════════════════════════════════════════════════
#  MOLIYAVIY PROGNOZ
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/prognoz", tags=["Hisobotlar"])
async def prognoz(uid: int = Depends(get_uid)):
    """Kelasi oy moliyaviy bashorat."""
    from shared.services.moliyaviy_prognoz import moliyaviy_prognoz
    async with rls_conn(uid) as c:
        return await moliyaviy_prognoz(c, uid)


# ════════════════════════════════════════════════════════════════
#  OMBOR PROGNOZ — tovar qachon tugadi bashorat
# ════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════
#  KLIENT CRM
# ════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════
#  CHEGIRMA
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/chegirma/qoidalar", tags=["Chegirma"])
async def chegirma_list(uid: int = Depends(get_uid)):
    """Chegirma qoidalari ro'yxati."""
    from shared.services.chegirma import chegirma_qoidalar_olish
    async with rls_conn(uid) as c:
        return await chegirma_qoidalar_olish(c, uid)


# ════════════════════════════════════════════════════════════════
#  RAQOBAT MONITORING
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/raqobat/tahlil", tags=["Raqobat"])
async def raqobat_tahlil_api(uid: int = Depends(get_uid)):
    """Raqobat narx tahlili."""
    from shared.services.raqobat_monitoring import raqobat_tahlil, raqobat_xulosa
    async with rls_conn(uid) as c:
        return {
            "tahlil": await raqobat_tahlil(c, uid),
            "xulosa": await raqobat_xulosa(c, uid),
        }


# ════════════════════════════════════════════════════════════════
#  CLICK / PAYME WEBHOOK
# ════════════════════════════════════════════════════════════════

@app.post("/webhook/click", tags=["To'lov"], include_in_schema=False)
async def webhook_click(request: Request):
    """Click.uz to'lov webhook."""
    from shared.services.tolov_integratsiya import ClickProvider
    data = await request.json() if request.headers.get("content-type", "").startswith("application/json") else dict(await request.form())

    provider = ClickProvider()
    if not provider.webhook_tekshirish(data, dict(request.headers)):
        log.warning("Click webhook: noto'g'ri imzo")
        return {"error": -1, "error_note": "Signature invalid"}

    natija = provider.webhook_parse(data)
    log.info("💳 Click webhook: %s summa=%s holat=%s",
             natija.tranzaksiya_id, natija.summa, natija.holat.value)

    # DB ga saqlash
    try:
        async with get_pool().acquire() as c:
            order_id = data.get("merchant_trans_id", "")
            await c.execute("""
                INSERT INTO tolov_tranzaksiyalar
                    (user_id, order_id, provider, tranzaksiya_id, summa, holat, meta)
                VALUES (0, $1, 'click', $2, $3, $4, $5::jsonb)
                ON CONFLICT DO NOTHING
            """, order_id, str(natija.tranzaksiya_id),
                natija.summa, natija.holat.value, json.dumps(data))
    except Exception as e:
        log.error("Click webhook DB: %s", e)

    return {"click_trans_id": data.get("click_trans_id"), "merchant_trans_id": data.get("merchant_trans_id"),
            "error": 0, "error_note": "Success"}


@app.post("/webhook/payme", tags=["To'lov"], include_in_schema=False)
async def webhook_payme(request: Request):
    """Payme.uz JSON-RPC webhook."""
    from shared.services.tolov_integratsiya import PaymeProvider
    data = await request.json()

    provider = PaymeProvider()
    if not provider.webhook_tekshirish(data, dict(request.headers)):
        return {"error": {"code": -32504, "message": "Auth failed"}, "id": data.get("id")}

    natija = provider.webhook_parse(data)
    log.info("💳 Payme webhook: method=%s holat=%s summa=%s",
             data.get("method"), natija.holat.value, natija.summa)

    # DB ga saqlash
    try:
        async with get_pool().acquire() as c:
            account = data.get("params", {}).get("account", {})
            order_id = account.get("order_id", "")
            await c.execute("""
                INSERT INTO tolov_tranzaksiyalar
                    (user_id, order_id, provider, tranzaksiya_id, summa, holat, meta)
                VALUES (0, $1, 'payme', $2, $3, $4, $5::jsonb)
                ON CONFLICT DO NOTHING
            """, str(order_id), str(natija.tranzaksiya_id),
                natija.summa, natija.holat.value, json.dumps(data))
    except Exception as e:
        log.error("Payme webhook DB: %s", e)

    # Payme JSON-RPC javob
    method = data.get("method", "")
    if method == "CheckPerformTransaction":
        return {"result": {"allow": True}, "id": data.get("id")}
    elif method == "CreateTransaction":
        return {"result": {"create_time": int(time.time()*1000),
                           "transaction": natija.tranzaksiya_id,
                           "state": 1}, "id": data.get("id")}
    elif method == "PerformTransaction":
        return {"result": {"transaction": natija.tranzaksiya_id,
                           "perform_time": int(time.time()*1000),
                           "state": 2}, "id": data.get("id")}
    return {"result": {"success": True}, "id": data.get("id")}


# ════════════════════════════════════════════════════════════════
#  NOTIFICATION API
# ════════════════════════════════════════════════════════════════

@app.get("/api/v1/notification/preview/{turi}", tags=["Notification"])
async def notification_preview(turi: str, uid: int = Depends(get_uid)):
    """Bildirishnoma oldindan ko'rish (ertalab/kechqurun/haftalik/critical)."""
    from shared.services.smart_notification import notification_dispatch
    async with rls_conn(uid) as c:
        matn = await notification_dispatch(c, uid, turi)
    if not matn:
        return {"matn": None, "sabab": "Ma'lumot yo'q yoki sotuv bo'lmagan"}
    return {"matn": matn, "turi": turi}
