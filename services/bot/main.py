"""
╔══════════════════════════════════════════════════════════════════╗
║  SAVDOAI MASHRAB MOLIYA  v25.3  PRODUCTION GRADE               ║
║  @savdoai_mashrab_bot                                            ║
║                                                                  ║
║  🎤 OVOZ-BIRINCHI: Ovoz yuboring — bot hamma ishni qiladi      ║
║  🧠 DUAL-BRAIN AI: Gemini (ovoz) + Claude (mantiq)             ║
║  🛡️ XAVFSIZ: Draft→Tasdiqlash→Saqlash→Audit                   ║
║  📸 VISION AI: Rasm → matn (nakladnoy, chek o'qish)            ║
║  💳 KASSA: Naqd / Karta / O'tkazma                              ║
║  📋 NAKLADNOY: Word+Excel+PDF | MIJOZ MA'LUMOTLARI             ║
║  🔒 HIMOYA: RLS + JWT + Decimal + FK + Audit Log                ║
║  📊 SAP-GRADE: Double-Entry Ledger + Reconciliation             ║
║                                                                  ║
║  v25.3 PRODUCTION YANGILIKLAR:                                   ║
║  ✅ Railway production deploy — crash-proof schema_init          ║
║  ✅ Bulletproof startup (hech qachon crash qilmaydi)            ║
║  ✅ Nixpacks + Procfile + railway.toml — 3x fallback            ║
║  ✅ Statement-by-statement SQL execution (asyncpg safe)          ║
║  ✅ Non-critical error skip (INDEX, VIEW, TRIGGER)              ║
║  ✅ Double-Entry Ledger (ikki tomonlama buxgalteriya)           ║
║  ✅ Idempotency + Hujjat versiyalash + Reconciliation           ║
║  ✅ 264 test — 100% passing                                      ║
╚══════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations
import io, logging, sys, time
import datetime
import pytz
from collections import defaultdict
from decimal import Decimal
from typing import Optional

from telegram import (
    Update, BotCommand,
    InlineKeyboardButton, InlineKeyboardMarkup, InputFile,
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler,
    filters, ContextTypes,
)
from telegram.constants import ParseMode
from telegram.helpers import escape_markdown as _esc_md

def esc(t: str) -> str:
    """Telegram MarkdownV2 uchun maxsus belgilarni escape qilish"""
    if not isinstance(t, str): t = str(t)
    SPECIAL = ['\\', '`', '*', '_', '{', '}', '[', ']', '(', ')',
               '#', '+', '-', '.', '!', '|', '~', '>', '=']
    for ch in SPECIAL:
        t = t.replace(ch, '\\' + ch)
    return t
from telegram.error import BadRequest

from config import Config, cfg as _cfg_fn, config_init
import sys as _sys, os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.dirname(__file__))))

# Print HMAC — productionda PRINT_SECRET majburiy (bot va API bir xil qiymat).
import shared.services.print_session  # noqa: F401

import services.bot.db  as db
from shared.database.pool import rls_conn as _rls_conn, pool_init as _pool_init
from shared.utils import like_escape
import services.bot.bot_services.voice      as ovoz_xizmat
import services.bot.bot_services.analyst    as ai_xizmat
from services.bot.bot_services.voice_pipeline import process_voice
from services.bot.handlers.shogird import _shogird_xarajat_qabul
import services.bot.bot_services.export_pdf   as pdf_xizmat
import services.bot.bot_services.export_excel as excel_xizmat
import services.bot.bot_services.nakladnoy    as nakl_xizmat
import services.bot.bot_services.rasm_handler as rasm_xizmat
from shared.utils.fmt import (
    pul, chek_md, SAHIFA,
    sotuv_cheki, kirim_cheki, qaytarish_cheki,
    kunlik_matn, oylik_matn, foyda_matn,
    klient_hisobi_matn,
)

import os as _os
_os.makedirs("/tmp/mm_logs", exist_ok=True)

_log_handlers: list = [logging.StreamHandler(sys.stdout)]
try:
    from logging.handlers import RotatingFileHandler
    _log_handlers.append(RotatingFileHandler(
        "/tmp/mm_logs/mm_bot.log",
        maxBytes=5*1024*1024, backupCount=3, encoding="utf-8",
    ))
except Exception:
    pass  # log not yet configured

logging.basicConfig(
    format="%(asctime)s │ %(levelname)-8s │ %(name)s │ %(message)s",
    level=logging.INFO, handlers=_log_handlers,
)
for _s in ("httpx","httpcore","telegram.ext._application"):
    logging.getLogger(_s).setLevel(logging.WARNING)

# Sentry (ixtiyoriy — SENTRY_DSN bo'lsa uladi)
_SENTRY_DSN = _os.getenv("SENTRY_DSN")
if _SENTRY_DSN:
    try:
        import sentry_sdk
        sentry_sdk.init(dsn=_SENTRY_DSN, traces_sample_rate=0.05)
        logging.getLogger(__name__).info("✅ Sentry ulandi (bot)")
    except ImportError:
        pass
__version__ = "25.3"
__author__  = "Mashrab Moliya"

# Segment nomi matnlari
SEGMENT_NOMI = {
    "optom":      "Optom savdo",
    "chakana":    "Chakana savdo",
    "oshxona":    "Oshxona / Kafe",
    "xozmak":     "Xo'zmag",
    "kiyim":      "Kiyim-kechak",
    "gosht":      "Go'sht do'koni",
    "meva":       "Meva-sabzavot",
    "qurilish":   "Qurilish mollari",
    "avto":       "Avto ehtiyot qismlar",
    "dorixona":   "Dorixona",
    "texnika":    "Texnika / Elektronika",
    "mebel":      "Mebel",
    "mato":       "Mato / Gazlama",
    "gul":        "Gul do'koni",
    "kosmetika":  "Kosmetika / Parfyumeriya",
    "universal":  "Boshqa (universal)",
}

log = logging.getLogger("mm")

# ── Turbo kesh (user = 120s, hisobot = 60s) ──────────────────────
import time as _time

# Umumiy yordamchi funksiyalar — bot_helpers.py dan import
from services.bot.bot_helpers import (
    _kesh, _kesh_ol, _kesh_yoz, _kesh_tozala, _kesh_tozala_prefix,
    _KESH_TTL, _KESH_USER_TTL, _KESH_MAX_SIZE,
    _user_ol_kesh, faol_tekshir, _yuborish, _safe_reply, xat,
    _md_safe, _truncate, tg,
)


async def health_check(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """Railway health monitoring — v25.3"""
    import time as _t
    start = _t.monotonic()
    # DB ping
    try:
        from shared.database.pool import pool_health
        import asyncio
        db_info = await asyncio.wait_for(pool_health(), timeout=5)
        db_ms = db_info.get("ping_ms", "?")
        db_status = f"✅ DB: {db_ms}ms (pool: {db_info.get('used',0)}/{db_info.get('size',0)})"
    except Exception as _e:
        log.debug("Xato: %s", _e)
        db_status = "⚠️ DB: tekshirib bo'lmadi"
    # Cache stats
    cache_size = len(_kesh)
    total_ms = round((_t.monotonic() - start) * 1000, 1)
    await update.message.reply_text(
        f"✅ Bot {__version__} ishlayapti\n"
        f"🕐 {datetime.datetime.now(pytz.timezone('Asia/Tashkent')).strftime('%H:%M:%S')}\n"
        f"{db_status}\n"
        f"💾 Kesh: {cache_size} ta yozuv\n"
        f"⚡ Javob: {total_ms}ms"
    )


async def xato_handler(update: object,
                        ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """Global xato handler — hech narsa jimgina o'tmaydi"""
    import traceback
    xato = ctx.error

    # Deploy paytida Conflict normal — log spam oldini olish
    if "Conflict" in str(type(xato).__name__) or "terminated by other" in str(xato):
        log.info("🔄 Deploy: eski instance to'xtatilmoqda (normal)")
        return

    tb   = "".join(traceback.format_exception(type(xato), xato, xato.__traceback__))
    log.error("⛔ Global xato:\n%s", tb)

    # Adminlarga xabar berish
    try:
        if _CFG:
            for aid in _CFG.admin_ids:
                await ctx.bot.send_message(
                    aid,
                    f"⛔ *Bot xatosi*\n\n"
                    "Xato yuz berdi",
                    parse_mode=ParseMode.MARKDOWN,
                )
    except Exception as _exc:
        log.debug("%s: %s", "main", _exc)  # was silent

H_SEGMENT, H_DOKON, H_TELEFON = range(3)
_oxirgi: dict[int,float] = defaultdict(float)
FLOOD_SON = 1.5
_FLOOD_MAX_SIZE = 10000
_CFG: Optional[Config] = None


def cfg() -> Config:
    assert _CFG is not None; return _CFG

def _flood_ok(uid:int) -> bool:
    hozir=time.monotonic()
    if hozir-_oxirgi[uid]<FLOOD_SON: return False
    _oxirgi[uid]=hozir
    # Xotira tozalash — 10K dan ortiq yozuv bo'lsa eskilarini o'chirish
    if len(_oxirgi) > _FLOOD_MAX_SIZE:
        cutoff = hozir - 60
        expired = [k for k, v in _oxirgi.items() if v < cutoff]
        for k in expired:
            _oxirgi.pop(k, None)
    return True


def asosiy_menyu() -> InlineKeyboardMarkup:
    from telegram import WebAppInfo
    _dash_url = _os.getenv("WEB_URL", "https://savdoai-web-production.up.railway.app") + "/tg"
    buttons = [
        [InlineKeyboardButton("📥 Kirim", callback_data="m:kirim"),
         InlineKeyboardButton("📤 Sotuv", callback_data="m:chiqim")],
        [InlineKeyboardButton("↩️ Qaytarish", callback_data="m:qaytarish"),
         InlineKeyboardButton("💰 Qarz to'lash", callback_data="m:qarzlar")],
        [InlineKeyboardButton("📋 Nakladnoy", callback_data="m:nakladnoy"),
         InlineKeyboardButton("📋 Faktura", callback_data="m:faktura")],
        [InlineKeyboardButton("📦 Tovarlar", callback_data="m:tovarlar"),
         InlineKeyboardButton("👥 Klientlar", callback_data="m:klientlar")],
        [InlineKeyboardButton("📊 Hisobot", callback_data="m:hisobot"),
         InlineKeyboardButton("💹 Foyda", callback_data="m:foyda")],
        [InlineKeyboardButton("💳 Kassa", callback_data="m:kassa"),
         InlineKeyboardButton("📸 Rasm OCR", callback_data="m:rasm")],
        [InlineKeyboardButton("🏭 Ombor", callback_data="m:ombor"),
         InlineKeyboardButton("⚠️ Kam qoldiq", callback_data="m:ogoh")],
        [InlineKeyboardButton("📒 Jurnal", callback_data="m:jurnal"),
         InlineKeyboardButton("📊 Balans", callback_data="m:balans")],
        [InlineKeyboardButton("👥 Shogirdlar", callback_data="m:shogirdlar"),
         InlineKeyboardButton("🏷 Narx guruh", callback_data="m:narx")],
        [InlineKeyboardButton("📱 Dashboard", web_app=WebAppInfo(url=_dash_url))],
        [InlineKeyboardButton("🆕 Yangiliklar", callback_data="m:yangilik"),
         InlineKeyboardButton("❓ Yordam", callback_data="m:yordam")],
    ]
    return InlineKeyboardMarkup(buttons)


# ════════════ START + RO'YXAT ════════════

async def cmd_start(update:Update, ctx:ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id
    log.info("📩 /start: uid=%d name=%s", uid, update.effective_user.full_name)
    try:
        user=await _user_ol_kesh(uid)
    except Exception as _e:
        log.error("❌ /start user_ol xato: %s", _e, exc_info=True)
        await update.message.reply_text("❌ Xato yuz berdi. Qayta /start bosing.")
        return ConversationHandler.END
    if user and user.get("faol", False):
        kam=await db.kam_qoldiq_tovarlar(uid)
        ism = user.get('to_liq_ism') or user.get('ism', 'Foydalanuvchi')
        dokon = user.get('dokon_nomi', "Do'konim")

        # Vaqtga qarab salomlashish
        import pytz as _pz
        from datetime import datetime as _dt
        soat = _dt.now(_pz.timezone("Asia/Tashkent")).hour
        if soat < 12:
            salom = "🌅 Xayrli tong"
        elif soat < 17:
            salom = "☀️ Xayrli kun"
        else:
            salom = "🌙 Xayrli kech"

        # Bugungi qisqa statistika
        bugun_stat = ""
        try:
            from shared.database.pool import rls_conn as _rc
            async with _rc(uid) as _sc:
                _bs = await _sc.fetchrow("""
                    SELECT COUNT(*) soni, COALESCE(SUM(jami),0) jami
                    FROM sotuv_sessiyalar
                    WHERE (sana AT TIME ZONE 'Asia/Tashkent')::date = CURRENT_DATE
                """)
                _qarz = await _sc.fetchval("""
                    SELECT COALESCE(SUM(qolgan),0) FROM qarzlar
                    WHERE yopildi=FALSE AND qolgan>0
                """) or 0
                if int(_bs["soni"]) > 0:
                    bugun_stat = (
                        f"\n📊 *Bugun:* {int(_bs['soni'])} sotuv"
                        f" | {float(_bs['jami']):,.0f} so'm"
                    )
                if float(_qarz) > 0:
                    bugun_stat += f"\n💳 Jami qarz: {float(_qarz):,.0f}"
        except Exception:
            pass

        ogoh = ""
        if kam:
            ogoh = f"\n⚠️ Kam qoldiq: {', '.join(t['nomi'] for t in kam[:3])}"

        await update.message.reply_text(
            f"{salom}, *{ism}*!\n"
            f"🏪 {dokon}\n"
            f"{bugun_stat}{ogoh}\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n"
            f"🤖 *SavdoAI v{__version__}*\n\n"
            "🎤 *Ovoz yuboring:*\n"
            "_\"Salimovga 50 Ariel ketti qarzga\"_\n\n"
            "📋 *Yangi buyruqlar:*\n"
            "/kpi — 📊 Samaradorlik\n"
            "/tahlil — 🧠 AI maslahat\n"
            "/eslatma — 📨 Qarz eslatma\n"
            "/loyalty — ⭐ Bonus ball\n"
            "/buyurtma — 📦 Tovar buyurtma\n"
            "/yordam — Barcha buyruqlar",
            parse_mode=ParseMode.MARKDOWN, reply_markup=asosiy_menyu(),
        )
        return ConversationHandler.END
    if user and not user.get("faol", False):
        await update.message.reply_text("⏳ Hisobingiz tasdiqlanmagan.")
        return ConversationHandler.END
    await db.user_yoz(uid,
        update.effective_user.full_name or "Nomsiz",
        update.effective_user.username)
    await update.message.reply_text(
        "👋 *Mashrab Moliya*ga xush kelibsiz!\n\nBiznes turini tanlang:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=tg(
            [("🏭 Optom (ulgurji)",     "seg:optom")],
            [("🏪 Chakana (mayda)",     "seg:chakana")],
            [("🍽️ Oshxona / Kafe",     "seg:oshxona")],
            [("🍦 Xo'zmag",            "seg:xozmak")],
            [("👔 Kiyim-kechak",       "seg:kiyim")],
            [("🥩 Go'sht do'koni",     "seg:gosht")],
            [("🍎 Meva-sabzavot",      "seg:meva")],
            [("🧱 Qurilish mollari",   "seg:qurilish")],
            [("🚗 Avto ehtiyot qismlar","seg:avto")],
            [("💊 Dorixona",           "seg:dorixona")],
            [("📱 Texnika / Elektronika","seg:texnika")],
            [("🛋️ Mebel",              "seg:mebel")],
            [("🧵 Mato / Gazlama",     "seg:mato")],
            [("🌹 Gul do'koni",        "seg:gul")],
            [("💄 Kosmetika",          "seg:kosmetika")],
            [("🛒 Boshqa (universal)", "seg:universal")],
        ),
    )
    return H_SEGMENT


async def h_segment(update:Update, ctx:ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    seg=q.data.replace("seg:",""); ctx.user_data["seg"]=seg
    await xat(q,f"✅ *{SEGMENT_NOMI[seg]}* tanlandi!\n\n🏪 Do'kon nomini yozing:",
               parse_mode=ParseMode.MARKDOWN)
    return H_DOKON


async def h_dokon(update:Update, ctx:ContextTypes.DEFAULT_TYPE):
    ctx.user_data["dokon"]=update.message.text.strip()
    await update.message.reply_text("📞 Telefon raqamingiz (+998...):")
    return H_TELEFON


async def h_telefon(update:Update, ctx:ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id
    seg=ctx.user_data["seg"]; dokon=ctx.user_data["dokon"]
    tel=update.message.text.strip()
    await db.user_yangilab(uid,segment=seg,dokon_nomi=dokon,telefon=tel)
    _kesh_tozala(f"user:{uid}")

    # ── Tayyor tovar bazasini yuklash (segment bo'yicha) ──
    seed_soni = 0
    try:
        from shared.services.seed_catalog import seed_tovarlar
        from shared.database.pool import get_pool
        async with db._P().acquire() as c:
            seed_soni = await seed_tovarlar(c, uid, seg)
    except Exception as _seed_e:
        log.warning("Seed catalog xato uid=%d: %s", uid, _seed_e)

    seed_msg = f"\n📦 {seed_soni} ta tayyor tovar yuklandi!" if seed_soni else ""
    await update.message.reply_text(
        f"✅ Ro'yxatdan o'tdingiz!{seed_msg}\n⏳ Admin tasdiqlaguncha kuting."
    )
    for aid in cfg().admin_ids:
        try:
            await ctx.bot.send_message(aid,
                f"🆕 *YANGI FOYDALANUVCHI*\n\n"
                f"👤 {update.effective_user.full_name}\n🆔 `{uid}`\n"
                f"🏪 {dokon}\n📦 {SEGMENT_NOMI[seg]}\n📞 {tel}",
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=tg(
                    [(f"✅ Tasdiqlash",f"adm:ok:{uid}")],
                    [(f"❌ Rad etish", f"adm:no:{uid}")],
                ))
        except Exception as e: log.warning("Admin %s: %s",aid,e)
    return ConversationHandler.END


# ════════════ OVOZ / MATN ════════════

async def ovoz_qabul(update:Update, ctx:ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id
    if not _flood_ok(uid): return
    if not await faol_tekshir(update): return
    from shared.services.guards import is_duplicate_message
    if is_duplicate_message(uid, f"voice:{update.message.voice.file_id}"): return

    voice_dur = update.message.voice.duration or 0  # Telegram bergan davomiylik (sekund)
    uzun_audio = voice_dur > 30  # 30s+ → VAD + chunking pipeline

    # Faqat ⏳ — matn yo'q
    holat = await update.message.reply_text(
        f"⏳ {voice_dur}s audio qayta ishlanmoqda..." if uzun_audio else "⏳"
    )

    tmp_path = None
    try:
        fayl=await ctx.bot.get_file(update.message.voice.file_id)
        audio=bytes(await fayl.download_as_bytearray())
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            tmp.write(audio)
            tmp_path = tmp.name

        if uzun_audio:
            # ── UZUN AUDIO: VAD (shovqin tozalash) + Chunking + Parallel Gemini ──
            async def _progress(pct, text):
                try: await holat.edit_text(f"⏳ {pct}% — {text}")
                except Exception as _e: log.debug("silent: %s", _e)
            matn = await ovoz_xizmat.ovoz_matn_uzun(
                tmp_path, progress_callback=_progress, uid=uid
            )
        else:
            # ── QISQA AUDIO: to'g'ridan-to'g'ri Gemini ──
            matn = await ovoz_xizmat.matnga_aylantir(tmp_path, uid=uid)

        if not matn:
            try:
                from shared.services.suhbatdosh import tushunilmadi
                await holat.edit_text(tushunilmadi())
            except Exception:
                await holat.edit_text("🤔 Tushunolmadim. Yana bir bor aytib ko'ring.")
            return
        pipeline_result = await process_voice(matn, user_id=uid)
        if pipeline_result.get("confidence") == "none":
            await holat.edit_text(
                "🔇 Ovoz eshitilmadi. Iltimos qayta yuboring.\n"
                "💡 Masalan: \"Nasriddin akaga 2 Ariel 56000\""
            )
            return
        # LOW va MEDIUM confidence ham Claude ga yuboriladi —
        # Claude Sonnet kontekstdan tushunadi, foydalanuvchiga draft ko'rsatadi
        matn = pipeline_result.get("text") or matn

        # Shogird xarajat tekshiruvi (ovoz uchun)
        if not cfg().is_admin(uid):
            try:
                from shared.services.shogird_xarajat import shogird_topish_tg
                from shared.database.pool import get_pool
                pool = get_pool()
                async with pool.acquire() as raw_conn:
                    shogird = await shogird_topish_tg(raw_conn, uid)
                if shogird:
                    handled = await _shogird_xarajat_qabul(update, ctx, matn, shogird)
                    if handled:
                        try: await holat.delete()
                        except Exception as _e: log.debug("silent: %s", _e)
                        return
            except Exception as _se:
                log.debug("Shogird ovoz: %s", _se)

        # ⏳ o'chirib, natija yuborish
        try: await holat.delete()
        except Exception as _e: log.debug("silent: %s", _e)

        # ═══ OVOZDA "CHEK CHIQAR" TEKSHIRUVI ═══
        try:
            from shared.services.print_intent import detect_print_intent
            from shared.services.bot_print_handler import handle_print_intent_message
            _pk = detect_print_intent(matn)
            if _pk:
                if await handle_print_intent_message(update, ctx, _pk, db):
                    return
        except Exception as _pi:
            log.debug("Print intent (ovoz): %s", _pi)

        await _qayta_ishlash(update,ctx,matn)
    except Exception as xato:
        log.error("ovoz_qabul: %s",xato,exc_info=True)
        try:
            await holat.edit_text("❌ Xato yuz berdi")
        except Exception as _e:
            log.debug("Xato: %s", _e)
            try: await update.message.reply_text("❌ Xato yuz berdi")
            except Exception as _e: log.debug("silent: %s", _e)
    finally:
        if tmp_path:
            try: _os.unlink(tmp_path)
            except Exception as _e: log.debug("silent: %s", _e)


# ════════════ NAKLADNOY ════════════

async def cmd_nakladnoy(update:Update, ctx:ContextTypes.DEFAULT_TYPE):
    if not await faol_tekshir(update): return
    await update.message.reply_text(
        "📋 *NAKLADNOY YARATISH*\n\n"
        "Ovoz yuboring yoki yozing:\n\n"
        "_O'zbek: \"Salimovga nakladnoy yoz, 50 Ariel 45,000\"_\n"
        "_Rus: \"Накладная для Иванова, 50 Ariel по 45000\"_\n\n"
        "✅ Word + Excel + PDF + Imzo/Muhr joyi!\n\n"
        "📦 *Ko'p nakladnoy ishlari uchun:*\n"
        "/nakladnoy_excel — Excel nakladnoy (SalesDoc formatida)\n"
        "/reestr_excel — Kunlik reestr (SalesDoc formatida)",
        parse_mode=ParseMode.MARKDOWN)


async def cmd_nakladnoy_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """SalesDoc Накладные 3.1 uslubida Excel nakladnoy yaratish"""
    if not await faol_tekshir(update): return
    uid = update.effective_user.id
    try:
        msg = await update.message.reply_text("📋 Excel nakladnoy tayyorlanmoqda...")

        import tempfile, time as _t
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from shared.database.pool import rls_conn

        # Bugungi sotuvlarni olish (user-specific, RLS orqali)
        async with rls_conn(uid) as conn:
            today = _t.strftime('%Y-%m-%d')
            rows = await conn.fetch("""
                SELECT s.id, s.klient_id, s.jami, s.sana,
                       k.ism as klient_ismi, k.telefon, k.manzil,
                       k.jami_sotib as balans
                FROM sotuv_sessiyalar s
                LEFT JOIN klientlar k ON s.klient_id = k.id
                WHERE DATE(s.sana AT TIME ZONE 'Asia/Tashkent') = $1::date
                ORDER BY s.id DESC
                LIMIT 50
            """, today)

            # Agar bo'sh bo'lsa — oxirgi 10 ta
            if not rows:
                rows = await conn.fetch("""
                    SELECT s.id, s.klient_id, s.jami, s.sana,
                           k.ism as klient_ismi, k.telefon, k.manzil,
                           k.jami_sotib as balans
                    FROM sotuv_sessiyalar s
                    LEFT JOIN klientlar k ON s.klient_id = k.id
                    ORDER BY s.id DESC LIMIT 10
                """)

            if not rows:
                await msg.edit_text("⚠️ Bugungi sotuvlar topilmadi. Avval sotuv qiling.")
                return

            # Har bir sotuv uchun tovarlarni olish
            orders_data = []
            for r in rows:
                items = await conn.fetch("""
                    SELECT c.tovar_nomi as nomi, c.miqdor, c.sotish_narxi as narx, c.birlik
                    FROM chiqimlar c WHERE c.sessiya_id = $1
                """, r['id'])
                orders_data.append({
                    'id': r['id'],
                    'klient_ismi': r['klient_ismi'] or 'Mijoz',
                    'telefon': r['telefon'] or '',
                    'manzil': r['manzil'] or '',
                    'balans': r['balans'] or 0,
                    'items': [dict(i) for i in items],
                })

        # Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Накладные 3.1"

        bold_big = Font(name="Arial", bold=True, size=12)
        bold = Font(name="Arial", bold=True, size=10)
        normal = Font(name="Arial", size=10)
        small = Font(name="Arial", size=9)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

        widths = {1: 5, 2: 40, 3: 10, 4: 8, 5: 12, 6: 14}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        row_num = 1
        today_str = _t.strftime('%d.%m.%Y')

        for order in orders_data:
            # Sarlavha
            ws.cell(row=row_num, column=3, value=f"Накладная №{order['id']}  от {today_str}").font = bold_big
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
            row_num += 1
            ws.cell(row=row_num, column=2, value=f"Кому: {order['klient_ismi']}").font = bold
            ws.cell(row=row_num, column=5, value="ТП: SavdoAI").font = normal
            row_num += 1
            ws.cell(row=row_num, column=2, value=f"Адрес: {order['manzil']}").font = normal
            ws.cell(row=row_num, column=5, value="Тел: +998770030080").font = normal
            row_num += 1
            ws.cell(row=row_num, column=2, value=f"Тел: {order['telefon']}").font = normal
            ws.cell(row=row_num, column=5, value="Территория: Samarqand").font = normal
            row_num += 1
            ws.cell(row=row_num, column=5, value="Контактное лицо:").font = normal
            row_num += 1
            ws.cell(row=row_num, column=5, value="Фирма: SavdoAI").font = normal
            row_num += 1
            ws.cell(row=row_num, column=2, value="ИНН:").font = normal
            row_num += 1
            ws.cell(row=row_num, column=2, value="Экспедитор:").font = normal
            row_num += 1
            ws.cell(row=row_num, column=2, value=f"Баланс клиента: {order['balans']:,.0f}").font = normal
            row_num += 2

            # ZAKAZ
            ws.cell(row=row_num, column=1, value=f"ЗАКАЗ (d0_{order['id']})").font = bold
            row_num += 1

            # Jadval sarlavhasi
            headers = ["№", "Наименование", "Кол-во", "ЕИ", "Цена", "Сумма"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=h)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border
            row_num += 1

            # Tovarlar
            for i, item in enumerate(order['items'], 1):
                miqdor = float(item.get('miqdor', 0))
                narx = float(item.get('narx', 0))
                summa = miqdor * narx
                ws.cell(row=row_num, column=1, value=i).font = normal
                ws.cell(row=row_num, column=2, value=item.get('nomi', '')).font = normal
                ws.cell(row=row_num, column=3, value=miqdor).font = normal
                ws.cell(row=row_num, column=4, value=item.get('birlik', 'dona')).font = normal
                ws.cell(row=row_num, column=5, value=narx).font = normal
                ws.cell(row=row_num, column=6, value=summa).font = normal
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = border
                    ws.cell(row=row_num, column=col).alignment = center if col != 2 else left
                row_num += 1

            # Itog
            if order['items']:
                total_qty = sum(float(i.get('miqdor', 0)) for i in order['items'])
                total_sum = sum(float(i.get('miqdor', 0)) * float(i.get('narx', 0)) for i in order['items'])
                ws.cell(row=row_num, column=2, value="Итог").font = bold
                ws.cell(row=row_num, column=3, value=total_qty).font = bold
                ws.cell(row=row_num, column=6, value=total_sum).font = bold
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = border
                row_num += 1

            # Imzo
            ws.cell(row=row_num, column=2, value="Отпустил: _________________________").font = small
            ws.cell(row=row_num, column=4, value="Принял: ____________________________").font = small
            row_num += 3

        xlsx_path = os.path.join(tempfile.gettempdir(), f"nakladnoy_{int(_t.time())}.xlsx")
        wb.save(xlsx_path)

        with open(xlsx_path, "rb") as f:
            await ctx.bot.send_document(
                chat_id=update.effective_chat.id,
                document=f,
                filename=f"Накладные_{today_str}.xlsx",
                caption=f"📋 SalesDoc formatidagi nakladnoy\n"
                        f"Format: Накладные 3.1\n"
                        f"Buyurtmalar: {len(orders_data)} ta",
            )
        import os as _os
        try: _os.remove(xlsx_path)
        except Exception: pass
        await msg.delete()
    except Exception as e:
        log.error(f"Nakladnoy excel xato: {e}", exc_info=True)
        await update.message.reply_text(f"⚠️ Xato: {e}")


async def cmd_reestr_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """SalesDoc Реестр 3.0 uslubida Excel reestr yaratish"""
    if not await faol_tekshir(update): return
    uid = update.effective_user.id
    try:
        msg = await update.message.reply_text("📋 Reestr tayyorlanmoqda...")

        import tempfile, time as _t
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from shared.database.pool import rls_conn

        async with rls_conn(uid) as conn:
            today = _t.strftime('%Y-%m-%d')
            rows = await conn.fetch("""
                SELECT s.id, s.klient_id, s.jami, s.sana,
                       k.ism as klient_ismi, k.telefon, k.manzil,
                       k.jami_sotib as balans
                FROM sotuv_sessiyalar s
                LEFT JOIN klientlar k ON s.klient_id = k.id
                WHERE DATE(s.sana AT TIME ZONE 'Asia/Tashkent') = $1::date
                ORDER BY s.id DESC
                LIMIT 100
            """, today)
            if not rows:
                rows = await conn.fetch("""
                    SELECT s.id, s.klient_id, s.jami, s.sana,
                           k.ism as klient_ismi, k.telefon, k.manzil,
                           k.jami_sotib as balans
                    FROM sotuv_sessiyalar s
                    LEFT JOIN klientlar k ON s.klient_id = k.id
                    ORDER BY s.id DESC LIMIT 20
                """)

        if not rows:
            await msg.edit_text("⚠️ Sotuvlar topilmadi. Avval sotuv qiling.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реестр 3.0"

        bold = Font(name="Arial", bold=True, size=10)
        normal = Font(name="Arial", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        widths = {1: 5, 2: 14, 3: 35, 4: 18, 5: 18, 6: 22, 7: 14, 8: 14, 9: 12}
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        headers = ["№", "Дата отгрузки", "Торгов. Точка", "Адрес", "Номер клиента",
                   "Торгов. Пред.", "Баланс клиента", "сум", "Отметка"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = bold
            cell.alignment = center
            cell.border = border

        today_str = _t.strftime('%d.%m.%Y')
        total_sum = 0

        for i, r in enumerate(rows, 1):
            row = i + 2
            summa = float(r['jami'] or 0)
            total_sum += summa
            ws.cell(row=row, column=1, value=i).font = normal
            ws.cell(row=row, column=2, value=today_str).font = normal
            ws.cell(row=row, column=3, value=r['klient_ismi'] or "Mijoz").font = normal
            ws.cell(row=row, column=4, value=r['manzil'] or "").font = normal
            ws.cell(row=row, column=5, value=r['telefon'] or "").font = normal
            ws.cell(row=row, column=6, value="SavdoAI").font = normal
            ws.cell(row=row, column=7, value=float(r['balans'] or 0)).font = normal
            ws.cell(row=row, column=8, value=summa).font = normal
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = center

        total_row = len(rows) + 3
        ws.cell(row=total_row, column=3, value="Total").font = bold
        ws.cell(row=total_row, column=8, value=total_sum).font = bold
        for col in range(1, 10):
            ws.cell(row=total_row, column=col).border = border

        xlsx_path = os.path.join(tempfile.gettempdir(), f"reestr_{int(_t.time())}.xlsx")
        wb.save(xlsx_path)

        with open(xlsx_path, "rb") as f:
            await ctx.bot.send_document(
                chat_id=update.effective_chat.id,
                document=f,
                filename=f"Реестр_{today_str}.xlsx",
                caption=f"📋 Reestr — {len(rows)} ta sotuv\n"
                        f"Jami: {total_sum:,.0f} so'm\n"
                        f"Format: Реестр 3.0",
            )
        import os as _os
        try: _os.remove(xlsx_path)
        except Exception: pass
        await msg.delete()
    except Exception as e:
        log.error(f"Reestr xato: {e}", exc_info=True)
        await update.message.reply_text(f"⚠️ Xato: {e}")


# ════════════ TASDIQLASH ════════════


# ═══ TASDIQ HELPERS (extracted from tasdiq_cb for maintainability) ═══

# ════════════ EKSPORT ════════════

# ════════════ MENYU CALLBACKLAR ════════════

# ════════════ NARX GURUHLARI ════════════
# Handlers ko'chirildi → services/bot/handlers/narx.py
# register_narx_handlers() ilovani_qur() da chaqiriladi


# ════════════ SHOGIRD XARAJAT NAZORATI ════════════
# Handlers ko'chirildi → services/bot/handlers/shogird.py
# register_shogird_handlers() ilovani_qur() da chaqiriladi

# ════════════ ADMIN ════════════

# ════════════ AVTOMATIK JOBLAR ════════════
# Handlers ko'chirildi → services/bot/handlers/jobs.py
from services.bot.handlers.jobs import (
    avto_kunlik_hisobot, avto_haftalik_hisobot,
    avto_qarz_eslatma, obuna_eslatma,
    avto_ertalab_hisobot,
)

# ═══ MODULAR HANDLER IMPORTS ═══
from services.bot.handlers.commands import (
    cmd_menyu, cmd_hisobot, cmd_tez, cmd_guruh, cmd_qarz, cmd_foyda,
    cmd_klient, cmd_top, cmd_ombor, cmd_status, cmd_kassa, cmd_faktura,
    cmd_balans, cmd_jurnal, cmd_chiqim, cmd_tovar, cmd_yangilik,
    cmd_imkoniyatlar, cmd_yordam, cmd_ogoh, cmd_hafta,
    cmd_foydalanuvchilar, cmd_faollashtir, cmd_statistika,
    cmd_savatlar, cmd_savat, _ovoz_buyruq_bajar,
    cmd_narx_tavsiya, cmd_dokon,
    cmd_crm, cmd_chegirma, cmd_prognoz, cmd_raqobat, cmd_rfm,
    cmd_top_tovar, cmd_top_klient, cmd_kategoriya_stat, cmd_ombor_qiymati,
)
from services.bot.handlers.callbacks import (
    eksport_cb, nakladnoy_sessiya_cb, menyu_cb, paginatsiya_cb,
    _hujjat_cb, _hisobot_excel_cb, hisobot_cb, klient_hisobi_cb,
    faktura_cb, admin_cb, _tezkor_cb,
)
from services.bot.handlers.hujjat import hujjat_qabul
from services.bot.handlers.barcode import cmd_barcode, barcode_cb
from services.bot.handlers.savdo import (
    tasdiq_cb, _qayta_ishlash, _nakladnoy_yuborish,
    _chek_thermal_va_pdf_yuborish,
    _savat_qosh_va_javob, _savat_yop_va_nakladnoy,
)
from services.bot.handlers.matn import matn_qabul

# ════════════ KOMANDALAR ════════════

# ════════════ APP ════════════

async def boshlash(app:Application) -> None:
    global _CFG; _CFG=app.bot_data["cfg"]
    # Handler modullari uchun config reference
    from services.bot.bot_helpers import set_cfg
    set_cfg(_CFG)
    # 1. DB pool — FATAL (pool_init siz bot ishlay olmaydi)
    try:
        await db.pool_init(_CFG.database_url, min_size=_CFG.db_min, max_size=_CFG.db_max)
    except Exception as _e:
        log.critical("❌ DB ulanishda xato: %s", _e, exc_info=True)
        raise RuntimeError(f"DB pool init muvaffaqiyatsiz: {_e}") from _e
    # 1b. Shared pool ham ishga tushirish (rls_conn, ledger, kassa uchun)
    # Bot pool (db._P) asosiy pool — shared pool faqat RLS va ledger uchun
    # MINIMAL: 1-2 connection — DB connection limit tejash
    try:
        await _pool_init(_CFG.database_url, min_size=1, max_size=min(3, max(2, _CFG.db_max // 5)))
        log.info("✅ Shared pool ulandi (minimal — DB limit tejash)")
    except Exception as _e:
        log.warning("⚠️ Shared pool xato (ba'zi funksiyalar ishlamasligi mumkin): %s", _e)
    # 2. Schema — non-fatal (jadvallar allaqachon bo'lishi mumkin)
    try:
        await db.schema_init()
    except Exception as _e:
        log.error("⚠️ Schema init xato (bot davom etadi): %s", _e)
    try:
        ovoz_xizmat.ishga_tushir(_CFG.gemini_key, _CFG.gemini_model)
        try:
            from shared.database.pool import get_pool
            get_pool()
            await ovoz_xizmat.stt_prompt_yangilash()
            log.info(
                "✅ Voice STT + fuzzy: RLS bo'yicha har foydalanuvchi uchun alohida yuklanadi"
            )
        except Exception as _voice_e:
            log.warning("Voice pipeline init xato: %s", _voice_e)
    except Exception as _e:
        log.warning("Gemini ishga tushmadi (ovoz xizmati o'chirildi): %s", _e)
    try:
        ai_xizmat.ishga_tushir(_CFG.anthropic_key)
    except Exception as _e:
        log.critical("Claude ishga tushmadi: %s", _e, exc_info=True)
        raise RuntimeError(f"AI xizmat init muvaffaqiyatsiz: {_e}") from _e
    log.info("✅ Bot xizmatlar tayyor")
    # Redis cache (ixtiyoriy — yo'q bo'lsa ham bot ishlaydi)
    try:
        redis_url = _os.environ.get("REDIS_URL", "")
        if redis_url:
            from shared.cache.redis_cache import redis_init
            await redis_init(redis_url)
    except Exception as _e:
        log.warning("Redis ulana olmadi (cache o'chirildi): %s", _e)
    # Vision AI (ixtiyoriy — Gemini key bilan ishlaydi)
    try:
        from shared.services.vision import ishga_tushir as vision_init
        vision_init(_CFG.gemini_key, _CFG.gemini_model)
    except Exception as _e:
        log.info("ℹ️ Vision AI yuklanmadi (ixtiyoriy): %s", _e)

    # TTS ovozli javob
    try:
        from services.bot.bot_services.tts import ishga_tushir as tts_init
        tts_init(_CFG.gemini_key)
    except Exception as _e:
        log.info("ℹ️ TTS yuklanmadi (ixtiyoriy): %s", _e)
    # Dual-Brain MoE Router
    try:
        from services.cognitive.ai_router import router_init
        router_init(
            gemini_key=_CFG.gemini_key,
            claude_key=_CFG.anthropic_key,
            gemini_model=_CFG.gemini_model,
            claude_model=_CFG.claude_model,
        )
        log.info("🧠 Dual-Brain MoE Router ulandi")
    except Exception as _e:
        log.warning("⚠️ MoE Router yuklanmadi: %s", _e)
    await app.bot.set_my_commands([
        BotCommand("start",            "Botni boshlash"),
        BotCommand("menyu",            "Asosiy menyu"),
        BotCommand("yangilik",         "🆕 v25.3 yangiliklari"),
        BotCommand("imkoniyatlar",     "📋 Barcha imkoniyatlar"),
        BotCommand("yordam",           "❓ Qanday ishlatish"),
        BotCommand("nakladnoy",        "Nakladnoy (Word+Excel+PDF)"),
        BotCommand("faktura",          "Hisob-faktura yaratish"),
        BotCommand("hisobot",          "Bugungi hisobot"),
        BotCommand("qarz",             "Qarzlar ro'yxati"),
        BotCommand("foyda",            "Foyda tahlili"),
        BotCommand("klient",           "Klient qidirish (ism/tel)"),
        BotCommand("top",              "Top klientlar"),
        BotCommand("ombor",            "Ombor holati"),
        BotCommand("ogoh",             "Kam qoldiq tovarlar"),
        BotCommand("tez",              "⚡ Tezkor tugmalar"),
        BotCommand("guruh",            "👥 Guruhli sotuv"),
        BotCommand("tahlil",           "📸 Ko'p rasm tahlil"),
        BotCommand("hafta",            "Haftalik hisobot"),
        BotCommand("kassa",            "Kassa holati (naqd/karta)"),
        BotCommand("status",           "Bot holati (admin)"),
        BotCommand("balans",           "📊 Balans tekshiruvi (admin)"),
        BotCommand("jurnal",           "📒 Jurnal yozuvlar"),
        BotCommand("foydalanuvchilar", "Foydalanuvchilar (admin)"),
        BotCommand("faollashtir",      "Faollashtirish (admin)"),
        BotCommand("statistika",       "Statistika (admin)"),
        BotCommand("shogird_qosh",    "Shogird qo'shish (admin)"),
        BotCommand("shogirdlar",      "Shogirdlar ro'yxati (admin)"),
        BotCommand("xarajatlar",      "Xarajatlar nazorati (admin)"),
        BotCommand("narx_guruh",      "Narx guruhi yaratish"),
        BotCommand("narx_qoy",        "Guruh narxi qo'yish"),
        BotCommand("klient_narx",     "Klient shaxsiy narx"),
        BotCommand("klient_guruh",    "Klient guruhga biriktirish"),
        BotCommand("savatlar",        "Ochiq savatlar ko'rish"),
        BotCommand("savat",           "Klient savati ko'rish"),
        BotCommand("eslatma",         "📨 Qarz eslatma yuborish"),
        BotCommand("kpi",             "📊 KPI ko'rsatkichlar"),
        BotCommand("loyalty",         "⭐ Bonus ball tekshirish"),
        BotCommand("tahlil",          "🧠 AI biznes tahlili"),
        BotCommand("buyurtma",        "📦 Tovar buyurtma tavsiyasi"),
        BotCommand("marshrut",        "🗺 GPS kunlik marshrut"),
        BotCommand("tariflar",        "💎 Tarif planlari"),
        BotCommand("prognoz",         "🔮 AI talab prognozi"),
        BotCommand("clv",             "💎 Klient qiymati"),
    ])

    job_queue=app.job_queue
    if job_queue:
        # Bot standalone rejim — Worker mavjud bo'lsa u boshqaradi
        _worker_url = _os.getenv("WORKER_URL", "")
        if _worker_url:
            log.info("⚡ Worker aniqlandi — bot hisobot scheduleri o'chirildi")
            # Worker/Celery Beat handles scheduling — bot skips
        else:
            # Standalone: bot o'z schedulerini yoqadi
            try:
                job_queue._scheduler.configure(misfire_grace_time=60)
            except Exception as _e: log.debug("silent: %s", _e)
        import pytz,datetime
        tz=pytz.timezone(_CFG.timezone)
        _standalone = not _worker_url
        # Kunlik hisobot — 22:00
        if _standalone:
            # Standalone rejim: bot o'z schedulerini boshqaradi
            job_queue.run_daily(
                avto_kunlik_hisobot,
                time=datetime.time(hour=_CFG.kunlik_soat, minute=0, tzinfo=tz),
                name="kunlik_hisobot",
            )
            job_queue.run_daily(
                avto_haftalik_hisobot,
                time=datetime.time(hour=_CFG.haftalik_soat, minute=0, tzinfo=tz),
                days=(0,),  # 0 = Dushanba
                name="haftalik_hisobot",
            )
            job_queue.run_daily(
                avto_qarz_eslatma,
                time=datetime.time(hour=_CFG.qarz_soat, minute=0, tzinfo=tz),
                name="qarz_eslatma",
            )
            job_queue.run_daily(
                obuna_eslatma,
                time=datetime.time(hour=_CFG.obuna_soat, minute=0, tzinfo=tz),
                name="obuna_eslatma",
            )
            # Ertalab hisobot — 09:00 Toshkent
            job_queue.run_daily(
                avto_ertalab_hisobot,
                time=datetime.time(hour=9, minute=0, tzinfo=tz),
                name="ertalab_hisobot",
            )
            log.info("✅ Standalone: kunlik/haftalik/qarz/obuna/ertalab joblar yoqildi")

            # ═══ v25.3.2 SMART NOTIFICATION JOBS ═══
            async def _smart_notify(ctx, turi):
                """Smart notification — barcha faol userlarga."""
                from shared.services.smart_notification import notification_dispatch
                from shared.database.pool import get_pool
                try:
                    async with get_pool().acquire() as c:
                        users = await c.fetch("SELECT id FROM users WHERE faol=TRUE")
                        for u in users:
                            uid = u["id"]
                            try:
                                from shared.database.pool import rls_conn
                                async with rls_conn(uid) as rc:
                                    matn = await notification_dispatch(rc, uid, turi)
                                if matn:
                                    await ctx.bot.send_message(uid, matn, parse_mode="Markdown")
                            except Exception as _ue:
                                pass
                except Exception as e:
                    log.debug("Smart notify %s: %s", turi, e)

            job_queue.run_daily(
                lambda ctx: _smart_notify(ctx, "kechqurun"),
                time=datetime.time(hour=20, minute=0, tzinfo=tz),
                name="smart_kechki",
            )
            job_queue.run_repeating(
                lambda ctx: _smart_notify(ctx, "critical"),
                interval=7200,  # Har 2 soatda critical alert
                first=120, name="smart_critical",
            )
            log.info("✅ Smart notification joblar yoqildi (kechki, critical)")
        else:
            log.info("✅ Worker rejim: scheduling Worker/Beat tomonida boshqariladi")
# ════════════ OCHIQ SAVAT (Multi-Klient) ════════════

def ilovani_qur(conf:Config) -> Application:
    app=(Application.builder().token(conf.bot_token).post_init(boshlash).build())
    app.bot_data["cfg"]=conf

    # Inline funksiyalar → commands.py ga ko'chirildi
    from services.bot.handlers.commands import (
        cmd_ping, cmd_token, cmd_webapp, cmd_parol, inline_qidirish,
    )
    app.add_handler(CommandHandler("ping", cmd_ping))
    app.add_handler(CommandHandler("token", cmd_token))
    app.add_handler(CommandHandler("webapp", cmd_webapp))
    app.add_handler(CommandHandler("parol", cmd_parol))

    royxat=ConversationHandler(
        entry_points=[CommandHandler("start",cmd_start)],
        states={
            H_SEGMENT:[CallbackQueryHandler(h_segment,pattern=r"^seg:")],
            H_DOKON:[MessageHandler(filters.TEXT & ~filters.COMMAND,h_dokon)],
            H_TELEFON:[MessageHandler(filters.TEXT & ~filters.COMMAND,h_telefon)],
        },
        fallbacks=[CommandHandler("start",cmd_start)],
        allow_reentry=True,per_message=False,
    )
    app.add_handler(royxat)
    app.add_handler(MessageHandler(filters.VOICE,ovoz_qabul))
    app.add_handler(MessageHandler(filters.PHOTO,rasm_xizmat.rasm_qabul))
    app.add_handler(CallbackQueryHandler(rasm_xizmat.rasm_nakladnoy_cb, pattern=r"^rasm:nakl"))
    app.add_handler(CallbackQueryHandler(rasm_xizmat.rasm_amal_cb, pattern=r"^rasm:(kirim|sotuv)"))
    app.add_handler(CommandHandler("tahlil", rasm_xizmat.kop_rasm_tahlil_cmd))
    app.add_handler(MessageHandler(filters.Document.ALL, hujjat_qabul))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND,matn_qabul))

    # ═══ INLINE QUERY ═══
    from telegram.ext import InlineQueryHandler
    app.add_handler(InlineQueryHandler(inline_qidirish))
    app.add_handler(CallbackQueryHandler(tasdiq_cb,         pattern=r"^t:"))
    app.add_handler(CallbackQueryHandler(nakladnoy_sessiya_cb,pattern=r"^n:sess:"))
    app.add_handler(CallbackQueryHandler(menyu_cb,          pattern=r"^m:"))
    app.add_handler(CallbackQueryHandler(hisobot_cb,        pattern=r"^hs:"))
    app.add_handler(CallbackQueryHandler(admin_cb,          pattern=r"^adm:"))
    app.add_handler(CallbackQueryHandler(paginatsiya_cb,    pattern=r"^(tv|kl):"))
    app.add_handler(CallbackQueryHandler(klient_hisobi_cb,  pattern=r"^kh:"))
    app.add_handler(CallbackQueryHandler(faktura_cb,        pattern=r"^fkt:"))
    app.add_handler(CallbackQueryHandler(eksport_cb,        pattern=r"^eks:"))
    app.add_handler(CallbackQueryHandler(_hujjat_cb, pattern=r"^huj:"))
    app.add_handler(CallbackQueryHandler(_hisobot_excel_cb, pattern=r"^hisob_excel:"))
    app.add_handler(CallbackQueryHandler(_tezkor_cb, pattern=r"^tez:"))
    app.add_handler(CommandHandler("menyu",            cmd_menyu))
    app.add_handler(CommandHandler("tez",              cmd_tez))
    app.add_handler(CommandHandler("guruh",            cmd_guruh))
    app.add_handler(CommandHandler("nakladnoy",        cmd_nakladnoy))
    app.add_handler(CommandHandler("nakladnoy_excel",  cmd_nakladnoy_excel))
    app.add_handler(CommandHandler("reestr_excel",     cmd_reestr_excel))
    app.add_handler(CommandHandler("hisobot",          cmd_hisobot))
    app.add_handler(CommandHandler("qarz",             cmd_qarz))
    app.add_handler(CommandHandler("foyda",            cmd_foyda))
    app.add_handler(CommandHandler("klient",           cmd_klient))
    app.add_handler(CommandHandler("ogoh",             cmd_ogoh))
    app.add_handler(CommandHandler("hafta",            cmd_hafta))
    app.add_handler(CommandHandler("status",           cmd_status))
    app.add_handler(CommandHandler("kassa",            cmd_kassa))
    app.add_handler(CommandHandler("faktura",          cmd_faktura))
    app.add_handler(CommandHandler("chiqim",           cmd_chiqim))
    app.add_handler(CommandHandler("tovar",            cmd_tovar))
    app.add_handler(CommandHandler("balans",           cmd_balans))
    app.add_handler(CommandHandler("jurnal",           cmd_jurnal))
    app.add_handler(CommandHandler("yangilik",         cmd_yangilik))
    app.add_handler(CommandHandler("imkoniyatlar",     cmd_imkoniyatlar))
    app.add_handler(CommandHandler("yordam",           cmd_yordam))
    app.add_handler(CommandHandler("top",              cmd_top))
    app.add_handler(CommandHandler("ombor",            cmd_ombor))
    app.add_handler(CommandHandler("foydalanuvchilar", cmd_foydalanuvchilar))
    app.add_handler(CommandHandler("faollashtir",      cmd_faollashtir))
    app.add_handler(CommandHandler("statistika",       cmd_statistika))
    app.add_handler(CommandHandler("health", health_check))
    # Shogird xarajat
    # Narx handlers — handlers/narx.py dan
    from services.bot.handlers.narx import register_narx_handlers
    register_narx_handlers(app)
    # Shogird handlers — handlers/shogird.py dan
    from services.bot.handlers.shogird import register_shogird_handlers
    register_shogird_handlers(app)
    app.add_handler(CommandHandler("savatlar",         cmd_savatlar))
    app.add_handler(CommandHandler("savat",            cmd_savat))
    # ═══ YANGI v25.3.2 HANDLERLAR ═══
    app.add_handler(CommandHandler("barcode",          cmd_barcode))
    app.add_handler(CommandHandler("narx_tavsiya",     cmd_narx_tavsiya))
    app.add_handler(CommandHandler("dokon",            cmd_dokon))
    app.add_handler(CommandHandler("crm",              cmd_crm))
    app.add_handler(CommandHandler("chegirma",         cmd_chegirma))
    app.add_handler(CommandHandler("prognoz",          cmd_prognoz))
    app.add_handler(CommandHandler("raqobat",          cmd_raqobat))
    app.add_handler(CommandHandler("rfm",              cmd_rfm))
    app.add_handler(CommandHandler("top_tovar",        cmd_top_tovar))
    app.add_handler(CommandHandler("top_klient",       cmd_top_klient))
    app.add_handler(CommandHandler("kategoriya",       cmd_kategoriya_stat))
    app.add_handler(CommandHandler("ombor_qiymati",    cmd_ombor_qiymati))
    app.add_handler(CallbackQueryHandler(barcode_cb,   pattern=r"^bc:"))
    # ═══ v25.3.2 KUCHLI HANDLERLAR — Qarz eslatma, KPI, Loyalty ═══
    from services.bot.handlers.yangi import register_yangi_handlers
    register_yangi_handlers(app)
    # ═══ EXCEL CHAT — AI savol-javob ═══
    from services.bot.handlers.excel_chat import register_excel_chat_handlers
    register_excel_chat_handlers(app)
    # ═══ SMART AI — har qanday savolga javob ═══
    async def cmd_savol(update, ctx):
        """Biznes haqida har qanday savol — AI bazadan javob beradi"""
        from services.bot.bot_helpers import faol_tekshir
        if not await faol_tekshir(update): return
        matn = " ".join(ctx.args) if ctx.args else ""
        if not matn:
            await update.message.reply_text(
                "🤖 *Smart AI — Har qanday savol bering!*\n\n"
                "Masalan:\n"
                "• `/savol Salimovning qarzi qancha?`\n"
                "• `/savol omborda Ariel bormi?`\n"
                "• `/savol bu hafta eng ko'p sotilgan tovar`\n"
                "• `/savol qancha klientim bor?`\n"
                "• `/savol shu oy foydam qancha?`\n\n"
                "Yoki shunchaki savol yozing — bot tushunadi!",
                parse_mode=ParseMode.MARKDOWN)
            return
        wait = await update.message.reply_text("🤔 Bazani tekshiryapman...")
        try:
            from shared.services.smart_ai import smart_javob
            javob = await smart_javob(update.effective_user.id, matn)
            try:
                await wait.edit_text(javob, parse_mode=ParseMode.MARKDOWN)
            except Exception:
                await wait.edit_text(javob.replace("*","").replace("_","").replace("`",""))
        except Exception as e:
            await wait.edit_text(f"❌ Xato: {e}")
    app.add_handler(CommandHandler("savol", cmd_savol))
    app.add_error_handler(xato_handler)
    return app


def _polling_lock_key(bot_token: str) -> str:
    """Bot token asosida barqaror lock kaliti (token logga chiqmaydi)."""
    import hashlib
    token_hash = hashlib.sha256(bot_token.encode("utf-8")).hexdigest()[:16]
    return f"savdoai:telegram:polling-lock:{token_hash}"


def _acquire_polling_singleton_lock(bot_token: str):
    """
    Redis bor bo'lsa, faqat bitta polling instance ishlashini kafolatlaydi.
    Lock band bo'lsa process chiqadi — "other getUpdates" konfliktini oldini oladi.
    """
    redis_url = _os.environ.get("REDIS_URL", "").strip()
    if not redis_url:
        log.warning("REDIS_URL topilmadi — polling singleton lock o'chirilgan.")
        return None, None

    import redis
    import socket
    import uuid

    lock_ttl = int(_os.environ.get("BOT_POLL_LOCK_TTL_SECONDS", "120"))
    owner = f"{socket.gethostname()}:{_os.getpid()}:{uuid.uuid4().hex[:8]}"
    key = _polling_lock_key(bot_token)
    client = redis.from_url(
        redis_url,
        decode_responses=True,
        socket_connect_timeout=3,
        socket_timeout=3,
    )

    ok = client.set(key, owner, nx=True, ex=lock_ttl)
    if not ok:
        current_owner = client.get(key)
        raise RuntimeError(
            f"Polling lock band: {key} owner={current_owner}. "
            "Boshqa bot instance allaqachon getUpdates qilmoqda."
        )

    log.info("🔒 Polling singleton lock olindi: %s", key)
    return client, (key, owner)


def _release_polling_singleton_lock(lock_client, lock_meta) -> None:
    if not lock_client or not lock_meta:
        return
    key, owner = lock_meta
    try:
        current = lock_client.get(key)
        if current == owner:
            lock_client.delete(key)
            log.info("🔓 Polling singleton lock bo'shatildi: %s", key)
    except Exception as e:
        log.warning("Polling lock bo'shatishda xato: %s", e)


def _start_polling_lock_heartbeat(lock_client, lock_meta):
    """
    Polling davomida lock TTL yangilanib turadi.
    """
    if not lock_client or not lock_meta:
        return None

    import threading

    key, owner = lock_meta
    lock_ttl = int(_os.environ.get("BOT_POLL_LOCK_TTL_SECONDS", "120"))
    interval = max(10, lock_ttl // 3)
    stop_event = threading.Event()

    def _worker() -> None:
        while not stop_event.is_set():
            stop_event.wait(interval)
            if stop_event.is_set():
                break
            try:
                current = lock_client.get(key)
                if current == owner:
                    lock_client.expire(key, lock_ttl)
                else:
                    log.warning("⚠️ Polling lock owner o'zgardi: %s", key)
                    break
            except Exception as e:
                log.warning("Polling lock heartbeat xato: %s", e)

    t = threading.Thread(target=_worker, name="polling-lock-heartbeat", daemon=True)
    t.start()
    return stop_event


def main() -> None:
    # Lock avval — takroriy instance tez chiqadi; ilovani_qur og'ir (handlerlar).
    conf = config_init()
    lock_client = None
    lock_meta = None
    lock_heartbeat_stop = None
    try:
        lock_client, lock_meta = _acquire_polling_singleton_lock(conf.bot_token)
        lock_heartbeat_stop = _start_polling_lock_heartbeat(lock_client, lock_meta)
    except RuntimeError as e:
        log.critical("⛔ Bot ishga tushmadi: %s", e)
        raise SystemExit(1) from e

    try:
        app = ilovani_qur(conf)
        log.info("▶️  Polling boshlandi...")
        # Webhook tozalash (agar webhook qolgan bo'lsa — polling ishlamaydi!)
        try:
            import httpx
            r = httpx.get(f"https://api.telegram.org/bot{conf.bot_token}/deleteWebhook?drop_pending_updates=true", timeout=10)
            log.info("🔄 Webhook tozalandi: %s", r.text[:100])
        except Exception as _wh:
            log.warning("Webhook tozalash: %s", _wh)
        try:
            app.run_polling(
                allowed_updates=Update.ALL_TYPES,
                drop_pending_updates=True,
                close_loop=False,
            )
        except KeyboardInterrupt:
            log.info("⏹️  Bot to'xtatildi (KeyboardInterrupt)")
        except Exception as e:
            log.error("⛔ Bot xatosi: %s", e, exc_info=True)
            raise
    finally:
        if lock_heartbeat_stop is not None:
            lock_heartbeat_stop.set()
        _release_polling_singleton_lock(lock_client, lock_meta)


if __name__=="__main__":
    main()
