"""Excel eksport — openpyxl"""
from __future__ import annotations
import io
import logging
from datetime import datetime
from typing import Any
import pytz
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

TZ     = pytz.timezone("Asia/Tashkent")
log    = logging.getLogger(__name__)
KOK    = "1a56db"
YASHIL = "059669"
SARIQ  = "d97706"
QIZIL  = "dc2626"
BINAF  = "7c3aed"
OCH    = "f7f9fc"
OQ     = "FFFFFF"


def _hozir() -> str:
    return datetime.now(TZ).strftime("%d.%m.%Y %H:%M")


def _pul(v: Any) -> str:
    try:
        return f"{float(v):,.0f} so'm"
    except Exception:
        return "0 so'm"


def _to_ldirish(rang: str) -> PatternFill:
    return PatternFill("solid", fgColor=rang)


def _shrift(qalin=False, rang="000000", o_lcham=10) -> Font:
    return Font(bold=qalin, color=rang, size=o_lcham, name="Calibri")


def _chegara() -> Border:
    t = Side(style="thin", color="d1d5db")
    return Border(left=t, right=t, top=t, bottom=t)


def _hizalash(gorizontal="left", vertikal="center",
               sindirishli=False) -> Alignment:
    return Alignment(
        horizontal=gorizontal, vertical=vertikal,
        wrap_text=sindirishli
    )


def _sarlavha(ws, qator: int, ustunlar: list[str],
               rang: str = KOK) -> None:
    for u, qiy in enumerate(ustunlar, 1):
        katak = ws.cell(row=qator, column=u, value=qiy)
        katak.font      = _shrift(qalin=True, rang=OQ)
        katak.fill      = _to_ldirish(rang)
        katak.border    = _chegara()
        katak.alignment = _hizalash("center")


def _ma_lumot(ws, qator: int, qiymatlar: list,
               soya: bool = False) -> None:
    to_ldirish = _to_ldirish(OCH) if soya else None
    for u, qiy in enumerate(qiymatlar, 1):
        katak           = ws.cell(row=qator, column=u, value=qiy)
        katak.font      = _shrift()
        katak.border    = _chegara()
        katak.alignment = _hizalash()
        if to_ldirish:
            katak.fill = to_ldirish


def _kenglikni_moslashtir(ws, min_k=8, max_k=40) -> None:
    for ustun in ws.columns:
        kenglik = min_k
        for katak in ustun:
            try:
                uzunlik = len(str(katak.value or ""))
                if uzunlik > kenglik:
                    kenglik = uzunlik
            except Exception as _exc:
                log.debug("%s: %s", "export_excel", _exc)  # was silent
        ws.column_dimensions[
            get_column_letter(ustun[0].column)
        ].width = min(kenglik + 3, max_k)


# ─── SOTUV EXCEL ──────────────────────────────────────────

def sotuv_excel(data: dict, dokon_nomi: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sotuv cheki"

    ws["A1"].value = "MASHRAB MOLIYA"
    ws["A1"].font  = _shrift(qalin=True, o_lcham=14, rang=KOK)
    ws["A2"].value = dokon_nomi
    ws["A3"].value = f"Sana: {_hozir()}"
    if data.get("klient"):
        ws["A4"].value = f"Klient: {data['klient']}"

    _sarlavha(ws, 6, ["#","Tovar nomi","Miqdor","Birlik","Narx (so'm)","Jami (so'm)"])
    for i, t in enumerate(data.get("tovarlar", []), 1):
        miq  = float(t.get("miqdor", 0))
        bir  = t.get("birlik", "dona")
        narx = float(t.get("narx",   0))
        jami = float(t.get("jami",   0))
        miq_s = f"{miq:.1f}".rstrip("0").rstrip(".")
        _ma_lumot(ws, 6 + i, [
            i,
            t.get("nomi", ""),
            f"{miq_s}g" if bir == "gramm" else miq_s,
            bir,
            f"{narx:,.0f}/kg" if (bir == "gramm" and narx) else (f"{narx:,.0f}" if narx else "—"),
            jami,
        ], soya=(i % 2 == 0))

    oxirgi = 6 + len(data.get("tovarlar", []))
    ws.cell(oxirgi + 2, 1).value = "JAMI:"
    ws.cell(oxirgi + 2, 1).font  = _shrift(qalin=True)
    ws.cell(oxirgi + 2, 6).value = float(data.get("jami_summa", 0))
    ws.cell(oxirgi + 2, 6).font  = _shrift(qalin=True)

    q = float(data.get("qarz", 0))
    if q > 0:
        tl = float(data.get("tolandan") or data.get("tolangan") or 0)
        ws.cell(oxirgi + 3, 1).value = "To'landi:"
        ws.cell(oxirgi + 3, 6).value = tl
        ws.cell(oxirgi + 4, 1).value = "QARZ:"
        ws.cell(oxirgi + 4, 1).font  = _shrift(qalin=True, rang=QIZIL)
        ws.cell(oxirgi + 4, 6).value = q
        ws.cell(oxirgi + 4, 6).font  = _shrift(qalin=True, rang=QIZIL)

    _kenglikni_moslashtir(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── KLIENT HISOBI EXCEL ──────────────────────────────────

def klient_hisobi_excel(data: dict, dokon_nomi: str) -> bytes:
    if not data:
        return b""
    k  = data["klient"]
    wb = Workbook()

    # 1-varaq: Umumiy
    ws1 = wb.active
    ws1.title = "Umumiy"
    ws1["A1"].value = "MASHRAB MOLIYA"
    ws1["A1"].font  = _shrift(qalin=True, o_lcham=14, rang=KOK)
    ws1["A2"].value = f"Klient hisobi: {k['ism']}"
    ws1["A2"].font  = _shrift(qalin=True, o_lcham=12)
    ws1["A3"].value = f"Hisobot: {_hozir()}"

    statistika = [
        ("Jami sotuv:",          float(data["jami_sotuv"])),
        ("To'langan:",           float(data["jami_tolangan"])),
        ("Qaytarilgan:",         float(data["jami_qaytarilgan"])),
        ("Qolgan qarz:",         float(data["faol_qarz"])),
        ("Sotuvlar soni:",       data["sotuv_soni"]),
    ]
    for i, (sarlavha, qiymat) in enumerate(statistika, 5):
        ws1.cell(i, 1).value = sarlavha
        ws1.cell(i, 1).font  = _shrift(qalin=True)
        katak = ws1.cell(i, 2, value=qiymat)
        if sarlavha == "Qolgan qarz:" and float(qiymat) > 0:
            katak.font = _shrift(qalin=True, rang=QIZIL)

    # 2-varaq: Sotuvlar
    ws2 = wb.create_sheet("Sotuvlar")
    _sarlavha(ws2, 1,
              ["Sana","Tovarlar","Jami","To'landi","Qarz"],
              YASHIL)
    for i, row in enumerate(data.get("sotuvlar", []), 2):
        tv_str = (row.get("tovarlar_str") or "")
        qarz_q = float(row["qarz"]) if float(row["qarz"]) > 0 else ""
        _ma_lumot(ws2, i, [
            str(row["sana"])[:16] if row.get("sana") else "",
            tv_str,
            float(row["jami"]),
            float(row["tolangan"]),
            qarz_q,
        ], soya=(i % 2 == 0))

    # 3-varaq: Qaytarishlar
    if data.get("qaytarishlar"):
        ws3 = wb.create_sheet("Qaytarishlar")
        _sarlavha(ws3, 1,
                  ["Sana","Tovar","Miqdor","Birlik","Summa","Sabab"],
                  SARIQ)
        for i, q in enumerate(data["qaytarishlar"], 2):
            _ma_lumot(ws3, i, [
                str(q["sana"])[:10],
                q.get("tovar_nomi", ""),
                float(q.get("miqdor") or 0),
                q.get("birlik", ""),
                float(q.get("jami") or 0),
                q.get("sabab") or "",
            ], soya=(i % 2 == 0))

    # 4-varaq: Qarzlar
    if data.get("qarzlar"):
        ws4 = wb.create_sheet("Qarzlar")
        _sarlavha(ws4, 1,
                  ["Sana","Dastlabki","To'landi","Qolgan","Holat"],
                  QIZIL)
        for i, q in enumerate(data["qarzlar"], 2):
            holat = "Yopildi" if q["yopildi"] else "Ochiq"
            _ma_lumot(ws4, i, [
                str(q["yaratilgan"])[:10],
                float(q["dastlabki_summa"]),
                float(q["tolangan"]),
                float(q["qolgan"]),
                holat,
            ], soya=(i % 2 == 0))

    for varaq in wb.worksheets:
        _kenglikni_moslashtir(varaq)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── KUNLIK HISOBOT EXCEL ─────────────────────────────────

def kunlik_excel(d: dict, dokon_nomi: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Kunlik hisobot"

    ws["A1"].value = "MASHRAB MOLIYA"
    ws["A1"].font  = _shrift(qalin=True, o_lcham=14, rang=KOK)
    ws["A2"].value = f"Kunlik hisobot: {d['kun']}"
    ws["A2"].font  = _shrift(qalin=True, o_lcham=12)
    ws["A3"].value = f"Yaratildi: {_hozir()}"

    _sarlavha(ws, 5, ["Ko'rsatkich", "Qiymat"], KOK)
    qatorlar = [
        ("Kirim soni",    d["kr_n"]),
        ("Kirim summasi", float(d["kr_jami"])),
        ("Sotuv soni",    d["ch_n"]),
        ("Sotuv summasi", float(d["ch_jami"])),
        ("To'landi",      float(d["tolangan"])),
        ("Yangi qarz",    float(d["yangi_qarz"])),
        ("SOF FOYDA",     float(d["foyda"])),
        ("Jami qarz",     float(d["jami_qarz"])),
    ]
    for i, (sarlavha, qiymat) in enumerate(qatorlar, 6):
        _ma_lumot(ws, i, [sarlavha, qiymat], soya=(i % 2 == 0))
        if sarlavha == "SOF FOYDA":
            ws.cell(i, 1).font = _shrift(qalin=True)
            ws.cell(i, 2).font = _shrift(qalin=True, rang=YASHIL)

    if d.get("by_kat"):
        bosh = len(qatorlar) + 8
        ws.cell(bosh, 1).value = "Kategoriyalar:"
        ws.cell(bosh, 1).font  = _shrift(qalin=True)
        _sarlavha(ws, bosh + 1, ["Kategoriya", "Jami", "Soni"], BINAF)
        for i, item in enumerate(d["by_kat"], bosh + 2):
            _ma_lumot(ws, i, [item["kategoriya"], float(item["j"]), item["n"]],
                      soya=(i % 2 == 0))

    _kenglikni_moslashtir(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def hisobot_excel(d: dict, dokon_nomi: str) -> bytes:
    """Universal hisobot Excel — kunlik, haftalik, oylik."""
    davr = d.get("davr", "kunlik")
    sarlavha_nom = {"kunlik": "Kunlik", "haftalik": "Haftalik", "oylik": "Oylik"}

    wb = Workbook()
    ws = wb.active
    ws.title = f"{sarlavha_nom.get(davr, '')} hisobot"

    ws["A1"].value = dokon_nomi or "MASHRAB MOLIYA"
    ws["A1"].font = _shrift(qalin=True, o_lcham=14, rang=KOK)
    ws["A2"].value = f"{sarlavha_nom.get(davr, '')} hisobot: {d.get('sana', '')}"
    ws["A2"].font = _shrift(qalin=True, o_lcham=12)
    ws["A3"].value = f"Yaratildi: {_hozir()}"

    _sarlavha(ws, 5, ["Ko'rsatkich", "Qiymat"], KOK)
    qatorlar = [
        ("Sotuv soni", d.get("sotuv_soni", 0)),
        ("Sotuv summasi", float(d.get("sotuv_jami", 0))),
        ("To'landi", float(d.get("tolangan", 0))),
        ("Yangi qarz", float(d.get("yangi_qarz", 0))),
        ("Kirim soni", d.get("kirim_soni", 0)),
        ("Kirim summasi", float(d.get("kirim_jami", 0))),
        ("Qaytarish", float(d.get("qaytarish_jami", 0))),
        ("SOF FOYDA", float(d.get("foyda", 0))),
        ("O'rtacha chek", float(d.get("ortacha_chek", 0))),
        ("Jami qarz", float(d.get("jami_qarz", 0))),
        ("Qarz nisbati %", d.get("qarz_nisbati", 0)),
    ]
    for i, (nomi, qiymat) in enumerate(qatorlar, 6):
        _ma_lumot(ws, i, [nomi, qiymat], soya=(i % 2 == 0))
        if nomi == "SOF FOYDA":
            ws.cell(i, 1).font = _shrift(qalin=True)
            rang = YASHIL if qiymat >= 0 else QIZIL
            ws.cell(i, 2).font = _shrift(qalin=True, rang=rang)

    if d.get("top5_tovar"):
        bosh = len(qatorlar) + 8
        ws.cell(bosh, 1).value = "ENG KO'P SOTILGAN:"
        ws.cell(bosh, 1).font = _shrift(qalin=True)
        _sarlavha(ws, bosh + 1, ["Tovar", "Miqdor", "Jami", "Foyda"], BINAF)
        for i, tv in enumerate(d["top5_tovar"], bosh + 2):
            _ma_lumot(ws, i, [
                tv["nomi"], tv["miqdor"],
                float(tv["jami"]), float(tv.get("foyda", 0))
            ], soya=(i % 2 == 0))

    if d.get("top5_klient"):
        bosh2 = len(qatorlar) + 8 + len(d.get("top5_tovar", [])) + 4
        ws.cell(bosh2, 1).value = "YIRIK KLIENTLAR:"
        ws.cell(bosh2, 1).font = _shrift(qalin=True)
        _sarlavha(ws, bosh2 + 1, ["Klient", "Jami sotuv", "Qarz"], BINAF)
        for i, kl in enumerate(d["top5_klient"], bosh2 + 2):
            _ma_lumot(ws, i, [
                kl["ism"], float(kl["jami_sotuv"]),
                float(kl.get("jami_qarz", 0))
            ], soya=(i % 2 == 0))

    _kenglikni_moslashtir(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
